import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import glob
import re
import sys
import csv
import openpyxl as op

from ctypes import alignment
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Font, numbers
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

from openpyxl.styles.alignment import Alignment

# パスで指定したファイルの一覧をリスト形式で取得
csv_files_ipan = glob.glob("/content/drive/MyDrive/data/一般管理/*.csv")
# CSVファイルの中身を追加していくリストを準備
data_list_ipan = []
# 読み込むファイルのリストをスキャンして取得
for file in csv_files_ipan:
    data_list_ipan.append(pd.read_csv(file))
# リストを全て列方向に結合
df_ipan_t = pd.concat(data_list_ipan, axis=1, sort=True)
# columnsパラメータで列名を設定
feature_ipan = [
    "役員1",
    "一般間接1",
    "一般間接2",
    "一般間接3",
    "一般間接6",
    "一般販売1",
    "一般販売2",
]
df_ipan_t.columns = feature_ipan
# 行名の設定
df_ipan_t = df_ipan_t.rename(
    {
        0: "在籍者",
        1: "在籍者主幹以下人数",
        2: "実在籍者",
        3: "有休時間",
        4: "有休時間在籍者平均",
        5: "欠勤時間",
        6: "勤務時間",
        7: "遅早時間",
        8: "出勤率",
        9: "実労働時間",
        10: "ズレ時間",
        11: "残業時間",
        12: "残業時間主幹以下平均",
        13: "法定外休出時間",
        14: "法定外休出主幹以下平均",
        15: "法定休出時間",
        16: "法定休出主幹以下平均",
        17: "時間外60時間超",
        18: "代休時間",
        19: "応援時間",
        20: "総労働時間",
        21: "基本給",
        22: "役職手当",
        23: "営業手当",
        24: "地域手当",
        25: "特別手当",
        26: "特別技技手当",
        27: "調整手当",
        28: "別居手当",
        29: "通勤手当",
        30: "小計1",
        31: "残業手当",
        32: "休出手当",
        33: "深夜勤務手当",
        34: "交替時差手当",
        35: "休業手当",
        36: "休業控除",
        37: "代休他",
        38: "欠勤・遅早控除",
        39: "精算分",
        40: "小計2",
        41: "総支給額",
        42: "応援時間額",
        43: "役員振替",
        44: "部門振替",
        45: "合計",
    }
)
# 整形---------------------
# ワークブックの生成
wb = Workbook()
# ワークシートの生成
ws = wb.active
ws.title = "一般管理"
# DataFrameを行単位のデータにする
rows = dataframe_to_rows(df_ipan_t, index=True, header=True)
# 1セルずつ処理を実行する
for row_no, row in enumerate(rows, 1):
    for col_no, value in enumerate(row, 1):
        # データを書き込む
        ws.cell(row=row_no, column=col_no, value=value)
# 不要な行の削除
ws.delete_rows(2)
# 表示倍率の設定
ws.sheet_view.zoomScale = 100
# 列幅の設定
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 12
ws.column_dimensions["J"].width = 12
ws.column_dimensions["K"].width = 12
ws.column_dimensions["L"].width = 12
# 表示形式
format = "#,##0"
for row in ws["B2:L2"]:
    for cell in row:
        cell.number_format = format

format = "0.00"
for row in ws["B4:L9"]:
    for cell in row:
        cell.number_format = format

format = "0.00%"
for row in ws["B10:L10"]:
    for cell in row:
        cell.number_format = format

format = "0.00"
for row in ws["B11:L12"]:
    for cell in row:
        cell.number_format = format

format = "0.00"
for row in ws["B13:L22"]:
    for cell in row:
        cell.number_format = format

format = "#,##0"
for row in ws["B23:L23"]:
    for cell in row:
        cell.number_format = format

format = "0.00"
for row in ws["B24:L26"]:
    for cell in row:
        cell.number_format = format

format = "#,##0"
for row in ws["B23:L48"]:
    for cell in row:
        cell.number_format = format
# ヘッダー行のスタイル設定
header = ws[1]
for header_cell in header:
    # フォントを設定する
    header_cell.fill = PatternFill(patternType="solid", fgColor="008000")
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
# 中央揃え
for row in ws["B1:H1"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous")
ws["L1"].alignment = Alignment(horizontal="centerContinuous")
# 指定した行の背景色を黄色にする
mylist = [21, 44, 45, 46]
# for list in mylist:
#     for row in ws.iter_rows():
#         for cell in row:
#             if cell.row == list:
#                 cell.fill = PatternFill(
#                     fgColor="FFFF00", bgColor="FFFF00", fill_type="solid"
#                 )
# 小計1
ws["B32"] = "=SUM(B23:B31)"
ws["C32"] = "=SUM(C23:C31)"
ws["D32"] = "=SUM(D23:D31)"
ws["E32"] = "=SUM(E23:E31)"
ws["F32"] = "=SUM(F23:F31)"
ws["G32"] = "=SUM(G23:G31)"
ws["H32"] = "=SUM(H23:H31)"

# 小計2
ws["B42"] = "=SUM(B33:B41)"
ws["C42"] = "=SUM(C33:C41)"
ws["D42"] = "=SUM(D33:D41)"
ws["E42"] = "=SUM(E33:E41)"
ws["F42"] = "=SUM(F33:F41)"
ws["G42"] = "=SUM(G33:G41)"
ws["H42"] = "=SUM(H33:H41)"

# 総支給額
ws["B43"] = "=B32 + B42"
ws["C43"] = "=C32 + C42"
ws["D43"] = "=D32 + D42"
ws["E43"] = "=E32 + E42"
ws["F43"] = "=F32 + F42"
ws["G43"] = "=G32 + G42"
ws["H43"] = "=H32 + H42"

# 列合計の埋め込み
ws["B47"] = "=B43+B44+B45+B46"
ws["C47"] = "=C43+C44+C45+C46"
ws["D47"] = "=D43+D44+D45+D46"
ws["E47"] = "=E43+E44+E45+E46"
ws["F47"] = "=F43+F44+F45+F46"
ws["G47"] = "=G43+G44+G45+G46"
ws["H47"] = "=H43+H44+H45+H46"

# 合計列の追加
ws["L1"] = "【合計】"

side1 = Side(border_style="thin", color="000000")
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)

for row in ws["A2:L47"]:
    for cell in row:
        cell.border = border_aro

# 行合計の埋め込み
ws["L2"] = "=SUM(B2:H2)"
ws["L3"] = "=SUM(B3:H3)"
ws["L4"] = "=SUM(B4:H4)"
ws["L5"] = "=SUM(B5:H5)"
ws["L6"] = "=SUM(B6:H6) / 7"
ws["L7"] = "=SUM(B7:H7)"
ws["L8"] = "=SUM(B8:H8)"
ws["L9"] = "=SUM(B9:H9)"
ws["L10"] = "=SUM(B10:H10) / 7"
ws["L11"] = "=SUM(B11:H11)"
ws["L12"] = "=SUM(B12:H12)"
ws["L13"] = "=SUM(B13:H13)"
ws["L14"] = "=SUM(B14:H14) / 7"
ws["L15"] = "=SUM(B15:H15)"
ws["L16"] = "=SUM(B16:H16) / 7"
ws["L17"] = "=SUM(B17:H17)"
ws["L18"] = "=SUM(B18:H18) / 7"
ws["L19"] = "=SUM(B19:H19)"
ws["L20"] = "=SUM(B20:H20)"
ws["L21"] = "=SUM(B21:H21)"
ws["L22"] = "=SUM(B22:H22)"
ws["L23"] = "=SUM(B23:H23)"
ws["L24"] = "=SUM(B24:H24)"
ws["L25"] = "=SUM(B25:H25)"
ws["L26"] = "=SUM(B26:H26)"
ws["L27"] = "=SUM(B27:H27)"
ws["L28"] = "=SUM(B28:H28)"
ws["L29"] = "=SUM(B29:H29)"
ws["L30"] = "=SUM(B30:H30)"
ws["L31"] = "=SUM(B31:H31)"
ws["L32"] = "=SUM(B32:H32)"
ws["L33"] = "=SUM(B33:H33)"
ws["L34"] = "=SUM(B34:H34)"
ws["L35"] = "=SUM(B35:H35)"
ws["L36"] = "=SUM(B36:H36)"
ws["L37"] = "=SUM(B37:H37)"
ws["L38"] = "=SUM(B38:H38)"
ws["L39"] = "=SUM(B39:H39)"
ws["L40"] = "=SUM(B40:H40)"
ws["L41"] = "=SUM(B41:H41)"
ws["L42"] = "=SUM(B42:H42)"
ws["L43"] = "=SUM(B43:H43)"
ws["L44"] = "=SUM(B44:H44)"
ws["L45"] = "=SUM(B45:H45)"
ws["L46"] = "=SUM(B46:H46)"
ws["L47"] = "=L43+L44+L45+L46"

ws.delete_rows(48, 49)

# Excelファイルを出力
wb.save("/content/drive/MyDrive/data/EXCEL/一般管理.xlsx")

# 鍛造
# パスで指定したファイルの一覧をリスト形式で取得
csv_files_tanzo = glob.glob("/content/drive/MyDrive/data/鍛造/*.csv")
# CSVファイルの中身を追加していくリストを表示
data_list_tanzo = []

# 読み込むファイルのリストをスキャン
for file in csv_files_tanzo:
    data_list_tanzo.append(pd.read_csv(file))

# リストを全て列方向に結合
df_tanzo_t = pd.concat(data_list_tanzo, axis=1, sort=True)
# print(df_tanzo_t)

# columnsパラメータで列名を設定
feature_tanzo = ["間接1", "間接2", "間接3", "間接4", "間接5", "間接6", "直接1", "直接2"]
df_tanzo_t.columns = feature_tanzo
# 行名の設定
df_tanzo_t = df_tanzo_t.rename(
    {
        0: "在籍者",
        1: "在籍者主幹以下人数",
        2: "実在籍者",
        3: "有休時間",
        4: "有休時間在籍者平均",
        5: "欠勤時間",
        6: "勤務時間",
        7: "遅早時間",
        8: "出勤率",
        9: "実労働時間",
        10: "ズレ時間",
        11: "残業時間",
        12: "残業時間主幹以下平均",
        13: "法定外休出時間",
        14: "法定外休出主幹以下平均",
        15: "法定休出時間",
        16: "法定休出主幹以下平均",
        17: "時間外60時間超",
        18: "代休時間",
        19: "応援時間",
        20: "総労働時間",
        21: "基本給",
        22: "役職手当",
        23: "営業手当",
        24: "地域手当",
        25: "特別手当",
        26: "特別技技手当",
        27: "調整手当",
        28: "別居手当",
        29: "通勤手当",
        30: "小計1",
        31: "残業手当",
        32: "休出手当",
        33: "深夜勤務手当",
        34: "交替時差手当",
        35: "休業手当",
        36: "休業控除",
        37: "代休他",
        38: "欠勤・遅早控除",
        39: "精算分",
        40: "小計2",
        41: "総支給額",
        42: "応援時間額",
        43: "役員振替",
        44: "部門振替",
        45: "合計",
    }
)
# 整形---------------------
# ワークブックの生成
wb = Workbook()
# ワークシートの生成
ws = wb.active
ws.title = "鍛造工場"
# DataFrameを行単位のデータにする
rows = dataframe_to_rows(df_tanzo_t, index=True, header=True)
# 1セルずつ処理を実行する
for row_no, row in enumerate(rows, 1):
    for col_no, value in enumerate(row, 1):
        # データを書き込む
        ws.cell(row=row_no, column=col_no, value=value)
# 不要な行の削除
ws.delete_rows(2)
# 列の追加
ws.insert_cols(8)
# 表示倍率の設定
ws.sheet_view.zoomScale = 100
# 列幅の設定
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 12
ws.column_dimensions["J"].width = 12
ws.column_dimensions["K"].width = 12
ws.column_dimensions["L"].width = 12
# 表示形式
format = "#,##0"
for row in ws["B2:L2"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B4:L9"]:
    for cell in row:
        cell.number_format = format
format = "0.00%"
for row in ws["B10:L10"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B11:L12"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B13:L22"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B23:L23"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B24:L26"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B27:L51"]:
    for cell in row:
        cell.number_format = format
# ヘッダー行のスタイル設定
header = ws[1]
for header_cell in header:
    # フォントを設定する
    header_cell.fill = PatternFill(patternType="solid", fgColor="008000")
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
# 中央揃え
for row in ws["B1:L1"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous")
# 指定した行の背景色を黄色にする
mylist = [21, 44, 45, 46]
# for list in mylist:
#     for row in ws.iter_rows():
#         for cell in row:
#             if cell.row == list:
#                 cell.fill = PatternFill(
#                     fgColor="FFFF00", bgColor="FFFF00", fill_type="solid"
#                 )
# 小計1
ws["B32"] = "=SUM(B23:B31)"
ws["C32"] = "=SUM(C23:C31)"
ws["D32"] = "=SUM(D23:D31)"
ws["E32"] = "=SUM(E23:E31)"
ws["F32"] = "=SUM(F23:F31)"
ws["G32"] = "=SUM(G23:G31)"
ws["H32"] = "=SUM(H23:H31)"
ws["I32"] = "=SUM(I23:I31)"
ws["J32"] = "=SUM(J23:J31)"
ws["K32"] = "=SUM(K23:K31)"
# 小計2
ws["B42"] = "=SUM(B33:B41)"
ws["C42"] = "=SUM(C33:C41)"
ws["D42"] = "=SUM(D33:D41)"
ws["E42"] = "=SUM(E33:E41)"
ws["F42"] = "=SUM(F33:F41)"
ws["G42"] = "=SUM(G33:G41)"
ws["H42"] = "=SUM(H33:H41)"
ws["I42"] = "=SUM(I33:I41)"
ws["J42"] = "=SUM(J33:J41)"
ws["K42"] = "=SUM(K33:K41)"
# 総支給額
ws["B43"] = "=B32 + B42"
ws["C43"] = "=C32 + C42"
ws["D43"] = "=D32 + D42"
ws["E43"] = "=E32 + E42"
ws["F43"] = "=F32 + F42"
ws["G43"] = "=G32 + G42"
ws["H43"] = "=H32 + H42"
ws["I43"] = "=I32 + I42"
ws["J43"] = "=J32 + J42"
ws["K43"] = "=K32 + K42"
# 列合計の埋め込み
ws["B47"] = "=B43+B44+B45+B46"
ws["C47"] = "=C43+C44+C45+C46"
ws["D47"] = "=D43+D44+D45+D46"
ws["E47"] = "=E43+E44+E45+E46"
ws["F47"] = "=F43+F44+F45+F46"
ws["G47"] = "=G43+G44+G45+G46"
ws["H47"] = "=H43+H44+H45+H46"
ws["I47"] = "=I43+I44+I45+I46"
ws["J47"] = "=J43+J44+J45+J46"
ws["K47"] = "=K43+K44+K45+L46"
# 間接計列の追加
ws["H1"] = "【間接計】"
# 間接計列への計算式の埋め込み
ws["H2"] = "=SUM(B2:G2)"
ws["H3"] = "=SUM(B3:G3)"
ws["H4"] = "=SUM(B4:G4)"
ws["H5"] = "=SUM(B5:G5)"
ws["H6"] = "=SUM(B6:G6) / 6"
ws["H7"] = "=SUM(B7:G7)"
ws["H8"] = "=SUM(B8:G8)"
ws["H9"] = "=SUM(B9:G9)"
ws["H10"] = "=SUM(B10:G10) / 6"
ws["H11"] = "=SUM(B11:G11)"
ws["H12"] = "=SUM(B12:G12)"
ws["H13"] = "=SUM(B13:G13)"
ws["H14"] = "=SUM(B14:G14) / 6"
ws["H15"] = "=SUM(B15:G15)"
ws["H16"] = "=SUM(B16:G16) / 6"
ws["H17"] = "=SUM(B17:G17)"
ws["H18"] = "=SUM(B18:G18) / 6"
ws["H19"] = "=SUM(B19:G19)"
ws["H20"] = "=SUM(B20:G20)"
ws["H21"] = "=SUM(B21:G21)"
ws["H22"] = "=SUM(B22:G22)"
ws["H23"] = "=SUM(B23:G23)"
ws["H24"] = "=SUM(B24:G24)"
ws["H25"] = "=SUM(B25:G25)"
ws["H26"] = "=SUM(B26:G26)"
ws["H27"] = "=SUM(B27:G27)"
ws["H28"] = "=SUM(B28:G28)"
ws["H29"] = "=SUM(B29:G29)"
ws["H30"] = "=SUM(B30:G30)"
ws["H31"] = "=SUM(B31:G31)"
ws["H32"] = "=SUM(B32:G32)"
ws["H33"] = "=SUM(B33:G33)"
ws["H34"] = "=SUM(B34:G34)"
ws["H35"] = "=SUM(B35:G35)"
ws["H36"] = "=SUM(B36:G36)"
ws["H37"] = "=SUM(B37:G37)"
ws["H38"] = "=SUM(B38:G38)"
ws["H39"] = "=SUM(B39:G39)"
ws["H40"] = "=SUM(B40:G40)"
ws["H41"] = "=SUM(B41:G41)"
ws["H42"] = "=SUM(B42:G42)"
ws["H43"] = "=SUM(B43:G43)"
ws["H44"] = "=SUM(B44:G44)"
ws["H45"] = "=SUM(B45:G45)"
ws["H46"] = "=SUM(B46:G46)"
ws["H47"] = "=H43+H44+H45+H46"
# 直接計列の追加
ws["K1"] = "【直接計】"
# 直接計列への計算式の埋め込み
ws["K2"] = "=SUM(I2:J2)"
ws["K3"] = "=SUM(I3:J3)"
ws["K4"] = "=SUM(I4:J4)"
ws["K5"] = "=SUM(I5:J5)"
ws["K6"] = "=SUM(I6:J6) / 2"
ws["K7"] = "=SUM(I7:J7)"
ws["K8"] = "=SUM(I8:J8)"
ws["K9"] = "=SUM(I9:J9)"
ws["K10"] = "=SUM(I10:J10) / 2"
ws["K11"] = "=SUM(I11:J11)"
ws["K12"] = "=SUM(I12:J12)"
ws["K13"] = "=SUM(I13:J13)"
ws["K14"] = "=SUM(I14:J14) / 2"
ws["K15"] = "=SUM(I15:J15)"
ws["K16"] = "=SUM(I16:J16) / 2"
ws["K17"] = "=SUM(I17:J17)"
ws["K18"] = "=SUM(I18:J18) / 2"
ws["K19"] = "=SUM(I19:J19)"
ws["K20"] = "=SUM(I20:J20)"
ws["K21"] = "=SUM(I21:J21)"
ws["K22"] = "=SUM(I22:J22)"
ws["K23"] = "=SUM(I23:J23)"
ws["K24"] = "=SUM(I24:J24)"
ws["K25"] = "=SUM(I25:J25)"
ws["K26"] = "=SUM(I26:J26)"
ws["K27"] = "=SUM(I27:J27)"
ws["K28"] = "=SUM(I28:J28)"
ws["K29"] = "=SUM(I29:J29)"
ws["K30"] = "=SUM(I30:J30)"
ws["K31"] = "=SUM(I31:J31)"
ws["K32"] = "=SUM(I32:J32)"
ws["K33"] = "=SUM(I33:J33)"
ws["K34"] = "=SUM(I34:J34)"
ws["K35"] = "=SUM(I35:J35)"
ws["K36"] = "=SUM(I36:J36)"
ws["K37"] = "=SUM(I37:J37)"
ws["K38"] = "=SUM(I38:J38)"
ws["K39"] = "=SUM(I39:J39)"
ws["K40"] = "=SUM(I40:J40)"
ws["K41"] = "=SUM(I41:J41)"
ws["K42"] = "=SUM(I42:J42)"
ws["K43"] = "=SUM(I43:J43)"
ws["K44"] = "=SUM(I44:J44)"
ws["K45"] = "=SUM(I45:J45)"
ws["K46"] = "=SUM(I46:J46)"
ws["K47"] = "=K43+K44+K45+K46"
# 直接列ヘッダーの書式
fill = PatternFill(patternType="solid", fgColor="008000")
ws["K1"].fill = fill
ws["K1"].font = Font(bold=True, color="FFFFFF")
# 合計列の追加とヘッダーの書式
ws["L1"] = "【合計】"
ws["L1"].fill = fill
ws["L1"].font = Font(bold=True, color="FFFFFF")
# 罫線
side1 = Side(border_style="thin", color="000000")
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
for row in ws["A2:L49"]:
    for cell in row:
        cell.border = border_aro
# 行合計の埋め込み
ws["L2"] = "=SUM(H2,K2)"
ws["L3"] = "=SUM(H3,K3)"
ws["L4"] = "=SUM(H4,K4)"
ws["L5"] = "=SUM(H5,K5)"
ws["L6"] = "=SUM(H6,K6) / 2"
ws["L7"] = "=SUM(H7,K7)"
ws["L8"] = "=SUM(H8,K8)"
ws["L9"] = "=SUM(H9,K9)"
ws["L10"] = "=(H10+K10) / 2"
ws["L11"] = "=SUM(H11,K11)"
ws["L12"] = "=SUM(K12,H12)"
ws["L13"] = "=SUM(H13,K13)"
ws["L14"] = "=SUM(H14,K14) / 2"
ws["L15"] = "=SUM(H15,K15)"
ws["L16"] = "=SUM(H16,K16) / 2"
ws["L17"] = "=SUM(H17,K17)"
ws["L18"] = "=SUM(H18,K18) / 2"
ws["L19"] = "=SUM(H19,K19)"
ws["L20"] = "=SUM(H20,K20)"
ws["L21"] = "=SUM(H21,K21)"
ws["L22"] = "=SUM(H22,K22)"
ws["L23"] = "=SUM(H23,K23)"
ws["L24"] = "=SUM(H24,K24)"
ws["L25"] = "=SUM(H25,K25)"
ws["L26"] = "=SUM(H26,K26)"
ws["L27"] = "=SUM(H27,K27)"
ws["L28"] = "=SUM(H28,K28)"
ws["L29"] = "=SUM(H29,K29)"
ws["L30"] = "=SUM(H30,K30)"
ws["L31"] = "=SUM(H31,K31)"
ws["L32"] = "=SUM(H32,K32)"
ws["L33"] = "=SUM(H33,K33)"
ws["L34"] = "=SUM(H34,K34)"
ws["L35"] = "=SUM(H35,K35)"
ws["L36"] = "=SUM(H36,K36)"
ws["L37"] = "=SUM(H37,K37)"
ws["L38"] = "=SUM(H38,K38)"
ws["L39"] = "=SUM(H39,K39)"
ws["L40"] = "=SUM(H40,K40)"
ws["L41"] = "=SUM(H41,K41)"
ws["L42"] = "=SUM(H42,K42)"
ws["L43"] = "=SUM(H43,K43)"
ws["L44"] = "=SUM(H44,K44)"
ws["L45"] = "=SUM(H45,K45)"
ws["L46"] = "=SUM(H46,K46)"
ws["L47"] = "=L43+L44+L45+L46"
ws.delete_rows(48, 49)
# Excelファイルを出力
wb.save("/content/drive/MyDrive/data/EXCEL/鍛造.xlsx")
# 切削
# パスで指定したファイルの一覧をリスト形式で取得
csv_files_sesaku = glob.glob("/content/drive/MyDrive/data/切削/*.csv")
# CSVファイルの中身を追加していくリストを表示
data_list_sesaku = []
# 読み込むファイルのリストをスキャン
for file in csv_files_sesaku:
    data_list_sesaku.append(pd.read_csv(file))
# リストを全て列方向に結合
df_sesaku_t = pd.concat(data_list_sesaku, axis=1, sort=True)
# columnsパラメータで列名を設定
feature_sesaku = [
    "間接1",
    "間接2",
    "間接4",
    "間接5",
    "間接6",
    "直接1",
    "直接2",
    "直接4",
]
df_sesaku_t.columns = feature_sesaku
# 行名の設定
df_sesaku_t = df_sesaku_t.rename(
    {
        0: "在籍者",
        1: "在籍者主幹以下人数",
        2: "実在籍者",
        3: "有休時間",
        4: "有休時間在籍者平均",
        5: "欠勤時間",
        6: "勤務時間",
        7: "遅早時間",
        8: "出勤率",
        9: "実労働時間",
        10: "ズレ時間",
        11: "残業時間",
        12: "残業時間主幹以下平均",
        13: "法定外休出時間",
        14: "法定外休出主幹以下平均",
        15: "法定休出時間",
        16: "法定休出主幹以下平均",
        17: "時間外60時間超",
        18: "代休時間",
        19: "応援時間",
        20: "総労働時間",
        21: "基本給",
        22: "役職手当",
        23: "営業手当",
        24: "地域手当",
        25: "特別手当",
        26: "特別技技手当",
        27: "調整手当",
        28: "別居手当",
        29: "通勤手当",
        30: "小計1",
        31: "残業手当",
        32: "休出手当",
        33: "深夜勤務手当",
        34: "交替時差手当",
        35: "休業手当",
        36: "休業控除",
        37: "代休他",
        38: "欠勤・遅早控除",
        39: "精算分",
        40: "小計2",
        41: "総支給額",
        42: "応援時間額",
        43: "役員振替",
        44: "部門振替",
        45: "合計",
    }
)
# 整形---------------------
# ワークブックの生成
wb = Workbook()
# ワークシートの生成
ws = wb.active
ws.title = "切削工場"
# DataFrameを行単位のデータにする
rows = dataframe_to_rows(df_sesaku_t, index=True, header=True)
# 1セルずつ処理を実行する
for row_no, row in enumerate(rows, 1):
    for col_no, value in enumerate(row, 1):
        # データを書き込む
        ws.cell(row=row_no, column=col_no, value=value)
# 不要な行の削除
ws.delete_rows(2)
# 列の追加
ws.insert_cols(7)
# 表示倍率の設定
ws.sheet_view.zoomScale = 100
# 列幅の設定
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 12
ws.column_dimensions["J"].width = 12
ws.column_dimensions["K"].width = 12
ws.column_dimensions["L"].width = 12
# 表示形式
format = "#,##0"
for row in ws["B2:L2"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B4:L9"]:
    for cell in row:
        cell.number_format = format
format = "0.00%"
for row in ws["B10:L10"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B11:L12"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B13:L22"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B23:L23"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B24:L26"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B25:L51"]:
    for cell in row:
        cell.number_format = format
# ヘッダー行のスタイル設定
header = ws[1]
for header_cell in header:
    # フォントを設定する
    header_cell.fill = PatternFill(patternType="solid", fgColor="008000")
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
# 中央揃え
for row in ws["B1:L1"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous")
# 指定した行の背景色を黄色にする
mylist = [21, 44, 45, 46]
# for list in mylist:
#     for row in ws.iter_rows():
#         for cell in row:
#             if cell.row == list:
#                 cell.fill = PatternFill(
#                     fgColor="FFFF00", bgColor="FFFF00", fill_type="solid"
#                )
# 小計1
ws["B32"] = "=SUM(B23:B31)"
ws["C32"] = "=SUM(C23:C31)"
ws["D32"] = "=SUM(D23:D31)"
ws["E32"] = "=SUM(E23:E31)"
ws["F32"] = "=SUM(F23:F31)"
ws["G32"] = "=SUM(G23:G31)"
ws["H32"] = "=SUM(H23:H31)"
ws["I32"] = "=SUM(I23:I31)"
ws["J32"] = "=SUM(J23:J31)"
ws["K32"] = "=SUM(K23:K31)"
# 小計2
ws["B42"] = "=SUM(B33:B41)"
ws["C42"] = "=SUM(C33:C41)"
ws["D42"] = "=SUM(D33:D41)"
ws["E42"] = "=SUM(E33:E41)"
ws["F42"] = "=SUM(F33:F41)"
ws["G42"] = "=SUM(G33:G41)"
ws["H42"] = "=SUM(H33:H41)"
ws["I42"] = "=SUM(I33:I41)"
ws["J42"] = "=SUM(J33:J41)"
ws["K42"] = "=SUM(K33:K41)"
# 総支給額
ws["B43"] = "=B32 + B42"
ws["C43"] = "=C32 + C42"
ws["D43"] = "=D32 + D42"
ws["E43"] = "=E32 + E42"
ws["F43"] = "=F32 + F42"
ws["G43"] = "=G32 + G42"
ws["H43"] = "=H32 + H42"
ws["I43"] = "=I32 + I42"
ws["J43"] = "=J32 + J42"
ws["K43"] = "=K32 + K42"
# 列合計の埋め込み
ws["B47"] = "=B43+B44+B45+B46"
ws["C47"] = "=C43+C44+C45+C46"
ws["D47"] = "=D43+D44+D45+D46"
ws["E47"] = "=E43+E44+E45+E46"
ws["F47"] = "=F43+F44+F45+F46"
ws["G47"] = "=G43+G44+G45+G46"
ws["H47"] = "=H43+H44+H45+H46"
ws["I47"] = "=I43+I44+I45+I46"
ws["J47"] = "=J43+J44+J45+J46"
ws["K47"] = "=K43+K44+K45+K46"
# 間接計列の追加
ws["G1"] = "【間接計】"
# 間接計列への計算式の埋め込み
ws["G2"] = "=SUM(B2:F2)"
ws["G3"] = "=SUM(B3:F3)"
ws["G4"] = "=SUM(B4:F4)"
ws["G5"] = "=SUM(B5:F5)"
ws["G6"] = "=SUM(B6:F6) / 5"
ws["G7"] = "=SUM(B7:F7)"
ws["G8"] = "=SUM(B8:F8)"
ws["G9"] = "=SUM(B9:F9)"
ws["G10"] = "=SUM(B10:F10) / 5"
ws["G11"] = "=SUM(B11:F11)"
ws["G12"] = "=SUM(B12:F12)"
ws["G13"] = "=SUM(B13:F13)"
ws["G14"] = "=SUM(B14:F14) / 5"
ws["G15"] = "=SUM(B15:F15)"
ws["G16"] = "=SUM(B16:F16) / 5"
ws["G17"] = "=SUM(B17:F17)"
ws["G18"] = "=SUM(B18:F18) / 5"
ws["G19"] = "=SUM(B19:F19)"
ws["G20"] = "=SUM(B20:F20)"
ws["G21"] = "=SUM(B21:F21)"
ws["G22"] = "=SUM(B22:F22)"
ws["G23"] = "=SUM(B23:F23)"
ws["G24"] = "=SUM(B24:F24)"
ws["G25"] = "=SUM(B25:F25)"
ws["G26"] = "=SUM(B26:F26)"
ws["G27"] = "=SUM(B27:F27)"
ws["G28"] = "=SUM(B28:F28)"
ws["G29"] = "=SUM(B29:F29)"
ws["G30"] = "=SUM(B30:F30)"
ws["G31"] = "=SUM(B31:F31)"
ws["G32"] = "=SUM(B32:F32)"
ws["G33"] = "=SUM(B33:F33)"
ws["G34"] = "=SUM(B34:F34)"
ws["G35"] = "=SUM(B35:F35)"
ws["G36"] = "=SUM(B36:F36)"
ws["G37"] = "=SUM(B37:F37)"
ws["G38"] = "=SUM(B38:F38)"
ws["G39"] = "=SUM(B39:F39)"
ws["G40"] = "=SUM(B40:F40)"
ws["G41"] = "=SUM(B41:F41)"
ws["G42"] = "=SUM(B42:F42)"
ws["G43"] = "=SUM(B43:F43)"
ws["G44"] = "=SUM(B44:F44)"
ws["G45"] = "=SUM(B45:F45)"
ws["G46"] = "=SUM(B46:F46)"
ws["G47"] = "=G43+G44+G45+G46"
# ヘッダーの書式
fill = PatternFill(patternType="solid", fgColor="008000")
ws["K1"].fill = fill
ws["K1"].font = Font(bold=True, color="FFFFFF")
# 合計列の追加とヘッダーの書式
ws["K1"] = "【合計】"
ws["K1"].fill = fill
ws["K1"].font = Font(bold=True, color="FFFFFF")
# 直接計列の追加
ws["K1"] = "【直接計】"
# 直接計列への計算式の埋め込み
ws["K2"] = "=SUM(H2:J2)"
ws["K3"] = "=SUM(H3:J3)"
ws["K4"] = "=SUM(H4:J4)"
ws["K5"] = "=SUM(H5:J5)"
ws["K6"] = "=SUM(H6:J6) / 3"
ws["K7"] = "=SUM(H7:J7)"
ws["K8"] = "=SUM(H8:J8)"
ws["K9"] = "=SUM(H9:J9)"
ws["K10"] = "=SUM(H10:J10) / 3"
ws["K11"] = "=SUM(H11:J11)"
ws["K12"] = "=SUM(H12:J12)"
ws["K13"] = "=SUM(H13:J13)"
ws["K14"] = "=SUM(H14:J14) / 3"
ws["K15"] = "=SUM(H15:J15)"
ws["K16"] = "=SUM(H16:J16) / 3"
ws["K17"] = "=SUM(H17:J17)"
ws["K18"] = "=SUM(H18:J18) / 3"
ws["K19"] = "=SUM(H19:J19)"
ws["K20"] = "=SUM(H20:J20)"
ws["K21"] = "=SUM(H21:J21)"
ws["K22"] = "=SUM(H22:J22)"
ws["K23"] = "=SUM(H23:J23)"
ws["K24"] = "=SUM(H24:J24)"
ws["K25"] = "=SUM(H25:J25)"
ws["K26"] = "=SUM(H26:J26)"
ws["K27"] = "=SUM(H27:J27)"
ws["K28"] = "=SUM(H28:J28)"
ws["K29"] = "=SUM(H29:J29)"
ws["K30"] = "=SUM(H30:J30)"
ws["K31"] = "=SUM(H31:J31)"
ws["K32"] = "=SUM(H32:J32)"
ws["K33"] = "=SUM(H33:J33)"
ws["K34"] = "=SUM(H34:J34)"
ws["K35"] = "=SUM(H35:J35)"
ws["K36"] = "=SUM(H36:J36)"
ws["K37"] = "=SUM(H37:J37)"
ws["K38"] = "=SUM(H38:J38)"
ws["K39"] = "=SUM(H39:J39)"
ws["K40"] = "=SUM(H40:J40)"
ws["K41"] = "=SUM(H41:J41)"
ws["K42"] = "=SUM(H42:J42)"
ws["K43"] = "=SUM(H43:J43)"
ws["K44"] = "=SUM(H44:J44)"
ws["K45"] = "=SUM(H45:J45)"
ws["K46"] = "=SUM(H46:J46)"
ws["K47"] = "=K43+K44+K45+K46"
# ヘッダーの書式
fill = PatternFill(patternType="solid", fgColor="008000")
ws["L1"].fill = fill
ws["L1"].font = Font(bold=True, color="FFFFFF")
# 合計列の追加とヘッダーの書式
ws["L1"] = "【合計】"
ws["L1"].fill = fill
ws["L1"].font = Font(bold=True, color="FFFFFF")
# 罫線
side1 = Side(border_style="thin", color="000000")
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
for row in ws["A2:L49"]:
    for cell in row:
        cell.border = border_aro
# 行合計の計算式埋め込み
ws["L2"] = "=G2+K2"
ws["L3"] = "=G3+K3"
ws["L4"] = "=G4+K4"
ws["L5"] = "=G5+K5"
ws["L6"] = "=(G6+K6) / 2"
ws["L7"] = "=G7+K7"
ws["L8"] = "=G8+K8"
ws["L9"] = "=G9+K9"
ws["L10"] = "=(G10+K10) / 2"
ws["L11"] = "=G11+K11"
ws["L12"] = "=G12+K12"
ws["L13"] = "=G13+K13"
ws["L14"] = "=(G14+K14) / 2"
ws["L15"] = "=G15+K15"
ws["L16"] = "=(G16+K16) / 2"
ws["L17"] = "=G17+K17"
ws["L18"] = "=(G18+K18) / 2"
ws["L19"] = "=G19+K19"
ws["L20"] = "=G20+K20"
ws["L21"] = "=G21+K21"
ws["L22"] = "=G22+K22"
ws["L23"] = "=G23+K23"
ws["L24"] = "=G24+K24"
ws["L25"] = "=G25+K25"
ws["L26"] = "=G26+K26"
ws["L27"] = "=G27+K27"
ws["L28"] = "=G28+K28"
ws["L29"] = "=G29+K29"
ws["L30"] = "=G30+K30"
ws["L31"] = "=G31+K31"
ws["L32"] = "=G32+K32"
ws["L33"] = "=G33+K33"
ws["L34"] = "=G34+K34"
ws["L35"] = "=G35+K35"
ws["L36"] = "=G36+K36"
ws["L37"] = "=G37+K37"
ws["L38"] = "=G38+K38"
ws["L39"] = "=G39+K39"
ws["L40"] = "=G40+K40"
ws["L41"] = "=G41+K41"
ws["L42"] = "=G42+K42"
ws["L43"] = "=G43+K43"
ws["L44"] = "=G44+K44"
ws["L45"] = "=G45+K45"
ws["L46"] = "=G46+K46"
ws["L47"] = "=L43+L44+L45+L46"
ws.delete_rows(48, 49)
# Excelファイルを出力
wb.save("/content/drive/MyDrive/data/EXCEL/切削.xlsx")
# AC
# パスで指定したファイルの一覧をリスト形式で取得
csv_files_ac = glob.glob("/content/drive/MyDrive/data/AC/*.csv")
# CSVファイルの中身を追加していくリストを表示
data_list_ac = []
# 読み込むファイルのリストをスキャン
for file in csv_files_ac:
    data_list_ac.append(pd.read_csv(file))
# リストを全て列方向に結合
df_ac_t = pd.concat(data_list_ac, axis=1, sort=True)
# columnsパラメータで列名を設定
feature_ac = ["間接1", "間接2", "間接4", "間接5", "直接1", "直接4"]
df_ac_t.columns = feature_ac
# 行名の設定
df_ac_t = df_ac_t.rename(
    {
        0: "在籍者",
        1: "在籍者主幹以下人数",
        2: "実在籍者",
        3: "有休時間",
        4: "有休時間在籍者平均",
        5: "欠勤時間",
        6: "勤務時間",
        7: "遅早時間",
        8: "出勤率",
        9: "実労働時間",
        10: "ズレ時間",
        11: "残業時間",
        12: "残業時間主幹以下平均",
        13: "法定外休出時間",
        14: "法定外休出主幹以下平均",
        15: "法定休出時間",
        16: "法定休出主幹以下平均",
        17: "時間外60時間超",
        18: "代休時間",
        19: "応援時間",
        20: "総労働時間",
        21: "基本給",
        22: "役職手当",
        23: "営業手当",
        24: "地域手当",
        25: "特別手当",
        26: "特別技技手当",
        27: "調整手当",
        28: "別居手当",
        29: "通勤手当",
        30: "小計1",
        31: "残業手当",
        32: "休出手当",
        33: "深夜勤務手当",
        34: "交替時差手当",
        35: "休業手当",
        36: "休業控除",
        37: "代休他",
        38: "欠勤・遅早控除",
        39: "精算分",
        40: "小計2",
        41: "総支給額",
        42: "応援時間額",
        43: "役員振替",
        44: "部門振替",
        45: "合計",
    }
)
# 整形---------------------
# ワークブックの生成
wb = Workbook()
# ワークシートの生成
ws = wb.active
ws.title = "AC工場"
# DataFrameを行単位のデータにする
rows = dataframe_to_rows(df_ac_t, index=True, header=True)
# 1セルずつ処理を実行する
for row_no, row in enumerate(rows, 1):
    for col_no, value in enumerate(row, 1):
        # データを書き込む
        ws.cell(row=row_no, column=col_no, value=value)
# 不要な行の削除
ws.delete_rows(2)
# 列の追加
ws.insert_cols(6)
# 表示倍率の設定
ws.sheet_view.zoomScale = 100
# 列幅の設定
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 12
ws.column_dimensions["J"].width = 12
ws.column_dimensions["K"].width = 12
ws.column_dimensions["L"].width = 12
# 表示形式
format = "#,##0"
for row in ws["B2:L2"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B4:L9"]:
    for cell in row:
        cell.number_format = format
format = "0.00%"
for row in ws["B10:L10"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B11:L12"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B13:L22"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B23:L23"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B24:L26"]:
    for cell in row:
        cell.number_format = format

format = "#,##0"
for row in ws["B23:L48"]:
    for cell in row:
        cell.number_format = format
# ヘッダー行のスタイル設定
header = ws[1]
for header_cell in header:
    # フォントを設定する
    header_cell.fill = PatternFill(patternType="solid", fgColor="008000")
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
# 中央揃え
for row in ws["B1:F1"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous")
ws["H1"].alignment = Alignment(horizontal="centerContinuous")
ws["G1"].alignment = Alignment(horizontal="centerContinuous")
ws["L1"].alignment = Alignment(horizontal="centerContinuous")
# 指定した行の背景色を黄色にする
mylist = [21, 44, 45, 46]

# for list in mylist:
#     for row in ws.iter_rows():
#         for cell in row:
#             if cell.row == list:
#                 cell.fill = PatternFill(
#                     fgColor="FFFF00", bgColor="FFFF00", fill_type="solid"
#                 )
# 小計1
ws["B32"] = "=SUM(B23:B31)"
ws["C32"] = "=SUM(C23:C31)"
ws["D32"] = "=SUM(D23:D31)"
ws["E32"] = "=SUM(E23:E31)"
ws["F32"] = "=SUM(F23:F31)"
ws["G32"] = "=SUM(G23:G31)"
ws["H32"] = "=SUM(H23:H31)"
# 小計2
ws["B42"] = "=SUM(B33:B41)"
ws["C42"] = "=SUM(C33:C41)"
ws["D42"] = "=SUM(D33:D41)"
ws["E42"] = "=SUM(E33:E41)"
ws["F42"] = "=SUM(F33:F41)"
ws["G42"] = "=SUM(G33:G41)"
ws["H42"] = "=SUM(H33:H41)"
# 総支給額
ws["B43"] = "=B32 + B42"
ws["C43"] = "=C32 + C42"
ws["D43"] = "=D32 + D42"
ws["E43"] = "=E32 + E42"
ws["F43"] = "=F32 + F42"
ws["G43"] = "=G32 + G42"
ws["H43"] = "=H32 + H42"
# 列合計の埋め込み
ws["B47"] = "=B43+B44+B45+B46"
ws["C47"] = "=C43+C44+C45+C46"
ws["D47"] = "=D43+D44+D45+D46"
ws["E47"] = "=E43+E44+E45+E46"
ws["F47"] = "=F43+F44+F45+F46"
ws["G47"] = "=G43+G44+G45+G46"
ws["H47"] = "=H43+H44+H45+H46"
# 間接計列の追加
ws["F1"] = "【間接計】"
# 間接計列への計算式の埋め込み
ws["F2"] = "=SUM(B2:E2)"
ws["F3"] = "=SUM(B3:E3)"
ws["F4"] = "=SUM(B4:E4)"
ws["F5"] = "=SUM(B5:E5)"
ws["F6"] = "=SUM(B6:E6) / 4"
ws["F7"] = "=SUM(B7:E7)"
ws["F8"] = "=SUM(B8:E8)"
ws["F9"] = "=SUM(B9:E9)"
ws["F10"] = "=SUM(B10:E10) / 4"
ws["F11"] = "=SUM(B11:E11)"
ws["F12"] = "=SUM(B12:E12)"
ws["F13"] = "=SUM(B13:E13)"
ws["F14"] = "=SUM(B14:E14) / 4"
ws["F15"] = "=SUM(B15:E15)"
ws["F16"] = "=SUM(B16:E16) / 4"
ws["F17"] = "=SUM(B17:E17)"
ws["F18"] = "=SUM(B18:E18) / 4"
ws["F19"] = "=SUM(B19:E19)"
ws["F20"] = "=SUM(B20:E20)"
ws["F21"] = "=SUM(B21:E21)"
ws["F22"] = "=SUM(B22:E22)"
ws["F23"] = "=SUM(B23:E23)"
ws["F24"] = "=SUM(B24:E24)"
ws["F25"] = "=SUM(B25:E25)"
ws["F26"] = "=SUM(B26:E26)"
ws["F27"] = "=SUM(B27:E27)"
ws["F28"] = "=SUM(B28:E28)"
ws["F29"] = "=SUM(B29:E29)"
ws["F30"] = "=SUM(B30:E30)"
ws["F31"] = "=SUM(B31:E31)"
ws["F32"] = "=SUM(B32:E32)"
ws["F33"] = "=SUM(B33:E33)"
ws["F34"] = "=SUM(B34:E34)"
ws["F35"] = "=SUM(B35:E35)"
ws["F36"] = "=SUM(B36:E36)"
ws["F37"] = "=SUM(B37:E37)"
ws["F38"] = "=SUM(B38:E38)"
ws["F39"] = "=SUM(B39:E39)"
ws["F40"] = "=SUM(B40:E40)"
ws["F41"] = "=SUM(B41:E41)"
ws["F42"] = "=SUM(B42:E42)"
ws["F43"] = "=SUM(B43:E43)"
ws["F44"] = "=SUM(B44:E44)"
ws["F45"] = "=SUM(B45:E45)"
ws["F46"] = "=SUM(B46:E46)"
ws["F47"] = "=E43+E44+E45+E46"
# ヘッダーの書式
fill = PatternFill(patternType="solid", fgColor="008000")
# 直接計列の追加
ws["I1"] = "【直接計】"
# 直接計列への計算式の埋め込み
ws["I2"] = "=SUM(G2:H2)"
ws["I3"] = "=SUM(G3:H3)"
ws["I4"] = "=SUM(G4:H4)"
ws["I5"] = "=SUM(G5:H5)"
ws["I6"] = "=SUM(G6:H6) / 2"
ws["I7"] = "=SUM(G7:H7)"
ws["I8"] = "=SUM(G8:H8)"
ws["I9"] = "=SUM(G9:H9)"
ws["I10"] = "=SUM(G10:H10) / 2"
ws["I11"] = "=SUM(G11:H11)"
ws["I12"] = "=SUM(G12:H12)"
ws["I13"] = "=SUM(G13:H13)"
ws["I14"] = "=SUM(G14:H14) / 2"
ws["I15"] = "=SUM(G15:H15)"
ws["I16"] = "=SUM(G16:H16) / 2"
ws["I17"] = "=SUM(G17:H17)"
ws["I18"] = "=SUM(G18:H18) / 2"
ws["I19"] = "=SUM(G19:H19)"
ws["I20"] = "=SUM(G20:H20)"
ws["I21"] = "=SUM(G21:H21)"
ws["I22"] = "=SUM(G22:H22)"
ws["I23"] = "=SUM(G23:H23)"
ws["I24"] = "=SUM(G24:H24)"
ws["I25"] = "=SUM(G25:H25)"
ws["I26"] = "=SUM(G26:H26)"
ws["I27"] = "=SUM(G27:H27)"
ws["I28"] = "=SUM(G28:H28)"
ws["I29"] = "=SUM(G29:H29)"
ws["I30"] = "=SUM(G30:H30)"
ws["I31"] = "=SUM(G31:H31)"
ws["I32"] = "=SUM(G32:H32)"
ws["I33"] = "=SUM(G33:H33)"
ws["I34"] = "=SUM(G34:H34)"
ws["I35"] = "=SUM(G35:H35)"
ws["I36"] = "=SUM(G36:H36)"
ws["I37"] = "=SUM(G37:H37)"
ws["I38"] = "=SUM(G38:H38)"
ws["I39"] = "=SUM(G39:H39)"
ws["I40"] = "=SUM(G40:H40)"
ws["I41"] = "=SUM(G41:H41)"
ws["I42"] = "=SUM(G42:H42)"
ws["I43"] = "=SUM(G43:H43)"
ws["I44"] = "=SUM(G44:H44)"
ws["I45"] = "=SUM(G45:H45)"
ws["I46"] = "=SUM(G46:H46)"
ws["I47"] = "=H43+H44+H45+H46"
# ヘッダーの書式
fill = PatternFill(patternType="solid", fgColor="008000")
ws["L1"].fill = fill
ws["L1"].font = Font(bold=True, color="FFFFFF")
# 合計列の追加とヘッダーの書式
ws["L1"] = "【合計】"
ws["L1"].fill = fill
ws["L1"].font = Font(bold=True, color="FFFFFF")
# 罫線
side1 = Side(border_style="thin", color="000000")
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
for row in ws["A2:L49"]:
    for cell in row:
        cell.border = border_aro
# 行合計の計算式埋め込み
ws["L2"] = "=F2+I2"
ws["L3"] = "=F3+I3"
ws["L4"] = "=F4+I4"
ws["L5"] = "=F5+I5"
ws["L6"] = "=(F6+I6) / 2"
ws["L7"] = "=F7+I7"
ws["L8"] = "=F8+I8"
ws["L9"] = "=F9+I9"
ws["L10"] = "=(F10+I10) / 2"
ws["L11"] = "=F11+I11"
ws["L12"] = "=F12+I12"
ws["L13"] = "=F13+I13"
ws["L14"] = "=(F14+I14) / 2"
ws["L15"] = "=F15+I15"
ws["L16"] = "=(F16+I16) / 2 "
ws["L17"] = "=F17+I17"
ws["L18"] = "=(F18+I18) / 2"
ws["L19"] = "=F19+I19"
ws["L20"] = "=F20+I20"
ws["L21"] = "=F21+I21"
ws["L22"] = "=F22+I22"
ws["L23"] = "=F23+I23"
ws["L24"] = "=F24+I24"
ws["L25"] = "=F25+I25"
ws["L26"] = "=F26+I26"
ws["L27"] = "=F27+I27"
ws["L28"] = "=F28+I28"
ws["L29"] = "=F29+I29"
ws["L30"] = "=F30+I30"
ws["L31"] = "=F31+I31"
ws["L32"] = "=F32+I32"
ws["L33"] = "=F33+I33"
ws["L34"] = "=F34+I34"
ws["L35"] = "=F35+I35"
ws["L36"] = "=F36+I36"
ws["L37"] = "=F37+I37"
ws["L38"] = "=F38+I38"
ws["L39"] = "=F39+I39"
ws["L40"] = "=F40+I40"
ws["L41"] = "=F41+I41"
ws["L42"] = "=F42+I42"
ws["L43"] = "=F43+I43"
ws["L44"] = "=F44+I44"
ws["L45"] = "=F45+I45"
ws["L46"] = "=F46+I46"
ws["L47"] = "=L43+L44+L45+L46"
ws.delete_rows(48, 49)
# Excelファイルを出力
wb.save("/content/drive/MyDrive/data/EXCEL/AC.xlsx")
# PC
# パスで指定したファイルの一覧をリスト形式で取得
csv_files_pc = glob.glob("/content/drive/MyDrive/data/PC/*.csv")
# CSVファイルの中身を追加していくリストを表示
data_list_pc = []
# 読み込むファイルのリストをスキャン
for file in csv_files_pc:
    data_list_pc.append(pd.read_csv(file))
# リストを全て列方向に結合
df_pc_t = pd.concat(data_list_pc, axis=1, sort=True)
# print(df_pc_t)
# columnsパラメータで列名を設定
feature_pc = ["間接1", "間接2", "間接4", "間接5", "間接6", "直接1", "直接4"]
df_pc_t.columns = feature_pc
# 行名の設定
df_pc_t = df_pc_t.rename(
    {
        0: "在籍者",
        1: "在籍者主幹以下人数",
        2: "実在籍者",
        3: "有休時間",
        4: "有休時間在籍者平均",
        5: "欠勤時間",
        6: "勤務時間",
        7: "遅早時間",
        8: "出勤率",
        9: "実労働時間",
        10: "ズレ時間",
        11: "残業時間",
        12: "残業時間主幹以下平均",
        13: "法定外休出時間",
        14: "法定外休出主幹以下平均",
        15: "法定休出時間",
        16: "法定休出主幹以下平均",
        17: "時間外60時間超",
        18: "代休時間",
        19: "応援時間",
        20: "総労働時間",
        21: "基本給",
        22: "役職手当",
        23: "営業手当",
        24: "地域手当",
        25: "特別手当",
        26: "特別技技手当",
        27: "調整手当",
        28: "別居手当",
        29: "通勤手当",
        30: "小計1",
        31: "残業手当",
        32: "休出手当",
        33: "深夜勤務手当",
        34: "交替時差手当",
        35: "休業手当",
        36: "休業控除",
        37: "代休他",
        38: "欠勤・遅早控除",
        39: "精算分",
        40: "小計2",
        41: "総支給額",
        42: "応援時間額",
        43: "役員振替",
        44: "部門振替",
        45: "合計",
    }
)

# 整形---------------------
# ワークブックの生成
wb = Workbook()
# ワークシートの生成
ws = wb.active
ws.title = "PC工場"
# DataFrameを行単位のデータにする
rows = dataframe_to_rows(df_pc_t, index=True, header=True)
# 1セルずつ処理を実行する
for row_no, row in enumerate(rows, 1):
    for col_no, value in enumerate(row, 1):
        # データを書き込む
        ws.cell(row=row_no, column=col_no, value=value)
# 不要な行の削除
ws.delete_rows(2)
# 列の追加
ws.insert_cols(7)
# 表示倍率の設定
ws.sheet_view.zoomScale = 100
# 列幅の設定
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 12
ws.column_dimensions["J"].width = 12
ws.column_dimensions["K"].width = 12
ws.column_dimensions["L"].width = 12
# 表示形式
format = "#,##0"
for row in ws["B2:L2"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B4:L9"]:
    for cell in row:
        cell.number_format = format
format = "0.00%"
for row in ws["B10:L10"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B11:L12"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B13:L22"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B23:L23"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B24:L26"]:
    for cell in row:
        cell.number_format = format

format = "#,##0"
for row in ws["B25:L49"]:
    for cell in row:
        cell.number_format = format
# ヘッダー行のスタイル設定
header = ws[1]
for header_cell in header:
    # フォントを設定する
    header_cell.fill = PatternFill(patternType="solid", fgColor="008000")
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
# 中央揃え
for row in ws["B1:J1"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous")
ws["L1"].alignment = Alignment(horizontal="centerContinuous")
# 指定した行の背景色を黄色にする
mylist = [21, 44, 45, 46]
# for list in mylist:
#    for row in ws.iter_rows():
#        for cell in row:
#            if cell.row == list:
#                cell.fill = PatternFill(
#                    fgColor="FFFF00", bgColor="FFFF00", fill_type="solid"
#                )
# 小計1
ws["B32"] = "=SUM(B23:B31)"
ws["C32"] = "=SUM(C23:C31)"
ws["D32"] = "=SUM(D23:D31)"
ws["E32"] = "=SUM(E23:E31)"
ws["F32"] = "=SUM(F23:F31)"
ws["G32"] = "=SUM(G23:G31)"
ws["H32"] = "=SUM(H23:H31)"
ws["I32"] = "=SUM(I23:I31)"
ws["J32"] = "=SUM(J23:J31)"
# 小計2
ws["B42"] = "=SUM(B33:B41)"
ws["C42"] = "=SUM(C33:C41)"
ws["D42"] = "=SUM(D33:D41)"
ws["E42"] = "=SUM(E33:E41)"
ws["F42"] = "=SUM(F33:F41)"
ws["G42"] = "=SUM(G33:G41)"
ws["H42"] = "=SUM(H33:H41)"
ws["I42"] = "=SUM(I33:I41)"
ws["J42"] = "=SUM(J33:J41)"
# 総支給額
ws["B43"] = "=B32 + B42"
ws["C43"] = "=C32 + C42"
ws["D43"] = "=D32 + D42"
ws["E43"] = "=E32 + E42"
ws["F43"] = "=F32 + F42"
ws["G43"] = "=G32 + G42"
ws["H43"] = "=H32 + H42"
ws["I43"] = "=I32 + I42"
ws["J43"] = "=J32 + J42"
# 列合計の埋め込み
ws["B47"] = "=B43+B44+B45+B46"
ws["C47"] = "=C43+C44+C45+C46"
ws["D47"] = "=D43+D44+D45+D46"
ws["E47"] = "=E43+E44+E45+E46"
ws["F47"] = "=F43+F44+F45+F46"
ws["G47"] = "=G43+G44+G45+G46"
ws["H47"] = "=H43+H44+H45+H46"
ws["I47"] = "=I43+I44+I45+I46"
ws["J47"] = "=J43+J44+J45+J46"
# 間接計列の追加
ws["G1"] = "【間接計】"
# 間接計列への計算式の埋め込み
ws["G2"] = "=SUM(B2:F2)"
ws["G3"] = "=SUM(B3:F3)"
ws["G4"] = "=SUM(B4:F4)"
ws["G5"] = "=SUM(B5:F5)"
ws["G6"] = "=SUM(B6:F6) / 5"
ws["G7"] = "=SUM(B7:F7)"
ws["G8"] = "=SUM(B8:F8)"
ws["G9"] = "=SUM(B9:F9)"
ws["G10"] = "=SUM(B10:F10) / 5"
ws["G11"] = "=SUM(B11:F11)"
ws["G12"] = "=SUM(B12:F12)"
ws["G13"] = "=SUM(B13:F13)"
ws["G14"] = "=SUM(B14:F14) / 5"
ws["G15"] = "=SUM(B15:F15)"
ws["G16"] = "=SUM(B16:F16) / 5"
ws["G17"] = "=SUM(B17:F17)"
ws["G18"] = "=SUM(B18:F18) / 5"
ws["G19"] = "=SUM(B19:F19)"
ws["G20"] = "=SUM(B20:F20)"
ws["G21"] = "=SUM(B21:F21)"
ws["G22"] = "=SUM(B22:F22)"
ws["G23"] = "=SUM(B23:F23)"
ws["G24"] = "=SUM(B24:F24)"
ws["G25"] = "=SUM(B25:F25)"
ws["G26"] = "=SUM(B26:F26)"
ws["G27"] = "=SUM(B27:F27)"
ws["G28"] = "=SUM(B28:F28)"
ws["G29"] = "=SUM(B29:F29)"
ws["G30"] = "=SUM(B30:F30)"
ws["G31"] = "=SUM(B31:F31)"
ws["G32"] = "=SUM(B32:F32)"
ws["G33"] = "=SUM(B33:F33)"
ws["G34"] = "=SUM(B34:F34)"
ws["G35"] = "=SUM(B35:F35)"
ws["G36"] = "=SUM(B36:F36)"
ws["G37"] = "=SUM(B37:F37)"
ws["G38"] = "=SUM(B38:F38)"
ws["G39"] = "=SUM(B39:F39)"
ws["G40"] = "=SUM(B40:F40)"
ws["G41"] = "=SUM(B41:F41)"
ws["G42"] = "=SUM(B42:F42)"
ws["G43"] = "=SUM(B43:F43)"
ws["G44"] = "=SUM(B44:F44)"
ws["G45"] = "=SUM(B45:F45)"
ws["G46"] = "=SUM(B46:F46)"
ws["G47"] = "=G43+G44+G45+G46"
# 直接計列の追加
fill = PatternFill(patternType="solid", fgColor="008000")
ws["J1"].fill = fill
ws["J1"].font = Font(bold=True, color="FFFFFF")
ws["J1"] = "【直接計】"
# 直接計列への計算式の埋め込み
ws["J2"] = "=SUM(H2:I2)"
ws["J3"] = "=SUM(H3:I3)"
ws["J4"] = "=SUM(H4:I4)"
ws["J5"] = "=SUM(H5:I5)"
ws["J6"] = "=SUM(H6:I6) / 2"
ws["J7"] = "=SUM(H7:I7)"
ws["J8"] = "=SUM(H8:I8)"
ws["J9"] = "=SUM(H9:I9)"
ws["J10"] = "=SUM(H10:I10) / 2"
ws["J11"] = "=SUM(H11:I11)"
ws["J12"] = "=SUM(H12:I12)"
ws["J13"] = "=SUM(H13:I13)"
ws["J14"] = "=SUM(H14:I14) / 2"
ws["J15"] = "=SUM(H15:I15)"
ws["J16"] = "=SUM(H16:I16) / 2"
ws["J17"] = "=SUM(H17:I17)"
ws["J18"] = "=SUM(H18:I18) / 2"
ws["J19"] = "=SUM(H19:I19)"
ws["J20"] = "=SUM(H20:I20)"
ws["J21"] = "=SUM(H21:I21)"
ws["J22"] = "=SUM(H22:I22)"
ws["J23"] = "=SUM(H23:I23)"
ws["J24"] = "=SUM(H24:I24)"
ws["J25"] = "=SUM(H25:I25)"
ws["J26"] = "=SUM(H26:I26)"
ws["J27"] = "=SUM(H27:I27)"
ws["J28"] = "=SUM(H28:I28)"
ws["J29"] = "=SUM(H29:I29)"
ws["J30"] = "=SUM(H30:I30)"
ws["J31"] = "=SUM(H31:I31)"
ws["J32"] = "=SUM(H32:I32)"
ws["J33"] = "=SUM(H33:I33)"
ws["J34"] = "=SUM(H34:I34)"
ws["J35"] = "=SUM(H35:I35)"
ws["J36"] = "=SUM(H36:I36)"
ws["J37"] = "=SUM(H37:I37)"
ws["J38"] = "=SUM(H38:I38)"
ws["J39"] = "=SUM(H39:I39)"
ws["J40"] = "=SUM(H40:I40)"
ws["J41"] = "=SUM(H41:I41)"
ws["J42"] = "=SUM(H42:I42)"
ws["J43"] = "=SUM(H43:I43)"
ws["J44"] = "=SUM(H44:I44)"
ws["J45"] = "=SUM(H45:I45)"
ws["J46"] = "=SUM(H46:I46)"
ws["J47"] = "=J43+J44+J45+J46"
# ヘッダーの書式
fill = PatternFill(patternType="solid", fgColor="008000")
ws["L1"].fill = fill
ws["L1"].font = Font(bold=True, color="FFFFFF")
# 合計列の追加とヘッダーの書式
ws["L1"] = "【合計】"
ws["L1"].fill = fill
ws["L1"].font = Font(bold=True, color="FFFFFF")
# 罫線
side1 = Side(border_style="thin", color="000000")
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
for row in ws["A2:L49"]:
    for cell in row:
        cell.border = border_aro
# 行合計の計算式埋め込み
ws["L2"] = "=G2+J2"
ws["L3"] = "=G3+J3"
ws["L4"] = "=G4+J4"
ws["L5"] = "=G5+J5"
ws["L6"] = "=(G6+J6) / 2"
ws["L7"] = "=G7+J7"
ws["L8"] = "=G8+J8"
ws["L9"] = "=G9+J9"
ws["L10"] = "=(G10+J10) / 2"
ws["L11"] = "=G11+J11"
ws["L12"] = "=G12+J12"
ws["L13"] = "=G13+J13"
ws["L14"] = "=(G14+J14) / 2"
ws["L15"] = "=G15+J15"
ws["L16"] = "=(G16+J16) / 2"
ws["L17"] = "=G17+J17"
ws["L18"] = "=(G18+J18) / 2"
ws["L19"] = "=G19+J19"
ws["L20"] = "=G20+J20"
ws["L21"] = "=G21+J21"
ws["L22"] = "=G22+J22"
ws["L23"] = "=G23+J23"
ws["L24"] = "=G24+J24"
ws["L25"] = "=G25+J25"
ws["L26"] = "=G26+J26"
ws["L27"] = "=G27+J27"
ws["L28"] = "=G28+J28"
ws["L29"] = "=G29+J29"
ws["L30"] = "=G30+J30"
ws["L31"] = "=G31+J31"
ws["L32"] = "=G32+J32"
ws["L33"] = "=G33+J33"
ws["L34"] = "=G34+J34"
ws["L35"] = "=G35+J35"
ws["L36"] = "=G36+J36"
ws["L37"] = "=G37+J37"
ws["L38"] = "=G38+J38"
ws["L39"] = "=G39+J39"
ws["L40"] = "=G40+J40"
ws["L41"] = "=G41+J41"
ws["L42"] = "=G42+J42"
ws["L43"] = "=G43+J43"
ws["L44"] = "=G44+J44"
ws["L45"] = "=G45+J45"
ws["L46"] = "=G46+J46"
ws["L47"] = "=L43+L44+L45+L46"
ws.delete_rows(48, 49)
# Excelファイルを出力
wb.save("/content/drive/MyDrive/data/EXCEL/PC.xlsx")
# 宮城
# パスで指定したファイルの一覧をリスト形式で取得
csv_files_miyagi = glob.glob("/content/drive/MyDrive/data/宮城/*.csv")
# CSVファイルの中身を追加していくリストを表示
data_list_miyagi = []
# 読み込むファイルのリストをスキャン
for file in csv_files_miyagi:
    data_list_miyagi.append(pd.read_csv(file))
# リストを全て列方向に結合
df_miyagi_t = pd.concat(data_list_miyagi, axis=1, sort=True)
# columnsパラメータで列名を設定
feature_miyagi = ["間接1", "間接2", "間接4", "間接6", "直接1"]
df_miyagi_t.columns = feature_miyagi
# 行名の設定
df_miyagi_t = df_miyagi_t.rename(
    {
        0: "在籍者",
        1: "在籍者主幹以下人数",
        2: "実在籍者",
        3: "有休時間",
        4: "有休時間在籍者平均",
        5: "欠勤時間",
        6: "勤務時間",
        7: "遅早時間",
        8: "出勤率",
        9: "実労働時間",
        10: "ズレ時間",
        11: "残業時間",
        12: "残業時間主幹以下平均",
        13: "法定外休出時間",
        14: "法定外休出主幹以下平均",
        15: "法定休出時間",
        16: "法定休出主幹以下平均",
        17: "時間外60時間超",
        18: "代休時間",
        19: "応援時間",
        20: "総労働時間",
        21: "基本給",
        22: "役職手当",
        23: "営業手当",
        24: "地域手当",
        25: "特別手当",
        26: "特別技技手当",
        27: "調整手当",
        28: "別居手当",
        29: "通勤手当",
        30: "小計1",
        31: "残業手当",
        32: "休出手当",
        33: "深夜勤務手当",
        34: "交替時差手当",
        35: "休業手当",
        36: "休業控除",
        37: "代休他",
        38: "欠勤・遅早控除",
        39: "精算分",
        40: "小計2",
        41: "総支給額",
        42: "応援時間額",
        43: "役員振替",
        44: "部門振替",
        45: "合計",
    }
)
# 整形---------------------
# ワークブックの生成
wb = Workbook()
# ワークシートの生成
ws = wb.active
ws.title = "宮城工場"
# DataFrameを行単位のデータにする
rows = dataframe_to_rows(df_miyagi_t, index=True, header=True)
# 1セルずつ処理を実行する
for row_no, row in enumerate(rows, 1):
    for col_no, value in enumerate(row, 1):
        # データを書き込む
        ws.cell(row=row_no, column=col_no, value=value)
# 不要な行の削除
ws.delete_rows(2)
# 列の追加
ws.insert_cols(6)
# 表示倍率の設定
ws.sheet_view.zoomScale = 100
# 列幅の設定
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 12
ws.column_dimensions["J"].width = 12
ws.column_dimensions["K"].width = 12
ws.column_dimensions["L"].width = 12
# 表示形式
format = "#,##0"
for row in ws["B2:L2"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B4:L9"]:
    for cell in row:
        cell.number_format = format
format = "0.00%"
for row in ws["B10:L10"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B11:L12"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B13:L22"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B23:L23"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B24:L26"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B25:L49"]:
    for cell in row:
        cell.number_format = format
# ヘッダー行のスタイル設定
header = ws[1]
for header_cell in header:
    # フォントを設定する
    header_cell.fill = PatternFill(patternType="solid", fgColor="008000")
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
# 中央揃え
for row in ws["B1:H1"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous")
ws["L1"].alignment = Alignment(horizontal="centerContinuous")
# 指定した行の背景色を黄色にする
mylist = [21, 44, 45, 46]
# for list in mylist:
#    for row in ws.iter_rows():
#        for cell in row:
#            if cell.row == list:
#                cell.fill = PatternFill(
#                    fgColor="FFFF00", bgColor="FFFF00", fill_type="solid"
#                )
# 小計1
ws["B32"] = "=SUM(B23:B31)"
ws["C32"] = "=SUM(C23:C31)"
ws["D32"] = "=SUM(D23:D31)"
ws["E32"] = "=SUM(E23:E31)"
ws["F32"] = "=SUM(F23:F31)"
ws["G32"] = "=SUM(G23:G31)"
ws["H32"] = "=SUM(H23:H31)"
# 小計2
ws["B42"] = "=SUM(B33:B41)"
ws["C42"] = "=SUM(C33:C41)"
ws["D42"] = "=SUM(D33:D41)"
ws["E42"] = "=SUM(E33:E41)"
ws["F42"] = "=SUM(F33:F41)"
ws["G42"] = "=SUM(G33:G41)"
ws["H42"] = "=SUM(H33:h41)"
# 総支給額
ws["B43"] = "=B32 + B42"
ws["C43"] = "=C32 + C42"
ws["D43"] = "=D32 + D42"
ws["E43"] = "=E32 + E42"
ws["F43"] = "=F32 + F42"
ws["G43"] = "=G32 + G42"
ws["H43"] = "=H32 + H42"
# 列合計の埋め込み
ws["B47"] = "=B43+B44+B45+B46"
ws["C47"] = "=C43+C44+C45+C46"
ws["D47"] = "=D43+D44+D45+D46"
ws["E47"] = "=E43+E44+E45+E46"
ws["F47"] = "=F43+F44+F45+F46"
ws["G47"] = "=G43+G44+G45+G46"
ws["H47"] = "=H43+H44+H45+H46"
# 間接計列の追加
ws["F1"] = "【間接計】"
# 間接計列への計算式の埋め込み
ws["F2"] = "=SUM(B2:E2)"
ws["F3"] = "=SUM(B3:E3)"
ws["F4"] = "=SUM(B4:E4)"
ws["F5"] = "=SUM(B5:E5)"
ws["F6"] = "=SUM(B6:E6) / 4"
ws["F7"] = "=SUM(B7:E7)"
ws["F8"] = "=SUM(B8:E8)"
ws["F9"] = "=SUM(B9:E9)"
ws["F10"] = "=SUM(B10:E10) / 4"
ws["F11"] = "=SUM(B11:E11)"
ws["F12"] = "=SUM(B12:E12)"
ws["F13"] = "=SUM(B13:E13)"
ws["F14"] = "=SUM(B14:E14) / 4"
ws["F15"] = "=SUM(B15:E15)"
ws["F16"] = "=SUM(B16:E16) / 4"
ws["F17"] = "=SUM(B17:E17)"
ws["F18"] = "=SUM(B18:E18) / 4"
ws["F19"] = "=SUM(B19:E19)"
ws["F20"] = "=SUM(B20:E20)"
ws["F21"] = "=SUM(B21:E21)"
ws["F22"] = "=SUM(B22:E22)"
ws["F23"] = "=SUM(B23:E23)"
ws["F24"] = "=SUM(B24:E24)"
ws["F25"] = "=SUM(B25:E25)"
ws["F26"] = "=SUM(B26:E26)"
ws["F27"] = "=SUM(B27:E27)"
ws["F28"] = "=SUM(B28:E28)"
ws["F29"] = "=SUM(B29:E29)"
ws["F30"] = "=SUM(B30:E30)"
ws["F31"] = "=SUM(B31:E31)"
ws["F32"] = "=SUM(B32:E32)"
ws["F33"] = "=SUM(B33:E33)"
ws["F34"] = "=SUM(B34:E34)"
ws["F35"] = "=SUM(B35:E35)"
ws["F36"] = "=SUM(B36:E36)"
ws["F37"] = "=SUM(B37:E37)"
ws["F38"] = "=SUM(B38:E38)"
ws["F39"] = "=SUM(B39:E39)"
ws["F40"] = "=SUM(B40:E40)"
ws["F41"] = "=SUM(B41:E41)"
ws["F42"] = "=SUM(B42:E42)"
ws["F43"] = "=SUM(B43:E43)"
ws["F44"] = "=SUM(B44:E44)"
ws["F45"] = "=SUM(B45:E45)"
ws["F46"] = "=SUM(B46:E46)"
ws["F47"] = "=F43+F44+F45+F46"
# 直接計列の追加
ws["H1"] = "【直接計】"
# 直接計列への計算式の埋め込み
ws["H2"] = "=G2"
ws["H3"] = "=G3"
ws["H4"] = "=G4"
ws["H5"] = "=G5"
ws["H6"] = "=G6"
ws["H7"] = "=G7"
ws["H8"] = "=G8"
ws["H9"] = "=G9"
ws["H10"] = "=G10"
ws["H11"] = "=G11"
ws["H12"] = "=G12"
ws["H13"] = "=G13"
ws["H14"] = "=G14"
ws["H15"] = "=G15"
ws["H16"] = "=G16"
ws["H17"] = "=G17"
ws["H18"] = "=G18"
ws["H19"] = "=G19"
ws["H20"] = "=G20"
ws["H21"] = "=G21"
ws["H22"] = "=G22"
ws["H23"] = "=G23"
ws["H24"] = "=G24"
ws["H25"] = "=G25"
ws["H26"] = "=G26"
ws["H27"] = "=G27"
ws["H28"] = "=G28"
ws["H29"] = "=G29"
ws["H30"] = "=G30"
ws["H31"] = "=G31"
ws["H32"] = "=G32"
ws["H33"] = "=G33"
ws["H34"] = "=G34"
ws["H35"] = "=G35"
ws["H36"] = "=G36"
ws["H37"] = "=G37"
ws["H38"] = "=G38"
ws["H39"] = "=G39"
ws["H40"] = "=G40"
ws["H41"] = "=G41"
ws["H42"] = "=G42"
ws["H43"] = "=G43"
ws["H44"] = "=G44"
ws["H45"] = "=G45"
ws["H46"] = "=G46"
ws["H47"] = "=H43+H44+H45+H46"
# 直接列ヘッダーの書式
fill = PatternFill(patternType="solid", fgColor="008000")
# 合計列の追加とヘッダーの書式
ws["L1"] = "【合計】"
ws["L1"].fill = fill
ws["L1"].font = Font(bold=True, color="FFFFFF")
# 罫線
side1 = Side(border_style="thin", color="000000")
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
for row in ws["A2:L49"]:
    for cell in row:
        cell.border = border_aro
# 行合計の埋め込み
ws["L2"] = "=SUM(F2,H2)"
ws["L3"] = "=SUM(F3,H3)"
ws["L4"] = "=SUM(F4,H4)"
ws["L5"] = "=SUM(F5,H5)"
ws["L6"] = "=SUM(F6,H6) / 2"
ws["L7"] = "=SUM(F7,H7)"
ws["L8"] = "=SUM(F8,H8)"
ws["L9"] = "=SUM(F9,H9)"
ws["L10"] = "=SUM(F10,H10) / 2"
ws["L11"] = "=SUM(F11,H11)"
ws["L12"] = "=SUM(F12,H12)"
ws["L13"] = "=SUM(F13,H13)"
ws["L14"] = "=SUM(F14,H14) / 2"
ws["L15"] = "=SUM(F15,H15)"
ws["L16"] = "=SUM(F16,H16) / 2"
ws["L17"] = "=SUM(F17,H17)"
ws["L18"] = "=SUM(F18,H18) / 2"
ws["L19"] = "=SUM(F19,H19)"
ws["L20"] = "=SUM(F20,H20)"
ws["L21"] = "=SUM(F21,H21)"
ws["L22"] = "=SUM(F22,H22)"
ws["L23"] = "=SUM(F23,H23)"
ws["L24"] = "=SUM(F24,H24)"
ws["L25"] = "=SUM(F25,H25)"
ws["L26"] = "=SUM(F26,H26)"
ws["L27"] = "=SUM(F27,H27)"
ws["L28"] = "=SUM(F28,H28)"
ws["L29"] = "=SUM(F29,H29)"
ws["L30"] = "=SUM(F30,H30)"
ws["L31"] = "=SUM(F31,H31)"
ws["L32"] = "=SUM(F32,H32)"
ws["L33"] = "=SUM(F33,H33)"
ws["L34"] = "=SUM(F34,H34)"
ws["L35"] = "=SUM(F35,H35)"
ws["L36"] = "=SUM(F36,H36)"
ws["L37"] = "=SUM(F37,H37)"
ws["L38"] = "=SUM(F38,H38)"
ws["L39"] = "=SUM(F39,H39)"
ws["L40"] = "=SUM(F40,H40)"
ws["L41"] = "=SUM(F41,H41)"
ws["L42"] = "=SUM(F42,H42)"
ws["L43"] = "=SUM(F43,H43)"
ws["L44"] = "=SUM(F44,H44)"
ws["L45"] = "=SUM(F45,H45)"
ws["L46"] = "=SUM(F46,H46)"
ws["L47"] = "=L43+L44+L45+L46"
ws.delete_rows(48, 49)
# Excelファイルを出力
wb.save("/content/drive/MyDrive/data/EXCEL/宮城.xlsx")
# 住設
# パスで指定したファイルの一覧をリスト形式で取得
csv_files_jyusetu = glob.glob("/content/drive/MyDrive/data/住設/*.csv")
# CSVファイルの中身を追加していくリストを表示
data_list_jyusetu = []
# 読み込むファイルのリストをスキャン
for file in csv_files_jyusetu:
    data_list_jyusetu.append(pd.read_csv(file))
# リストを全て列方向に結合
df_jyusetu_t = pd.concat(data_list_jyusetu, axis=1, sort=True)
# columnsパラメータで列名を設定
feature_jyusetu = ["間接2", "間接4", "間接6"]
df_jyusetu_t.columns = feature_jyusetu
# 行名の設定
df_jyusetu_t = df_jyusetu_t.rename(
    {
        0: "在籍者",
        1: "在籍者主幹以下人数",
        2: "実在籍者",
        3: "有休時間",
        4: "有休時間在籍者平均",
        5: "欠勤時間",
        6: "勤務時間",
        7: "遅早時間",
        8: "出勤率",
        9: "実労働時間",
        10: "ズレ時間",
        11: "残業時間",
        12: "残業時間主幹以下平均",
        13: "法定外休出時間",
        14: "法定外休出主幹以下平均",
        15: "法定休出時間",
        16: "法定休出主幹以下平均",
        17: "時間外60時間超",
        18: "代休時間",
        19: "応援時間",
        20: "総労働時間",
        21: "基本給",
        22: "役職手当",
        23: "営業手当",
        24: "地域手当",
        25: "特別手当",
        26: "特別技技手当",
        27: "調整手当",
        28: "別居手当",
        29: "通勤手当",
        30: "小計1",
        31: "残業手当",
        32: "休出手当",
        33: "深夜勤務手当",
        34: "交替時差手当",
        35: "休業手当",
        36: "休業控除",
        37: "代休他",
        38: "欠勤・遅早控除",
        39: "精算分",
        40: "小計2",
        41: "総支給額",
        42: "応援時間額",
        43: "役員振替",
        44: "部門振替",
        45: "合計",
    }
)

# 整形---------------------
# ワークブックの生成
wb = Workbook()
# ワークシートの生成
ws = wb.active
ws.title = "住設"
# DataFrameを行単位のデータにする
rows = dataframe_to_rows(df_jyusetu_t, index=True, header=True)
# 1セルずつ処理を実行する
for row_no, row in enumerate(rows, 1):
    for col_no, value in enumerate(row, 1):
        # データを書き込む
        ws.cell(row=row_no, column=col_no, value=value)
# 不要な行の削除
ws.delete_rows(2)
# 表示倍率の設定
ws.sheet_view.zoomScale = 100
# 列幅の設定
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 12
ws.column_dimensions["J"].width = 12
ws.column_dimensions["K"].width = 12
ws.column_dimensions["L"].width = 12
# 表示形式
format = "#,##0"
for row in ws["B2:L2"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B4:L9"]:
    for cell in row:
        cell.number_format = format
format = "0.00%"
for row in ws["B10:L10"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B11:L12"]:
    for cell in row:
        cell.number_format = format
format = "0.00"
for row in ws["B13:L22"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B23:L23"]:
    for cell in row:
        cell.number_format = format
format = "#,##0"
for row in ws["B24:L26"]:
    for cell in row:
        cell.number_format = format

format = "#,##0"
for row in ws["B25:L49"]:
    for cell in row:
        cell.number_format = format
# ヘッダー行のスタイル設定
header = ws[1]
for header_cell in header:
    # フォントを設定する
    header_cell.fill = PatternFill(patternType="solid", fgColor="008000")
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
# 中央揃え
for row in ws["B1:E1"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous")
ws["L1"].alignment = Alignment(horizontal="centerContinuous")
# 指定した行の背景色を黄色にする
mylist = [21, 44, 45, 46]
# for list in mylist:
#    for row in ws.iter_rows():
#        for cell in row:
#            if cell.row == list:
#                cell.fill = PatternFill(
#                   fgColor="FFFF00", bgColor="FFFF00", fill_type="solid"
#                )
# 小計1
ws["B32"] = "=SUM(B23:B31)"
ws["C32"] = "=SUM(C23:C31)"
ws["D32"] = "=SUM(D23:D31)"
ws["E32"] = "=SUM(E23:E31)"
# 小計2
ws["B42"] = "=SUM(B33:B41)"
ws["C42"] = "=SUM(C33:C41)"
ws["D42"] = "=SUM(D33:D41)"
ws["E42"] = "=SUM(E33:E41)"
# 総支給額
ws["B43"] = "=B32 + B42"
ws["C43"] = "=C32 + C42"
ws["D43"] = "=D32 + D42"
ws["E43"] = "=E32 + E42"
# 列合計の埋め込み
ws["B47"] = "=B43+B44+B45+B46"
ws["C47"] = "=C43+C44+C45+C46"
ws["D47"] = "=D43+D44+D45+D46"
ws["E47"] = "=E43+E44+E45+E46"
# 間接計列の追加
ws["E1"] = "【間接計】"
# 間接計列への計算式の埋め込み
ws["E2"] = "=SUM(B2:D2)"
ws["E3"] = "=SUM(B3:D3)"
ws["E4"] = "=SUM(B4:D4)"
ws["E5"] = "=SUM(B5:D5)"
ws["E6"] = "=SUM(B6:D6) / 3"
ws["E7"] = "=SUM(B7:D7)"
ws["E8"] = "=SUM(B8:D8)"
ws["E9"] = "=SUM(B9:D9)"
ws["E10"] = "=SUM(B10:D10) / 3"
ws["E11"] = "=SUM(B11:D11)"
ws["E12"] = "=SUM(B12:D12)"
ws["E13"] = "=SUM(B13:D13)"
ws["E14"] = "=SUM(B14:D14) / 3"
ws["E15"] = "=SUM(B15:D15)"
ws["E16"] = "=SUM(B16:D16) / 3"
ws["E17"] = "=SUM(B17:D17)"
ws["E18"] = "=SUM(B18:D18) / 3"
ws["E19"] = "=SUM(B19:D19)"
ws["E20"] = "=SUM(B20:D20)"
ws["E21"] = "=SUM(B21:D21)"
ws["E22"] = "=SUM(B22:D22)"
ws["E23"] = "=SUM(B23:D23)"
ws["E24"] = "=SUM(B24:D24)"
ws["E25"] = "=SUM(B25:D25)"
ws["E26"] = "=SUM(B26:D26)"
ws["E27"] = "=SUM(B27:D27)"
ws["E28"] = "=SUM(B28:D28)"
ws["E29"] = "=SUM(B29:D29)"
ws["E30"] = "=SUM(B30:D30)"
ws["E31"] = "=SUM(B31:D31)"
ws["E32"] = "=SUM(B32:D32)"
ws["E33"] = "=SUM(B33:D33)"
ws["E34"] = "=SUM(B34:D34)"
ws["E35"] = "=SUM(B35:D35)"
ws["E36"] = "=SUM(B36:D36)"
ws["E37"] = "=SUM(B37:D37)"
ws["E38"] = "=SUM(B38:D38)"
ws["E39"] = "=SUM(B39:D39)"
ws["E40"] = "=SUM(B40:D40)"
ws["E41"] = "=SUM(B41:D41)"
ws["E42"] = "=SUM(B42:D42)"
ws["E43"] = "=SUM(B43:D43)"
ws["E44"] = "=SUM(B44:D44)"
ws["E45"] = "=SUM(B45:D45)"
ws["E46"] = "=SUM(B46:D46)"
ws["E47"] = "=E43+E44+E45+E46"
# 合計項目の追加とヘッダーの書式
# ws['A51'] = '合  計'
# ヘッダーの書式
fill = PatternFill(patternType="solid", fgColor="008000")
# 合計列の追加とヘッダーの書式
ws["L1"] = "【合計】"
ws["L1"].fill = fill
ws["L1"].font = Font(bold=True, color="FFFFFF")
# 罫線
side1 = Side(border_style="thin", color="000000")
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
for row in ws["A2:L49"]:
    for cell in row:
        cell.border = border_aro
# 行合計の埋め込み
ws["L2"] = "=E2"
ws["L3"] = "=E3"
ws["L4"] = "=E4"
ws["L5"] = "=E5"
ws["L6"] = "=E6"
ws["L7"] = "=E7"
ws["L8"] = "=E8"
ws["L9"] = "=E9"
ws["L10"] = "=E10"
ws["L11"] = "=E11"
ws["L12"] = "=E12"
ws["L13"] = "=E13"
ws["L14"] = "=E14"
ws["L15"] = "=E15"
ws["L16"] = "=E16"
ws["L17"] = "=E17"
ws["L18"] = "=E18"
ws["L19"] = "=E19"
ws["L20"] = "=E20"
ws["L21"] = "=E21"
ws["L22"] = "=E22"
ws["L23"] = "=E23"
ws["L24"] = "=E24"
ws["L25"] = "=E25"
ws["L26"] = "=E26"
ws["L27"] = "=E27"
ws["L28"] = "=E28"
ws["L29"] = "=E29"
ws["L30"] = "=E30"
ws["L31"] = "=E31"
ws["L32"] = "=E32"
ws["L33"] = "=E33"
ws["L34"] = "=E34"
ws["L35"] = "=E35"
ws["L36"] = "=E36"
ws["L37"] = "=E37"
ws["L38"] = "=E38"
ws["L39"] = "=E39"
ws["L40"] = "=E40"
ws["L41"] = "=E41"
ws["L42"] = "=E42"
ws["L43"] = "=E43"
ws["L44"] = "=E44"
ws["L45"] = "=E45"
ws["L46"] = "=E46"
ws["L47"] = "=L43+L44+L45+L46"

ws.delete_rows(48, 49)

# Excelファイルを出力
wb.save("/content/drive/MyDrive/data/EXCEL/住設.xlsx")