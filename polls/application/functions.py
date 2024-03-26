import pandas as pd
import numpy as np
import sklearn, csv, re, os
import xlrd
import glob
import csv
import sys
import openpyxl as op

from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from ctypes import alignment
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Font, numbers
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.alignment import Alignment


# ファイルを指定してデータフレーム化
def df_change(data):
    df_jinji = pd.read_excel("c:/test/jinji.XLS")
    df_jinji = df_jinji.drop(columns=["社員名", "集計区分－１        "])
    df_syain = pd.read_excel("c:/test/syain.XLS")
    df_syain = df_syain.drop(
        columns=[
            "氏 名",
            "役職ｺｰﾄﾞ",
            "役職",
            "所属分類1名",
            "所属分類2名",
            "所属分類3名",
            "所 .3",
            "所属分類4名",
            "所 .4",
            "所属分類5名",
            "備考",
            "在籍区分",
            "在籍",
            "退職年",
            "退職月",
            "退職日",
        ]
    )
    # 人事ファイルと社員ファイルを結合してデータフレーム化
    df_m = pd.merge(df_jinji, df_syain, how="inner", on="社員ｺｰﾄﾞ")
    # 役員データフレームの作成
    df_yakuin_kbn = df_m[
        (df_m["所 "] == 0) & (df_m["所 .1"] == 0) & (df_m["所 .2"] == 0)
    ]
    df_yakuin_kbn = df_yakuin_kbn.astype("object")
    df_yakuin_kbn["区分"] = "役員1"
    df_yakuin_kbn = df_yakuin_kbn.rename(
        columns={"所 ": "所属1", "所 .1": "所属2", "所 .2": "所属3"}
    )
    # 一般間接1データフレームの作成
    df_ipan_kbn103 = df_m[(df_m["集計区分－２        "] == 103)]
    df_ipan_kbn103 = df_ipan_kbn103.astype("object")
    df_ipan_kbn103["区分"] = "一般間接1"
    # 一般間接2データフレームの作成
    df_ipan_kbn104 = df_m[(df_m["集計区分－２        "] == 104)]
    df_ipan_kbn104 = df_ipan_kbn104.astype("object")
    df_ipan_kbn104["区分"] = "一般間接2"
    # 一般間接3データフレームの作成
    df_ipan_kbn105 = df_m[(df_m["集計区分－２        "] == 105)]
    df_ipan_kbn105 = df_ipan_kbn105.astype("object")
    df_ipan_kbn105["区分"] = "一般間接3"
    # 一般間接6データフレームの作成
    df_ipan_kbn108 = df_m[(df_m["集計区分－２        "] == 108)]
    df_ipan_kbn108 = df_ipan_kbn108.astype("object")
    df_ipan_kbn108["区分"] = "一般間接6"
    # 一般販売1データフレームの作成
    df_ipan_kbn109 = df_m[(df_m["集計区分－２        "] == 109)]
    df_ipan_kbn109 = df_ipan_kbn109.astype("object")
    df_ipan_kbn109["区分"] = "一般販売1"
    # 一般販売2データフレームの作成
    df_ipan_kbn110 = df_m[(df_m["集計区分－２        "] == 110)]
    df_ipan_kbn110 = df_ipan_kbn110.astype("object")
    df_ipan_kbn110["区分"] = "一般販売2"
    # 一般管理データフレームの作成
    df_ipan_kbn = pd.concat(
        [
            df_ipan_kbn103,
            df_ipan_kbn104,
            df_ipan_kbn105,
            df_ipan_kbn108,
            df_ipan_kbn109,
            df_ipan_kbn110,
        ],
        axis=0,
        ignore_index=True,
    )
    df_ipan_kbn = df_ipan_kbn.rename(
        columns={"所 ": "所属1", "所 .1": "所属2", "所 .2": "所属3"}
    )
    # 鍛造間接1のデータフレーム作成
    df_tanzo_kbn211 = df_m[(df_m["集計区分－２        "] == 211)]
    df_tanzo_kbn211 = df_tanzo_kbn211.astype("object")
    df_tanzo_kbn211["区分"] = "間接1"
    # 鍛造間接2のデータフレーム作成
    df_tanzo_kbn212 = df_m[(df_m["集計区分－２        "] == 212)]
    df_tanzo_kbn212 = df_tanzo_kbn212.astype("object")
    df_tanzo_kbn212["区分"] = "間接2"
    # 鍛造間接3のデータフレーム作成
    df_tanzo_kbn213 = df_m[(df_m["集計区分－２        "] == 213)]
    df_tanzo_kbn213 = df_tanzo_kbn213.astype("object")
    df_tanzo_kbn213["区分"] = "間接3"
    # 鍛造間接4のデータフレーム作成
    df_tanzo_kbn214 = df_m[(df_m["集計区分－２        "] == 214)]
    df_tanzo_kbn214 = df_tanzo_kbn214.astype("object")
    df_tanzo_kbn214["区分"] = "間接4"
    # 鍛造間接5のデータフレーム作成
    df_tanzo_kbn215 = df_m[(df_m["集計区分－２        "] == 215)]
    df_tanzo_kbn215 = df_tanzo_kbn215.astype("object")
    df_tanzo_kbn215["区分"] = "間接5"
    # 鍛造間接6のデータフレーム作成
    df_tanzo_kbn216 = df_m[(df_m["集計区分－２        "] == 216)]
    df_tanzo_kbn216 = df_tanzo_kbn216.astype("object")
    df_tanzo_kbn216["区分"] = "間接6"
    # 鍛造直接1のデータフレーム作成
    df_tanzo_kbn218 = df_m[(df_m["集計区分－２        "] == 218)]
    df_tanzo_kbn218 = df_tanzo_kbn218.astype("object")
    df_tanzo_kbn218["区分"] = "直接1"
    # 鍛造直接4のデータフレーム作成
    df_tanzo_kbn221 = df_m[(df_m["集計区分－２        "] == 221)]
    df_tanzo_kbn221 = df_tanzo_kbn221.astype("object")
    df_tanzo_kbn221["区分"] = "直接4"
    # 鍛造のデータフレーム作成
    df_tanzo_kbn = pd.concat(
        [
            df_tanzo_kbn211,
            df_tanzo_kbn212,
            df_tanzo_kbn213,
            df_tanzo_kbn214,
            df_tanzo_kbn215,
            df_tanzo_kbn216,
            df_tanzo_kbn218,
            df_tanzo_kbn221,
        ],
        axis=0,
        ignore_index=True,
    )
    df_tanzo_kbn = df_tanzo_kbn.rename(
        columns={"所 ": "所属1", "所 .1": "所属2", "所 .2": "所属3"}
    )
    # 切削間接1のデータフレーム作成
    df_sesaku_kbn311 = df_m[(df_m["集計区分－２        "] == 311)]
    df_sesaku_kbn311 = df_sesaku_kbn311.astype("object")
    df_sesaku_kbn311["区分"] = "間接1"
    # 切削間接2のデータフレーム作成
    df_sesaku_kbn312 = df_m[(df_m["集計区分－２        "] == 312)]
    df_sesaku_kbn312 = df_sesaku_kbn312.astype("object")
    df_sesaku_kbn312["区分"] = "間接2"
    # 切削間接4のデータフレーム作成
    df_sesaku_kbn314 = df_m[(df_m["集計区分－２        "] == 314)]
    df_sesaku_kbn314 = df_sesaku_kbn314.astype("object")
    df_sesaku_kbn314["区分"] = "間接4"
    # 切削間接5のデータフレーム作成
    df_sesaku_kbn315 = df_m[(df_m["集計区分－２        "] == 315)]
    df_sesaku_kbn315 = df_sesaku_kbn315.astype("object")
    df_sesaku_kbn315["区分"] = "間接5"
    # 切削間接6のデータフレーム作成
    df_sesaku_kbn316 = df_m[(df_m["集計区分－２        "] == 316)]
    df_sesaku_kbn316 = df_sesaku_kbn316.astype("object")
    df_sesaku_kbn316["区分"] = "間接6"
    # 切削直接1のデータフレーム作成
    df_sesaku_kbn318 = df_m[(df_m["集計区分－２        "] == 318)]
    df_sesaku_kbn318 = df_sesaku_kbn318.astype("object")
    df_sesaku_kbn318["区分"] = "直接1"
    # 切削直接2のデータフレーム作成
    df_sesaku_kbn319 = df_m[(df_m["集計区分－２        "] == 319)]
    df_sesaku_kbn319 = df_sesaku_kbn319.astype("object")
    df_sesaku_kbn319["区分"] = "直接2"
    # 切削直接4のデータフレーム作成
    df_sesaku_kbn321 = df_m[(df_m["集計区分－２        "] == 321)]
    df_sesaku_kbn321 = df_sesaku_kbn321.astype("object")
    df_sesaku_kbn321["区分"] = "直接4"
    # 切削のデータフレーム作成
    df_sesaku_kbn = pd.concat(
        [
            df_sesaku_kbn311,
            df_sesaku_kbn312,
            df_sesaku_kbn314,
            df_sesaku_kbn315,
            df_sesaku_kbn316,
            df_sesaku_kbn318,
            df_sesaku_kbn319,
            df_sesaku_kbn321,
        ],
        axis=0,
        ignore_index=True,
    )
    df_sesaku_kbn = df_sesaku_kbn.rename(
        columns={"所 ": "所属1", "所 .1": "所属2", "所 .2": "所属3"}
    )
    # AC間接1のデータフレーム作成
    df_ac_kbn411 = df_m[(df_m["集計区分－２        "] == 411)]
    df_ac_kbn411 = df_ac_kbn411.astype("object")
    df_ac_kbn411["区分"] = "間接1"
    # AC間接2のデータフレーム作成
    df_ac_kbn412 = df_m[(df_m["集計区分－２        "] == 412)]
    df_ac_kbn412 = df_ac_kbn412.astype("object")
    df_ac_kbn412["区分"] = "間接2"
    # AC間接4のデータフレーム作成
    df_ac_kbn414 = df_m[(df_m["集計区分－２        "] == 414)]
    df_ac_kbn414 = df_ac_kbn414.astype("object")
    df_ac_kbn414["区分"] = "間接4"
    # AC間接5のデータフレーム作成
    df_ac_kbn415 = df_m[(df_m["集計区分－２        "] == 415)]
    df_ac_kbn415 = df_ac_kbn415.astype("object")
    df_ac_kbn415["区分"] = "間接5"
    # AC直接1のデータフレーム作成
    df_ac_kbn418 = df_m[(df_m["集計区分－２        "] == 418)]
    df_ac_kbn418 = df_ac_kbn418.astype("object")
    df_ac_kbn418["区分"] = "直接1"
    # AC直接4のデータフレーム作成
    df_ac_kbn421 = df_m[(df_m["集計区分－２        "] == 421)]
    df_ac_kbn421 = df_ac_kbn421.astype("object")
    df_ac_kbn421["区分"] = "直接4"
    # ACのデータフレーム作成
    df_ac_kbn = pd.concat(
        [
            df_ac_kbn411,
            df_ac_kbn412,
            df_ac_kbn414,
            df_ac_kbn415,
            df_ac_kbn418,
            df_ac_kbn421,
        ],
        axis=0,
        ignore_index=True,
    )
    df_ac_kbn = df_ac_kbn.rename(
        columns={"所 ": "所属1", "所 .1": "所属2", "所 .2": "所属3"}
    )
    # PC間接1のデータフレーム作成
    df_pc_kbn511 = df_m[(df_m["集計区分－２        "] == 511)]
    df_pc_kbn511 = df_pc_kbn511.astype("object")
    df_pc_kbn511["区分"] = "間接1"
    # PC間接2のデータフレーム作成
    df_pc_kbn512 = df_m[(df_m["集計区分－２        "] == 512)]
    df_pc_kbn512 = df_pc_kbn512.astype("object")
    df_pc_kbn512["区分"] = "間接2"
    # PC間接4のデータフレーム作成
    df_pc_kbn514 = df_m[(df_m["集計区分－２        "] == 514)]
    df_pc_kbn514 = df_pc_kbn514.astype("object")
    df_pc_kbn514["区分"] = "間接4"
    # PC間接5のデータフレーム作成
    df_pc_kbn515 = df_m[(df_m["集計区分－２        "] == 515)]
    df_pc_kbn515 = df_pc_kbn515.astype("object")
    df_pc_kbn515["区分"] = "間接5"
    # PC間接6のデータフレーム作成
    df_pc_kbn516 = df_m[(df_m["集計区分－２        "] == 516)]
    df_pc_kbn516 = df_pc_kbn516.astype("object")
    df_pc_kbn516["区分"] = "間接6"
    # PC直接1のデータフレーム作成
    df_pc_kbn518 = df_m[(df_m["集計区分－２        "] == 518)]
    df_pc_kbn518 = df_pc_kbn518.astype("object")
    df_pc_kbn518["区分"] = "直接1"
    # PC直接4のデータフレーム作成
    df_pc_kbn521 = df_m[(df_m["集計区分－２        "] == 521)]
    df_pc_kbn521 = df_pc_kbn521.astype("object")
    df_pc_kbn521["区分"] = "直接4"
    # PCのデータフレーム作成
    df_pc_kbn = pd.concat(
        [
            df_pc_kbn511,
            df_pc_kbn512,
            df_pc_kbn514,
            df_pc_kbn515,
            df_pc_kbn516,
            df_pc_kbn518,
            df_pc_kbn521,
        ],
        axis=0,
        ignore_index=True,
    )
    df_pc_kbn = df_pc_kbn.rename(
        columns={"所 ": "所属1", "所 .1": "所属2", "所 .2": "所属3"}
    )
    # 宮城間接1のデータフレーム作成
    df_miyagi_kbn611 = df_m[(df_m["集計区分－２        "] == 611)]
    df_miyagi_kbn611 = df_miyagi_kbn611.astype("object")
    df_miyagi_kbn611["区分"] = "間接1"
    # 宮城間接2のデータフレーム作成
    df_miyagi_kbn612 = df_m[(df_m["集計区分－２        "] == 612)]
    df_miyagi_kbn612 = df_miyagi_kbn612.astype("object")
    df_miyagi_kbn612["区分"] = "間接2"
    # 宮城間接4のデータフレーム作成
    df_miyagi_kbn614 = df_m[(df_m["集計区分－２        "] == 614)]
    df_miyagi_kbn614 = df_miyagi_kbn614.astype("object")
    df_miyagi_kbn614["区分"] = "間接4"
    # 宮城間接6のデータフレーム作成
    df_miyagi_kbn616 = df_m[(df_m["集計区分－２        "] == 616)]
    df_miyagi_kbn616 = df_miyagi_kbn616.astype("object")
    df_miyagi_kbn616["区分"] = "間接6"
    # 宮城直接1のデータフレーム作成
    df_miyagi_kbn618 = df_m[(df_m["集計区分－２        "] == 618)]
    df_miyagi_kbn618 = df_miyagi_kbn618.astype("object")
    df_miyagi_kbn618["区分"] = "直接1"
    # 宮城のデータフレーム作成
    df_miyagi_kbn = pd.concat(
        [
            df_miyagi_kbn611,
            df_miyagi_kbn612,
            df_miyagi_kbn614,
            df_miyagi_kbn616,
            df_miyagi_kbn618,
        ],
        axis=0,
        ignore_index=True,
    )
    df_miyagi_kbn = df_miyagi_kbn.rename(
        columns={"所 ": "所属1", "所 .1": "所属2", "所 .2": "所属3"}
    )
    # 住設間接2のデータフレーム作成
    df_jyusetu_kbn712 = df_m[(df_m["集計区分－２        "] == 712)]
    df_jyusetu_kbn712 = df_jyusetu_kbn712.astype("object")
    df_jyusetu_kbn712["区分"] = "間接2"
    # 住設間接4のデータフレーム作成
    df_jyusetu_kbn714 = df_m[(df_m["集計区分－２        "] == 714)]
    df_jyusetu_kbn714 = df_jyusetu_kbn714.astype("object")
    df_jyusetu_kbn714["区分"] = "間接4"
    # 住設間接6のデータフレーム作成
    df_jyusetu_kbn716 = df_m[(df_m["集計区分－２        "] == 716)]
    df_jyusetu_kbn716 = df_jyusetu_kbn716.astype("object")
    df_jyusetu_kbn716["区分"] = "間接6"
    # 住設のデータフレーム作成
    df_jyusetu_kbn = pd.concat(
        [df_jyusetu_kbn712, df_jyusetu_kbn714, df_jyusetu_kbn716],
        axis=0,
        ignore_index=True,
    )
    df_jyusetu_kbn = df_jyusetu_kbn.rename(
        columns={"所 ": "所属1", "所 .1": "所属2", "所 .2": "所属3"}
    )
    # 給与データファイルのデータフレーム化
    df_kinsi = pd.read_excel("c:/test/kinsi.XLS")
    df_kinsi = df_kinsi.drop(columns=df_kinsi.columns[[16, 41, 61]], axis=1)
    df_kinsi = df_kinsi.dropna(how="any")
    df_kinsi.drop(
        labels=[
            "所属4",
            "所属5",
            "社員名",
            "支給日[年]",
            "支給日[月]",
            "支給日[日]",
            "【 勤怠 】  ",
            "深夜時間    ",
            "休業日数    ",
            "【 支給 】  ",
            "課税支給額  ",
            "【 控除 】  ",
            "健康保険    ",
            "厚生年金保険",
            "雇用保険    ",
            "社会保険計  ",
            "(内)介護保険",
            "課税対象額  ",
            "所得税      ",
            "住民税      ",
            "財形預金    ",
            "生命保険    ",
            "自動車保険  ",
            "社員積立    ",
            "労金        ",
            "その他控除  ",
            "控除合計額  ",
            "【 補助 】  ",
            "職能等級    ",
            "基本給１    ",
            "基本給２    ",
            "【 合計 】  ",
            "差引支給額  ",
            "銀行振込１  ",
            "銀行振込２  ",
            "銀行振込３  ",
            "現金支給額  ",
        ],
        axis=1,
        inplace=True,
    )
    df_kinsi_s = df_kinsi.set_index(["所属1", "所属2", "所属3"])
    df_kinsi_s = df_kinsi_s.rename(columns={"            .1": "支給額"})

    # 給与データフレームと各区分データフレームを結合する
    # 役員
    df_yakuin_m = pd.merge(df_yakuin_kbn, df_kinsi_s, how="inner", on="社員ｺｰﾄﾞ")
    # 一般管理
    df_ipan_m = pd.merge(df_ipan_kbn, df_kinsi_s, how="inner", on="社員ｺｰﾄﾞ")
    # 鍛造
    df_tanzo_m = pd.merge(df_tanzo_kbn, df_kinsi_s, how="inner", on="社員ｺｰﾄﾞ")
    # 切削
    df_sesaku_m = pd.merge(df_sesaku_kbn, df_kinsi_s, how="inner", on="社員ｺｰﾄﾞ")
    # AC
    df_ac_m = pd.merge(df_ac_kbn, df_kinsi_s, how="inner", on="社員ｺｰﾄﾞ")
    # PC
    df_pc_m = pd.merge(df_pc_kbn, df_kinsi_s, how="inner", on="社員ｺｰﾄﾞ")
    # 宮城
    df_miyagi_m = pd.merge(df_miyagi_kbn, df_kinsi_s, how="inner", on="社員ｺｰﾄﾞ")
    # 住設
    df_jyusetu_m = pd.merge(df_jyusetu_kbn, df_kinsi_s, how="inner", on="社員ｺｰﾄﾞ")

    # ----集計----
    # 一般管理_役員
    df_yakuin_1 = df_yakuin_m.groupby("区分").get_group("役員1")
    df_yakuin_1 = df_yakuin_1.drop("所属1", axis=1)
    member = df_yakuin_1["所属2"] == 0
    df_yakuin_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_yakuin_1["所属3"] > 0
    df_yakuin_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_yakuin_1["出勤日数    "]
    df_yakuin_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_yakuin_1["有休日数    "] * 8
    df_yakuin_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    df_yakuin_1.insert(4, "有休時間在籍者平均", ave_yukyu.round(2))
    abs_time = df_yakuin_1["欠勤日数    "] * 8
    df_yakuin_1.insert(5, "欠勤時間", abs_time)
    work_time = df_yakuin_1["勤務時間    "]
    df_yakuin_1.insert(6, "勤務時間", work_time)
    late_early_time = df_yakuin_1["遅早時間    "]
    df_yakuin_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_yakuin_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    df_yakuin_1.insert(16, "出勤率", work_rate / member)
    real_work_time = real_member * 8
    df_yakuin_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_yakuin_1["ズレ時間    "]
    df_yakuin_1.insert(18, "ズレ時間", zure_time)
    overtime = df_yakuin_1["残業時間    "] + df_yakuin_1["深夜残業    "]
    df_yakuin_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_yakuin_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_yakuin_1["法外休出    "]
    df_yakuin_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_yakuin_1.insert(27, "法定外休出主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_yakuin_1["法定休出    "]
    df_yakuin_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_yakuin_1.insert(29, "法定休出主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_yakuin_1["６０Ｈ超    "]
    df_yakuin_1.insert(30, "時間外60時間超", overtime_60)
    holiday_time = df_yakuin_1["代休時間    "] + df_yakuin_1["深夜代休    "]
    df_yakuin_1.insert(31, "代休時間", holiday_time)
    df_yakuin_1.insert(32, "応援時間", 0)
    total_work_time = df_yakuin_1["勤務時間    "] + df_yakuin_1["残業時間    "]
    df_yakuin_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_yakuin_1["基 本 給    "] + df_yakuin_1["支給額"]
    df_yakuin_1.insert(35, "基本給", basic_salary)
    post_allowance = df_yakuin_1["役職手当    "]
    df_yakuin_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_yakuin_1["営業手当    "]
    df_yakuin_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_yakuin_1["地域手当    "]
    df_yakuin_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_yakuin_1["特殊手当    "]
    df_yakuin_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_yakuin_1["特別技技手当"]
    df_yakuin_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_yakuin_1["調整手当    "]
    df_yakuin_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_yakuin_1["別居手当    "]
    df_yakuin_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_yakuin_1["通勤手当    "]
    df_yakuin_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + spe_tec_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_yakuin_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_yakuin_1["残業手当    "]
    df_yakuin_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_yakuin_1["休出手当    "]
    df_yakuin_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_yakuin_1["深夜勤務手当"]
    df_yakuin_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_yakuin_1["交替時差手当"]
    df_yakuin_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_yakuin_1["休業手当    "]
    df_yakuin_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_yakuin_1["休業控除    "]
    df_yakuin_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_yakuin_1["代 休 他    "]
    df_yakuin_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_yakuin_1["欠勤控除    "] + df_yakuin_1["遅早控除    "]
    df_yakuin_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_yakuin_1["精 算 分    "]
    df_yakuin_1.insert(53, "精算分", settlement)
    sub_total_2 = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
        + closed_deduction
        + compny_leave_etc
        + abs_early_deduction
        + settlement
    )
    df_yakuin_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 + sub_total_2
    df_yakuin_1.insert(55, "総支給額", total)
    df_yakuin_1.insert(56, "応援時間額", 0)
    df_yakuin_1.insert(57, "役員振替", 0)
    df_yakuin_1.insert(58, "部門振替", 0)
    df_yakuin_1.insert(59, "合計", 0)

    # 不要フィールドの削除
    df_yakuin_1 = df_yakuin_1.drop("所属2", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("所属3", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("区分", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("出勤日数    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("有休日数    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("欠勤日数    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("残業時間    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("深夜残業    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("法外休出    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("法定休出    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("代休時間    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("深夜代休    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("６０Ｈ超    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("別居手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("通勤手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("特別技技手当", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("特殊手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("地域手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("営業手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("役職手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("調整手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("基 本 給    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("残業手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("休出手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("深夜勤務手当", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("交替時差手当", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("休業手当    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("代 休 他    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("欠勤控除    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("遅早控除    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("精 算 分    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("支給合計額  ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("休業控除    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("支給額", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("ズレ時間    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("雑費・食事代", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("雑費・衣靴代", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("雑費        ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("受診料・他  ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("雑費・会費等", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("勤務時間    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("遅早時間    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("特休日数    ", axis=1)
    df_yakuin_1 = df_yakuin_1.drop("集計区分－２        ", axis=1)
    df_yakuin_1 = df_yakuin_1.sum()
    # 仮CSVファイルの出力（Excel出力のため)
    df_yakuin_1.to_csv(
        "c:/test/一般管理/A.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 一般管理_一般間接1
    df_ipan_1 = df_ipan_m.groupby("区分").get_group("一般間接1")
    df_ipan_1 = df_ipan_1.drop("所属1", axis=1)
    member = df_ipan_1["所属2"] > 0
    df_ipan_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ipan_1["所属3"] > 0
    df_ipan_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ipan_1["出勤日数    "]
    df_ipan_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_ipan_1["有休日数    "] * 8
    df_ipan_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ipan_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ipan_1["欠勤日数    "] * 8
    df_ipan_1.insert(5, "欠勤時間", abs_time)
    work_time = df_ipan_1["勤務時間    "]
    df_ipan_1.insert(6, "勤務時間", work_time)
    late_early_time = df_ipan_1["遅早時間    "]
    df_ipan_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ipan_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ipan_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ipan_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_ipan_1["ズレ時間    "]
    df_ipan_1.insert(18, "ズレ時間", zure_time)
    overtime = df_ipan_1["残業時間    "] + df_ipan_1["深夜残業    "]
    df_ipan_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ipan_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ipan_1["法外休出    "]
    df_ipan_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ipan_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ipan_1["法定休出    "]
    df_ipan_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ipan_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ipan_1["６０Ｈ超    "]
    df_ipan_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ipan_1["代休時間    "] + df_ipan_1["深夜代休    "]
    df_ipan_1.insert(31, "代休時間", holiday_time)
    df_ipan_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_ipan_1["勤務時間    "]
        + df_ipan_1["残業時間    "]
        + df_ipan_1["法外休出    "]
        + df_ipan_1["法定休出    "]
    )
    df_ipan_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ipan_1["基 本 給    "] + df_ipan_1["支給額"]
    df_ipan_1.insert(35, "基本給", basic_salary)
    post_allowance = df_ipan_1["役職手当    "]
    df_ipan_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ipan_1["営業手当    "]
    df_ipan_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ipan_1["地域手当    "]
    df_ipan_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ipan_1["特殊手当    "]
    df_ipan_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ipan_1["特別技技手当"]
    df_ipan_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ipan_1["調整手当    "]
    df_ipan_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ipan_1["別居手当    "]
    df_ipan_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ipan_1["通勤手当    "]
    df_ipan_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + spe_tec_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    sub_total_1 = sub_total_1.sum()
    sub_total_1 = sub_total_1 / member
    sub_total_1 = sub_total_1.astype(int)
    df_ipan_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ipan_1["残業手当    "]
    df_ipan_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ipan_1["休出手当    "]
    df_ipan_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ipan_1["深夜勤務手当"]
    df_ipan_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ipan_1["交替時差手当"]
    df_ipan_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ipan_1["休業手当    "]
    df_ipan_1.insert(49, "休業手当", closed_allowance)
    teate_total = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
    )
    teate_total = teate_total.sum()
    closed_deduction = df_ipan_1["休業控除    "]
    df_ipan_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ipan_1["代 休 他    "]
    df_ipan_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ipan_1["欠勤控除    "] + df_ipan_1["遅早控除    "]
    df_ipan_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    kojyo_total = closed_deduction + compny_leave_etc + abs_early_deduction
    kojyo_total = kojyo_total.sum()
    settlement = df_ipan_1["精 算 分    "]
    df_ipan_1.insert(53, "精算分", settlement)
    sub_total_2 = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
        + closed_deduction
        + compny_leave_etc
        + abs_early_deduction
        + settlement
    )
    df_ipan_1.insert(54, "小計 2", sub_total_2)
    total = (sub_total_1 + sub_total_2) - kojyo_total
    df_ipan_1.insert(55, "総支給額", total)
    df_ipan_1.insert(56, "応援時間額", 0)
    df_ipan_1.insert(57, "役員振替", 0)
    df_ipan_1.insert(58, "部門振替", 0)
    df_ipan_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ipan_1 = df_ipan_1.drop("所属2", axis=1)
    df_ipan_1 = df_ipan_1.drop("所属3", axis=1)
    df_ipan_1 = df_ipan_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_ipan_1 = df_ipan_1.drop("区分", axis=1)
    df_ipan_1 = df_ipan_1.drop("出勤日数    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("有休日数    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("欠勤日数    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("残業時間    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("深夜残業    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("法外休出    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("法定休出    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("代休時間    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("深夜代休    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("６０Ｈ超    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("別居手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("通勤手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("特別技技手当", axis=1)
    df_ipan_1 = df_ipan_1.drop("特殊手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("地域手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("営業手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("役職手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("調整手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("基 本 給    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("残業手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("休出手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("深夜勤務手当", axis=1)
    df_ipan_1 = df_ipan_1.drop("交替時差手当", axis=1)
    df_ipan_1 = df_ipan_1.drop("休業手当    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("代 休 他    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("欠勤控除    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("遅早控除    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("精 算 分    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("支給合計額  ", axis=1)
    df_ipan_1 = df_ipan_1.drop("休業控除    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("支給額", axis=1)
    df_ipan_1 = df_ipan_1.drop("ズレ時間    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("雑費・食事代", axis=1)
    df_ipan_1 = df_ipan_1.drop("雑費・衣靴代", axis=1)
    df_ipan_1 = df_ipan_1.drop("雑費        ", axis=1)
    df_ipan_1 = df_ipan_1.drop("受診料・他  ", axis=1)
    df_ipan_1 = df_ipan_1.drop("雑費・会費等", axis=1)
    df_ipan_1 = df_ipan_1.drop("勤務時間    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("遅早時間    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("特休日数    ", axis=1)
    df_ipan_1 = df_ipan_1.drop("集計区分－２        ", axis=1)
    df_ipan_1 = df_ipan_1.sum()
    df_ipan_1.to_csv(
        "c:/test/一般管理/B.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 一般管理_一般間接2
    df_ipan_2 = df_ipan_m.groupby("区分").get_group("一般間接2")
    df_ipan_2 = df_ipan_2.drop("所属1", axis=1)
    member = df_ipan_2["所属2"] > 0
    df_ipan_2.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ipan_2["所属3"] > 0
    df_ipan_2.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ipan_2["出勤日数    "]
    df_ipan_2.insert(2, "実在籍者", real_member)
    time_yukyu = df_ipan_2["有休日数    "] * 8
    df_ipan_2.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ipan_2.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ipan_2["欠勤日数    "] * 8
    df_ipan_2.insert(5, "欠勤時間", abs_time)
    work_time = df_ipan_2["勤務時間    "]
    df_ipan_2.insert(6, "勤務時間", work_time)
    late_early_time = df_ipan_2["遅早時間    "]
    df_ipan_2.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ipan_2["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ipan_2.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ipan_2.insert(17, "実労働時間", real_work_time)
    zure_time = df_ipan_2["ズレ時間    "]
    df_ipan_2.insert(18, "ズレ時間", zure_time)
    overtime = df_ipan_2["残業時間    "] + df_ipan_2["深夜残業    "]
    df_ipan_2.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ipan_2.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ipan_2["法外休出    "]
    df_ipan_2.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ipan_2.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ipan_2["法定休出    "]
    df_ipan_2.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ipan_2.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ipan_2["６０Ｈ超    "]
    df_ipan_2.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ipan_2["代休時間    "] + df_ipan_2["深夜代休    "]
    df_ipan_2.insert(31, "代休時間", holiday_time)
    df_ipan_2.insert(32, "応援時間", 0)
    total_work_time = (
        df_ipan_2["勤務時間    "]
        + df_ipan_2["残業時間    "]
        + df_ipan_2["法外休出    "]
        + df_ipan_2["法定休出    "]
    )
    df_ipan_2.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ipan_2["基 本 給    "] + df_ipan_2["支給額"]
    df_ipan_2.insert(35, "基本給", basic_salary)
    post_allowance = df_ipan_2["役職手当    "]
    df_ipan_2.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ipan_2["営業手当    "]
    df_ipan_2.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ipan_2["地域手当    "]
    df_ipan_2.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ipan_2["特殊手当    "]
    df_ipan_2.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ipan_2["特別技技手当"]
    df_ipan_2.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ipan_2["調整手当    "]
    df_ipan_2.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ipan_2["別居手当    "]
    df_ipan_2.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ipan_2["通勤手当    "]
    df_ipan_2.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + spe_tec_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    sub_total_1 = sub_total_1.sum()
    sub_total_1 = sub_total_1 / member
    sub_total_1 = sub_total_1.astype(int)
    df_ipan_2.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ipan_2["残業手当    "]
    df_ipan_2.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ipan_2["休出手当    "]
    df_ipan_2.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ipan_2["深夜勤務手当"]
    df_ipan_2.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ipan_2["交替時差手当"]
    df_ipan_2.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ipan_2["休業手当    "]
    df_ipan_2.insert(49, "休業手当", closed_allowance)
    teate_total = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
    )
    teate_total = teate_total.sum()
    closed_deduction = df_ipan_2["休業控除    "]
    df_ipan_2.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ipan_2["代 休 他    "]
    df_ipan_2.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ipan_2["欠勤控除    "] + df_ipan_2["遅早控除    "]
    df_ipan_2.insert(52, "欠勤・遅早控除", abs_early_deduction)
    kojyo_total = closed_deduction + compny_leave_etc + abs_early_deduction
    kojyo_total = kojyo_total.sum()
    settlement = df_ipan_2["精 算 分    "]
    df_ipan_2.insert(53, "精算分", settlement)
    sub_total_2 = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
        + closed_deduction
        + compny_leave_etc
        + abs_early_deduction
        + settlement
    )
    df_ipan_2.insert(54, "小計 2", sub_total_2)
    total = (sub_total_1 + sub_total_2) - kojyo_total
    df_ipan_2.insert(55, "総支給額", total)
    df_ipan_2.insert(56, "応援時間額", 0)
    df_ipan_2.insert(57, "役員振替", 0)
    df_ipan_2.insert(58, "部門振替", 0)
    df_ipan_2.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ipan_2 = df_ipan_2.drop("所属2", axis=1)
    df_ipan_2 = df_ipan_2.drop("所属3", axis=1)
    df_ipan_2 = df_ipan_2.drop("社員ｺｰﾄﾞ", axis=1)
    df_ipan_2 = df_ipan_2.drop("区分", axis=1)
    df_ipan_2 = df_ipan_2.drop("出勤日数    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("有休日数    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("欠勤日数    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("残業時間    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("深夜残業    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("法外休出    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("法定休出    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("代休時間    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("深夜代休    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("６０Ｈ超    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("別居手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("通勤手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("特別技技手当", axis=1)
    df_ipan_2 = df_ipan_2.drop("特殊手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("地域手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("営業手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("役職手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("調整手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("基 本 給    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("残業手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("休出手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("深夜勤務手当", axis=1)
    df_ipan_2 = df_ipan_2.drop("交替時差手当", axis=1)
    df_ipan_2 = df_ipan_2.drop("休業手当    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("代 休 他    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("欠勤控除    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("遅早控除    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("精 算 分    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("支給合計額  ", axis=1)
    df_ipan_2 = df_ipan_2.drop("休業控除    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("支給額", axis=1)
    df_ipan_2 = df_ipan_2.drop("ズレ時間    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("雑費・食事代", axis=1)
    df_ipan_2 = df_ipan_2.drop("雑費・衣靴代", axis=1)
    df_ipan_2 = df_ipan_2.drop("雑費        ", axis=1)
    df_ipan_2 = df_ipan_2.drop("受診料・他  ", axis=1)
    df_ipan_2 = df_ipan_2.drop("雑費・会費等", axis=1)
    df_ipan_2 = df_ipan_2.drop("勤務時間    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("遅早時間    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("特休日数    ", axis=1)
    df_ipan_2 = df_ipan_2.drop("集計区分－２        ", axis=1)
    df_ipan_2 = df_ipan_2.sum()
    df_ipan_2.to_csv(
        "c:/test/一般管理/C.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 一般管理_一般間接3
    df_ipan_3 = df_ipan_m.groupby("区分").get_group("一般間接3")
    df_ipan_3 = df_ipan_3.drop("所属1", axis=1)
    member = df_ipan_3["所属2"] > 0
    df_ipan_3.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ipan_3["所属3"] > 0
    df_ipan_3.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ipan_3["出勤日数    "]
    df_ipan_3.insert(2, "実在籍者", real_member)
    time_yukyu = df_ipan_3["有休日数    "] * 8
    df_ipan_3.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ipan_3.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ipan_3["欠勤日数    "] * 8
    df_ipan_3.insert(5, "欠勤時間", abs_time)
    work_time = df_ipan_3["勤務時間    "]
    df_ipan_3.insert(6, "勤務時間", work_time)
    late_early_time = df_ipan_3["遅早時間    "]
    df_ipan_3.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ipan_3["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ipan_3.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ipan_3.insert(17, "実労働時間", real_work_time)
    zure_time = df_ipan_3["ズレ時間    "]
    df_ipan_3.insert(18, "ズレ時間", zure_time)
    overtime = df_ipan_3["残業時間    "] + df_ipan_3["深夜残業    "]
    df_ipan_3.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ipan_3.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ipan_3["法外休出    "]
    df_ipan_3.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ipan_3.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ipan_3["法定休出    "]
    df_ipan_3.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ipan_3.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ipan_3["６０Ｈ超    "]
    df_ipan_3.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ipan_3["代休時間    "] + df_ipan_3["深夜代休    "]
    df_ipan_3.insert(31, "代休時間", holiday_time)
    df_ipan_3.insert(32, "応援時間", 0)
    total_work_time = (
        df_ipan_3["勤務時間    "]
        + df_ipan_3["残業時間    "]
        + df_ipan_3["法外休出    "]
        + df_ipan_3["法定休出    "]
    )
    df_ipan_3.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ipan_3["基 本 給    "] + df_ipan_3["支給額"]
    df_ipan_3.insert(35, "基本給", basic_salary)
    post_allowance = df_ipan_3["役職手当    "]
    df_ipan_3.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ipan_3["営業手当    "]
    df_ipan_3.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ipan_3["地域手当    "]
    df_ipan_3.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ipan_3["特殊手当    "]
    df_ipan_3.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ipan_3["特別技技手当"]
    df_ipan_3.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ipan_3["調整手当    "]
    df_ipan_3.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ipan_3["別居手当    "]
    df_ipan_3.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ipan_3["通勤手当    "]
    df_ipan_3.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + spe_tec_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    sub_total_1 = sub_total_1.sum()
    sub_total_1 = sub_total_1 / member
    sub_total_1 = sub_total_1.astype(int)
    df_ipan_3.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ipan_3["残業手当    "]
    df_ipan_3.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ipan_3["休出手当    "]
    df_ipan_3.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ipan_3["深夜勤務手当"]
    df_ipan_3.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ipan_3["交替時差手当"]
    df_ipan_3.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ipan_3["休業手当    "]
    df_ipan_3.insert(49, "休業手当", closed_allowance)
    teate_total = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
    )
    teate_total = teate_total.sum()
    closed_deduction = df_ipan_3["休業控除    "]
    df_ipan_3.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ipan_3["代 休 他    "]
    df_ipan_3.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ipan_3["欠勤控除    "] + df_ipan_3["遅早控除    "]
    df_ipan_3.insert(52, "欠勤・遅早控除", abs_early_deduction)
    kojyo_total = closed_deduction + compny_leave_etc + abs_early_deduction
    kojyo_total = kojyo_total.sum()
    settlement = df_ipan_3["精 算 分    "]
    df_ipan_3.insert(53, "精算分", settlement)
    sub_total_2 = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
        + closed_deduction
        + compny_leave_etc
        + abs_early_deduction
        + settlement
    )
    df_ipan_3.insert(54, "小計 2", sub_total_2)
    total = (sub_total_1 + sub_total_2) - kojyo_total
    df_ipan_3.insert(55, "総支給額", total)
    df_ipan_3.insert(56, "応援時間額", 0)
    df_ipan_3.insert(57, "役員振替", 0)
    df_ipan_3.insert(58, "部門振替", 0)
    df_ipan_3.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ipan_3 = df_ipan_3.drop("所属2", axis=1)
    df_ipan_3 = df_ipan_3.drop("所属3", axis=1)
    df_ipan_3 = df_ipan_3.drop("社員ｺｰﾄﾞ", axis=1)
    df_ipan_3 = df_ipan_3.drop("区分", axis=1)
    df_ipan_3 = df_ipan_3.drop("出勤日数    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("有休日数    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("欠勤日数    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("残業時間    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("深夜残業    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("法外休出    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("法定休出    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("代休時間    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("深夜代休    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("６０Ｈ超    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("別居手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("通勤手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("特別技技手当", axis=1)
    df_ipan_3 = df_ipan_3.drop("特殊手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("地域手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("営業手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("役職手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("調整手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("基 本 給    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("残業手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("休出手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("深夜勤務手当", axis=1)
    df_ipan_3 = df_ipan_3.drop("交替時差手当", axis=1)
    df_ipan_3 = df_ipan_3.drop("休業手当    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("代 休 他    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("欠勤控除    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("遅早控除    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("精 算 分    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("支給合計額  ", axis=1)
    df_ipan_3 = df_ipan_3.drop("休業控除    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("支給額", axis=1)
    df_ipan_3 = df_ipan_3.drop("ズレ時間    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("雑費・食事代", axis=1)
    df_ipan_3 = df_ipan_3.drop("雑費・衣靴代", axis=1)
    df_ipan_3 = df_ipan_3.drop("雑費        ", axis=1)
    df_ipan_3 = df_ipan_3.drop("受診料・他  ", axis=1)
    df_ipan_3 = df_ipan_3.drop("雑費・会費等", axis=1)
    df_ipan_3 = df_ipan_3.drop("勤務時間    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("遅早時間    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("特休日数    ", axis=1)
    df_ipan_3 = df_ipan_3.drop("集計区分－２        ", axis=1)
    df_ipan_3 = df_ipan_3.sum()
    df_ipan_3.to_csv(
        "c:/test/一般管理/D.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 一般管理_一般間接6
    df_ipan_6 = df_ipan_m.groupby("区分").get_group("一般間接6")
    df_ipan_6 = df_ipan_6.drop("所属1", axis=1)
    member = df_ipan_6["所属2"] > 0
    df_ipan_6.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ipan_6["所属3"] > 0
    df_ipan_6.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ipan_6["出勤日数    "]
    df_ipan_6.insert(2, "実在籍者", real_member)
    time_yukyu = df_ipan_6["有休日数    "] * 8
    df_ipan_6.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ipan_6.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ipan_6["欠勤日数    "] * 8
    df_ipan_6.insert(5, "欠勤時間", abs_time)
    work_time = df_ipan_6["勤務時間    "]
    df_ipan_6.insert(6, "勤務時間", work_time)
    late_early_time = df_ipan_6["遅早時間    "]
    df_ipan_6.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ipan_6["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ipan_6.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ipan_6.insert(17, "実労働時間", real_work_time)
    zure_time = df_ipan_6["ズレ時間    "]
    df_ipan_6.insert(18, "ズレ時間", zure_time)
    overtime = df_ipan_6["残業時間    "] + df_ipan_6["深夜残業    "]
    df_ipan_6.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ipan_6.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ipan_6["法外休出    "]
    df_ipan_6.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ipan_6.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ipan_6["法定休出    "]
    df_ipan_6.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ipan_6.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ipan_6["６０Ｈ超    "]
    df_ipan_6.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ipan_6["代休時間    "] + df_ipan_6["深夜代休    "]
    df_ipan_6.insert(31, "代休時間", holiday_time)
    df_ipan_6.insert(32, "応援時間", 0)
    total_work_time = (
        df_ipan_6["勤務時間    "]
        + df_ipan_6["残業時間    "]
        + df_ipan_6["法外休出    "]
        + df_ipan_6["法定休出    "]
    )
    df_ipan_6.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ipan_6["基 本 給    "] + df_ipan_6["支給額"]
    df_ipan_6.insert(35, "基本給", basic_salary)
    post_allowance = df_ipan_6["役職手当    "]
    df_ipan_6.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ipan_6["営業手当    "]
    df_ipan_6.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ipan_6["地域手当    "]
    df_ipan_6.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ipan_6["特殊手当    "]
    df_ipan_6.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ipan_6["特別技技手当"]
    df_ipan_6.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ipan_6["調整手当    "]
    df_ipan_6.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ipan_6["別居手当    "]
    df_ipan_6.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ipan_6["通勤手当    "]
    df_ipan_6.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + spe_tec_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    sub_total_1 = sub_total_1.sum()
    sub_total_1 = sub_total_1 / member
    sub_total_1 = sub_total_1.astype(int)
    df_ipan_6.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ipan_6["残業手当    "]
    df_ipan_6.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ipan_6["休出手当    "]
    df_ipan_6.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ipan_6["深夜勤務手当"]
    df_ipan_6.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ipan_6["交替時差手当"]
    df_ipan_6.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ipan_6["休業手当    "]
    df_ipan_6.insert(49, "休業手当", closed_allowance)
    teate_total = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
    )
    teate_total = teate_total.sum()
    closed_deduction = df_ipan_6["休業控除    "]
    df_ipan_6.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ipan_6["代 休 他    "]
    df_ipan_6.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ipan_6["欠勤控除    "] + df_ipan_6["遅早控除    "]
    df_ipan_6.insert(52, "欠勤・遅早控除", abs_early_deduction)
    kojyo_total = closed_deduction + compny_leave_etc + abs_early_deduction
    kojyo_total = kojyo_total.sum()
    settlement = df_ipan_6["精 算 分    "]
    df_ipan_6.insert(53, "精算分", settlement)
    sub_total_2 = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
        + closed_deduction
        + compny_leave_etc
        + abs_early_deduction
        + settlement
    )
    df_ipan_6.insert(54, "小計 2", sub_total_2)
    total = (sub_total_1 + sub_total_2) - kojyo_total
    df_ipan_6.insert(55, "総支給額", total)
    df_ipan_6.insert(56, "応援時間額", 0)
    df_ipan_6.insert(57, "役員振替", 0)
    df_ipan_6.insert(58, "部門振替", 0)
    df_ipan_6.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ipan_6 = df_ipan_6.drop("所属2", axis=1)
    df_ipan_6 = df_ipan_6.drop("所属3", axis=1)
    df_ipan_6 = df_ipan_6.drop("社員ｺｰﾄﾞ", axis=1)
    df_ipan_6 = df_ipan_6.drop("区分", axis=1)
    df_ipan_6 = df_ipan_6.drop("出勤日数    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("有休日数    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("欠勤日数    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("残業時間    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("深夜残業    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("法外休出    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("法定休出    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("代休時間    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("深夜代休    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("６０Ｈ超    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("別居手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("通勤手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("特別技技手当", axis=1)
    df_ipan_6 = df_ipan_6.drop("特殊手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("地域手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("営業手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("役職手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("調整手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("基 本 給    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("残業手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("休出手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("深夜勤務手当", axis=1)
    df_ipan_6 = df_ipan_6.drop("交替時差手当", axis=1)
    df_ipan_6 = df_ipan_6.drop("休業手当    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("代 休 他    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("欠勤控除    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("遅早控除    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("精 算 分    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("支給合計額  ", axis=1)
    df_ipan_6 = df_ipan_6.drop("休業控除    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("支給額", axis=1)
    df_ipan_6 = df_ipan_6.drop("ズレ時間    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("雑費・食事代", axis=1)
    df_ipan_6 = df_ipan_6.drop("雑費・衣靴代", axis=1)
    df_ipan_6 = df_ipan_6.drop("雑費        ", axis=1)
    df_ipan_6 = df_ipan_6.drop("受診料・他  ", axis=1)
    df_ipan_6 = df_ipan_6.drop("雑費・会費等", axis=1)
    df_ipan_6 = df_ipan_6.drop("勤務時間    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("遅早時間    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("特休日数    ", axis=1)
    df_ipan_6 = df_ipan_6.drop("集計区分－２        ", axis=1)
    df_ipan_6 = df_ipan_6.sum()
    df_ipan_6.to_csv(
        "c:/test/一般管理/E.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 一般管理_一般販売1
    df_ipan_h_1 = df_ipan_m.groupby("区分").get_group("一般販売1")
    df_ipan_h_1 = df_ipan_h_1.drop("所属1", axis=1)
    member = df_ipan_h_1["所属2"] > 0
    df_ipan_h_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ipan_h_1["所属3"] > 0
    df_ipan_h_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ipan_h_1["出勤日数    "]
    df_ipan_h_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_ipan_h_1["有休日数    "] * 8
    df_ipan_h_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ipan_h_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ipan_h_1["欠勤日数    "] * 8
    df_ipan_h_1.insert(5, "欠勤時間", abs_time)
    work_time = df_ipan_h_1["勤務時間    "]
    df_ipan_h_1.insert(6, "勤務時間", work_time)
    late_early_time = df_ipan_h_1["遅早時間    "]
    df_ipan_h_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ipan_h_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ipan_h_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ipan_h_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_ipan_h_1["ズレ時間    "]
    df_ipan_h_1.insert(18, "ズレ時間", zure_time)
    overtime = df_ipan_h_1["残業時間    "] + df_ipan_h_1["深夜残業    "]
    df_ipan_h_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ipan_h_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ipan_h_1["法外休出    "]
    df_ipan_h_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ipan_h_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ipan_h_1["法定休出    "]
    df_ipan_h_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ipan_h_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ipan_h_1["６０Ｈ超    "]
    df_ipan_h_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ipan_h_1["代休時間    "] + df_ipan_h_1["深夜代休    "]
    df_ipan_h_1.insert(31, "代休時間", holiday_time)
    df_ipan_h_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_ipan_h_1["勤務時間    "]
        + df_ipan_h_1["残業時間    "]
        + df_ipan_h_1["法外休出    "]
        + df_ipan_h_1["法定休出    "]
    )
    df_ipan_h_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ipan_h_1["基 本 給    "] + df_ipan_h_1["支給額"]
    df_ipan_h_1.insert(35, "基本給", basic_salary)
    post_allowance = df_ipan_h_1["役職手当    "]
    df_ipan_h_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ipan_h_1["営業手当    "]
    df_ipan_h_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ipan_h_1["地域手当    "]
    df_ipan_h_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ipan_h_1["特殊手当    "]
    df_ipan_h_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ipan_h_1["特別技技手当"]
    df_ipan_h_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ipan_h_1["調整手当    "]
    df_ipan_h_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ipan_h_1["別居手当    "]
    df_ipan_h_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ipan_h_1["通勤手当    "]
    df_ipan_h_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + spe_tec_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    sub_total_1 = sub_total_1.sum()
    sub_total_1 = sub_total_1 / member
    sub_total_1 = sub_total_1.astype(int)
    df_ipan_h_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ipan_h_1["残業手当    "]
    df_ipan_h_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ipan_h_1["休出手当    "]
    df_ipan_h_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ipan_h_1["深夜勤務手当"]
    df_ipan_h_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ipan_h_1["交替時差手当"]
    df_ipan_h_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ipan_h_1["休業手当    "]
    df_ipan_h_1.insert(49, "休業手当", closed_allowance)
    teate_total = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
    )
    teate_total = teate_total.sum()
    closed_deduction = df_ipan_h_1["休業控除    "]
    df_ipan_h_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ipan_h_1["代 休 他    "]
    df_ipan_h_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ipan_h_1["欠勤控除    "] + df_ipan_h_1["遅早控除    "]
    df_ipan_h_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    kojyo_total = closed_deduction + compny_leave_etc + abs_early_deduction
    kojyo_total = kojyo_total.sum()
    settlement = df_ipan_h_1["精 算 分    "]
    df_ipan_h_1.insert(53, "精算分", settlement)
    sub_total_2 = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
        + closed_deduction
        + compny_leave_etc
        + abs_early_deduction
        + settlement
    )
    df_ipan_h_1.insert(54, "小計 2", sub_total_2)
    total = (sub_total_1 + sub_total_2) - kojyo_total
    df_ipan_h_1.insert(55, "総支給額", total)
    df_ipan_h_1.insert(56, "応援時間額", 0)
    df_ipan_h_1.insert(57, "役員振替", 0)
    df_ipan_h_1.insert(58, "部門振替", 0)
    df_ipan_h_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ipan_h_1 = df_ipan_h_1.drop("所属2", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("所属3", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("区分", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("出勤日数    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("有休日数    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("欠勤日数    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("残業時間    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("深夜残業    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("法外休出    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("法定休出    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("代休時間    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("深夜代休    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("６０Ｈ超    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("別居手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("通勤手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("特別技技手当", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("特殊手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("地域手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("営業手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("役職手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("調整手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("基 本 給    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("残業手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("休出手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("深夜勤務手当", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("交替時差手当", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("休業手当    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("代 休 他    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("欠勤控除    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("遅早控除    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("精 算 分    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("支給合計額  ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("休業控除    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("支給額", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("ズレ時間    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("雑費・食事代", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("雑費・衣靴代", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("雑費        ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("受診料・他  ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("雑費・会費等", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("勤務時間    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("遅早時間    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("特休日数    ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.drop("集計区分－２        ", axis=1)
    df_ipan_h_1 = df_ipan_h_1.sum()
    df_ipan_h_1.to_csv(
        "c:/test/一般管理/F.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 一般管理_一般販売2
    df_ipan_h_2 = df_ipan_m.groupby("区分").get_group("一般販売2")
    df_ipan_h_2 = df_ipan_h_2.drop("所属1", axis=1)
    member = df_ipan_h_2["所属2"] > 0
    df_ipan_h_2.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ipan_h_2["所属3"] > 0
    df_ipan_h_2.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ipan_h_2["出勤日数    "]
    df_ipan_h_2.insert(2, "実在籍者", real_member)
    time_yukyu = df_ipan_h_2["有休日数    "] * 8
    df_ipan_h_2.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ipan_h_2.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ipan_h_2["欠勤日数    "] * 8
    df_ipan_h_2.insert(5, "欠勤時間", abs_time)
    work_time = df_ipan_h_2["勤務時間    "]
    df_ipan_h_2.insert(6, "勤務時間", work_time)
    late_early_time = df_ipan_h_2["遅早時間    "]
    df_ipan_h_2.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ipan_h_2["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ipan_h_2.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ipan_h_2.insert(17, "実労働時間", real_work_time)
    zure_time = df_ipan_h_2["ズレ時間    "]
    df_ipan_h_2.insert(18, "ズレ時間", zure_time)
    overtime = df_ipan_h_2["残業時間    "] + df_ipan_h_2["深夜残業    "]
    df_ipan_h_2.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ipan_h_2.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ipan_h_2["法外休出    "]
    df_ipan_h_2.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ipan_h_2.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ipan_h_2["法定休出    "]
    df_ipan_h_2.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ipan_h_2.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ipan_h_2["６０Ｈ超    "]
    df_ipan_h_2.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ipan_h_2["代休時間    "] + df_ipan_h_2["深夜代休    "]
    df_ipan_h_2.insert(31, "代休時間", holiday_time)
    df_ipan_h_2.insert(32, "応援時間", 0)
    total_work_time = (
        df_ipan_h_2["勤務時間    "]
        + df_ipan_h_2["残業時間    "]
        + df_ipan_h_2["法外休出    "]
        + df_ipan_h_2["法定休出    "]
    )
    df_ipan_h_2.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ipan_h_2["基 本 給    "] + df_ipan_h_2["支給額"]
    df_ipan_h_2.insert(35, "基本給", basic_salary)
    post_allowance = df_ipan_h_2["役職手当    "]
    df_ipan_h_2.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ipan_h_2["営業手当    "]
    df_ipan_h_2.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ipan_h_2["地域手当    "]
    df_ipan_h_2.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ipan_h_2["特殊手当    "]
    df_ipan_h_2.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ipan_h_2["特別技技手当"]
    df_ipan_h_2.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ipan_h_2["調整手当    "]
    df_ipan_h_2.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ipan_h_2["別居手当    "]
    df_ipan_h_2.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ipan_h_2["通勤手当    "]
    df_ipan_h_2.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + spe_tec_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    sub_total_1 = sub_total_1.sum()
    sub_total_1 = sub_total_1 / member
    sub_total_1 = sub_total_1.astype(int)
    df_ipan_h_2.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ipan_h_2["残業手当    "]
    df_ipan_h_2.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ipan_h_2["休出手当    "]
    df_ipan_h_2.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ipan_h_2["深夜勤務手当"]
    df_ipan_h_2.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ipan_h_2["交替時差手当"]
    df_ipan_h_2.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ipan_h_2["休業手当    "]
    df_ipan_h_2.insert(49, "休業手当", closed_allowance)
    teate_total = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
    )
    teate_total = teate_total.sum()
    closed_deduction = df_ipan_h_2["休業控除    "]
    df_ipan_h_2.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ipan_h_2["代 休 他    "]
    df_ipan_h_2.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ipan_h_2["欠勤控除    "] + df_ipan_h_2["遅早控除    "]
    df_ipan_h_2.insert(52, "欠勤・遅早控除", abs_early_deduction)
    kojyo_total = closed_deduction + compny_leave_etc + abs_early_deduction
    kojyo_total = kojyo_total.sum()
    settlement = df_ipan_h_2["精 算 分    "]
    df_ipan_h_2.insert(53, "精算分", settlement)
    sub_total_2 = (
        overtime_allowance
        + vacation_allowance
        + night_work_allowance
        + time_difference_allowance
        + closed_allowance
        + closed_deduction
        + compny_leave_etc
        + abs_early_deduction
        + settlement
    )
    df_ipan_h_2.insert(54, "小計 2", sub_total_2)
    total = (sub_total_1 + sub_total_2) - kojyo_total
    df_ipan_h_2.insert(55, "総支給額", total)
    df_ipan_h_2.insert(56, "応援時間額", 0)
    df_ipan_h_2.insert(57, "役員振替", 0)
    df_ipan_h_2.insert(58, "部門振替", 0)
    df_ipan_h_2.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ipan_h_2 = df_ipan_h_2.drop("所属2", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("所属3", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("社員ｺｰﾄﾞ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("区分", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("出勤日数    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("有休日数    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("欠勤日数    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("残業時間    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("深夜残業    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("法外休出    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("法定休出    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("代休時間    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("深夜代休    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("６０Ｈ超    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("別居手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("通勤手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("特別技技手当", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("特殊手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("地域手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("営業手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("役職手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("調整手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("基 本 給    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("残業手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("休出手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("深夜勤務手当", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("交替時差手当", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("休業手当    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("代 休 他    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("欠勤控除    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("遅早控除    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("精 算 分    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("支給合計額  ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("休業控除    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("支給額", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("ズレ時間    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("雑費・食事代", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("雑費・衣靴代", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("雑費        ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("受診料・他  ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("雑費・会費等", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("勤務時間    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("遅早時間    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("特休日数    ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.drop("集計区分－２        ", axis=1)
    df_ipan_h_2 = df_ipan_h_2.sum()
    df_ipan_h_2.to_csv(
        "c:/test/一般管理/G.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 鍛造_間接1
    df_tanzo_k_1 = df_tanzo_m.groupby("区分").get_group("間接1")
    df_tanzo_k_1 = df_tanzo_k_1.drop("所属1", axis=1)
    member = df_tanzo_k_1["所属2"] > 0
    df_tanzo_k_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_tanzo_k_1["所属3"] > 0
    df_tanzo_k_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_tanzo_k_1["出勤日数    "]
    df_tanzo_k_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_tanzo_k_1["有休日数    "] * 8
    df_tanzo_k_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_tanzo_k_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_tanzo_k_1["欠勤日数    "] * 8
    df_tanzo_k_1.insert(5, "欠勤時間", abs_time)
    work_time = df_tanzo_k_1["勤務時間    "]
    df_tanzo_k_1.insert(6, "勤務時間", work_time)
    late_early_time = df_tanzo_k_1["遅早時間    "]
    df_tanzo_k_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_tanzo_k_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_tanzo_k_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_tanzo_k_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_tanzo_k_1["ズレ時間    "]
    df_tanzo_k_1.insert(18, "ズレ時間", zure_time)
    overtime = df_tanzo_k_1["残業時間    "] + df_tanzo_k_1["深夜残業    "]
    df_tanzo_k_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_tanzo_k_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_tanzo_k_1["法外休出    "]
    df_tanzo_k_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_tanzo_k_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_tanzo_k_1["法定休出    "]
    df_tanzo_k_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_tanzo_k_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_tanzo_k_1["６０Ｈ超    "]
    df_tanzo_k_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_tanzo_k_1["代休時間    "] + df_tanzo_k_1["深夜代休    "]
    df_tanzo_k_1.insert(31, "代休時間", holiday_time)
    df_tanzo_k_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_tanzo_k_1["勤務時間    "]
        + df_tanzo_k_1["残業時間    "]
        + df_tanzo_k_1["法外休出    "]
        + df_tanzo_k_1["法定休出    "]
    )
    df_tanzo_k_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_tanzo_k_1["基 本 給    "] + df_tanzo_k_1["支給額"]
    df_tanzo_k_1.insert(35, "基本給", basic_salary)
    post_allowance = df_tanzo_k_1["役職手当    "]
    df_tanzo_k_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_tanzo_k_1["営業手当    "]
    df_tanzo_k_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_tanzo_k_1["地域手当    "]
    df_tanzo_k_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_tanzo_k_1["特殊手当    "]
    df_tanzo_k_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_tanzo_k_1["特別技技手当"]
    df_tanzo_k_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_tanzo_k_1["調整手当    "]
    df_tanzo_k_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_tanzo_k_1["別居手当    "]
    df_tanzo_k_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_tanzo_k_1["通勤手当    "]
    df_tanzo_k_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_tanzo_k_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_tanzo_k_1["残業手当    "]
    df_tanzo_k_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_tanzo_k_1["休出手当    "]
    df_tanzo_k_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_tanzo_k_1["深夜勤務手当"]
    df_tanzo_k_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_tanzo_k_1["交替時差手当"]
    df_tanzo_k_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_tanzo_k_1["休業手当    "]
    df_tanzo_k_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_tanzo_k_1["休業控除    "]
    df_tanzo_k_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_tanzo_k_1["代 休 他    "]
    df_tanzo_k_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_tanzo_k_1["欠勤控除    "] + df_tanzo_k_1["遅早控除    "]
    df_tanzo_k_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_tanzo_k_1["精 算 分    "]
    df_tanzo_k_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_tanzo_k_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_tanzo_k_1.insert(55, "総支給額", total)
    df_tanzo_k_1.insert(56, "応援時間額", 0)
    df_tanzo_k_1.insert(57, "役員振替", 0)
    df_tanzo_k_1.insert(58, "部門振替", 0)
    df_tanzo_k_1.insert(59, "合計", 0)
    # フィールドの削除
    df_tanzo_k_1 = df_tanzo_k_1.drop("所属2", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("所属3", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("区分", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("出勤日数    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("有休日数    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("欠勤日数    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("残業時間    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("深夜残業    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("法外休出    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("法定休出    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("代休時間    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("深夜代休    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("６０Ｈ超    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("別居手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("通勤手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("特別技技手当", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("特殊手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("地域手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("営業手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("役職手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("調整手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("基 本 給    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("残業手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("休出手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("深夜勤務手当", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("交替時差手当", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("休業手当    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("代 休 他    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("欠勤控除    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("遅早控除    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("精 算 分    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("支給合計額  ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("休業控除    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("支給額", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("ズレ時間    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("雑費・食事代", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("雑費・衣靴代", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("雑費        ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("受診料・他  ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("雑費・会費等", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("勤務時間    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("遅早時間    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("特休日数    ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.drop("集計区分－２        ", axis=1)
    df_tanzo_k_1 = df_tanzo_k_1.sum()
    df_tanzo_k_1.to_csv(
        "c:/test/鍛造/A.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 鍛造_間接2
    df_tanzo_k_2 = df_tanzo_m.groupby("区分").get_group("間接2")
    df_tanzo_k_2 = df_tanzo_k_2.drop("所属1", axis=1)
    member = df_tanzo_k_2["所属2"] > 0
    df_tanzo_k_2.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_tanzo_k_2["所属3"] > 0
    df_tanzo_k_2.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_tanzo_k_2["出勤日数    "]
    df_tanzo_k_2.insert(2, "実在籍者", real_member)
    time_yukyu = df_tanzo_k_2["有休日数    "] * 8
    df_tanzo_k_2.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_tanzo_k_2.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_tanzo_k_2["欠勤日数    "] * 8
    df_tanzo_k_2.insert(5, "欠勤時間", abs_time)
    work_time = df_tanzo_k_2["勤務時間    "]
    df_tanzo_k_2.insert(6, "勤務時間", work_time)
    late_early_time = df_tanzo_k_2["遅早時間    "]
    df_tanzo_k_2.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_tanzo_k_2["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_tanzo_k_2.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_tanzo_k_2.insert(17, "実労働時間", real_work_time)
    zure_time = df_tanzo_k_2["ズレ時間    "]
    df_tanzo_k_2.insert(18, "ズレ時間", zure_time)
    overtime = df_tanzo_k_2["残業時間    "] + df_tanzo_k_2["深夜残業    "]
    df_tanzo_k_2.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_tanzo_k_2.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_tanzo_k_2["法外休出    "]
    df_tanzo_k_2.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_tanzo_k_2.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_tanzo_k_2["法定休出    "]
    df_tanzo_k_2.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_tanzo_k_2.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_tanzo_k_2["６０Ｈ超    "]
    df_tanzo_k_2.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_tanzo_k_2["代休時間    "] + df_tanzo_k_2["深夜代休    "]
    df_tanzo_k_2.insert(31, "代休時間", holiday_time)
    df_tanzo_k_2.insert(32, "応援時間", 0)
    total_work_time = (
        df_tanzo_k_2["勤務時間    "]
        + df_tanzo_k_2["残業時間    "]
        + df_tanzo_k_2["法外休出    "]
        + df_tanzo_k_2["法定休出    "]
    )
    df_tanzo_k_2.insert(33, "総労働時間", total_work_time)
    basic_salary = df_tanzo_k_2["基 本 給    "] + df_tanzo_k_2["支給額"]
    df_tanzo_k_2.insert(35, "基本給", basic_salary)
    post_allowance = df_tanzo_k_2["役職手当    "]
    df_tanzo_k_2.insert(36, "役職手当", post_allowance)
    sales_allowance = df_tanzo_k_2["営業手当    "]
    df_tanzo_k_2.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_tanzo_k_2["地域手当    "]
    df_tanzo_k_2.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_tanzo_k_2["特殊手当    "]
    df_tanzo_k_2.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_tanzo_k_2["特別技技手当"]
    df_tanzo_k_2.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_tanzo_k_2["調整手当    "]
    df_tanzo_k_2.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_tanzo_k_2["別居手当    "]
    df_tanzo_k_2.insert(42, "別居手当", sep_allowance)
    com_allowance = df_tanzo_k_2["通勤手当    "]
    df_tanzo_k_2.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_tanzo_k_2.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_tanzo_k_2["残業手当    "]
    df_tanzo_k_2.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_tanzo_k_2["休出手当    "]
    df_tanzo_k_2.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_tanzo_k_2["深夜勤務手当"]
    df_tanzo_k_2.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_tanzo_k_2["交替時差手当"]
    df_tanzo_k_2.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_tanzo_k_2["休業手当    "]
    df_tanzo_k_2.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_tanzo_k_2["休業控除    "]
    df_tanzo_k_2.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_tanzo_k_2["代 休 他    "]
    df_tanzo_k_2.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_tanzo_k_2["欠勤控除    "] + df_tanzo_k_2["遅早控除    "]
    df_tanzo_k_2.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_tanzo_k_2["精 算 分    "]
    df_tanzo_k_2.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_tanzo_k_2.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_tanzo_k_2.insert(55, "総支給額", total)
    df_tanzo_k_2.insert(56, "応援時間額", 0)
    df_tanzo_k_2.insert(57, "役員振替", 0)
    df_tanzo_k_2.insert(58, "部門振替", 0)
    df_tanzo_k_2.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_tanzo_k_2 = df_tanzo_k_2.drop("所属2", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("所属3", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("社員ｺｰﾄﾞ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("区分", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("出勤日数    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("有休日数    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("欠勤日数    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("残業時間    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("深夜残業    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("法外休出    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("法定休出    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("代休時間    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("深夜代休    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("６０Ｈ超    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("別居手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("通勤手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("特別技技手当", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("特殊手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("地域手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("営業手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("役職手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("調整手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("基 本 給    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("残業手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("休出手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("深夜勤務手当", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("交替時差手当", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("休業手当    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("代 休 他    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("欠勤控除    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("遅早控除    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("精 算 分    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("支給合計額  ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("休業控除    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("支給額", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("ズレ時間    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("雑費・食事代", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("雑費・衣靴代", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("雑費        ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("受診料・他  ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("雑費・会費等", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("勤務時間    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("遅早時間    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("特休日数    ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.drop("集計区分－２        ", axis=1)
    df_tanzo_k_2 = df_tanzo_k_2.sum()
    df_tanzo_k_2.to_csv(
        "c:/test/鍛造/B.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 鍛造_間接3
    df_tanzo_k_3 = df_tanzo_m.groupby("区分").get_group("間接3")
    df_tanzo_k_3 = df_tanzo_k_3.drop("所属1", axis=1)
    member = df_tanzo_k_3["所属2"] > 0
    df_tanzo_k_3.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_tanzo_k_3["所属3"] > 0
    df_tanzo_k_3.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_tanzo_k_3["出勤日数    "]
    df_tanzo_k_3.insert(2, "実在籍者", real_member)
    time_yukyu = df_tanzo_k_3["有休日数    "] * 8
    df_tanzo_k_3.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_tanzo_k_3.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_tanzo_k_3["欠勤日数    "] * 8
    df_tanzo_k_3.insert(5, "欠勤時間", abs_time)
    work_time = df_tanzo_k_3["勤務時間    "]
    df_tanzo_k_3.insert(6, "勤務時間", work_time)
    late_early_time = df_tanzo_k_3["遅早時間    "]
    df_tanzo_k_3.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_tanzo_k_3["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_tanzo_k_3.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_tanzo_k_3.insert(17, "実労働時間", real_work_time)
    zure_time = df_tanzo_k_3["ズレ時間    "]
    df_tanzo_k_3.insert(18, "ズレ時間", zure_time)
    overtime = df_tanzo_k_3["残業時間    "] + df_tanzo_k_3["深夜残業    "]
    df_tanzo_k_3.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_tanzo_k_3.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_tanzo_k_3["法外休出    "]
    df_tanzo_k_3.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_tanzo_k_3.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_tanzo_k_3["法定休出    "]
    df_tanzo_k_3.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_tanzo_k_3.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_tanzo_k_3["６０Ｈ超    "]
    df_tanzo_k_3.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_tanzo_k_3["代休時間    "] + df_tanzo_k_3["深夜代休    "]
    df_tanzo_k_3.insert(31, "代休時間", holiday_time)
    df_tanzo_k_3.insert(32, "応援時間", 0)
    total_work_time = (
        df_tanzo_k_3["勤務時間    "]
        + df_tanzo_k_3["残業時間    "]
        + df_tanzo_k_3["法外休出    "]
        + df_tanzo_k_3["法定休出    "]
    )
    df_tanzo_k_3.insert(33, "総労働時間", total_work_time)
    basic_salary = df_tanzo_k_3["基 本 給    "] + df_tanzo_k_3["支給額"]
    df_tanzo_k_3.insert(35, "基本給", basic_salary)
    post_allowance = df_tanzo_k_3["役職手当    "]
    df_tanzo_k_3.insert(36, "役職手当", post_allowance)
    sales_allowance = df_tanzo_k_3["営業手当    "]
    df_tanzo_k_3.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_tanzo_k_3["地域手当    "]
    df_tanzo_k_3.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_tanzo_k_3["特殊手当    "]
    df_tanzo_k_3.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_tanzo_k_3["特別技技手当"]
    df_tanzo_k_3.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_tanzo_k_3["調整手当    "]
    df_tanzo_k_3.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_tanzo_k_3["別居手当    "]
    df_tanzo_k_3.insert(42, "別居手当", sep_allowance)
    com_allowance = df_tanzo_k_3["通勤手当    "]
    df_tanzo_k_3.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_tanzo_k_3.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_tanzo_k_3["残業手当    "]
    df_tanzo_k_3.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_tanzo_k_3["休出手当    "]
    df_tanzo_k_3.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_tanzo_k_3["深夜勤務手当"]
    df_tanzo_k_3.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_tanzo_k_3["交替時差手当"]
    df_tanzo_k_3.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_tanzo_k_3["休業手当    "]
    df_tanzo_k_3.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_tanzo_k_3["休業控除    "]
    df_tanzo_k_3.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_tanzo_k_3["代 休 他    "]
    df_tanzo_k_3.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_tanzo_k_3["欠勤控除    "] + df_tanzo_k_3["遅早控除    "]
    df_tanzo_k_3.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_tanzo_k_3["精 算 分    "]
    df_tanzo_k_3.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_tanzo_k_3.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_tanzo_k_3.insert(55, "総支給額", total)
    df_tanzo_k_3.insert(56, "応援時間額", 0)
    df_tanzo_k_3.insert(57, "役員振替", 0)
    df_tanzo_k_3.insert(58, "部門振替", 0)
    df_tanzo_k_3.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_tanzo_k_3 = df_tanzo_k_3.drop("所属2", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("所属3", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("社員ｺｰﾄﾞ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("区分", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("出勤日数    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("有休日数    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("欠勤日数    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("残業時間    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("深夜残業    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("法外休出    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("法定休出    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("代休時間    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("深夜代休    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("６０Ｈ超    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("別居手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("通勤手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("特別技技手当", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("特殊手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("地域手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("営業手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("役職手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("調整手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("基 本 給    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("残業手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("休出手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("深夜勤務手当", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("交替時差手当", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("休業手当    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("代 休 他    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("欠勤控除    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("遅早控除    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("精 算 分    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("支給合計額  ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("休業控除    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("支給額", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("ズレ時間    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("雑費・食事代", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("雑費・衣靴代", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("雑費        ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("受診料・他  ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("雑費・会費等", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("勤務時間    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("遅早時間    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("特休日数    ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.drop("集計区分－２        ", axis=1)
    df_tanzo_k_3 = df_tanzo_k_3.sum()
    df_tanzo_k_3.to_csv(
        "c:/test/鍛造/C.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 鍛造_間接4
    df_tanzo_k_4 = df_tanzo_m.groupby("区分").get_group("間接4")
    df_tanzo_k_4 = df_tanzo_k_4.drop("所属1", axis=1)
    member = df_tanzo_k_4["所属2"] > 0
    df_tanzo_k_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_tanzo_k_4["所属3"] > 0
    df_tanzo_k_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_tanzo_k_4["出勤日数    "]
    df_tanzo_k_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_tanzo_k_4["有休日数    "] * 8
    df_tanzo_k_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_tanzo_k_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_tanzo_k_4["欠勤日数    "] * 8
    df_tanzo_k_4.insert(5, "欠勤時間", abs_time)
    work_time = df_tanzo_k_4["勤務時間    "]
    df_tanzo_k_4.insert(6, "勤務時間", work_time)
    late_early_time = df_tanzo_k_4["遅早時間    "]
    df_tanzo_k_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_tanzo_k_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_tanzo_k_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_tanzo_k_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_tanzo_k_4["ズレ時間    "]
    df_tanzo_k_4.insert(18, "ズレ時間", zure_time)
    overtime = df_tanzo_k_4["残業時間    "] + df_tanzo_k_4["深夜残業    "]
    df_tanzo_k_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_tanzo_k_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_tanzo_k_4["法外休出    "]
    df_tanzo_k_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_tanzo_k_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_tanzo_k_4["法定休出    "]
    df_tanzo_k_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_tanzo_k_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_tanzo_k_4["６０Ｈ超    "]
    df_tanzo_k_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_tanzo_k_4["代休時間    "] + df_tanzo_k_4["深夜代休    "]
    df_tanzo_k_4.insert(31, "代休時間", holiday_time)
    df_tanzo_k_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_tanzo_k_4["勤務時間    "]
        + df_tanzo_k_4["残業時間    "]
        + df_tanzo_k_4["法外休出    "]
        + df_tanzo_k_4["法定休出    "]
    )
    df_tanzo_k_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_tanzo_k_4["基 本 給    "] + df_tanzo_k_4["支給額"]
    df_tanzo_k_4.insert(35, "基本給", basic_salary)
    post_allowance = df_tanzo_k_4["役職手当    "]
    df_tanzo_k_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_tanzo_k_4["営業手当    "]
    df_tanzo_k_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_tanzo_k_4["地域手当    "]
    df_tanzo_k_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_tanzo_k_4["特殊手当    "]
    df_tanzo_k_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_tanzo_k_4["特別技技手当"]
    df_tanzo_k_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_tanzo_k_4["調整手当    "]
    df_tanzo_k_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_tanzo_k_4["別居手当    "]
    df_tanzo_k_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_tanzo_k_4["通勤手当    "]
    df_tanzo_k_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_tanzo_k_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_tanzo_k_4["残業手当    "]
    df_tanzo_k_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_tanzo_k_4["休出手当    "]
    df_tanzo_k_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_tanzo_k_4["深夜勤務手当"]
    df_tanzo_k_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_tanzo_k_4["交替時差手当"]
    df_tanzo_k_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_tanzo_k_4["休業手当    "]
    df_tanzo_k_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_tanzo_k_4["休業控除    "]
    df_tanzo_k_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_tanzo_k_4["代 休 他    "]
    df_tanzo_k_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_tanzo_k_4["欠勤控除    "] + df_tanzo_k_4["遅早控除    "]
    df_tanzo_k_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_tanzo_k_4["精 算 分    "]
    df_tanzo_k_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_tanzo_k_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_tanzo_k_4.insert(55, "総支給額", total)
    df_tanzo_k_4.insert(56, "応援時間額", 0)
    df_tanzo_k_4.insert(57, "役員振替", 0)
    df_tanzo_k_4.insert(58, "部門振替", 0)
    df_tanzo_k_4.insert(59, "合計", 0)
    df_tanzo_k_4 = df_tanzo_k_4.drop("所属2", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("所属3", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("区分", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("出勤日数    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("有休日数    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("欠勤日数    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("残業時間    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("深夜残業    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("法外休出    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("法定休出    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("代休時間    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("深夜代休    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("６０Ｈ超    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("別居手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("通勤手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("特別技技手当", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("特殊手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("地域手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("営業手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("役職手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("調整手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("基 本 給    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("残業手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("休出手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("深夜勤務手当", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("交替時差手当", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("休業手当    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("代 休 他    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("欠勤控除    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("遅早控除    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("精 算 分    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("支給合計額  ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("休業控除    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("支給額", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("ズレ時間    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("雑費・食事代", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("雑費・衣靴代", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("雑費        ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("受診料・他  ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("雑費・会費等", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("勤務時間    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("遅早時間    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("特休日数    ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.drop("集計区分－２        ", axis=1)
    df_tanzo_k_4 = df_tanzo_k_4.sum()
    df_tanzo_k_4.to_csv(
        "c:/test/鍛造/D.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 鍛造_間接5
    df_tanzo_k_5 = df_tanzo_m.groupby("区分").get_group("間接5")
    df_tanzo_k_5 = df_tanzo_k_5.drop("所属1", axis=1)
    member = df_tanzo_k_5["所属2"] > 0
    df_tanzo_k_5.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_tanzo_k_5["所属3"] > 0
    df_tanzo_k_5.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_tanzo_k_5["出勤日数    "]
    df_tanzo_k_5.insert(2, "実在籍者", real_member)
    time_yukyu = df_tanzo_k_5["有休日数    "] * 8
    df_tanzo_k_5.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_tanzo_k_5.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_tanzo_k_5["欠勤日数    "] * 8
    df_tanzo_k_5.insert(5, "欠勤時間", abs_time)
    work_time = df_tanzo_k_5["勤務時間    "]
    df_tanzo_k_5.insert(6, "勤務時間", work_time)
    late_early_time = df_tanzo_k_5["遅早時間    "]
    df_tanzo_k_5.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_tanzo_k_5["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_tanzo_k_5.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_tanzo_k_5.insert(17, "実労働時間", real_work_time)
    zure_time = df_tanzo_k_5["ズレ時間    "]
    df_tanzo_k_5.insert(18, "ズレ時間", zure_time)
    overtime = df_tanzo_k_5["残業時間    "] + df_tanzo_k_5["深夜残業    "]
    df_tanzo_k_5.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_tanzo_k_5.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_tanzo_k_5["法外休出    "]
    df_tanzo_k_5.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_tanzo_k_5.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_tanzo_k_5["法定休出    "]
    df_tanzo_k_5.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_tanzo_k_5.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_tanzo_k_5["６０Ｈ超    "]
    df_tanzo_k_5.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_tanzo_k_5["代休時間    "] + df_tanzo_k_5["深夜代休    "]
    df_tanzo_k_5.insert(31, "代休時間", holiday_time)
    df_tanzo_k_5.insert(32, "応援時間", 0)
    total_work_time = (
        df_tanzo_k_5["勤務時間    "]
        + df_tanzo_k_5["残業時間    "]
        + df_tanzo_k_5["法外休出    "]
        + df_tanzo_k_5["法定休出    "]
    )
    df_tanzo_k_5.insert(33, "総労働時間", total_work_time)
    basic_salary = df_tanzo_k_5["基 本 給    "] + df_tanzo_k_5["支給額"]
    df_tanzo_k_5.insert(35, "基本給", basic_salary)
    post_allowance = df_tanzo_k_5["役職手当    "]
    df_tanzo_k_5.insert(36, "役職手当", post_allowance)
    sales_allowance = df_tanzo_k_5["営業手当    "]
    df_tanzo_k_5.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_tanzo_k_5["地域手当    "]
    df_tanzo_k_5.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_tanzo_k_5["特殊手当    "]
    df_tanzo_k_5.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_tanzo_k_5["特別技技手当"]
    df_tanzo_k_5.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_tanzo_k_5["調整手当    "]
    df_tanzo_k_5.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_tanzo_k_5["別居手当    "]
    df_tanzo_k_5.insert(42, "別居手当", sep_allowance)
    com_allowance = df_tanzo_k_5["通勤手当    "]
    df_tanzo_k_5.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_tanzo_k_5.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_tanzo_k_5["残業手当    "]
    df_tanzo_k_5.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_tanzo_k_5["休出手当    "]
    df_tanzo_k_5.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_tanzo_k_5["深夜勤務手当"]
    df_tanzo_k_5.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_tanzo_k_5["交替時差手当"]
    df_tanzo_k_5.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_tanzo_k_5["休業手当    "]
    df_tanzo_k_5.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_tanzo_k_5["休業控除    "]
    df_tanzo_k_5.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_tanzo_k_5["代 休 他    "]
    df_tanzo_k_5.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_tanzo_k_5["欠勤控除    "] + df_tanzo_k_5["遅早控除    "]
    df_tanzo_k_5.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_tanzo_k_5["精 算 分    "]
    df_tanzo_k_5.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_tanzo_k_5.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_tanzo_k_5.insert(55, "総支給額", total)
    df_tanzo_k_5.insert(56, "応援時間額", 0)
    df_tanzo_k_5.insert(57, "役員振替", 0)
    df_tanzo_k_5.insert(58, "部門振替", 0)
    df_tanzo_k_5.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_tanzo_k_5 = df_tanzo_k_5.drop("所属2", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("所属3", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("社員ｺｰﾄﾞ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("区分", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("出勤日数    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("有休日数    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("欠勤日数    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("残業時間    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("深夜残業    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("法外休出    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("法定休出    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("代休時間    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("深夜代休    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("６０Ｈ超    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("別居手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("通勤手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("特別技技手当", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("特殊手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("地域手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("営業手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("役職手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("調整手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("基 本 給    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("残業手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("休出手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("深夜勤務手当", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("交替時差手当", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("休業手当    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("代 休 他    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("欠勤控除    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("遅早控除    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("精 算 分    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("支給合計額  ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("休業控除    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("支給額", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("ズレ時間    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("雑費・食事代", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("雑費・衣靴代", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("雑費        ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("受診料・他  ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("雑費・会費等", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("勤務時間    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("遅早時間    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("特休日数    ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.drop("集計区分－２        ", axis=1)
    df_tanzo_k_5 = df_tanzo_k_5.sum()
    df_tanzo_k_5.to_csv(
        "c:/test/鍛造/E.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 鍛造_間接6
    df_tanzo_k_6 = df_tanzo_m.groupby("区分").get_group("間接6")
    df_tanzo_k_6 = df_tanzo_k_6.drop("所属1", axis=1)
    member = df_tanzo_k_6["所属2"] > 0
    df_tanzo_k_6.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_tanzo_k_6["所属3"] > 0
    df_tanzo_k_6.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_tanzo_k_6["出勤日数    "]
    df_tanzo_k_6.insert(2, "実在籍者", real_member)
    time_yukyu = df_tanzo_k_6["有休日数    "] * 8
    df_tanzo_k_6.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_tanzo_k_6.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_tanzo_k_6["欠勤日数    "] * 8
    df_tanzo_k_6.insert(5, "欠勤時間", abs_time)
    work_time = df_tanzo_k_6["勤務時間    "]
    df_tanzo_k_6.insert(6, "勤務時間", work_time)
    late_early_time = df_tanzo_k_6["遅早時間    "]
    df_tanzo_k_6.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_tanzo_k_6["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_tanzo_k_6.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_tanzo_k_6.insert(17, "実労働時間", real_work_time)
    zure_time = df_tanzo_k_6["ズレ時間    "]
    df_tanzo_k_6.insert(18, "ズレ時間", zure_time)
    overtime = df_tanzo_k_6["残業時間    "] + df_tanzo_k_6["深夜残業    "]
    df_tanzo_k_6.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_tanzo_k_6.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_tanzo_k_6["法外休出    "]
    df_tanzo_k_6.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_tanzo_k_6.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_tanzo_k_6["法定休出    "]
    df_tanzo_k_6.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_tanzo_k_6.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_tanzo_k_6["６０Ｈ超    "]
    df_tanzo_k_6.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_tanzo_k_6["代休時間    "] + df_tanzo_k_6["深夜代休    "]
    df_tanzo_k_6.insert(31, "代休時間", holiday_time)
    df_tanzo_k_6.insert(32, "応援時間", 0)
    total_work_time = (
        df_tanzo_k_6["勤務時間    "]
        + df_tanzo_k_6["残業時間    "]
        + df_tanzo_k_6["法外休出    "]
        + df_tanzo_k_6["法定休出    "]
    )
    df_tanzo_k_6.insert(33, "総労働時間", total_work_time)
    basic_salary = df_tanzo_k_6["基 本 給    "] + df_tanzo_k_6["支給額"]
    df_tanzo_k_6.insert(35, "基本給", basic_salary)
    post_allowance = df_tanzo_k_6["役職手当    "]
    df_tanzo_k_6.insert(36, "役職手当", post_allowance)
    sales_allowance = df_tanzo_k_6["営業手当    "]
    df_tanzo_k_6.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_tanzo_k_6["地域手当    "]
    df_tanzo_k_6.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_tanzo_k_6["特殊手当    "]
    df_tanzo_k_6.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_tanzo_k_6["特別技技手当"]
    df_tanzo_k_6.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_tanzo_k_6["調整手当    "]
    df_tanzo_k_6.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_tanzo_k_6["別居手当    "]
    df_tanzo_k_6.insert(42, "別居手当", sep_allowance)
    com_allowance = df_tanzo_k_6["通勤手当    "]
    df_tanzo_k_6.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_tanzo_k_6.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_tanzo_k_6["残業手当    "]
    df_tanzo_k_6.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_tanzo_k_6["休出手当    "]
    df_tanzo_k_6.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_tanzo_k_6["深夜勤務手当"]
    df_tanzo_k_6.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_tanzo_k_6["交替時差手当"]
    df_tanzo_k_6.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_tanzo_k_6["休業手当    "]
    df_tanzo_k_6.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_tanzo_k_6["休業控除    "]
    df_tanzo_k_6.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_tanzo_k_6["代 休 他    "]
    df_tanzo_k_6.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_tanzo_k_6["欠勤控除    "] + df_tanzo_k_6["遅早控除    "]
    df_tanzo_k_6.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_tanzo_k_6["精 算 分    "]
    df_tanzo_k_6.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_tanzo_k_6.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_tanzo_k_6.insert(55, "総支給額", total)
    df_tanzo_k_6.insert(56, "応援時間額", 0)
    df_tanzo_k_6.insert(57, "役員振替", 0)
    df_tanzo_k_6.insert(58, "部門振替", 0)
    df_tanzo_k_6.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_tanzo_k_6 = df_tanzo_k_6.drop("所属2", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("所属3", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("社員ｺｰﾄﾞ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("区分", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("出勤日数    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("有休日数    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("欠勤日数    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("残業時間    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("深夜残業    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("法外休出    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("法定休出    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("代休時間    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("深夜代休    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("６０Ｈ超    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("別居手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("通勤手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("特別技技手当", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("特殊手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("地域手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("営業手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("役職手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("調整手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("基 本 給    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("残業手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("休出手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("深夜勤務手当", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("交替時差手当", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("休業手当    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("代 休 他    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("欠勤控除    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("遅早控除    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("精 算 分    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("支給合計額  ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("休業控除    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("支給額", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("ズレ時間    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("雑費・食事代", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("雑費・衣靴代", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("雑費        ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("受診料・他  ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("雑費・会費等", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("勤務時間    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("遅早時間    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("特休日数    ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.drop("集計区分－２        ", axis=1)
    df_tanzo_k_6 = df_tanzo_k_6.sum()
    df_tanzo_k_6.to_csv(
        "c:/test/鍛造/F.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 鍛造_直接1
    df_tanzo_t_1 = df_tanzo_m.groupby("区分").get_group("直接1")
    df_tanzo_t_1 = df_tanzo_t_1.drop("所属1", axis=1)
    member = df_tanzo_t_1["所属2"] > 0
    df_tanzo_t_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_tanzo_t_1["所属3"] > 0
    df_tanzo_t_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_tanzo_t_1["出勤日数    "]
    df_tanzo_t_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_tanzo_t_1["有休日数    "] * 8
    df_tanzo_t_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_tanzo_t_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_tanzo_t_1["欠勤日数    "] * 8
    df_tanzo_t_1.insert(5, "欠勤時間", abs_time)
    work_time = df_tanzo_t_1["勤務時間    "]
    df_tanzo_t_1.insert(6, "勤務時間", work_time)
    late_early_time = df_tanzo_t_1["遅早時間    "]
    df_tanzo_t_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_tanzo_t_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_tanzo_t_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_tanzo_t_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_tanzo_t_1["ズレ時間    "]
    df_tanzo_t_1.insert(18, "ズレ時間", zure_time)
    overtime = df_tanzo_t_1["残業時間    "] + df_tanzo_t_1["深夜残業    "]
    df_tanzo_t_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_tanzo_t_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_tanzo_t_1["法外休出    "]
    df_tanzo_t_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_tanzo_t_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_tanzo_t_1["法定休出    "]
    df_tanzo_t_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_tanzo_t_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_tanzo_t_1["６０Ｈ超    "]
    df_tanzo_t_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_tanzo_t_1["代休時間    "] + df_tanzo_t_1["深夜代休    "]
    df_tanzo_t_1.insert(31, "代休時間", holiday_time)
    df_tanzo_t_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_tanzo_t_1["勤務時間    "]
        + df_tanzo_t_1["残業時間    "]
        + df_tanzo_t_1["法外休出    "]
        + df_tanzo_t_1["法定休出    "]
    )
    df_tanzo_t_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_tanzo_t_1["基 本 給    "] + df_tanzo_t_1["支給額"]
    df_tanzo_t_1.insert(35, "基本給", basic_salary)
    post_allowance = df_tanzo_t_1["役職手当    "]
    df_tanzo_t_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_tanzo_t_1["営業手当    "]
    df_tanzo_t_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_tanzo_t_1["地域手当    "]
    df_tanzo_t_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_tanzo_t_1["特殊手当    "]
    df_tanzo_t_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_tanzo_t_1["特別技技手当"]
    df_tanzo_t_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_tanzo_t_1["調整手当    "]
    df_tanzo_t_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_tanzo_t_1["別居手当    "]
    df_tanzo_t_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_tanzo_t_1["通勤手当    "]
    df_tanzo_t_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_tanzo_t_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_tanzo_t_1["残業手当    "]
    df_tanzo_t_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_tanzo_t_1["休出手当    "]
    df_tanzo_t_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_tanzo_t_1["深夜勤務手当"]
    df_tanzo_t_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_tanzo_t_1["交替時差手当"]
    df_tanzo_t_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_tanzo_t_1["休業手当    "]
    df_tanzo_t_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_tanzo_t_1["休業控除    "]
    df_tanzo_t_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_tanzo_t_1["代 休 他    "]
    df_tanzo_t_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_tanzo_t_1["欠勤控除    "] + df_tanzo_t_1["遅早控除    "]
    df_tanzo_t_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_tanzo_t_1["精 算 分    "]
    df_tanzo_t_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_tanzo_t_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_tanzo_t_1.insert(55, "総支給額", total)
    df_tanzo_t_1.insert(56, "応援時間額", 0)
    df_tanzo_t_1.insert(57, "役員振替", 0)
    df_tanzo_t_1.insert(58, "部門振替", 0)
    df_tanzo_t_1.insert(59, "合計", 0)
    df_tanzo_t_1 = df_tanzo_t_1.drop("所属2", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("所属3", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("区分", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("出勤日数    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("有休日数    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("欠勤日数    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("残業時間    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("深夜残業    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("法外休出    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("法定休出    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("代休時間    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("深夜代休    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("６０Ｈ超    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("別居手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("通勤手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("特別技技手当", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("特殊手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("地域手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("営業手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("役職手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("調整手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("基 本 給    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("残業手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("休出手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("深夜勤務手当", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("交替時差手当", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("休業手当    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("代 休 他    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("欠勤控除    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("遅早控除    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("精 算 分    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("支給合計額  ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("休業控除    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("支給額", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("ズレ時間    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("雑費・食事代", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("雑費・衣靴代", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("雑費        ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("受診料・他  ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("雑費・会費等", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("勤務時間    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("遅早時間    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("特休日数    ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.drop("集計区分－２        ", axis=1)
    df_tanzo_t_1 = df_tanzo_t_1.sum()
    df_tanzo_t_1.to_csv(
        "c:/test/鍛造/G.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 鍛造_直接4
    df_tanzo_t_4 = df_tanzo_m.groupby("区分").get_group("直接4")
    df_tanzo_t_4 = df_tanzo_t_4.drop("所属1", axis=1)
    member = df_tanzo_t_4["所属2"] > 0
    df_tanzo_t_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_tanzo_t_4["所属3"] > 0
    df_tanzo_t_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_tanzo_t_4["出勤日数    "]
    df_tanzo_t_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_tanzo_t_4["有休日数    "] * 8
    df_tanzo_t_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_tanzo_t_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_tanzo_t_4["欠勤日数    "] * 8
    df_tanzo_t_4.insert(5, "欠勤時間", abs_time)
    work_time = df_tanzo_t_4["勤務時間    "]
    df_tanzo_t_4.insert(6, "勤務時間", work_time)
    late_early_time = df_tanzo_t_4["遅早時間    "]
    df_tanzo_t_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_tanzo_t_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_tanzo_t_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_tanzo_t_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_tanzo_t_4["ズレ時間    "]
    df_tanzo_t_4.insert(18, "ズレ時間", zure_time)
    overtime = df_tanzo_t_4["残業時間    "] + df_tanzo_t_4["深夜残業    "]
    df_tanzo_t_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_tanzo_t_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_tanzo_t_4["法外休出    "]
    df_tanzo_t_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_tanzo_t_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_tanzo_t_4["法定休出    "]
    df_tanzo_t_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_tanzo_t_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_tanzo_t_4["６０Ｈ超    "]
    df_tanzo_t_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_tanzo_t_4["代休時間    "] + df_tanzo_t_4["深夜代休    "]
    df_tanzo_t_4.insert(31, "代休時間", holiday_time)
    df_tanzo_t_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_tanzo_t_4["勤務時間    "]
        + df_tanzo_t_4["残業時間    "]
        + df_tanzo_t_4["法外休出    "]
        + df_tanzo_t_4["法定休出    "]
    )
    df_tanzo_t_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_tanzo_t_4["基 本 給    "] + df_tanzo_t_4["支給額"]
    df_tanzo_t_4.insert(35, "基本給", basic_salary)
    post_allowance = df_tanzo_t_4["役職手当    "]
    df_tanzo_t_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_tanzo_t_4["営業手当    "]
    df_tanzo_t_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_tanzo_t_4["地域手当    "]
    df_tanzo_t_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_tanzo_t_4["特殊手当    "]
    df_tanzo_t_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_tanzo_t_4["特別技技手当"]
    df_tanzo_t_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_tanzo_t_4["調整手当    "]
    df_tanzo_t_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_tanzo_t_4["別居手当    "]
    df_tanzo_t_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_tanzo_t_4["通勤手当    "]
    df_tanzo_t_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_tanzo_t_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_tanzo_t_4["残業手当    "]
    df_tanzo_t_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_tanzo_t_4["休出手当    "]
    df_tanzo_t_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_tanzo_t_4["深夜勤務手当"]
    df_tanzo_t_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_tanzo_t_4["交替時差手当"]
    df_tanzo_t_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_tanzo_t_4["休業手当    "]
    df_tanzo_t_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_tanzo_t_4["休業控除    "]
    df_tanzo_t_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_tanzo_t_4["代 休 他    "]
    df_tanzo_t_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_tanzo_t_4["欠勤控除    "] + df_tanzo_t_4["遅早控除    "]
    df_tanzo_t_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_tanzo_t_4["精 算 分    "]
    df_tanzo_t_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_tanzo_t_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_tanzo_t_4.insert(55, "総支給額", total)
    df_tanzo_t_4.insert(56, "応援時間額", 0)
    df_tanzo_t_4.insert(57, "役員振替", 0)
    df_tanzo_t_4.insert(58, "部門振替", 0)
    df_tanzo_t_4.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_tanzo_t_4 = df_tanzo_t_4.drop("所属2", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("所属3", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("区分", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("出勤日数    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("有休日数    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("欠勤日数    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("残業時間    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("深夜残業    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("法外休出    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("法定休出    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("代休時間    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("深夜代休    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("６０Ｈ超    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("別居手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("通勤手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("特別技技手当", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("特殊手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("地域手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("営業手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("役職手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("調整手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("基 本 給    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("残業手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("休出手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("深夜勤務手当", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("交替時差手当", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("休業手当    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("代 休 他    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("欠勤控除    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("遅早控除    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("精 算 分    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("支給合計額  ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("休業控除    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("支給額", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("ズレ時間    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("雑費・食事代", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("雑費・衣靴代", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("雑費        ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("受診料・他  ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("雑費・会費等", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("勤務時間    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("遅早時間    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("特休日数    ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.drop("集計区分－２        ", axis=1)
    df_tanzo_t_4 = df_tanzo_t_4.sum()
    df_tanzo_t_4.to_csv(
        "c:/test/鍛造/H.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 切削_間接1
    df_sesaku_k_1 = df_sesaku_m.groupby("区分").get_group("間接1")
    df_sesaku_k_1 = df_sesaku_k_1.drop("所属1", axis=1)
    member = df_sesaku_k_1["所属2"] > 0
    df_sesaku_k_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_sesaku_k_1["所属3"] > 0
    df_sesaku_k_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_sesaku_k_1["出勤日数    "]
    df_sesaku_k_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_sesaku_k_1["有休日数    "] * 8
    df_sesaku_k_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_sesaku_k_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_sesaku_k_1["欠勤日数    "] * 8
    df_sesaku_k_1.insert(5, "欠勤時間", abs_time)
    work_time = df_sesaku_k_1["勤務時間    "]
    df_sesaku_k_1.insert(6, "勤務時間", work_time)
    late_early_time = df_sesaku_k_1["遅早時間    "]
    df_sesaku_k_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_sesaku_k_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_sesaku_k_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_sesaku_k_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_sesaku_k_1["ズレ時間    "]
    df_sesaku_k_1.insert(18, "ズレ時間", zure_time)
    overtime = df_sesaku_k_1["残業時間    "] + df_sesaku_k_1["深夜残業    "]
    df_sesaku_k_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_sesaku_k_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_sesaku_k_1["法外休出    "]
    df_sesaku_k_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_sesaku_k_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_sesaku_k_1["法定休出    "]
    df_sesaku_k_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_sesaku_k_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_sesaku_k_1["６０Ｈ超    "]
    df_sesaku_k_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_sesaku_k_1["代休時間    "] + df_sesaku_k_1["深夜代休    "]
    df_sesaku_k_1.insert(31, "代休時間", holiday_time)
    df_sesaku_k_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_sesaku_k_1["勤務時間    "]
        + df_sesaku_k_1["残業時間    "]
        + df_sesaku_k_1["法外休出    "]
        + df_sesaku_k_1["法定休出    "]
    )
    df_sesaku_k_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_sesaku_k_1["基 本 給    "] + df_sesaku_k_1["支給額"]
    df_sesaku_k_1.insert(35, "基本給", basic_salary)
    post_allowance = df_sesaku_k_1["役職手当    "]
    df_sesaku_k_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_sesaku_k_1["営業手当    "]
    df_sesaku_k_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_sesaku_k_1["地域手当    "]
    df_sesaku_k_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_sesaku_k_1["特殊手当    "]
    df_sesaku_k_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_sesaku_k_1["特別技技手当"]
    df_sesaku_k_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_sesaku_k_1["調整手当    "]
    df_sesaku_k_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_sesaku_k_1["別居手当    "]
    df_sesaku_k_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_sesaku_k_1["通勤手当    "]
    df_sesaku_k_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_sesaku_k_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_sesaku_k_1["残業手当    "]
    df_sesaku_k_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_sesaku_k_1["休出手当    "]
    df_sesaku_k_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_sesaku_k_1["深夜勤務手当"]
    df_sesaku_k_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_sesaku_k_1["交替時差手当"]
    df_sesaku_k_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_sesaku_k_1["休業手当    "]
    df_sesaku_k_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_sesaku_k_1["休業控除    "]
    df_sesaku_k_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_sesaku_k_1["代 休 他    "]
    df_sesaku_k_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_sesaku_k_1["欠勤控除    "] + df_sesaku_k_1["遅早控除    "]
    df_sesaku_k_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_sesaku_k_1["精 算 分    "]
    df_sesaku_k_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_sesaku_k_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_sesaku_k_1.insert(55, "総支給額", total)
    df_sesaku_k_1.insert(56, "応援時間額", 0)
    df_sesaku_k_1.insert(57, "役員振替", 0)
    df_sesaku_k_1.insert(58, "部門振替", 0)
    df_sesaku_k_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_sesaku_k_1 = df_sesaku_k_1.drop("所属2", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("所属3", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("区分", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("出勤日数    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("有休日数    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("欠勤日数    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("残業時間    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("深夜残業    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("法外休出    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("法定休出    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("代休時間    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("深夜代休    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("６０Ｈ超    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("別居手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("通勤手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("特別技技手当", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("特殊手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("地域手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("営業手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("役職手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("調整手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("基 本 給    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("残業手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("休出手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("深夜勤務手当", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("交替時差手当", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("休業手当    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("代 休 他    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("欠勤控除    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("遅早控除    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("精 算 分    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("支給合計額  ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("休業控除    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("支給額", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("ズレ時間    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("雑費・食事代", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("雑費・衣靴代", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("雑費        ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("受診料・他  ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("雑費・会費等", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("勤務時間    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("遅早時間    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("特休日数    ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.drop("集計区分－２        ", axis=1)
    df_sesaku_k_1 = df_sesaku_k_1.sum()
    df_sesaku_k_1.to_csv(
        "c:/test/切削/A.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    df_sesaku_k_2 = df_sesaku_m.groupby("区分").get_group("間接2")
    df_sesaku_k_2 = df_sesaku_k_2.drop("所属1", axis=1)
    member = df_sesaku_k_2["所属2"] > 0
    df_sesaku_k_2.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_sesaku_k_2["所属3"] > 0
    df_sesaku_k_2.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_sesaku_k_2["出勤日数    "]
    df_sesaku_k_2.insert(2, "実在籍者", real_member)
    time_yukyu = df_sesaku_k_2["有休日数    "] * 8
    df_sesaku_k_2.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_sesaku_k_2.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_sesaku_k_2["欠勤日数    "] * 8
    df_sesaku_k_2.insert(5, "欠勤時間", abs_time)
    work_time = df_sesaku_k_2["勤務時間    "]
    df_sesaku_k_2.insert(6, "勤務時間", work_time)
    late_early_time = df_sesaku_k_2["遅早時間    "]
    df_sesaku_k_2.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_sesaku_k_2["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_sesaku_k_2.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_sesaku_k_2.insert(17, "実労働時間", real_work_time)
    zure_time = df_sesaku_k_2["ズレ時間    "]
    df_sesaku_k_2.insert(18, "ズレ時間", zure_time)
    overtime = df_sesaku_k_2["残業時間    "] + df_sesaku_k_2["深夜残業    "]
    df_sesaku_k_2.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_sesaku_k_2.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_sesaku_k_2["法外休出    "]
    df_sesaku_k_2.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_sesaku_k_2.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_sesaku_k_2["法定休出    "]
    df_sesaku_k_2.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_sesaku_k_2.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_sesaku_k_2["６０Ｈ超    "]
    df_sesaku_k_2.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_sesaku_k_2["代休時間    "] + df_sesaku_k_2["深夜代休    "]
    df_sesaku_k_2.insert(31, "代休時間", holiday_time)
    df_sesaku_k_2.insert(32, "応援時間", 0)
    total_work_time = (
        df_sesaku_k_2["勤務時間    "]
        + df_sesaku_k_2["残業時間    "]
        + df_sesaku_k_2["法外休出    "]
        + df_sesaku_k_2["法定休出    "]
    )
    df_sesaku_k_2.insert(33, "総労働時間", total_work_time)
    basic_salary = df_sesaku_k_2["基 本 給    "] + df_sesaku_k_2["支給額"]
    df_sesaku_k_2.insert(35, "基本給", basic_salary)
    post_allowance = df_sesaku_k_2["役職手当    "]
    df_sesaku_k_2.insert(36, "役職手当", post_allowance)
    sales_allowance = df_sesaku_k_2["営業手当    "]
    df_sesaku_k_2.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_sesaku_k_2["地域手当    "]
    df_sesaku_k_2.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_sesaku_k_2["特殊手当    "]
    df_sesaku_k_2.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_sesaku_k_2["特別技技手当"]
    df_sesaku_k_2.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_sesaku_k_2["調整手当    "]
    df_sesaku_k_2.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_sesaku_k_2["別居手当    "]
    df_sesaku_k_2.insert(42, "別居手当", sep_allowance)
    com_allowance = df_sesaku_k_2["通勤手当    "]
    df_sesaku_k_2.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_sesaku_k_2.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_sesaku_k_2["残業手当    "]
    df_sesaku_k_2.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_sesaku_k_2["休出手当    "]
    df_sesaku_k_2.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_sesaku_k_2["深夜勤務手当"]
    df_sesaku_k_2.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_sesaku_k_2["交替時差手当"]
    df_sesaku_k_2.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_sesaku_k_2["休業手当    "]
    df_sesaku_k_2.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_sesaku_k_2["休業控除    "]
    df_sesaku_k_2.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_sesaku_k_2["代 休 他    "]
    df_sesaku_k_2.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_sesaku_k_2["欠勤控除    "] + df_sesaku_k_2["遅早控除    "]
    df_sesaku_k_2.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_sesaku_k_2["精 算 分    "]
    df_sesaku_k_2.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_sesaku_k_2.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_sesaku_k_2.insert(55, "総支給額", total)
    df_sesaku_k_2.insert(56, "応援時間額", 0)
    df_sesaku_k_2.insert(57, "役員振替", 0)
    df_sesaku_k_2.insert(58, "部門振替", 0)
    df_sesaku_k_2.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_sesaku_k_2 = df_sesaku_k_2.drop("所属2", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("所属3", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("社員ｺｰﾄﾞ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("区分", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("出勤日数    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("有休日数    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("欠勤日数    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("残業時間    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("深夜残業    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("法外休出    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("法定休出    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("代休時間    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("深夜代休    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("６０Ｈ超    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("別居手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("通勤手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("特別技技手当", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("特殊手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("地域手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("営業手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("役職手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("調整手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("基 本 給    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("残業手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("休出手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("深夜勤務手当", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("交替時差手当", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("休業手当    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("代 休 他    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("欠勤控除    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("遅早控除    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("精 算 分    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("支給合計額  ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("休業控除    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("支給額", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("ズレ時間    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("雑費・食事代", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("雑費・衣靴代", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("雑費        ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("受診料・他  ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("雑費・会費等", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("勤務時間    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("遅早時間    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("特休日数    ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.drop("集計区分－２        ", axis=1)
    df_sesaku_k_2 = df_sesaku_k_2.sum()
    df_sesaku_k_2.to_csv(
        "c:/test/切削/B.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 切削_間接4
    df_sesaku_k_4 = df_sesaku_m.groupby("区分").get_group("間接4")
    df_sesaku_k_4 = df_sesaku_k_4.drop("所属1", axis=1)
    member = df_sesaku_k_4["所属2"] > 0
    df_sesaku_k_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_sesaku_k_4["所属3"] > 0
    df_sesaku_k_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_sesaku_k_4["出勤日数    "]
    df_sesaku_k_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_sesaku_k_4["有休日数    "] * 8
    df_sesaku_k_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_sesaku_k_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_sesaku_k_4["欠勤日数    "] * 8
    df_sesaku_k_4.insert(5, "欠勤時間", abs_time)
    work_time = df_sesaku_k_4["勤務時間    "]
    df_sesaku_k_4.insert(6, "勤務時間", work_time)
    late_early_time = df_sesaku_k_4["遅早時間    "]
    df_sesaku_k_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_sesaku_k_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_sesaku_k_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_sesaku_k_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_sesaku_k_4["ズレ時間    "]
    df_sesaku_k_4.insert(18, "ズレ時間", zure_time)
    overtime = df_sesaku_k_4["残業時間    "] + df_sesaku_k_4["深夜残業    "]
    df_sesaku_k_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_sesaku_k_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_sesaku_k_4["法外休出    "]
    df_sesaku_k_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_sesaku_k_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_sesaku_k_4["法定休出    "]
    df_sesaku_k_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_sesaku_k_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_sesaku_k_4["６０Ｈ超    "]
    df_sesaku_k_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_sesaku_k_4["代休時間    "] + df_sesaku_k_4["深夜代休    "]
    df_sesaku_k_4.insert(31, "代休時間", holiday_time)
    df_sesaku_k_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_sesaku_k_4["勤務時間    "]
        + df_sesaku_k_4["残業時間    "]
        + df_sesaku_k_4["法外休出    "]
        + df_sesaku_k_4["法定休出    "]
    )
    df_sesaku_k_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_sesaku_k_4["基 本 給    "] + df_sesaku_k_4["支給額"]
    df_sesaku_k_4.insert(35, "基本給", basic_salary)
    post_allowance = df_sesaku_k_4["役職手当    "]
    df_sesaku_k_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_sesaku_k_4["営業手当    "]
    df_sesaku_k_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_sesaku_k_4["地域手当    "]
    df_sesaku_k_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_sesaku_k_4["特殊手当    "]
    df_sesaku_k_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_sesaku_k_4["特別技技手当"]
    df_sesaku_k_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_sesaku_k_4["調整手当    "]
    df_sesaku_k_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_sesaku_k_4["別居手当    "]
    df_sesaku_k_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_sesaku_k_4["通勤手当    "]
    df_sesaku_k_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_sesaku_k_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_sesaku_k_4["残業手当    "]
    df_sesaku_k_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_sesaku_k_4["休出手当    "]
    df_sesaku_k_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_sesaku_k_4["深夜勤務手当"]
    df_sesaku_k_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_sesaku_k_4["交替時差手当"]
    df_sesaku_k_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_sesaku_k_4["休業手当    "]
    df_sesaku_k_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_sesaku_k_4["休業控除    "]
    df_sesaku_k_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_sesaku_k_4["代 休 他    "]
    df_sesaku_k_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_sesaku_k_4["欠勤控除    "] + df_sesaku_k_4["遅早控除    "]
    df_sesaku_k_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_sesaku_k_4["精 算 分    "]
    df_sesaku_k_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_sesaku_k_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_sesaku_k_4.insert(55, "総支給額", total)
    df_sesaku_k_4.insert(56, "応援時間額", 0)
    df_sesaku_k_4.insert(57, "役員振替", 0)
    df_sesaku_k_4.insert(58, "部門振替", 0)
    df_sesaku_k_4.insert(59, "合計", 0)
    df_sesaku_k_4 = df_sesaku_k_4.drop("所属2", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("所属3", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("区分", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("出勤日数    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("有休日数    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("欠勤日数    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("残業時間    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("深夜残業    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("法外休出    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("法定休出    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("代休時間    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("深夜代休    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("６０Ｈ超    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("別居手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("通勤手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("特別技技手当", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("特殊手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("地域手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("営業手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("役職手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("調整手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("基 本 給    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("残業手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("休出手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("深夜勤務手当", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("交替時差手当", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("休業手当    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("代 休 他    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("欠勤控除    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("遅早控除    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("精 算 分    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("支給合計額  ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("休業控除    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("支給額", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("ズレ時間    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("雑費・食事代", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("雑費・衣靴代", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("雑費        ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("受診料・他  ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("雑費・会費等", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("勤務時間    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("遅早時間    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("特休日数    ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.drop("集計区分－２        ", axis=1)
    df_sesaku_k_4 = df_sesaku_k_4.sum()
    df_sesaku_k_4.to_csv(
        "c:/test/切削/C.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 切削_間接5
    df_sesaku_k_5 = df_sesaku_m.groupby("区分").get_group("間接5")
    df_sesaku_k_5 = df_sesaku_k_5.drop("所属1", axis=1)
    member = df_sesaku_k_5["所属2"] > 0
    df_sesaku_k_5.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_sesaku_k_5["所属3"] > 0
    df_sesaku_k_5.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_sesaku_k_5["出勤日数    "]
    df_sesaku_k_5.insert(2, "実在籍者", real_member)
    time_yukyu = df_sesaku_k_5["有休日数    "] * 8
    df_sesaku_k_5.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_sesaku_k_5.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_sesaku_k_5["欠勤日数    "] * 8
    df_sesaku_k_5.insert(5, "欠勤時間", abs_time)
    work_time = df_sesaku_k_5["勤務時間    "]
    df_sesaku_k_5.insert(6, "勤務時間", work_time)
    late_early_time = df_sesaku_k_5["遅早時間    "]
    df_sesaku_k_5.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_sesaku_k_5["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_sesaku_k_5.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_sesaku_k_5.insert(17, "実労働時間", real_work_time)
    zure_time = df_sesaku_k_5["ズレ時間    "]
    df_sesaku_k_5.insert(18, "ズレ時間", zure_time)
    overtime = df_sesaku_k_5["残業時間    "] + df_sesaku_k_5["深夜残業    "]
    df_sesaku_k_5.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_sesaku_k_5.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_sesaku_k_5["法外休出    "]
    df_sesaku_k_5.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_sesaku_k_5.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_sesaku_k_5["法定休出    "]
    df_sesaku_k_5.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_sesaku_k_5.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_sesaku_k_5["６０Ｈ超    "]
    df_sesaku_k_5.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_sesaku_k_5["代休時間    "] + df_sesaku_k_5["深夜代休    "]
    df_sesaku_k_5.insert(31, "代休時間", holiday_time)
    df_sesaku_k_5.insert(32, "応援時間", 0)
    total_work_time = (
        df_sesaku_k_5["勤務時間    "]
        + df_sesaku_k_5["残業時間    "]
        + df_sesaku_k_5["法外休出    "]
        + df_sesaku_k_5["法定休出    "]
    )
    df_sesaku_k_5.insert(33, "総労働時間", total_work_time)
    basic_salary = df_sesaku_k_5["基 本 給    "] + df_sesaku_k_5["支給額"]
    df_sesaku_k_5.insert(35, "基本給", basic_salary)
    post_allowance = df_sesaku_k_5["役職手当    "]
    df_sesaku_k_5.insert(36, "役職手当", post_allowance)
    sales_allowance = df_sesaku_k_5["営業手当    "]
    df_sesaku_k_5.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_sesaku_k_5["地域手当    "]
    df_sesaku_k_5.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_sesaku_k_5["特殊手当    "]
    df_sesaku_k_5.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_sesaku_k_5["特別技技手当"]
    df_sesaku_k_5.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_sesaku_k_5["調整手当    "]
    df_sesaku_k_5.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_sesaku_k_5["別居手当    "]
    df_sesaku_k_5.insert(42, "別居手当", sep_allowance)
    com_allowance = df_sesaku_k_5["通勤手当    "]
    df_sesaku_k_5.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_sesaku_k_5.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_sesaku_k_5["残業手当    "]
    df_sesaku_k_5.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_sesaku_k_5["休出手当    "]
    df_sesaku_k_5.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_sesaku_k_5["深夜勤務手当"]
    df_sesaku_k_5.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_sesaku_k_5["交替時差手当"]
    df_sesaku_k_5.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_sesaku_k_5["休業手当    "]
    df_sesaku_k_5.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_sesaku_k_5["休業控除    "]
    df_sesaku_k_5.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_sesaku_k_5["代 休 他    "]
    df_sesaku_k_5.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_sesaku_k_5["欠勤控除    "] + df_sesaku_k_5["遅早控除    "]
    df_sesaku_k_5.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_sesaku_k_5["精 算 分    "]
    df_sesaku_k_5.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_sesaku_k_5.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_sesaku_k_5.insert(55, "総支給額", total)
    df_sesaku_k_5.insert(56, "応援時間額", 0)
    df_sesaku_k_5.insert(57, "役員振替", 0)
    df_sesaku_k_5.insert(58, "部門振替", 0)
    df_sesaku_k_5.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_sesaku_k_5 = df_sesaku_k_5.drop("所属2", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("所属3", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("社員ｺｰﾄﾞ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("区分", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("出勤日数    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("有休日数    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("欠勤日数    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("残業時間    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("深夜残業    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("法外休出    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("法定休出    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("代休時間    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("深夜代休    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("６０Ｈ超    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("別居手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("通勤手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("特別技技手当", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("特殊手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("地域手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("営業手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("役職手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("調整手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("基 本 給    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("残業手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("休出手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("深夜勤務手当", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("交替時差手当", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("休業手当    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("代 休 他    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("欠勤控除    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("遅早控除    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("精 算 分    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("支給合計額  ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("休業控除    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("支給額", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("ズレ時間    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("雑費・食事代", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("雑費・衣靴代", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("雑費        ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("受診料・他  ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("雑費・会費等", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("勤務時間    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("遅早時間    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("特休日数    ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.drop("集計区分－２        ", axis=1)
    df_sesaku_k_5 = df_sesaku_k_5.sum()
    df_sesaku_k_5.to_csv(
        "c:/test/切削/D.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 切削_間接6
    df_sesaku_k_6 = df_sesaku_m.groupby("区分").get_group("間接6")
    df_sesaku_k_6 = df_sesaku_k_6.drop("所属1", axis=1)
    member = df_sesaku_k_6["所属2"] > 0
    df_sesaku_k_6.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_sesaku_k_6["所属3"] > 0
    df_sesaku_k_6.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_sesaku_k_6["出勤日数    "]
    df_sesaku_k_6.insert(2, "実在籍者", real_member)
    time_yukyu = df_sesaku_k_6["有休日数    "] * 8
    df_sesaku_k_6.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_sesaku_k_6.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_sesaku_k_6["欠勤日数    "] * 8
    df_sesaku_k_6.insert(5, "欠勤時間", abs_time)
    work_time = df_sesaku_k_6["勤務時間    "]
    df_sesaku_k_6.insert(6, "勤務時間", work_time)
    late_early_time = df_sesaku_k_6["遅早時間    "]
    df_sesaku_k_6.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_sesaku_k_6["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_sesaku_k_6.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_sesaku_k_6.insert(17, "実労働時間", real_work_time)
    zure_time = df_sesaku_k_6["ズレ時間    "]
    df_sesaku_k_6.insert(18, "ズレ時間", zure_time)
    overtime = df_sesaku_k_6["残業時間    "] + df_sesaku_k_6["深夜残業    "]
    df_sesaku_k_6.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_sesaku_k_6.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_sesaku_k_6["法外休出    "]
    df_sesaku_k_6.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_sesaku_k_6.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_sesaku_k_6["法定休出    "]
    df_sesaku_k_6.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_sesaku_k_6.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_sesaku_k_6["６０Ｈ超    "]
    df_sesaku_k_6.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_sesaku_k_6["代休時間    "] + df_sesaku_k_6["深夜代休    "]
    df_sesaku_k_6.insert(31, "代休時間", holiday_time)
    df_sesaku_k_6.insert(32, "応援時間", 0)
    total_work_time = (
        df_sesaku_k_6["勤務時間    "]
        + df_sesaku_k_6["残業時間    "]
        + df_sesaku_k_6["法外休出    "]
        + df_sesaku_k_6["法定休出    "]
    )
    df_sesaku_k_6.insert(33, "総労働時間", total_work_time)
    basic_salary = df_sesaku_k_6["基 本 給    "] + df_sesaku_k_6["支給額"]
    df_sesaku_k_6.insert(35, "基本給", basic_salary)
    post_allowance = df_sesaku_k_6["役職手当    "]
    df_sesaku_k_6.insert(36, "役職手当", post_allowance)
    sales_allowance = df_sesaku_k_6["営業手当    "]
    df_sesaku_k_6.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_sesaku_k_6["地域手当    "]
    df_sesaku_k_6.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_sesaku_k_6["特殊手当    "]
    df_sesaku_k_6.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_sesaku_k_6["特別技技手当"]
    df_sesaku_k_6.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_sesaku_k_6["調整手当    "]
    df_sesaku_k_6.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_sesaku_k_6["別居手当    "]
    df_sesaku_k_6.insert(42, "別居手当", sep_allowance)
    com_allowance = df_sesaku_k_6["通勤手当    "]
    df_sesaku_k_6.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_sesaku_k_6.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_sesaku_k_6["残業手当    "]
    df_sesaku_k_6.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_sesaku_k_6["休出手当    "]
    df_sesaku_k_6.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_sesaku_k_6["深夜勤務手当"]
    df_sesaku_k_6.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_sesaku_k_6["交替時差手当"]
    df_sesaku_k_6.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_sesaku_k_6["休業手当    "]
    df_sesaku_k_6.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_sesaku_k_6["休業控除    "]
    df_sesaku_k_6.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_sesaku_k_6["代 休 他    "]
    df_sesaku_k_6.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_sesaku_k_6["欠勤控除    "] + df_sesaku_k_6["遅早控除    "]
    df_sesaku_k_6.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_sesaku_k_6["精 算 分    "]
    df_sesaku_k_6.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_sesaku_k_6.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_sesaku_k_6.insert(55, "総支給額", total)
    df_sesaku_k_6.insert(56, "応援時間額", 0)
    df_sesaku_k_6.insert(57, "役員振替", 0)
    df_sesaku_k_6.insert(58, "部門振替", 0)
    df_sesaku_k_6.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_sesaku_k_6 = df_sesaku_k_6.drop("所属2", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("所属3", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("社員ｺｰﾄﾞ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("区分", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("出勤日数    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("有休日数    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("欠勤日数    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("残業時間    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("深夜残業    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("法外休出    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("法定休出    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("代休時間    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("深夜代休    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("６０Ｈ超    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("別居手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("通勤手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("特別技技手当", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("特殊手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("地域手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("営業手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("役職手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("調整手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("基 本 給    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("残業手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("休出手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("深夜勤務手当", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("交替時差手当", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("休業手当    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("代 休 他    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("欠勤控除    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("遅早控除    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("精 算 分    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("支給合計額  ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("休業控除    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("支給額", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("ズレ時間    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("雑費・食事代", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("雑費・衣靴代", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("雑費        ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("受診料・他  ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("雑費・会費等", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("勤務時間    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("遅早時間    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("特休日数    ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.drop("集計区分－２        ", axis=1)
    df_sesaku_k_6 = df_sesaku_k_6.sum()
    df_sesaku_k_6.to_csv(
        "c:/test/切削/E.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 切削_直接1
    df_sesaku_t_1 = df_sesaku_m.groupby("区分").get_group("直接1")
    df_sesaku_t_1 = df_sesaku_t_1.drop("所属1", axis=1)
    member = df_sesaku_t_1["所属2"] > 0
    df_sesaku_t_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_sesaku_t_1["所属3"] > 0
    df_sesaku_t_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_sesaku_t_1["出勤日数    "]
    df_sesaku_t_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_sesaku_t_1["有休日数    "] * 8
    df_sesaku_t_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_sesaku_t_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_sesaku_t_1["欠勤日数    "] * 8
    df_sesaku_t_1.insert(5, "欠勤時間", abs_time)
    work_time = df_sesaku_t_1["勤務時間    "]
    df_sesaku_t_1.insert(6, "勤務時間", work_time)
    late_early_time = df_sesaku_t_1["遅早時間    "]
    df_sesaku_t_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_sesaku_t_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_sesaku_t_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_sesaku_t_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_sesaku_t_1["ズレ時間    "]
    df_sesaku_t_1.insert(18, "ズレ時間", zure_time)
    overtime = df_sesaku_t_1["残業時間    "] + df_sesaku_t_1["深夜残業    "]
    df_sesaku_t_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_sesaku_t_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_sesaku_t_1["法外休出    "]
    df_sesaku_t_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_sesaku_t_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_sesaku_t_1["法定休出    "]
    df_sesaku_t_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_sesaku_t_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_sesaku_t_1["６０Ｈ超    "]
    df_sesaku_t_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_sesaku_t_1["代休時間    "] + df_sesaku_t_1["深夜代休    "]
    df_sesaku_t_1.insert(31, "代休時間", holiday_time)
    df_sesaku_t_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_sesaku_t_1["勤務時間    "]
        + df_sesaku_t_1["残業時間    "]
        + df_sesaku_t_1["法外休出    "]
        + df_sesaku_t_1["法定休出    "]
    )
    df_sesaku_t_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_sesaku_t_1["基 本 給    "] + df_sesaku_t_1["支給額"]
    df_sesaku_t_1.insert(35, "基本給", basic_salary)
    post_allowance = df_sesaku_t_1["役職手当    "]
    df_sesaku_t_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_sesaku_t_1["営業手当    "]
    df_sesaku_t_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_sesaku_t_1["地域手当    "]
    df_sesaku_t_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_sesaku_t_1["特殊手当    "]
    df_sesaku_t_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_sesaku_t_1["特別技技手当"]
    df_sesaku_t_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_sesaku_t_1["調整手当    "]
    df_sesaku_t_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_sesaku_t_1["別居手当    "]
    df_sesaku_t_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_sesaku_t_1["通勤手当    "]
    df_sesaku_t_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_sesaku_t_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_sesaku_t_1["残業手当    "]
    df_sesaku_t_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_sesaku_t_1["休出手当    "]
    df_sesaku_t_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_sesaku_t_1["深夜勤務手当"]
    df_sesaku_t_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_sesaku_t_1["交替時差手当"]
    df_sesaku_t_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_sesaku_t_1["休業手当    "]
    df_sesaku_t_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_sesaku_t_1["休業控除    "]
    df_sesaku_t_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_sesaku_t_1["代 休 他    "]
    df_sesaku_t_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_sesaku_t_1["欠勤控除    "] + df_sesaku_t_1["遅早控除    "]
    df_sesaku_t_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_sesaku_t_1["精 算 分    "]
    df_sesaku_t_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_sesaku_t_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_sesaku_t_1.insert(55, "総支給額", total)
    df_sesaku_t_1.insert(56, "応援時間額", 0)
    df_sesaku_t_1.insert(57, "役員振替", 0)
    df_sesaku_t_1.insert(58, "部門振替", 0)
    df_sesaku_t_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_sesaku_t_1 = df_sesaku_t_1.drop("所属2", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("所属3", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("区分", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("出勤日数    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("有休日数    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("欠勤日数    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("残業時間    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("深夜残業    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("法外休出    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("法定休出    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("代休時間    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("深夜代休    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("６０Ｈ超    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("別居手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("通勤手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("特別技技手当", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("特殊手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("地域手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("営業手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("役職手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("調整手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("基 本 給    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("残業手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("休出手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("深夜勤務手当", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("交替時差手当", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("休業手当    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("代 休 他    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("欠勤控除    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("遅早控除    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("精 算 分    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("支給合計額  ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("休業控除    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("支給額", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("ズレ時間    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("雑費・食事代", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("雑費・衣靴代", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("雑費        ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("受診料・他  ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("雑費・会費等", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("勤務時間    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("遅早時間    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("特休日数    ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.drop("集計区分－２        ", axis=1)
    df_sesaku_t_1 = df_sesaku_t_1.sum()
    df_sesaku_t_1.to_csv(
        "c:/test/切削/F.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 切削_直接2
    df_sesaku_t_2 = df_sesaku_m.groupby("区分").get_group("直接2")
    df_sesaku_t_2 = df_sesaku_t_2.drop("所属1", axis=1)
    member = df_sesaku_t_2["所属2"] > 0
    df_sesaku_t_2.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_sesaku_t_2["所属3"] > 0
    df_sesaku_t_2.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_sesaku_t_2["出勤日数    "]
    df_sesaku_t_2.insert(2, "実在籍者", real_member)
    time_yukyu = df_sesaku_t_2["有休日数    "] * 8
    df_sesaku_t_2.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_sesaku_t_2.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_sesaku_t_2["欠勤日数    "] * 8
    df_sesaku_t_2.insert(5, "欠勤時間", abs_time)
    work_time = df_sesaku_t_2["勤務時間    "]
    df_sesaku_t_2.insert(6, "勤務時間", work_time)
    late_early_time = df_sesaku_t_2["遅早時間    "]
    df_sesaku_t_2.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_sesaku_t_2["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_sesaku_t_2.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_sesaku_t_2.insert(17, "実労働時間", real_work_time)
    zure_time = df_sesaku_t_2["ズレ時間    "]
    df_sesaku_t_2.insert(18, "ズレ時間", zure_time)
    overtime = df_sesaku_t_2["残業時間    "] + df_sesaku_t_2["深夜残業    "]
    df_sesaku_t_2.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_sesaku_t_2.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_sesaku_t_2["法外休出    "]
    df_sesaku_t_2.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_sesaku_t_2.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_sesaku_t_2["法定休出    "]
    df_sesaku_t_2.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_sesaku_t_2.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_sesaku_t_2["６０Ｈ超    "]
    df_sesaku_t_2.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_sesaku_t_2["代休時間    "] + df_sesaku_t_2["深夜代休    "]
    df_sesaku_t_2.insert(31, "代休時間", holiday_time)
    df_sesaku_t_2.insert(32, "応援時間", 0)
    total_work_time = (
        df_sesaku_t_2["勤務時間    "]
        + df_sesaku_t_2["残業時間    "]
        + df_sesaku_t_2["法外休出    "]
        + df_sesaku_t_2["法定休出    "]
    )
    df_sesaku_t_2.insert(33, "総労働時間", total_work_time)
    basic_salary = df_sesaku_t_2["基 本 給    "] + df_sesaku_t_2["支給額"]
    df_sesaku_t_2.insert(35, "基本給", basic_salary)
    post_allowance = df_sesaku_t_2["役職手当    "]
    df_sesaku_t_2.insert(36, "役職手当", post_allowance)
    sales_allowance = df_sesaku_t_2["営業手当    "]
    df_sesaku_t_2.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_sesaku_t_2["地域手当    "]
    df_sesaku_t_2.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_sesaku_t_2["特殊手当    "]
    df_sesaku_t_2.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_sesaku_t_2["特別技技手当"]
    df_sesaku_t_2.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_sesaku_t_2["調整手当    "]
    df_sesaku_t_2.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_sesaku_t_2["別居手当    "]
    df_sesaku_t_2.insert(42, "別居手当", sep_allowance)
    com_allowance = df_sesaku_t_2["通勤手当    "]
    df_sesaku_t_2.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_sesaku_t_2.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_sesaku_t_2["残業手当    "]
    df_sesaku_t_2.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_sesaku_t_2["休出手当    "]
    df_sesaku_t_2.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_sesaku_t_2["深夜勤務手当"]
    df_sesaku_t_2.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_sesaku_t_2["交替時差手当"]
    df_sesaku_t_2.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_sesaku_t_2["休業手当    "]
    df_sesaku_t_2.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_sesaku_t_2["休業控除    "]
    df_sesaku_t_2.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_sesaku_t_2["代 休 他    "]
    df_sesaku_t_2.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_sesaku_t_2["欠勤控除    "] + df_sesaku_t_2["遅早控除    "]
    df_sesaku_t_2.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_sesaku_t_2["精 算 分    "]
    df_sesaku_t_2.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_sesaku_t_2.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_sesaku_t_2.insert(55, "総支給額", total)
    df_sesaku_t_2.insert(56, "応援時間額", 0)
    df_sesaku_t_2.insert(57, "役員振替", 0)
    df_sesaku_t_2.insert(58, "部門振替", 0)
    df_sesaku_t_2.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_sesaku_t_2 = df_sesaku_t_2.drop("所属2", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("所属3", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("社員ｺｰﾄﾞ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("区分", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("出勤日数    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("有休日数    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("欠勤日数    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("残業時間    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("深夜残業    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("法外休出    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("法定休出    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("代休時間    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("深夜代休    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("６０Ｈ超    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("別居手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("通勤手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("特別技技手当", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("特殊手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("地域手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("営業手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("役職手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("調整手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("基 本 給    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("残業手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("休出手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("深夜勤務手当", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("交替時差手当", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("休業手当    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("代 休 他    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("欠勤控除    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("遅早控除    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("精 算 分    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("支給合計額  ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("休業控除    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("支給額", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("ズレ時間    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("雑費・食事代", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("雑費・衣靴代", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("雑費        ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("受診料・他  ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("雑費・会費等", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("勤務時間    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("遅早時間    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("特休日数    ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.drop("集計区分－２        ", axis=1)
    df_sesaku_t_2 = df_sesaku_t_2.sum()
    df_sesaku_t_2.to_csv(
        "c:/test/切削/G.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 切削_直接4
    df_sesaku_t_4 = df_sesaku_m.groupby("区分").get_group("直接4")
    df_sesaku_t_4 = df_sesaku_t_4.drop("所属1", axis=1)
    member = df_sesaku_t_4["所属2"] > 0
    df_sesaku_t_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_sesaku_t_4["所属3"] > 0
    df_sesaku_t_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_sesaku_t_4["出勤日数    "]
    df_sesaku_t_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_sesaku_t_4["有休日数    "] * 8
    df_sesaku_t_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_sesaku_t_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_sesaku_t_4["欠勤日数    "] * 8
    df_sesaku_t_4.insert(5, "欠勤時間", abs_time)
    work_time = df_sesaku_t_4["勤務時間    "]
    df_sesaku_t_4.insert(6, "勤務時間", work_time)
    late_early_time = df_sesaku_t_4["遅早時間    "]
    df_sesaku_t_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_sesaku_t_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_sesaku_t_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_sesaku_t_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_sesaku_t_4["ズレ時間    "]
    df_sesaku_t_4.insert(18, "ズレ時間", zure_time)
    overtime = df_sesaku_t_4["残業時間    "] + df_sesaku_t_4["深夜残業    "]
    df_sesaku_t_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_sesaku_t_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_sesaku_t_4["法外休出    "]
    df_sesaku_t_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_sesaku_t_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_sesaku_t_4["法定休出    "]
    df_sesaku_t_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_sesaku_t_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_sesaku_t_4["６０Ｈ超    "]
    df_sesaku_t_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_sesaku_t_4["代休時間    "] + df_sesaku_t_4["深夜代休    "]
    df_sesaku_t_4.insert(31, "代休時間", holiday_time)
    df_sesaku_t_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_sesaku_t_4["勤務時間    "]
        + df_sesaku_t_4["残業時間    "]
        + df_sesaku_t_4["法外休出    "]
        + df_sesaku_t_4["法定休出    "]
    )
    df_sesaku_t_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_sesaku_t_4["基 本 給    "] + df_sesaku_t_4["支給額"]
    df_sesaku_t_4.insert(35, "基本給", basic_salary)
    post_allowance = df_sesaku_t_4["役職手当    "]
    df_sesaku_t_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_sesaku_t_4["営業手当    "]
    df_sesaku_t_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_sesaku_t_4["地域手当    "]
    df_sesaku_t_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_sesaku_t_4["特殊手当    "]
    df_sesaku_t_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_sesaku_t_4["特別技技手当"]
    df_sesaku_t_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_sesaku_t_4["調整手当    "]
    df_sesaku_t_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_sesaku_t_4["別居手当    "]
    df_sesaku_t_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_sesaku_t_4["通勤手当    "]
    df_sesaku_t_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_sesaku_t_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_sesaku_t_4["残業手当    "]
    df_sesaku_t_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_sesaku_t_4["休出手当    "]
    df_sesaku_t_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_sesaku_t_4["深夜勤務手当"]
    df_sesaku_t_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_sesaku_t_4["交替時差手当"]
    df_sesaku_t_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_sesaku_t_4["休業手当    "]
    df_sesaku_t_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_sesaku_t_4["休業控除    "]
    df_sesaku_t_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_sesaku_t_4["代 休 他    "]
    df_sesaku_t_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_sesaku_t_4["欠勤控除    "] + df_sesaku_t_4["遅早控除    "]
    df_sesaku_t_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_sesaku_t_4["精 算 分    "]
    df_sesaku_t_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_sesaku_t_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_sesaku_t_4.insert(55, "総支給額", total)
    df_sesaku_t_4.insert(56, "応援時間額", 0)
    df_sesaku_t_4.insert(57, "役員振替", 0)
    df_sesaku_t_4.insert(58, "部門振替", 0)
    df_sesaku_t_4.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_sesaku_t_4 = df_sesaku_t_4.drop("所属2", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("所属3", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("区分", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("出勤日数    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("有休日数    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("欠勤日数    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("残業時間    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("深夜残業    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("法外休出    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("法定休出    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("代休時間    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("深夜代休    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("６０Ｈ超    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("別居手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("通勤手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("特別技技手当", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("特殊手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("地域手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("営業手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("役職手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("調整手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("基 本 給    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("残業手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("休出手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("深夜勤務手当", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("交替時差手当", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("休業手当    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("代 休 他    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("欠勤控除    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("遅早控除    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("精 算 分    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("支給合計額  ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("休業控除    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("支給額", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("ズレ時間    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("雑費・食事代", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("雑費・衣靴代", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("雑費        ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("受診料・他  ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("雑費・会費等", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("勤務時間    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("遅早時間    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("特休日数    ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.drop("集計区分－２        ", axis=1)
    df_sesaku_t_4 = df_sesaku_t_4.sum()
    df_sesaku_t_4.to_csv(
        "c:/test/切削/H.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # AC_間接1
    df_ac_k_1 = df_ac_m.groupby("区分").get_group("間接1")
    df_ac_k_1 = df_ac_k_1.drop("所属1", axis=1)
    member = df_ac_k_1["所属2"] > 0
    df_ac_k_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ac_k_1["所属3"] > 0
    df_ac_k_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ac_k_1["出勤日数    "]
    df_ac_k_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_ac_k_1["有休日数    "] * 8
    df_ac_k_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ac_k_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ac_k_1["欠勤日数    "] * 8
    df_ac_k_1.insert(5, "欠勤時間", abs_time)
    work_time = df_ac_k_1["勤務時間    "]
    df_ac_k_1.insert(6, "勤務時間", work_time)
    late_early_time = df_ac_k_1["遅早時間    "]
    df_ac_k_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ac_k_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ac_k_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ac_k_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_ac_k_1["ズレ時間    "]
    df_ac_k_1.insert(18, "ズレ時間", zure_time)
    overtime = df_ac_k_1["残業時間    "] + df_ac_k_1["深夜残業    "]
    df_ac_k_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ac_k_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ac_k_1["法外休出    "]
    df_ac_k_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ac_k_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ac_k_1["法定休出    "]
    df_ac_k_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ac_k_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ac_k_1["６０Ｈ超    "]
    df_ac_k_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ac_k_1["代休時間    "] + df_ac_k_1["深夜代休    "]
    df_ac_k_1.insert(31, "代休時間", holiday_time)
    df_ac_k_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_ac_k_1["勤務時間    "]
        + df_ac_k_1["残業時間    "]
        + df_ac_k_1["法外休出    "]
        + df_ac_k_1["法定休出    "]
    )
    df_ac_k_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ac_k_1["基 本 給    "] + df_ac_k_1["支給額"]
    df_ac_k_1.insert(35, "基本給", basic_salary)
    post_allowance = df_ac_k_1["役職手当    "]
    df_ac_k_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ac_k_1["営業手当    "]
    df_ac_k_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ac_k_1["地域手当    "]
    df_ac_k_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ac_k_1["特殊手当    "]
    df_ac_k_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ac_k_1["特別技技手当"]
    df_ac_k_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ac_k_1["調整手当    "]
    df_ac_k_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ac_k_1["別居手当    "]
    df_ac_k_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ac_k_1["通勤手当    "]
    df_ac_k_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_ac_k_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ac_k_1["残業手当    "]
    df_ac_k_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ac_k_1["休出手当    "]
    df_ac_k_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ac_k_1["深夜勤務手当"]
    df_ac_k_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ac_k_1["交替時差手当"]
    df_ac_k_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ac_k_1["休業手当    "]
    df_ac_k_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_ac_k_1["休業控除    "]
    df_ac_k_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ac_k_1["代 休 他    "]
    df_ac_k_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ac_k_1["欠勤控除    "] + df_ac_k_1["遅早控除    "]
    df_ac_k_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_ac_k_1["精 算 分    "]
    df_ac_k_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_ac_k_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_ac_k_1.insert(55, "総支給額", total)
    df_ac_k_1.insert(56, "応援時間額", 0)
    df_ac_k_1.insert(57, "役員振替", 0)
    df_ac_k_1.insert(58, "部門振替", 0)
    df_ac_k_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ac_k_1 = df_ac_k_1.drop("所属2", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("所属3", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("区分", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("出勤日数    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("有休日数    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("欠勤日数    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("残業時間    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("深夜残業    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("法外休出    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("法定休出    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("代休時間    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("深夜代休    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("６０Ｈ超    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("別居手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("通勤手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("特別技技手当", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("特殊手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("地域手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("営業手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("役職手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("調整手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("基 本 給    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("残業手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("休出手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("深夜勤務手当", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("交替時差手当", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("休業手当    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("代 休 他    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("欠勤控除    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("遅早控除    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("精 算 分    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("支給合計額  ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("休業控除    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("支給額", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("ズレ時間    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("雑費・食事代", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("雑費・衣靴代", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("雑費        ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("受診料・他  ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("雑費・会費等", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("勤務時間    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("遅早時間    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("特休日数    ", axis=1)
    df_ac_k_1 = df_ac_k_1.drop("集計区分－２        ", axis=1)
    df_ac_k_1 = df_ac_k_1.sum()
    df_ac_k_1.to_csv(
        "c:/test/AC/A.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # AC_間接2
    df_ac_k_2 = df_ac_m.groupby("区分").get_group("間接2")
    df_ac_k_2 = df_ac_k_2.drop("所属1", axis=1)
    member = df_ac_k_2["所属2"] > 0
    df_ac_k_2.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ac_k_2["所属3"] > 0
    df_ac_k_2.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ac_k_2["出勤日数    "]
    df_ac_k_2.insert(2, "実在籍者", real_member)
    time_yukyu = df_ac_k_2["有休日数    "] * 8
    df_ac_k_2.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ac_k_2.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ac_k_2["欠勤日数    "] * 8
    df_ac_k_2.insert(5, "欠勤時間", abs_time)
    work_time = df_ac_k_2["勤務時間    "]
    df_ac_k_2.insert(6, "勤務時間", work_time)
    late_early_time = df_ac_k_2["遅早時間    "]
    df_ac_k_2.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ac_k_2["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ac_k_2.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ac_k_2.insert(17, "実労働時間", real_work_time)
    zure_time = df_ac_k_2["ズレ時間    "]
    df_ac_k_2.insert(18, "ズレ時間", zure_time)
    overtime = df_ac_k_2["残業時間    "] + df_ac_k_2["深夜残業    "]
    df_ac_k_2.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ac_k_2.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ac_k_2["法外休出    "]
    df_ac_k_2.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ac_k_2.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ac_k_2["法定休出    "]
    df_ac_k_2.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ac_k_2.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ac_k_2["６０Ｈ超    "]
    df_ac_k_2.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ac_k_2["代休時間    "] + df_ac_k_2["深夜代休    "]
    df_ac_k_2.insert(31, "代休時間", holiday_time)
    df_ac_k_2.insert(32, "応援時間", 0)
    total_work_time = (
        df_ac_k_2["勤務時間    "]
        + df_ac_k_2["残業時間    "]
        + df_ac_k_2["法外休出    "]
        + df_ac_k_2["法定休出    "]
    )
    df_ac_k_2.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ac_k_2["基 本 給    "] + df_ac_k_2["支給額"]
    df_ac_k_2.insert(35, "基本給", basic_salary)
    post_allowance = df_ac_k_2["役職手当    "]
    df_ac_k_2.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ac_k_2["営業手当    "]
    df_ac_k_2.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ac_k_2["地域手当    "]
    df_ac_k_2.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ac_k_2["特殊手当    "]
    df_ac_k_2.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ac_k_2["特別技技手当"]
    df_ac_k_2.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ac_k_2["調整手当    "]
    df_ac_k_2.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ac_k_2["別居手当    "]
    df_ac_k_2.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ac_k_2["通勤手当    "]
    df_ac_k_2.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_ac_k_2.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ac_k_2["残業手当    "]
    df_ac_k_2.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ac_k_2["休出手当    "]
    df_ac_k_2.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ac_k_2["深夜勤務手当"]
    df_ac_k_2.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ac_k_2["交替時差手当"]
    df_ac_k_2.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ac_k_2["休業手当    "]
    df_ac_k_2.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_ac_k_2["休業控除    "]
    df_ac_k_2.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ac_k_2["代 休 他    "]
    df_ac_k_2.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ac_k_2["欠勤控除    "] + df_ac_k_2["遅早控除    "]
    df_ac_k_2.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_ac_k_2["精 算 分    "]
    df_ac_k_2.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_ac_k_2.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_ac_k_2.insert(55, "総支給額", total)
    df_ac_k_2.insert(56, "応援時間額", 0)
    df_ac_k_2.insert(57, "役員振替", 0)
    df_ac_k_2.insert(58, "部門振替", 0)
    df_ac_k_2.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ac_k_2 = df_ac_k_2.drop("所属2", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("所属3", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("社員ｺｰﾄﾞ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("区分", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("出勤日数    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("有休日数    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("欠勤日数    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("残業時間    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("深夜残業    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("法外休出    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("法定休出    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("代休時間    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("深夜代休    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("６０Ｈ超    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("別居手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("通勤手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("特別技技手当", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("特殊手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("地域手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("営業手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("役職手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("調整手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("基 本 給    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("残業手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("休出手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("深夜勤務手当", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("交替時差手当", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("休業手当    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("代 休 他    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("欠勤控除    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("遅早控除    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("精 算 分    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("支給合計額  ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("休業控除    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("支給額", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("ズレ時間    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("雑費・食事代", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("雑費・衣靴代", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("雑費        ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("受診料・他  ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("雑費・会費等", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("勤務時間    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("遅早時間    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("特休日数    ", axis=1)
    df_ac_k_2 = df_ac_k_2.drop("集計区分－２        ", axis=1)
    df_ac_k_2 = df_ac_k_2.sum()
    df_ac_k_2.to_csv(
        "c:/test/AC/B.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # AC_間接4
    df_ac_k_4 = df_ac_m.groupby("区分").get_group("間接4")
    df_ac_k_4 = df_ac_k_4.drop("所属1", axis=1)
    member = df_ac_k_4["所属2"] > 0
    df_ac_k_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ac_k_4["所属3"] > 0
    df_ac_k_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ac_k_4["出勤日数    "]
    df_ac_k_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_ac_k_4["有休日数    "] * 8
    df_ac_k_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ac_k_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ac_k_4["欠勤日数    "] * 8
    df_ac_k_4.insert(5, "欠勤時間", abs_time)
    work_time = df_ac_k_4["勤務時間    "]
    df_ac_k_4.insert(6, "勤務時間", work_time)
    late_early_time = df_ac_k_4["遅早時間    "]
    df_ac_k_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ac_k_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ac_k_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ac_k_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_ac_k_4["ズレ時間    "]
    df_ac_k_4.insert(18, "ズレ時間", zure_time)
    overtime = df_ac_k_4["残業時間    "] + df_ac_k_4["深夜残業    "]
    df_ac_k_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ac_k_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ac_k_4["法外休出    "]
    df_ac_k_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ac_k_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ac_k_4["法定休出    "]
    df_ac_k_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ac_k_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ac_k_4["６０Ｈ超    "]
    df_ac_k_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ac_k_4["代休時間    "] + df_ac_k_4["深夜代休    "]
    df_ac_k_4.insert(31, "代休時間", holiday_time)
    df_ac_k_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_ac_k_4["勤務時間    "]
        + df_ac_k_4["残業時間    "]
        + df_ac_k_4["法外休出    "]
        + df_ac_k_4["法定休出    "]
    )
    df_ac_k_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ac_k_4["基 本 給    "] + df_ac_k_4["支給額"]
    df_ac_k_4.insert(35, "基本給", basic_salary)
    post_allowance = df_ac_k_4["役職手当    "]
    df_ac_k_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ac_k_4["営業手当    "]
    df_ac_k_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ac_k_4["地域手当    "]
    df_ac_k_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ac_k_4["特殊手当    "]
    df_ac_k_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ac_k_4["特別技技手当"]
    df_ac_k_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ac_k_4["調整手当    "]
    df_ac_k_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ac_k_4["別居手当    "]
    df_ac_k_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ac_k_4["通勤手当    "]
    df_ac_k_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_ac_k_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ac_k_4["残業手当    "]
    df_ac_k_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ac_k_4["休出手当    "]
    df_ac_k_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ac_k_4["深夜勤務手当"]
    df_ac_k_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ac_k_4["交替時差手当"]
    df_ac_k_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ac_k_4["休業手当    "]
    df_ac_k_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_ac_k_4["休業控除    "]
    df_ac_k_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ac_k_4["代 休 他    "]
    df_ac_k_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ac_k_4["欠勤控除    "] + df_ac_k_4["遅早控除    "]
    df_ac_k_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_ac_k_4["精 算 分    "]
    df_ac_k_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_ac_k_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_ac_k_4.insert(55, "総支給額", total)
    df_ac_k_4.insert(56, "応援時間額", 0)
    df_ac_k_4.insert(57, "役員振替", 0)
    df_ac_k_4.insert(58, "部門振替", 0)
    df_ac_k_4.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ac_k_4 = df_ac_k_4.drop("所属2", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("所属3", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("区分", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("出勤日数    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("有休日数    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("欠勤日数    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("残業時間    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("深夜残業    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("法外休出    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("法定休出    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("代休時間    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("深夜代休    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("６０Ｈ超    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("別居手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("通勤手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("特別技技手当", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("特殊手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("地域手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("営業手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("役職手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("調整手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("基 本 給    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("残業手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("休出手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("深夜勤務手当", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("交替時差手当", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("休業手当    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("代 休 他    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("欠勤控除    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("遅早控除    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("精 算 分    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("支給合計額  ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("休業控除    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("支給額", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("ズレ時間    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("雑費・食事代", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("雑費・衣靴代", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("雑費        ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("受診料・他  ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("雑費・会費等", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("勤務時間    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("遅早時間    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("特休日数    ", axis=1)
    df_ac_k_4 = df_ac_k_4.drop("集計区分－２        ", axis=1)
    df_ac_k_4 = df_ac_k_4.sum()
    df_ac_k_4.to_csv(
        "c:/test/AC/C.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # AC_間接5
    df_ac_k_5 = df_ac_m.groupby("区分").get_group("間接5")
    df_ac_k_5 = df_ac_k_5.drop("所属1", axis=1)
    member = df_ac_k_5["所属2"] > 0
    df_ac_k_5.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ac_k_5["所属3"] > 0
    df_ac_k_5.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ac_k_5["出勤日数    "]
    df_ac_k_5.insert(2, "実在籍者", real_member)
    time_yukyu = df_ac_k_5["有休日数    "] * 8
    df_ac_k_5.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ac_k_5.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ac_k_5["欠勤日数    "] * 8
    df_ac_k_5.insert(5, "欠勤時間", abs_time)
    work_time = df_ac_k_5["勤務時間    "]
    df_ac_k_5.insert(6, "勤務時間", work_time)
    late_early_time = df_ac_k_5["遅早時間    "]
    df_ac_k_5.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ac_k_5["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ac_k_5.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ac_k_5.insert(17, "実労働時間", real_work_time)
    zure_time = df_ac_k_5["ズレ時間    "]
    df_ac_k_5.insert(18, "ズレ時間", zure_time)
    overtime = df_ac_k_5["残業時間    "] + df_ac_k_5["深夜残業    "]
    df_ac_k_5.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ac_k_5.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ac_k_5["法外休出    "]
    df_ac_k_5.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ac_k_5.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ac_k_5["法定休出    "]
    df_ac_k_5.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ac_k_5.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ac_k_5["６０Ｈ超    "]
    df_ac_k_5.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ac_k_5["代休時間    "] + df_ac_k_5["深夜代休    "]
    df_ac_k_5.insert(31, "代休時間", holiday_time)
    df_ac_k_5.insert(32, "応援時間", 0)
    total_work_time = (
        df_ac_k_5["勤務時間    "]
        + df_ac_k_5["残業時間    "]
        + df_ac_k_5["法外休出    "]
        + df_ac_k_5["法定休出    "]
    )
    df_ac_k_5.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ac_k_5["基 本 給    "] + df_ac_k_5["支給額"]
    df_ac_k_5.insert(35, "基本給", basic_salary)
    post_allowance = df_ac_k_5["役職手当    "]
    df_ac_k_5.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ac_k_5["営業手当    "]
    df_ac_k_5.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ac_k_5["地域手当    "]
    df_ac_k_5.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ac_k_5["特殊手当    "]
    df_ac_k_5.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ac_k_5["特別技技手当"]
    df_ac_k_5.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ac_k_5["調整手当    "]
    df_ac_k_5.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ac_k_5["別居手当    "]
    df_ac_k_5.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ac_k_5["通勤手当    "]
    df_ac_k_5.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_ac_k_5.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ac_k_5["残業手当    "]
    df_ac_k_5.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ac_k_5["休出手当    "]
    df_ac_k_5.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ac_k_5["深夜勤務手当"]
    df_ac_k_5.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ac_k_5["交替時差手当"]
    df_ac_k_5.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ac_k_5["休業手当    "]
    df_ac_k_5.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_ac_k_5["休業控除    "]
    df_ac_k_5.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ac_k_5["代 休 他    "]
    df_ac_k_5.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ac_k_5["欠勤控除    "] + df_ac_k_5["遅早控除    "]
    df_ac_k_5.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_ac_k_5["精 算 分    "]
    df_ac_k_5.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_ac_k_5.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_ac_k_5.insert(55, "総支給額", total)
    df_ac_k_5.insert(56, "応援時間額", 0)
    df_ac_k_5.insert(57, "役員振替", 0)
    df_ac_k_5.insert(58, "部門振替", 0)
    df_ac_k_5.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ac_k_5 = df_ac_k_5.drop("所属2", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("所属3", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("社員ｺｰﾄﾞ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("区分", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("出勤日数    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("有休日数    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("欠勤日数    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("残業時間    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("深夜残業    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("法外休出    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("法定休出    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("代休時間    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("深夜代休    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("６０Ｈ超    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("別居手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("通勤手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("特別技技手当", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("特殊手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("地域手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("営業手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("役職手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("調整手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("基 本 給    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("残業手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("休出手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("深夜勤務手当", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("交替時差手当", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("休業手当    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("代 休 他    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("欠勤控除    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("遅早控除    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("精 算 分    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("支給合計額  ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("休業控除    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("支給額", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("ズレ時間    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("雑費・食事代", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("雑費・衣靴代", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("雑費        ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("受診料・他  ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("雑費・会費等", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("勤務時間    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("遅早時間    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("特休日数    ", axis=1)
    df_ac_k_5 = df_ac_k_5.drop("集計区分－２        ", axis=1)
    df_ac_k_5 = df_ac_k_5.sum()
    df_ac_k_5.to_csv(
        "c:/test/AC/D.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # AC_直接1
    df_ac_t_1 = df_ac_m.groupby("区分").get_group("直接1")
    df_ac_t_1 = df_ac_t_1.drop("所属1", axis=1)
    member = df_ac_t_1["所属2"] > 0
    df_ac_t_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ac_t_1["所属3"] > 0
    df_ac_t_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ac_t_1["出勤日数    "]
    df_ac_t_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_ac_t_1["有休日数    "] * 8
    df_ac_t_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ac_t_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ac_t_1["欠勤日数    "] * 8
    df_ac_t_1.insert(5, "欠勤時間", abs_time)
    work_time = df_ac_t_1["勤務時間    "]
    df_ac_t_1.insert(6, "勤務時間", work_time)
    late_early_time = df_ac_t_1["遅早時間    "]
    df_ac_t_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ac_t_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ac_t_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ac_t_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_ac_t_1["ズレ時間    "]
    df_ac_t_1.insert(18, "ズレ時間", zure_time)
    overtime = df_ac_t_1["残業時間    "] + df_ac_t_1["深夜残業    "]
    df_ac_t_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ac_t_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ac_t_1["法外休出    "]
    df_ac_t_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ac_t_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ac_t_1["法定休出    "]
    df_ac_t_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ac_t_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ac_t_1["６０Ｈ超    "]
    df_ac_t_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ac_t_1["代休時間    "] + df_ac_t_1["深夜代休    "]
    df_ac_t_1.insert(31, "代休時間", holiday_time)
    df_ac_t_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_ac_t_1["勤務時間    "]
        + df_ac_t_1["残業時間    "]
        + df_ac_t_1["法外休出    "]
        + df_ac_t_1["法定休出    "]
    )
    df_ac_t_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ac_t_1["基 本 給    "] + df_ac_t_1["支給額"]
    df_ac_t_1.insert(35, "基本給", basic_salary)
    post_allowance = df_ac_t_1["役職手当    "]
    df_ac_t_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ac_t_1["営業手当    "]
    df_ac_t_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ac_t_1["地域手当    "]
    df_ac_t_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ac_t_1["特殊手当    "]
    df_ac_t_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ac_t_1["特別技技手当"]
    df_ac_t_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ac_t_1["調整手当    "]
    df_ac_t_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ac_t_1["別居手当    "]
    df_ac_t_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ac_t_1["通勤手当    "]
    df_ac_t_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_ac_t_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ac_t_1["残業手当    "]
    df_ac_t_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ac_t_1["休出手当    "]
    df_ac_t_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ac_t_1["深夜勤務手当"]
    df_ac_t_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ac_t_1["交替時差手当"]
    df_ac_t_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ac_t_1["休業手当    "]
    df_ac_t_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_ac_t_1["休業控除    "]
    df_ac_t_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ac_t_1["代 休 他    "]
    df_ac_t_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ac_t_1["欠勤控除    "] + df_ac_t_1["遅早控除    "]
    df_ac_t_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_ac_t_1["精 算 分    "]
    df_ac_t_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_ac_t_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_ac_t_1.insert(55, "総支給額", total)
    df_ac_t_1.insert(56, "応援時間額", 0)
    df_ac_t_1.insert(57, "役員振替", 0)
    df_ac_t_1.insert(58, "部門振替", 0)
    df_ac_t_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ac_t_1 = df_ac_t_1.drop("所属2", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("所属3", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("区分", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("出勤日数    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("有休日数    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("欠勤日数    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("残業時間    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("深夜残業    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("法外休出    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("法定休出    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("代休時間    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("深夜代休    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("６０Ｈ超    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("別居手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("通勤手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("特別技技手当", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("特殊手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("地域手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("営業手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("役職手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("調整手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("基 本 給    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("残業手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("休出手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("深夜勤務手当", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("交替時差手当", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("休業手当    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("代 休 他    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("欠勤控除    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("遅早控除    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("精 算 分    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("支給合計額  ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("休業控除    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("支給額", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("ズレ時間    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("雑費・食事代", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("雑費・衣靴代", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("雑費        ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("受診料・他  ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("雑費・会費等", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("勤務時間    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("遅早時間    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("特休日数    ", axis=1)
    df_ac_t_1 = df_ac_t_1.drop("集計区分－２        ", axis=1)
    df_ac_t_1 = df_ac_t_1.sum()
    df_ac_t_1.to_csv(
        "c:/test/AC/E.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # AC_直接4
    df_ac_t_4 = df_ac_m.groupby("区分").get_group("直接4")
    df_ac_t_4 = df_ac_t_4.drop("所属1", axis=1)
    member = df_ac_t_4["所属2"] > 0
    df_ac_t_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_ac_t_4["所属3"] > 0
    df_ac_t_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_ac_t_4["出勤日数    "]
    df_ac_t_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_ac_t_4["有休日数    "] * 8
    df_ac_t_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_ac_t_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_ac_t_4["欠勤日数    "] * 8
    df_ac_t_4.insert(5, "欠勤時間", abs_time)
    work_time = df_ac_t_4["勤務時間    "]
    df_ac_t_4.insert(6, "勤務時間", work_time)
    late_early_time = df_ac_t_4["遅早時間    "]
    df_ac_t_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_ac_t_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_ac_t_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_ac_t_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_ac_t_4["ズレ時間    "]
    df_ac_t_4.insert(18, "ズレ時間", zure_time)
    overtime = df_ac_t_4["残業時間    "] + df_ac_t_4["深夜残業    "]
    df_ac_t_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_ac_t_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_ac_t_4["法外休出    "]
    df_ac_t_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_ac_t_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_ac_t_4["法定休出    "]
    df_ac_t_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_ac_t_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_ac_t_4["６０Ｈ超    "]
    df_ac_t_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_ac_t_4["代休時間    "] + df_ac_t_4["深夜代休    "]
    df_ac_t_4.insert(31, "代休時間", holiday_time)
    df_ac_t_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_ac_t_4["勤務時間    "]
        + df_ac_t_4["残業時間    "]
        + df_ac_t_4["法外休出    "]
        + df_ac_t_4["法定休出    "]
    )
    df_ac_t_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_ac_t_4["基 本 給    "] + df_ac_t_4["支給額"]
    df_ac_t_4.insert(35, "基本給", basic_salary)
    post_allowance = df_ac_t_4["役職手当    "]
    df_ac_t_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_ac_t_4["営業手当    "]
    df_ac_t_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_ac_t_4["地域手当    "]
    df_ac_t_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_ac_t_4["特殊手当    "]
    df_ac_t_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_ac_t_4["特別技技手当"]
    df_ac_t_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_ac_t_4["調整手当    "]
    df_ac_t_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_ac_t_4["別居手当    "]
    df_ac_t_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_ac_t_4["通勤手当    "]
    df_ac_t_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_ac_t_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_ac_t_4["残業手当    "]
    df_ac_t_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_ac_t_4["休出手当    "]
    df_ac_t_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_ac_t_4["深夜勤務手当"]
    df_ac_t_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_ac_t_4["交替時差手当"]
    df_ac_t_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_ac_t_4["休業手当    "]
    df_ac_t_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_ac_t_4["休業控除    "]
    df_ac_t_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_ac_t_4["代 休 他    "]
    df_ac_t_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_ac_t_4["欠勤控除    "] + df_ac_t_4["遅早控除    "]
    df_ac_t_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_ac_t_4["精 算 分    "]
    df_ac_t_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_ac_t_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_ac_t_4.insert(55, "総支給額", total)
    df_ac_t_4.insert(56, "応援時間額", 0)
    df_ac_t_4.insert(57, "役員振替", 0)
    df_ac_t_4.insert(58, "部門振替", 0)
    df_ac_t_4.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_ac_t_4 = df_ac_t_4.drop("所属2", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("所属3", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("区分", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("出勤日数    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("有休日数    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("欠勤日数    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("残業時間    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("深夜残業    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("法外休出    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("法定休出    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("代休時間    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("深夜代休    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("６０Ｈ超    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("別居手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("通勤手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("特別技技手当", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("特殊手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("地域手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("営業手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("役職手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("調整手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("基 本 給    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("残業手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("休出手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("深夜勤務手当", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("交替時差手当", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("休業手当    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("代 休 他    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("欠勤控除    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("遅早控除    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("精 算 分    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("支給合計額  ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("休業控除    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("支給額", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("ズレ時間    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("雑費・食事代", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("雑費・衣靴代", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("雑費        ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("受診料・他  ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("雑費・会費等", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("勤務時間    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("遅早時間    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("特休日数    ", axis=1)
    df_ac_t_4 = df_ac_t_4.drop("集計区分－２        ", axis=1)
    df_ac_t_4 = df_ac_t_4.sum()
    df_ac_t_4.to_csv(
        "c:/test/AC/F.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # PC_間接1
    df_pc_k_1 = df_pc_m.groupby("区分").get_group("間接1")
    df_pc_k_1 = df_pc_k_1.drop("所属1", axis=1)
    member = df_pc_k_1["所属2"] > 0
    df_pc_k_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_pc_k_1["所属3"] > 0
    df_pc_k_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_pc_k_1["出勤日数    "]
    df_pc_k_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_pc_k_1["有休日数    "] * 8
    df_pc_k_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_pc_k_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_pc_k_1["欠勤日数    "] * 8
    df_pc_k_1.insert(5, "欠勤時間", abs_time)
    work_time = df_pc_k_1["勤務時間    "]
    df_pc_k_1.insert(6, "勤務時間", work_time)
    late_early_time = df_pc_k_1["遅早時間    "]
    df_pc_k_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_pc_k_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_pc_k_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_pc_k_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_pc_k_1["ズレ時間    "]
    df_pc_k_1.insert(18, "ズレ時間", zure_time)
    overtime = df_pc_k_1["残業時間    "] + df_pc_k_1["深夜残業    "]
    df_pc_k_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_pc_k_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_pc_k_1["法外休出    "]
    df_pc_k_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_pc_k_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_pc_k_1["法定休出    "]
    df_pc_k_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_pc_k_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_pc_k_1["６０Ｈ超    "]
    df_pc_k_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_pc_k_1["代休時間    "] + df_pc_k_1["深夜代休    "]
    df_pc_k_1.insert(31, "代休時間", holiday_time)
    df_pc_k_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_pc_k_1["勤務時間    "]
        + df_pc_k_1["残業時間    "]
        + df_pc_k_1["法外休出    "]
        + df_pc_k_1["法定休出    "]
    )
    df_pc_k_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_pc_k_1["基 本 給    "] + df_pc_k_1["支給額"]
    df_pc_k_1.insert(35, "基本給", basic_salary)
    post_allowance = df_pc_k_1["役職手当    "]
    df_pc_k_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_pc_k_1["営業手当    "]
    df_pc_k_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_pc_k_1["地域手当    "]
    df_pc_k_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_pc_k_1["特殊手当    "]
    df_pc_k_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_pc_k_1["特別技技手当"]
    df_pc_k_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_pc_k_1["調整手当    "]
    df_pc_k_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_pc_k_1["別居手当    "]
    df_pc_k_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_pc_k_1["通勤手当    "]
    df_pc_k_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_pc_k_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_pc_k_1["残業手当    "]
    df_pc_k_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_pc_k_1["休出手当    "]
    df_pc_k_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_pc_k_1["深夜勤務手当"]
    df_pc_k_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_pc_k_1["交替時差手当"]
    df_pc_k_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_pc_k_1["休業手当    "]
    df_pc_k_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_pc_k_1["休業控除    "]
    df_pc_k_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_pc_k_1["代 休 他    "]
    df_pc_k_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_pc_k_1["欠勤控除    "] + df_pc_k_1["遅早控除    "]
    df_pc_k_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_pc_k_1["精 算 分    "]
    df_pc_k_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_pc_k_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_pc_k_1.insert(55, "総支給額", total)
    df_pc_k_1.insert(56, "応援時間額", 0)
    df_pc_k_1.insert(57, "役員振替", 0)
    df_pc_k_1.insert(58, "部門振替", 0)
    df_pc_k_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_pc_k_1 = df_pc_k_1.drop("所属2", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("所属3", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("区分", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("出勤日数    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("有休日数    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("欠勤日数    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("残業時間    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("深夜残業    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("法外休出    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("法定休出    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("代休時間    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("深夜代休    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("６０Ｈ超    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("別居手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("通勤手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("特別技技手当", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("特殊手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("地域手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("営業手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("役職手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("調整手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("基 本 給    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("残業手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("休出手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("深夜勤務手当", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("交替時差手当", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("休業手当    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("代 休 他    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("欠勤控除    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("遅早控除    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("精 算 分    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("支給合計額  ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("休業控除    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("支給額", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("ズレ時間    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("雑費・食事代", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("雑費・衣靴代", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("雑費        ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("受診料・他  ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("雑費・会費等", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("勤務時間    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("遅早時間    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("特休日数    ", axis=1)
    df_pc_k_1 = df_pc_k_1.drop("集計区分－２        ", axis=1)
    df_pc_k_1 = df_pc_k_1.sum()
    df_pc_k_1.to_csv(
        "c:/test/PC/A.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # PC_間接2
    df_pc_k_2 = df_pc_m.groupby("区分").get_group("間接2")
    df_pc_k_2 = df_pc_k_2.drop("所属1", axis=1)
    member = df_pc_k_2["所属2"] > 0
    df_pc_k_2.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_pc_k_2["所属3"] > 0
    df_pc_k_2.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_pc_k_2["出勤日数    "]
    df_pc_k_2.insert(2, "実在籍者", real_member)
    time_yukyu = df_pc_k_2["有休日数    "] * 8
    df_pc_k_2.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_pc_k_2.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_pc_k_2["欠勤日数    "] * 8
    df_pc_k_2.insert(5, "欠勤時間", abs_time)
    work_time = df_pc_k_2["勤務時間    "]
    df_pc_k_2.insert(6, "勤務時間", work_time)
    late_early_time = df_pc_k_2["遅早時間    "]
    df_pc_k_2.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_pc_k_2["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_pc_k_2.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_pc_k_2.insert(17, "実労働時間", real_work_time)
    zure_time = df_pc_k_2["ズレ時間    "]
    df_pc_k_2.insert(18, "ズレ時間", zure_time)
    overtime = df_pc_k_2["残業時間    "] + df_pc_k_2["深夜残業    "]
    df_pc_k_2.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_pc_k_2.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_pc_k_2["法外休出    "]
    df_pc_k_2.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_pc_k_2.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_pc_k_2["法定休出    "]
    df_pc_k_2.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_pc_k_2.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_pc_k_2["６０Ｈ超    "]
    df_pc_k_2.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_pc_k_2["代休時間    "] + df_pc_k_2["深夜代休    "]
    df_pc_k_2.insert(31, "代休時間", holiday_time)
    df_pc_k_2.insert(32, "応援時間", 0)
    total_work_time = (
        df_pc_k_2["勤務時間    "]
        + df_pc_k_2["残業時間    "]
        + df_pc_k_2["法外休出    "]
        + df_pc_k_2["法定休出    "]
    )
    df_pc_k_2.insert(33, "総労働時間", total_work_time)
    basic_salary = df_pc_k_2["基 本 給    "] + df_pc_k_2["支給額"]
    df_pc_k_2.insert(35, "基本給", basic_salary)
    post_allowance = df_pc_k_2["役職手当    "]
    df_pc_k_2.insert(36, "役職手当", post_allowance)
    sales_allowance = df_pc_k_2["営業手当    "]
    df_pc_k_2.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_pc_k_2["地域手当    "]
    df_pc_k_2.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_pc_k_2["特殊手当    "]
    df_pc_k_2.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_pc_k_2["特別技技手当"]
    df_pc_k_2.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_pc_k_2["調整手当    "]
    df_pc_k_2.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_pc_k_2["別居手当    "]
    df_pc_k_2.insert(42, "別居手当", sep_allowance)
    com_allowance = df_pc_k_2["通勤手当    "]
    df_pc_k_2.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_pc_k_2.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_pc_k_2["残業手当    "]
    df_pc_k_2.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_pc_k_2["休出手当    "]
    df_pc_k_2.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_pc_k_2["深夜勤務手当"]
    df_pc_k_2.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_pc_k_2["交替時差手当"]
    df_pc_k_2.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_pc_k_2["休業手当    "]
    df_pc_k_2.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_pc_k_2["休業控除    "]
    df_pc_k_2.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_pc_k_2["代 休 他    "]
    df_pc_k_2.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_pc_k_2["欠勤控除    "] + df_pc_k_2["遅早控除    "]
    df_pc_k_2.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_pc_k_2["精 算 分    "]
    df_pc_k_2.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_pc_k_2.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_pc_k_2.insert(55, "総支給額", total)
    df_pc_k_2.insert(56, "応援時間額", 0)
    df_pc_k_2.insert(57, "役員振替", 0)
    df_pc_k_2.insert(58, "部門振替", 0)
    df_pc_k_2.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_pc_k_2 = df_pc_k_2.drop("所属2", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("所属3", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("社員ｺｰﾄﾞ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("区分", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("出勤日数    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("有休日数    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("欠勤日数    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("残業時間    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("深夜残業    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("法外休出    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("法定休出    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("代休時間    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("深夜代休    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("６０Ｈ超    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("別居手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("通勤手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("特別技技手当", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("特殊手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("地域手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("営業手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("役職手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("調整手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("基 本 給    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("残業手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("休出手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("深夜勤務手当", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("交替時差手当", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("休業手当    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("代 休 他    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("欠勤控除    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("遅早控除    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("精 算 分    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("支給合計額  ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("休業控除    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("支給額", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("ズレ時間    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("雑費・食事代", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("雑費・衣靴代", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("雑費        ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("受診料・他  ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("雑費・会費等", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("勤務時間    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("遅早時間    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("特休日数    ", axis=1)
    df_pc_k_2 = df_pc_k_2.drop("集計区分－２        ", axis=1)
    df_pc_k_2 = df_pc_k_2.sum()
    df_pc_k_2.to_csv(
        "c:/test/PC/B.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # PC_間接4
    df_pc_k_4 = df_pc_m.groupby("区分").get_group("間接4")
    df_pc_k_4 = df_pc_k_4.drop("所属1", axis=1)
    member = df_pc_k_4["所属2"] > 0
    df_pc_k_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_pc_k_4["所属3"] > 0
    df_pc_k_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_pc_k_4["出勤日数    "]
    df_pc_k_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_pc_k_4["有休日数    "] * 8
    df_pc_k_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_pc_k_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_pc_k_4["欠勤日数    "] * 8
    df_pc_k_4.insert(5, "欠勤時間", abs_time)
    work_time = df_pc_k_4["勤務時間    "]
    df_pc_k_4.insert(6, "勤務時間", work_time)
    late_early_time = df_pc_k_4["遅早時間    "]
    df_pc_k_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_pc_k_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_pc_k_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_pc_k_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_pc_k_4["ズレ時間    "]
    df_pc_k_4.insert(18, "ズレ時間", zure_time)
    overtime = df_pc_k_4["残業時間    "] + df_pc_k_4["深夜残業    "]
    df_pc_k_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_pc_k_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_pc_k_4["法外休出    "]
    df_pc_k_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_pc_k_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_pc_k_4["法定休出    "]
    df_pc_k_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_pc_k_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_pc_k_4["６０Ｈ超    "]
    df_pc_k_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_pc_k_4["代休時間    "] + df_pc_k_4["深夜代休    "]
    df_pc_k_4.insert(31, "代休時間", holiday_time)
    df_pc_k_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_pc_k_4["勤務時間    "]
        + df_pc_k_4["残業時間    "]
        + df_pc_k_4["法外休出    "]
        + df_pc_k_4["法定休出    "]
    )
    df_pc_k_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_pc_k_4["基 本 給    "] + df_pc_k_4["支給額"]
    df_pc_k_4.insert(35, "基本給", basic_salary)
    post_allowance = df_pc_k_4["役職手当    "]
    df_pc_k_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_pc_k_4["営業手当    "]
    df_pc_k_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_pc_k_4["地域手当    "]
    df_pc_k_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_pc_k_4["特殊手当    "]
    df_pc_k_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_pc_k_4["特別技技手当"]
    df_pc_k_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_pc_k_4["調整手当    "]
    df_pc_k_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_pc_k_4["別居手当    "]
    df_pc_k_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_pc_k_4["通勤手当    "]
    df_pc_k_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_pc_k_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_pc_k_4["残業手当    "]
    df_pc_k_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_pc_k_4["休出手当    "]
    df_pc_k_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_pc_k_4["深夜勤務手当"]
    df_pc_k_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_pc_k_4["交替時差手当"]
    df_pc_k_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_pc_k_4["休業手当    "]
    df_pc_k_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_pc_k_4["休業控除    "]
    df_pc_k_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_pc_k_4["代 休 他    "]
    df_pc_k_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_pc_k_4["欠勤控除    "] + df_pc_k_4["遅早控除    "]
    df_pc_k_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_pc_k_4["精 算 分    "]
    df_pc_k_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_pc_k_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_pc_k_4.insert(55, "総支給額", total)
    df_pc_k_4.insert(56, "応援時間額", 0)
    df_pc_k_4.insert(57, "役員振替", 0)
    df_pc_k_4.insert(58, "部門振替", 0)
    df_pc_k_4.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_pc_k_4 = df_pc_k_4.drop("所属2", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("所属3", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("区分", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("出勤日数    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("有休日数    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("欠勤日数    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("残業時間    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("深夜残業    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("法外休出    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("法定休出    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("代休時間    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("深夜代休    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("６０Ｈ超    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("別居手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("通勤手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("特別技技手当", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("特殊手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("地域手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("営業手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("役職手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("調整手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("基 本 給    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("残業手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("休出手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("深夜勤務手当", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("交替時差手当", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("休業手当    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("代 休 他    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("欠勤控除    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("遅早控除    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("精 算 分    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("支給合計額  ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("休業控除    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("支給額", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("ズレ時間    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("雑費・食事代", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("雑費・衣靴代", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("雑費        ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("受診料・他  ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("雑費・会費等", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("勤務時間    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("遅早時間    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("特休日数    ", axis=1)
    df_pc_k_4 = df_pc_k_4.drop("集計区分－２        ", axis=1)
    df_pc_k_4 = df_pc_k_4.sum()
    df_pc_k_4.to_csv(
        "c:/test/PC/C.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # PC_間接5
    df_pc_k_5 = df_pc_m.groupby("区分").get_group("間接5")
    df_pc_k_5 = df_pc_k_5.drop("所属1", axis=1)
    member = df_pc_k_5["所属2"] > 0
    df_pc_k_5.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_pc_k_5["所属3"] > 0
    df_pc_k_5.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_pc_k_5["出勤日数    "]
    df_pc_k_5.insert(2, "実在籍者", real_member)
    time_yukyu = df_pc_k_5["有休日数    "] * 8
    df_pc_k_5.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_pc_k_5.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_pc_k_5["欠勤日数    "] * 8
    df_pc_k_5.insert(5, "欠勤時間", abs_time)
    work_time = df_pc_k_5["勤務時間    "]
    df_pc_k_5.insert(6, "勤務時間", work_time)
    late_early_time = df_pc_k_5["遅早時間    "]
    df_pc_k_5.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_pc_k_5["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_pc_k_5.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_pc_k_5.insert(17, "実労働時間", real_work_time)
    zure_time = df_pc_k_5["ズレ時間    "]
    df_pc_k_5.insert(18, "ズレ時間", zure_time)
    overtime = df_pc_k_5["残業時間    "] + df_pc_k_5["深夜残業    "]
    df_pc_k_5.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_pc_k_5.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_pc_k_5["法外休出    "]
    df_pc_k_5.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_pc_k_5.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_pc_k_5["法定休出    "]
    df_pc_k_5.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_pc_k_5.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_pc_k_5["６０Ｈ超    "]
    df_pc_k_5.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_pc_k_5["代休時間    "] + df_pc_k_5["深夜代休    "]
    df_pc_k_5.insert(31, "代休時間", holiday_time)
    df_pc_k_5.insert(32, "応援時間", 0)
    total_work_time = (
        df_pc_k_5["勤務時間    "]
        + df_pc_k_5["残業時間    "]
        + df_pc_k_5["法外休出    "]
        + df_pc_k_5["法定休出    "]
    )
    df_pc_k_5.insert(33, "総労働時間", total_work_time)
    basic_salary = df_pc_k_5["基 本 給    "] + df_pc_k_5["支給額"]
    df_pc_k_5.insert(35, "基本給", basic_salary)
    post_allowance = df_pc_k_5["役職手当    "]
    df_pc_k_5.insert(36, "役職手当", post_allowance)
    sales_allowance = df_pc_k_5["営業手当    "]
    df_pc_k_5.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_pc_k_5["地域手当    "]
    df_pc_k_5.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_pc_k_5["特殊手当    "]
    df_pc_k_5.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_pc_k_5["特別技技手当"]
    df_pc_k_5.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_pc_k_5["調整手当    "]
    df_pc_k_5.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_pc_k_5["別居手当    "]
    df_pc_k_5.insert(42, "別居手当", sep_allowance)
    com_allowance = df_pc_k_5["通勤手当    "]
    df_pc_k_5.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_pc_k_5.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_pc_k_5["残業手当    "]
    df_pc_k_5.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_pc_k_5["休出手当    "]
    df_pc_k_5.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_pc_k_5["深夜勤務手当"]
    df_pc_k_5.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_pc_k_5["交替時差手当"]
    df_pc_k_5.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_pc_k_5["休業手当    "]
    df_pc_k_5.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_pc_k_5["休業控除    "]
    df_pc_k_5.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_pc_k_5["代 休 他    "]
    df_pc_k_5.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_pc_k_5["欠勤控除    "] + df_pc_k_5["遅早控除    "]
    df_pc_k_5.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_pc_k_5["精 算 分    "]
    df_pc_k_5.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_pc_k_5.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_pc_k_5.insert(55, "総支給額", total)
    df_pc_k_5.insert(56, "応援時間額", 0)
    df_pc_k_5.insert(57, "役員振替", 0)
    df_pc_k_5.insert(58, "部門振替", 0)
    df_pc_k_5.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_pc_k_5 = df_pc_k_5.drop("所属2", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("所属3", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("社員ｺｰﾄﾞ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("区分", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("出勤日数    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("有休日数    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("欠勤日数    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("残業時間    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("深夜残業    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("法外休出    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("法定休出    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("代休時間    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("深夜代休    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("６０Ｈ超    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("別居手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("通勤手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("特別技技手当", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("特殊手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("地域手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("営業手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("役職手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("調整手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("基 本 給    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("残業手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("休出手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("深夜勤務手当", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("交替時差手当", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("休業手当    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("代 休 他    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("欠勤控除    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("遅早控除    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("精 算 分    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("支給合計額  ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("休業控除    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("支給額", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("ズレ時間    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("雑費・食事代", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("雑費・衣靴代", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("雑費        ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("受診料・他  ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("雑費・会費等", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("勤務時間    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("遅早時間    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("特休日数    ", axis=1)
    df_pc_k_5 = df_pc_k_5.drop("集計区分－２        ", axis=1)
    df_pc_k_5 = df_pc_k_5.sum()
    df_pc_k_5.to_csv(
        "c:/test/PC/D.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # PC_間接6
    df_pc_k_6 = df_pc_m.groupby("区分").get_group("間接6")
    df_pc_k_6 = df_pc_k_6.drop("所属1", axis=1)
    member = df_pc_k_6["所属2"] > 0
    df_pc_k_6.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_pc_k_6["所属3"] > 0
    df_pc_k_6.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_pc_k_6["出勤日数    "]
    df_pc_k_6.insert(2, "実在籍者", real_member)
    time_yukyu = df_pc_k_6["有休日数    "] * 8
    df_pc_k_6.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_pc_k_6.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_pc_k_6["欠勤日数    "] * 8
    df_pc_k_6.insert(5, "欠勤時間", abs_time)
    work_time = df_pc_k_6["勤務時間    "]
    df_pc_k_6.insert(6, "勤務時間", work_time)
    late_early_time = df_pc_k_6["遅早時間    "]
    df_pc_k_6.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_pc_k_6["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_pc_k_6.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_pc_k_6.insert(17, "実労働時間", real_work_time)
    zure_time = df_pc_k_6["ズレ時間    "]
    df_pc_k_6.insert(18, "ズレ時間", zure_time)
    overtime = df_pc_k_6["残業時間    "] + df_pc_k_6["深夜残業    "]
    df_pc_k_6.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_pc_k_6.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_pc_k_6["法外休出    "]
    df_pc_k_6.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_pc_k_6.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_pc_k_6["法定休出    "]
    df_pc_k_6.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_pc_k_6.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_pc_k_6["６０Ｈ超    "]
    df_pc_k_6.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_pc_k_6["代休時間    "] + df_pc_k_6["深夜代休    "]
    df_pc_k_6.insert(31, "代休時間", holiday_time)
    df_pc_k_6.insert(32, "応援時間", 0)
    total_work_time = (
        df_pc_k_6["勤務時間    "]
        + df_pc_k_6["残業時間    "]
        + df_pc_k_6["法外休出    "]
        + df_pc_k_6["法定休出    "]
    )
    df_pc_k_6.insert(33, "総労働時間", total_work_time)
    basic_salary = df_pc_k_6["基 本 給    "] + df_pc_k_6["支給額"]
    df_pc_k_6.insert(35, "基本給", basic_salary)
    post_allowance = df_pc_k_6["役職手当    "]
    df_pc_k_6.insert(36, "役職手当", post_allowance)
    sales_allowance = df_pc_k_6["営業手当    "]
    df_pc_k_6.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_pc_k_6["地域手当    "]
    df_pc_k_6.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_pc_k_6["特殊手当    "]
    df_pc_k_6.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_pc_k_6["特別技技手当"]
    df_pc_k_6.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_pc_k_6["調整手当    "]
    df_pc_k_6.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_pc_k_6["別居手当    "]
    df_pc_k_6.insert(42, "別居手当", sep_allowance)
    com_allowance = df_pc_k_6["通勤手当    "]
    df_pc_k_6.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_pc_k_6.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_pc_k_6["残業手当    "]
    df_pc_k_6.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_pc_k_6["休出手当    "]
    df_pc_k_6.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_pc_k_6["深夜勤務手当"]
    df_pc_k_6.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_pc_k_6["交替時差手当"]
    df_pc_k_6.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_pc_k_6["休業手当    "]
    df_pc_k_6.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_pc_k_6["休業控除    "]
    df_pc_k_6.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_pc_k_6["代 休 他    "]
    df_pc_k_6.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_pc_k_6["欠勤控除    "] + df_pc_k_6["遅早控除    "]
    df_pc_k_6.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_pc_k_6["精 算 分    "]
    df_pc_k_6.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_pc_k_6.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_pc_k_6.insert(55, "総支給額", total)
    df_pc_k_6.insert(56, "応援時間額", 0)
    df_pc_k_6.insert(57, "役員振替", 0)
    df_pc_k_6.insert(58, "部門振替", 0)
    df_pc_k_6.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_pc_k_6 = df_pc_k_6.drop("所属2", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("所属3", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("社員ｺｰﾄﾞ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("区分", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("出勤日数    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("有休日数    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("欠勤日数    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("残業時間    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("深夜残業    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("法外休出    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("法定休出    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("代休時間    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("深夜代休    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("６０Ｈ超    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("別居手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("通勤手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("特別技技手当", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("特殊手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("地域手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("営業手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("役職手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("調整手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("基 本 給    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("残業手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("休出手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("深夜勤務手当", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("交替時差手当", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("休業手当    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("代 休 他    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("欠勤控除    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("遅早控除    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("精 算 分    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("支給合計額  ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("休業控除    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("支給額", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("ズレ時間    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("雑費・食事代", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("雑費・衣靴代", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("雑費        ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("受診料・他  ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("雑費・会費等", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("勤務時間    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("遅早時間    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("特休日数    ", axis=1)
    df_pc_k_6 = df_pc_k_6.drop("集計区分－２        ", axis=1)
    df_pc_k_6 = df_pc_k_6.sum()
    df_pc_k_6.to_csv(
        "c:/test/PC/E.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # PC_直接1
    df_pc_t_1 = df_pc_m.groupby("区分").get_group("直接1")
    df_pc_t_1 = df_pc_t_1.drop("所属1", axis=1)
    member = df_pc_t_1["所属2"] > 0
    df_pc_t_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_pc_t_1["所属3"] > 0
    df_pc_t_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_pc_t_1["出勤日数    "]
    df_pc_t_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_pc_t_1["有休日数    "] * 8
    df_pc_t_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_pc_t_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_pc_t_1["欠勤日数    "] * 8
    df_pc_t_1.insert(5, "欠勤時間", abs_time)
    work_time = df_pc_t_1["勤務時間    "]
    df_pc_t_1.insert(6, "勤務時間", work_time)
    late_early_time = df_pc_t_1["遅早時間    "]
    df_pc_t_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_pc_t_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_pc_t_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_pc_t_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_pc_t_1["ズレ時間    "]
    df_pc_t_1.insert(18, "ズレ時間", zure_time)
    overtime = df_pc_t_1["残業時間    "] + df_pc_t_1["深夜残業    "]
    df_pc_t_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_pc_t_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_pc_t_1["法外休出    "]
    df_pc_t_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_pc_t_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_pc_t_1["法定休出    "]
    df_pc_t_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_pc_t_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_pc_t_1["６０Ｈ超    "]
    df_pc_t_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_pc_t_1["代休時間    "] + df_pc_t_1["深夜代休    "]
    df_pc_t_1.insert(31, "代休時間", holiday_time)
    df_pc_t_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_pc_t_1["勤務時間    "]
        + df_pc_t_1["残業時間    "]
        + df_pc_t_1["法外休出    "]
        + df_pc_t_1["法定休出    "]
    )
    df_pc_t_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_pc_t_1["基 本 給    "] + df_pc_t_1["支給額"]
    df_pc_t_1.insert(35, "基本給", basic_salary)
    post_allowance = df_pc_t_1["役職手当    "]
    df_pc_t_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_pc_t_1["営業手当    "]
    df_pc_t_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_pc_t_1["地域手当    "]
    df_pc_t_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_pc_t_1["特殊手当    "]
    df_pc_t_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_pc_t_1["特別技技手当"]
    df_pc_t_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_pc_t_1["調整手当    "]
    df_pc_t_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_pc_t_1["別居手当    "]
    df_pc_t_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_pc_t_1["通勤手当    "]
    df_pc_t_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_pc_t_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_pc_t_1["残業手当    "]
    df_pc_t_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_pc_t_1["休出手当    "]
    df_pc_t_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_pc_t_1["深夜勤務手当"]
    df_pc_t_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_pc_t_1["交替時差手当"]
    df_pc_t_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_pc_t_1["休業手当    "]
    df_pc_t_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_pc_t_1["休業控除    "]
    df_pc_t_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_pc_t_1["代 休 他    "]
    df_pc_t_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_pc_t_1["欠勤控除    "] + df_pc_t_1["遅早控除    "]
    df_pc_t_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_pc_t_1["精 算 分    "]
    df_pc_t_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_pc_t_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_pc_t_1.insert(55, "総支給額", total)
    df_pc_t_1.insert(56, "応援時間額", 0)
    df_pc_t_1.insert(57, "役員振替", 0)
    df_pc_t_1.insert(58, "部門振替", 0)
    df_pc_t_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_pc_t_1 = df_pc_t_1.drop("所属2", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("所属3", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("区分", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("出勤日数    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("有休日数    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("欠勤日数    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("残業時間    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("深夜残業    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("法外休出    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("法定休出    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("代休時間    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("深夜代休    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("６０Ｈ超    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("別居手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("通勤手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("特別技技手当", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("特殊手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("地域手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("営業手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("役職手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("調整手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("基 本 給    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("残業手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("休出手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("深夜勤務手当", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("交替時差手当", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("休業手当    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("代 休 他    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("欠勤控除    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("遅早控除    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("精 算 分    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("支給合計額  ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("休業控除    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("支給額", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("ズレ時間    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("雑費・食事代", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("雑費・衣靴代", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("雑費        ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("受診料・他  ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("雑費・会費等", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("勤務時間    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("遅早時間    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("特休日数    ", axis=1)
    df_pc_t_1 = df_pc_t_1.drop("集計区分－２        ", axis=1)
    df_pc_t_1 = df_pc_t_1.sum()
    df_pc_t_1.to_csv(
        "c:/test/PC/F.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # PC_直接4
    df_pc_t_4 = df_pc_m.groupby("区分").get_group("直接4")
    df_pc_t_4 = df_pc_t_4.drop("所属1", axis=1)
    member = df_pc_t_4["所属2"] > 0
    df_pc_t_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_pc_t_4["所属3"] > 0
    df_pc_t_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_pc_t_4["出勤日数    "]
    df_pc_t_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_pc_t_4["有休日数    "] * 8
    df_pc_t_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_pc_t_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_pc_t_4["欠勤日数    "] * 8
    df_pc_t_4.insert(5, "欠勤時間", abs_time)
    work_time = df_pc_t_4["勤務時間    "]
    df_pc_t_4.insert(6, "勤務時間", work_time)
    late_early_time = df_pc_t_4["遅早時間    "]
    df_pc_t_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_pc_t_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_pc_t_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_pc_t_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_pc_t_4["ズレ時間    "]
    df_pc_t_4.insert(18, "ズレ時間", zure_time)
    overtime = df_pc_t_4["残業時間    "] + df_pc_t_4["深夜残業    "]
    df_pc_t_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_pc_t_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_pc_t_4["法外休出    "]
    df_pc_t_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_pc_t_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_pc_t_4["法定休出    "]
    df_pc_t_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_pc_t_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_pc_t_4["６０Ｈ超    "]
    df_pc_t_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_pc_t_4["代休時間    "] + df_pc_t_4["深夜代休    "]
    df_pc_t_4.insert(31, "代休時間", holiday_time)
    df_pc_t_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_pc_t_4["勤務時間    "]
        + df_pc_t_4["残業時間    "]
        + df_pc_t_4["法外休出    "]
        + df_pc_t_4["法定休出    "]
    )
    df_pc_t_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_pc_t_4["基 本 給    "] + df_pc_t_4["支給額"]
    df_pc_t_4.insert(35, "基本給", basic_salary)
    post_allowance = df_pc_t_4["役職手当    "]
    df_pc_t_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_pc_t_4["営業手当    "]
    df_pc_t_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_pc_t_4["地域手当    "]
    df_pc_t_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_pc_t_4["特殊手当    "]
    df_pc_t_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_pc_t_4["特別技技手当"]
    df_pc_t_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_pc_t_4["調整手当    "]
    df_pc_t_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_pc_t_4["別居手当    "]
    df_pc_t_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_pc_t_4["通勤手当    "]
    df_pc_t_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_pc_t_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_pc_t_4["残業手当    "]
    df_pc_t_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_pc_t_4["休出手当    "]
    df_pc_t_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_pc_t_4["深夜勤務手当"]
    df_pc_t_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_pc_t_4["交替時差手当"]
    df_pc_t_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_pc_t_4["休業手当    "]
    df_pc_t_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_pc_t_4["休業控除    "]
    df_pc_t_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_pc_t_4["代 休 他    "]
    df_pc_t_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_pc_t_4["欠勤控除    "] + df_pc_t_4["遅早控除    "]
    df_pc_t_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_pc_t_4["精 算 分    "]
    df_pc_t_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_pc_t_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_pc_t_4.insert(55, "総支給額", total)
    df_pc_t_4.insert(56, "応援時間額", 0)
    df_pc_t_4.insert(57, "役員振替", 0)
    df_pc_t_4.insert(58, "部門振替", 0)
    df_pc_t_4.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_pc_t_4 = df_pc_t_4.drop("所属2", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("所属3", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("区分", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("出勤日数    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("有休日数    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("欠勤日数    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("残業時間    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("深夜残業    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("法外休出    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("法定休出    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("代休時間    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("深夜代休    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("６０Ｈ超    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("別居手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("通勤手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("特別技技手当", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("特殊手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("地域手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("営業手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("役職手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("調整手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("基 本 給    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("残業手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("休出手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("深夜勤務手当", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("交替時差手当", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("休業手当    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("代 休 他    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("欠勤控除    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("遅早控除    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("精 算 分    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("支給合計額  ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("休業控除    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("支給額", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("ズレ時間    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("雑費・食事代", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("雑費・衣靴代", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("雑費        ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("受診料・他  ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("雑費・会費等", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("勤務時間    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("遅早時間    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("特休日数    ", axis=1)
    df_pc_t_4 = df_pc_t_4.drop("集計区分－２        ", axis=1)
    df_pc_t_4 = df_pc_t_4.sum()
    df_pc_t_4.to_csv(
        "c:/test/PC/G.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 宮城_間接1
    df_miyagi_k_1 = df_miyagi_m.groupby("区分").get_group("間接1")
    df_miyagi_k_1 = df_miyagi_k_1.drop("所属1", axis=1)
    member = df_miyagi_k_1["所属2"] > 0
    df_miyagi_k_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_miyagi_k_1["所属3"] > 0
    df_miyagi_k_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_miyagi_k_1["出勤日数    "]
    df_miyagi_k_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_miyagi_k_1["有休日数    "] * 8
    df_miyagi_k_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_miyagi_k_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_miyagi_k_1["欠勤日数    "] * 8
    df_miyagi_k_1.insert(5, "欠勤時間", abs_time)
    work_time = df_miyagi_k_1["勤務時間    "]
    df_miyagi_k_1.insert(6, "勤務時間", work_time)
    late_early_time = df_miyagi_k_1["遅早時間    "]
    df_miyagi_k_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_miyagi_k_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_miyagi_k_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_miyagi_k_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_miyagi_k_1["ズレ時間    "]
    df_miyagi_k_1.insert(18, "ズレ時間", zure_time)
    overtime = df_miyagi_k_1["残業時間    "] + df_miyagi_k_1["深夜残業    "]
    df_miyagi_k_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_miyagi_k_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_miyagi_k_1["法外休出    "]
    df_miyagi_k_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_miyagi_k_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_miyagi_k_1["法定休出    "]
    df_miyagi_k_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_miyagi_k_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_miyagi_k_1["６０Ｈ超    "]
    df_miyagi_k_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_miyagi_k_1["代休時間    "] + df_miyagi_k_1["深夜代休    "]
    df_miyagi_k_1.insert(31, "代休時間", holiday_time)
    df_miyagi_k_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_miyagi_k_1["勤務時間    "]
        + df_miyagi_k_1["残業時間    "]
        + df_miyagi_k_1["法外休出    "]
        + df_miyagi_k_1["法定休出    "]
    )
    df_miyagi_k_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_miyagi_k_1["基 本 給    "] + df_miyagi_k_1["支給額"]
    df_miyagi_k_1.insert(35, "基本給", basic_salary)
    post_allowance = df_miyagi_k_1["役職手当    "]
    df_miyagi_k_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_miyagi_k_1["営業手当    "]
    df_miyagi_k_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_miyagi_k_1["地域手当    "]
    df_miyagi_k_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_miyagi_k_1["特殊手当    "]
    df_miyagi_k_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_miyagi_k_1["特別技技手当"]
    df_miyagi_k_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_miyagi_k_1["調整手当    "]
    df_miyagi_k_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_miyagi_k_1["別居手当    "]
    df_miyagi_k_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_miyagi_k_1["通勤手当    "]
    df_miyagi_k_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_miyagi_k_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_miyagi_k_1["残業手当    "]
    df_miyagi_k_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_miyagi_k_1["休出手当    "]
    df_miyagi_k_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_miyagi_k_1["深夜勤務手当"]
    df_miyagi_k_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_miyagi_k_1["交替時差手当"]
    df_miyagi_k_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_miyagi_k_1["休業手当    "]
    df_miyagi_k_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_miyagi_k_1["休業控除    "]
    df_miyagi_k_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_miyagi_k_1["代 休 他    "]
    df_miyagi_k_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_miyagi_k_1["欠勤控除    "] + df_miyagi_k_1["遅早控除    "]
    df_miyagi_k_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_miyagi_k_1["精 算 分    "]
    df_miyagi_k_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_miyagi_k_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_miyagi_k_1.insert(55, "総支給額", total)
    df_miyagi_k_1.insert(56, "応援時間額", 0)
    df_miyagi_k_1.insert(57, "役員振替", 0)
    df_miyagi_k_1.insert(58, "部門振替", 0)
    df_miyagi_k_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_miyagi_k_1 = df_miyagi_k_1.drop("所属2", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("所属3", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("区分", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("出勤日数    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("有休日数    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("欠勤日数    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("残業時間    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("深夜残業    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("法外休出    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("法定休出    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("代休時間    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("深夜代休    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("６０Ｈ超    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("別居手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("通勤手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("特別技技手当", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("特殊手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("地域手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("営業手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("役職手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("調整手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("基 本 給    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("残業手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("休出手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("深夜勤務手当", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("交替時差手当", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("休業手当    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("代 休 他    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("欠勤控除    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("遅早控除    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("精 算 分    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("支給合計額  ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("休業控除    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("支給額", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("ズレ時間    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("雑費・食事代", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("雑費・衣靴代", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("雑費        ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("受診料・他  ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("雑費・会費等", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("勤務時間    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("遅早時間    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("特休日数    ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.drop("集計区分－２        ", axis=1)
    df_miyagi_k_1 = df_miyagi_k_1.sum()
    df_miyagi_k_1.to_csv(
        "c:/test/宮城/A.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 宮城_間接2
    df_miyagi_k_2 = df_miyagi_m.groupby("区分").get_group("間接2")
    df_miyagi_k_2 = df_miyagi_k_2.drop("所属1", axis=1)
    member = df_miyagi_k_2["所属2"] > 0
    df_miyagi_k_2.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_miyagi_k_2["所属3"] > 0
    df_miyagi_k_2.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_miyagi_k_2["出勤日数    "]
    df_miyagi_k_2.insert(2, "実在籍者", real_member)
    time_yukyu = df_miyagi_k_2["有休日数    "] * 8
    df_miyagi_k_2.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_miyagi_k_2.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_miyagi_k_2["欠勤日数    "] * 8
    df_miyagi_k_2.insert(5, "欠勤時間", abs_time)
    work_time = df_miyagi_k_2["勤務時間    "]
    df_miyagi_k_2.insert(6, "勤務時間", work_time)
    late_early_time = df_miyagi_k_2["遅早時間    "]
    df_miyagi_k_2.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_miyagi_k_2["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_miyagi_k_2.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_miyagi_k_2.insert(17, "実労働時間", real_work_time)
    zure_time = df_miyagi_k_2["ズレ時間    "]
    df_miyagi_k_2.insert(18, "ズレ時間", zure_time)
    overtime = df_miyagi_k_2["残業時間    "] + df_miyagi_k_2["深夜残業    "]
    df_miyagi_k_2.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_miyagi_k_2.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_miyagi_k_2["法外休出    "]
    df_miyagi_k_2.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_miyagi_k_2.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_miyagi_k_2["法定休出    "]
    df_miyagi_k_2.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_miyagi_k_2.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_miyagi_k_2["６０Ｈ超    "]
    df_miyagi_k_2.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_miyagi_k_2["代休時間    "] + df_miyagi_k_2["深夜代休    "]
    df_miyagi_k_2.insert(31, "代休時間", holiday_time)
    df_miyagi_k_2.insert(32, "応援時間", 0)
    total_work_time = (
        df_miyagi_k_2["勤務時間    "]
        + df_miyagi_k_2["残業時間    "]
        + df_miyagi_k_2["法外休出    "]
        + df_miyagi_k_2["法定休出    "]
    )
    df_miyagi_k_2.insert(33, "総労働時間", total_work_time)
    basic_salary = df_miyagi_k_2["基 本 給    "] + df_miyagi_k_2["支給額"]
    df_miyagi_k_2.insert(35, "基本給", basic_salary)
    post_allowance = df_miyagi_k_2["役職手当    "]
    df_miyagi_k_2.insert(36, "役職手当", post_allowance)
    sales_allowance = df_miyagi_k_2["営業手当    "]
    df_miyagi_k_2.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_miyagi_k_2["地域手当    "]
    df_miyagi_k_2.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_miyagi_k_2["特殊手当    "]
    df_miyagi_k_2.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_miyagi_k_2["特別技技手当"]
    df_miyagi_k_2.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_miyagi_k_2["調整手当    "]
    df_miyagi_k_2.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_miyagi_k_2["別居手当    "]
    df_miyagi_k_2.insert(42, "別居手当", sep_allowance)
    com_allowance = df_miyagi_k_2["通勤手当    "]
    df_miyagi_k_2.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_miyagi_k_2.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_miyagi_k_2["残業手当    "]
    df_miyagi_k_2.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_miyagi_k_2["休出手当    "]
    df_miyagi_k_2.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_miyagi_k_2["深夜勤務手当"]
    df_miyagi_k_2.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_miyagi_k_2["交替時差手当"]
    df_miyagi_k_2.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_miyagi_k_2["休業手当    "]
    df_miyagi_k_2.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_miyagi_k_2["休業控除    "]
    df_miyagi_k_2.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_miyagi_k_2["代 休 他    "]
    df_miyagi_k_2.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_miyagi_k_2["欠勤控除    "] + df_miyagi_k_2["遅早控除    "]
    df_miyagi_k_2.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_miyagi_k_2["精 算 分    "]
    df_miyagi_k_2.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_miyagi_k_2.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_miyagi_k_2.insert(55, "総支給額", total)
    df_miyagi_k_2.insert(56, "応援時間額", 0)
    df_miyagi_k_2.insert(57, "役員振替", 0)
    df_miyagi_k_2.insert(58, "部門振替", 0)
    df_miyagi_k_2.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_miyagi_k_2 = df_miyagi_k_2.drop("所属2", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("所属3", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("社員ｺｰﾄﾞ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("区分", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("出勤日数    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("有休日数    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("欠勤日数    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("残業時間    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("深夜残業    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("法外休出    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("法定休出    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("代休時間    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("深夜代休    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("６０Ｈ超    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("別居手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("通勤手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("特別技技手当", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("特殊手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("地域手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("営業手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("役職手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("調整手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("基 本 給    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("残業手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("休出手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("深夜勤務手当", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("交替時差手当", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("休業手当    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("代 休 他    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("欠勤控除    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("遅早控除    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("精 算 分    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("支給合計額  ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("休業控除    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("支給額", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("ズレ時間    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("雑費・食事代", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("雑費・衣靴代", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("雑費        ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("受診料・他  ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("雑費・会費等", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("勤務時間    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("遅早時間    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("特休日数    ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.drop("集計区分－２        ", axis=1)
    df_miyagi_k_2 = df_miyagi_k_2.sum()
    df_miyagi_k_2.to_csv(
        "c:/test/宮城/B.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 宮城_間接4
    df_miyagi_k_4 = df_miyagi_m.groupby("区分").get_group("間接4")
    df_miyagi_k_4 = df_miyagi_k_4.drop("所属1", axis=1)
    member = df_miyagi_k_4["所属2"] > 0
    df_miyagi_k_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_miyagi_k_4["所属3"] > 0
    df_miyagi_k_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_miyagi_k_4["出勤日数    "]
    df_miyagi_k_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_miyagi_k_4["有休日数    "] * 8
    df_miyagi_k_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_miyagi_k_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_miyagi_k_4["欠勤日数    "] * 8
    df_miyagi_k_4.insert(5, "欠勤時間", abs_time)
    work_time = df_miyagi_k_4["勤務時間    "]
    df_miyagi_k_4.insert(6, "勤務時間", work_time)
    late_early_time = df_miyagi_k_4["遅早時間    "]
    df_miyagi_k_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_miyagi_k_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_miyagi_k_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_miyagi_k_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_miyagi_k_4["ズレ時間    "]
    df_miyagi_k_4.insert(18, "ズレ時間", zure_time)
    overtime = df_miyagi_k_4["残業時間    "] + df_miyagi_k_4["深夜残業    "]
    df_miyagi_k_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_miyagi_k_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_miyagi_k_4["法外休出    "]
    df_miyagi_k_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_miyagi_k_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_miyagi_k_4["法定休出    "]
    df_miyagi_k_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_miyagi_k_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_miyagi_k_4["６０Ｈ超    "]
    df_miyagi_k_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_miyagi_k_4["代休時間    "] + df_miyagi_k_4["深夜代休    "]
    df_miyagi_k_4.insert(31, "代休時間", holiday_time)
    df_miyagi_k_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_miyagi_k_4["勤務時間    "]
        + df_miyagi_k_4["残業時間    "]
        + df_miyagi_k_4["法外休出    "]
        + df_miyagi_k_4["法定休出    "]
    )
    df_miyagi_k_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_miyagi_k_4["基 本 給    "] + df_miyagi_k_4["支給額"]
    df_miyagi_k_4.insert(35, "基本給", basic_salary)
    post_allowance = df_miyagi_k_4["役職手当    "]
    df_miyagi_k_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_miyagi_k_4["営業手当    "]
    df_miyagi_k_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_miyagi_k_4["地域手当    "]
    df_miyagi_k_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_miyagi_k_4["特殊手当    "]
    df_miyagi_k_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_miyagi_k_4["特別技技手当"]
    df_miyagi_k_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_miyagi_k_4["調整手当    "]
    df_miyagi_k_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_miyagi_k_4["別居手当    "]
    df_miyagi_k_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_miyagi_k_4["通勤手当    "]
    df_miyagi_k_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_miyagi_k_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_miyagi_k_4["残業手当    "]
    df_miyagi_k_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_miyagi_k_4["休出手当    "]
    df_miyagi_k_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_miyagi_k_4["深夜勤務手当"]
    df_miyagi_k_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_miyagi_k_4["交替時差手当"]
    df_miyagi_k_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_miyagi_k_4["休業手当    "]
    df_miyagi_k_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_miyagi_k_4["休業控除    "]
    df_miyagi_k_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_miyagi_k_4["代 休 他    "]
    df_miyagi_k_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_miyagi_k_4["欠勤控除    "] + df_miyagi_k_4["遅早控除    "]
    df_miyagi_k_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_miyagi_k_4["精 算 分    "]
    df_miyagi_k_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_miyagi_k_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_miyagi_k_4.insert(55, "総支給額", total)
    df_miyagi_k_4.insert(56, "応援時間額", 0)
    df_miyagi_k_4.insert(57, "役員振替", 0)
    df_miyagi_k_4.insert(58, "部門振替", 0)
    df_miyagi_k_4.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_miyagi_k_4 = df_miyagi_k_4.drop("所属2", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("所属3", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("区分", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("出勤日数    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("有休日数    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("欠勤日数    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("残業時間    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("深夜残業    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("法外休出    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("法定休出    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("代休時間    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("深夜代休    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("６０Ｈ超    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("別居手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("通勤手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("特別技技手当", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("特殊手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("地域手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("営業手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("役職手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("調整手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("基 本 給    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("残業手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("休出手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("深夜勤務手当", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("交替時差手当", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("休業手当    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("代 休 他    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("欠勤控除    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("遅早控除    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("精 算 分    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("支給合計額  ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("休業控除    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("支給額", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("ズレ時間    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("雑費・食事代", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("雑費・衣靴代", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("雑費        ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("受診料・他  ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("雑費・会費等", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("勤務時間    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("遅早時間    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("特休日数    ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.drop("集計区分－２        ", axis=1)
    df_miyagi_k_4 = df_miyagi_k_4.sum()
    df_miyagi_k_4.to_csv(
        "c:/test/宮城/C.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 宮城_間接6
    df_miyagi_k_6 = df_miyagi_m.groupby("区分").get_group("間接6")
    df_miyagi_k_6 = df_miyagi_k_6.drop("所属1", axis=1)
    member = df_miyagi_k_6["所属2"] > 0
    df_miyagi_k_6.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_miyagi_k_6["所属3"] > 0
    df_miyagi_k_6.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_miyagi_k_6["出勤日数    "]
    df_miyagi_k_6.insert(2, "実在籍者", real_member)
    time_yukyu = df_miyagi_k_6["有休日数    "] * 8
    df_miyagi_k_6.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_miyagi_k_6.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_miyagi_k_6["欠勤日数    "] * 8
    df_miyagi_k_6.insert(5, "欠勤時間", abs_time)
    work_time = df_miyagi_k_6["勤務時間    "]
    df_miyagi_k_6.insert(6, "勤務時間", work_time)
    late_early_time = df_miyagi_k_6["遅早時間    "]
    df_miyagi_k_6.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_miyagi_k_6["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_miyagi_k_6.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_miyagi_k_6.insert(17, "実労働時間", real_work_time)
    zure_time = df_miyagi_k_6["ズレ時間    "]
    df_miyagi_k_6.insert(18, "ズレ時間", zure_time)
    overtime = df_miyagi_k_6["残業時間    "] + df_miyagi_k_6["深夜残業    "]
    df_miyagi_k_6.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_miyagi_k_6.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_miyagi_k_6["法外休出    "]
    df_miyagi_k_6.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_miyagi_k_6.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_miyagi_k_6["法定休出    "]
    df_miyagi_k_6.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_miyagi_k_6.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_miyagi_k_6["６０Ｈ超    "]
    df_miyagi_k_6.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_miyagi_k_6["代休時間    "] + df_miyagi_k_6["深夜代休    "]
    df_miyagi_k_6.insert(31, "代休時間", holiday_time)
    df_miyagi_k_6.insert(32, "応援時間", 0)
    total_work_time = (
        df_miyagi_k_6["勤務時間    "]
        + df_miyagi_k_6["残業時間    "]
        + df_miyagi_k_6["法外休出    "]
        + df_miyagi_k_6["法定休出    "]
    )
    df_miyagi_k_6.insert(33, "総労働時間", total_work_time)
    basic_salary = df_miyagi_k_6["基 本 給    "] + df_miyagi_k_6["支給額"]
    df_miyagi_k_6.insert(35, "基本給", basic_salary)
    post_allowance = df_miyagi_k_6["役職手当    "]
    df_miyagi_k_6.insert(36, "役職手当", post_allowance)
    sales_allowance = df_miyagi_k_6["営業手当    "]
    df_miyagi_k_6.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_miyagi_k_6["地域手当    "]
    df_miyagi_k_6.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_miyagi_k_6["特殊手当    "]
    df_miyagi_k_6.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_miyagi_k_6["特別技技手当"]
    df_miyagi_k_6.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_miyagi_k_6["調整手当    "]
    df_miyagi_k_6.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_miyagi_k_6["別居手当    "]
    df_miyagi_k_6.insert(42, "別居手当", sep_allowance)
    com_allowance = df_miyagi_k_6["通勤手当    "]
    df_miyagi_k_6.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_miyagi_k_6.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_miyagi_k_6["残業手当    "]
    df_miyagi_k_6.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_miyagi_k_6["休出手当    "]
    df_miyagi_k_6.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_miyagi_k_6["深夜勤務手当"]
    df_miyagi_k_6.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_miyagi_k_6["交替時差手当"]
    df_miyagi_k_6.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_miyagi_k_6["休業手当    "]
    df_miyagi_k_6.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_miyagi_k_6["休業控除    "]
    df_miyagi_k_6.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_miyagi_k_6["代 休 他    "]
    df_miyagi_k_6.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_miyagi_k_6["欠勤控除    "] + df_miyagi_k_6["遅早控除    "]
    df_miyagi_k_6.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_miyagi_k_6["精 算 分    "]
    df_miyagi_k_6.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_miyagi_k_6.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_miyagi_k_6.insert(55, "総支給額", total)
    df_miyagi_k_6.insert(56, "応援時間額", 0)
    df_miyagi_k_6.insert(57, "役員振替", 0)
    df_miyagi_k_6.insert(58, "部門振替", 0)
    df_miyagi_k_6.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_miyagi_k_6 = df_miyagi_k_6.drop("所属2", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("所属3", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("社員ｺｰﾄﾞ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("区分", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("出勤日数    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("有休日数    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("欠勤日数    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("残業時間    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("深夜残業    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("法外休出    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("法定休出    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("代休時間    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("深夜代休    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("６０Ｈ超    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("別居手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("通勤手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("特別技技手当", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("特殊手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("地域手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("営業手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("役職手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("調整手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("基 本 給    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("残業手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("休出手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("深夜勤務手当", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("交替時差手当", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("休業手当    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("代 休 他    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("欠勤控除    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("遅早控除    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("精 算 分    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("支給合計額  ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("休業控除    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("支給額", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("ズレ時間    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("雑費・食事代", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("雑費・衣靴代", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("雑費        ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("受診料・他  ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("雑費・会費等", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("勤務時間    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("遅早時間    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("特休日数    ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.drop("集計区分－２        ", axis=1)
    df_miyagi_k_6 = df_miyagi_k_6.sum()
    df_miyagi_k_6.to_csv(
        "c:/test/宮城/D.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 宮城_直接1
    df_miyagi_t_1 = df_miyagi_m.groupby("区分").get_group("直接1")
    df_miyagi_t_1 = df_miyagi_t_1.drop("所属1", axis=1)
    member = df_miyagi_t_1["所属2"] > 0
    df_miyagi_t_1.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_miyagi_t_1["所属3"] > 0
    df_miyagi_t_1.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_miyagi_t_1["出勤日数    "]
    df_miyagi_t_1.insert(2, "実在籍者", real_member)
    time_yukyu = df_miyagi_t_1["有休日数    "] * 8
    df_miyagi_t_1.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_miyagi_t_1.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_miyagi_t_1["欠勤日数    "] * 8
    df_miyagi_t_1.insert(5, "欠勤時間", abs_time)
    work_time = df_miyagi_t_1["勤務時間    "] * 8
    df_miyagi_t_1.insert(6, "勤務時間", work_time)
    late_early_time = df_miyagi_t_1["遅早時間    "]
    df_miyagi_t_1.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_miyagi_t_1["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_miyagi_t_1.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_miyagi_t_1.insert(17, "実労働時間", real_work_time)
    zure_time = df_miyagi_t_1["ズレ時間    "]
    df_miyagi_t_1.insert(18, "ズレ時間", zure_time)
    overtime = df_miyagi_t_1["残業時間    "] + df_miyagi_t_1["深夜残業    "]
    df_miyagi_t_1.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_miyagi_t_1.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_miyagi_t_1["法外休出    "]
    df_miyagi_t_1.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_miyagi_t_1.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_miyagi_t_1["法定休出    "]
    df_miyagi_t_1.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_miyagi_t_1.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_miyagi_t_1["６０Ｈ超    "]
    df_miyagi_t_1.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_miyagi_t_1["代休時間    "] + df_miyagi_t_1["深夜代休    "]
    df_miyagi_t_1.insert(31, "代休時間", holiday_time)
    df_miyagi_t_1.insert(32, "応援時間", 0)
    total_work_time = (
        df_miyagi_t_1["勤務時間    "]
        + df_miyagi_t_1["残業時間    "]
        + df_miyagi_t_1["法外休出    "]
        + df_miyagi_t_1["法定休出    "]
    )
    df_miyagi_t_1.insert(33, "総労働時間", total_work_time)
    basic_salary = df_miyagi_t_1["基 本 給    "] + df_miyagi_t_1["支給額"]
    df_miyagi_t_1.insert(35, "基本給", basic_salary)
    post_allowance = df_miyagi_t_1["役職手当    "]
    df_miyagi_t_1.insert(36, "役職手当", post_allowance)
    sales_allowance = df_miyagi_t_1["営業手当    "]
    df_miyagi_t_1.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_miyagi_t_1["地域手当    "]
    df_miyagi_t_1.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_miyagi_t_1["特殊手当    "]
    df_miyagi_t_1.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_miyagi_t_1["特別技技手当"]
    df_miyagi_t_1.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_miyagi_t_1["調整手当    "]
    df_miyagi_t_1.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_miyagi_t_1["別居手当    "]
    df_miyagi_t_1.insert(42, "別居手当", sep_allowance)
    com_allowance = df_miyagi_t_1["通勤手当    "]
    df_miyagi_t_1.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_miyagi_t_1.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_miyagi_t_1["残業手当    "]
    df_miyagi_t_1.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_miyagi_t_1["休出手当    "]
    df_miyagi_t_1.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_miyagi_t_1["深夜勤務手当"]
    df_miyagi_t_1.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_miyagi_t_1["交替時差手当"]
    df_miyagi_t_1.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_miyagi_t_1["休業手当    "]
    df_miyagi_t_1.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_miyagi_t_1["休業控除    "]
    df_miyagi_t_1.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_miyagi_t_1["代 休 他    "]
    df_miyagi_t_1.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = df_miyagi_t_1["欠勤控除    "] + df_miyagi_t_1["遅早控除    "]
    df_miyagi_t_1.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_miyagi_t_1["精 算 分    "]
    df_miyagi_t_1.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_miyagi_t_1.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_miyagi_t_1.insert(55, "総支給額", total)
    df_miyagi_t_1.insert(56, "応援時間額", 0)
    df_miyagi_t_1.insert(57, "役員振替", 0)
    df_miyagi_t_1.insert(58, "部門振替", 0)
    df_miyagi_t_1.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_miyagi_t_1 = df_miyagi_t_1.drop("所属2", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("所属3", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("社員ｺｰﾄﾞ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("区分", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("出勤日数    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("有休日数    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("欠勤日数    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("残業時間    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("深夜残業    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("法外休出    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("法定休出    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("代休時間    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("深夜代休    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("６０Ｈ超    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("別居手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("通勤手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("特別技技手当", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("特殊手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("地域手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("営業手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("役職手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("調整手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("基 本 給    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("残業手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("休出手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("深夜勤務手当", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("交替時差手当", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("休業手当    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("代 休 他    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("欠勤控除    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("遅早控除    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("精 算 分    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("支給合計額  ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("休業控除    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("支給額", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("ズレ時間    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("雑費・食事代", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("雑費・衣靴代", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("雑費        ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("受診料・他  ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("雑費・会費等", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("勤務時間    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("遅早時間    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("特休日数    ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.drop("集計区分－２        ", axis=1)
    df_miyagi_t_1 = df_miyagi_t_1.sum()
    df_miyagi_t_1.to_csv(
        "c:/test/宮城/E.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 住設_間接2
    df_jyusetu_k_2 = df_jyusetu_m.groupby("区分").get_group("間接2")
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("所属1", axis=1)
    member = df_jyusetu_k_2["所属2"] > 0
    df_jyusetu_k_2.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_jyusetu_k_2["所属3"] > 0
    df_jyusetu_k_2.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_jyusetu_k_2["出勤日数    "]
    df_jyusetu_k_2.insert(2, "実在籍者", real_member)
    time_yukyu = df_jyusetu_k_2["有休日数    "] * 8
    df_jyusetu_k_2.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_jyusetu_k_2.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_jyusetu_k_2["欠勤日数    "] * 8
    df_jyusetu_k_2.insert(5, "欠勤時間", abs_time)
    work_time = df_jyusetu_k_2["勤務時間    "]
    df_jyusetu_k_2.insert(6, "勤務時間", work_time)
    late_early_time = df_jyusetu_k_2["遅早時間    "]
    df_jyusetu_k_2.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_jyusetu_k_2["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_jyusetu_k_2.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_jyusetu_k_2.insert(17, "実労働時間", real_work_time)
    zure_time = df_jyusetu_k_2["ズレ時間    "]
    df_jyusetu_k_2.insert(18, "ズレ時間", zure_time)
    overtime = df_jyusetu_k_2["残業時間    "] + df_jyusetu_k_2["深夜残業    "]
    df_jyusetu_k_2.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_jyusetu_k_2.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_jyusetu_k_2["法外休出    "]
    df_jyusetu_k_2.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_jyusetu_k_2.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_jyusetu_k_2["法定休出    "]
    df_jyusetu_k_2.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_jyusetu_k_2.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_jyusetu_k_2["６０Ｈ超    "]
    df_jyusetu_k_2.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_jyusetu_k_2["代休時間    "] + df_jyusetu_k_2["深夜代休    "]
    df_jyusetu_k_2.insert(31, "代休時間", holiday_time)
    df_jyusetu_k_2.insert(32, "応援時間", 0)
    total_work_time = (
        df_jyusetu_k_2["勤務時間    "]
        + df_jyusetu_k_2["残業時間    "]
        + df_jyusetu_k_2["法外休出    "]
        + df_jyusetu_k_2["法定休出    "]
    )
    df_jyusetu_k_2.insert(33, "総労働時間", total_work_time)
    basic_salary = df_jyusetu_k_2["基 本 給    "] + df_jyusetu_k_2["支給額"]
    df_jyusetu_k_2.insert(35, "基本給", basic_salary)
    post_allowance = df_jyusetu_k_2["役職手当    "]
    df_jyusetu_k_2.insert(36, "役職手当", post_allowance)
    sales_allowance = df_jyusetu_k_2["営業手当    "]
    df_jyusetu_k_2.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_jyusetu_k_2["地域手当    "]
    df_jyusetu_k_2.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_jyusetu_k_2["特殊手当    "]
    df_jyusetu_k_2.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_jyusetu_k_2["特別技技手当"]
    df_jyusetu_k_2.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_jyusetu_k_2["調整手当    "]
    df_jyusetu_k_2.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_jyusetu_k_2["別居手当    "]
    df_jyusetu_k_2.insert(42, "別居手当", sep_allowance)
    com_allowance = df_jyusetu_k_2["通勤手当    "]
    df_jyusetu_k_2.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_jyusetu_k_2.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_jyusetu_k_2["残業手当    "]
    df_jyusetu_k_2.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_jyusetu_k_2["休出手当    "]
    df_jyusetu_k_2.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_jyusetu_k_2["深夜勤務手当"]
    df_jyusetu_k_2.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_jyusetu_k_2["交替時差手当"]
    df_jyusetu_k_2.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_jyusetu_k_2["休業手当    "]
    df_jyusetu_k_2.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_jyusetu_k_2["休業控除    "]
    df_jyusetu_k_2.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_jyusetu_k_2["代 休 他    "]
    df_jyusetu_k_2.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = (
        df_jyusetu_k_2["欠勤控除    "] + df_jyusetu_k_2["遅早控除    "]
    )
    df_jyusetu_k_2.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_jyusetu_k_2["精 算 分    "]
    df_jyusetu_k_2.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_jyusetu_k_2.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_jyusetu_k_2.insert(55, "総支給額", total)
    df_jyusetu_k_2.insert(56, "応援時間額", 0)
    df_jyusetu_k_2.insert(57, "役員振替", 0)
    df_jyusetu_k_2.insert(58, "部門振替", 0)
    df_jyusetu_k_2.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("所属2", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("所属3", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("社員ｺｰﾄﾞ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("区分", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("出勤日数    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("有休日数    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("欠勤日数    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("残業時間    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("深夜残業    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("法外休出    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("法定休出    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("代休時間    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("深夜代休    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("６０Ｈ超    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("別居手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("通勤手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("特別技技手当", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("特殊手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("地域手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("営業手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("役職手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("調整手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("基 本 給    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("残業手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("休出手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("深夜勤務手当", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("交替時差手当", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("休業手当    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("代 休 他    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("欠勤控除    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("遅早控除    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("精 算 分    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("支給合計額  ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("休業控除    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("支給額", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("ズレ時間    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("雑費・食事代", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("雑費・衣靴代", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("雑費        ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("受診料・他  ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("雑費・会費等", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("勤務時間    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("遅早時間    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("特休日数    ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.drop("集計区分－２        ", axis=1)
    df_jyusetu_k_2 = df_jyusetu_k_2.sum()
    df_jyusetu_k_2.to_csv(
        "c:/test/住設/A.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 住設_間接4
    df_jyusetu_k_4 = df_jyusetu_m.groupby("区分").get_group("間接4")
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("所属1", axis=1)
    member = df_jyusetu_k_4["所属2"] > 0
    df_jyusetu_k_4.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_jyusetu_k_4["所属3"] > 0
    df_jyusetu_k_4.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_jyusetu_k_4["出勤日数    "]
    df_jyusetu_k_4.insert(2, "実在籍者", real_member)
    time_yukyu = df_jyusetu_k_4["有休日数    "] * 8
    df_jyusetu_k_4.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_jyusetu_k_4.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_jyusetu_k_4["欠勤日数    "] * 8
    df_jyusetu_k_4.insert(5, "欠勤時間", abs_time)
    work_time = df_jyusetu_k_4["勤務時間    "]
    df_jyusetu_k_4.insert(6, "勤務時間", work_time)
    late_early_time = df_jyusetu_k_4["遅早時間    "]
    df_jyusetu_k_4.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_jyusetu_k_4["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_jyusetu_k_4.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_jyusetu_k_4.insert(17, "実労働時間", real_work_time)
    zure_time = df_jyusetu_k_4["ズレ時間    "]
    df_jyusetu_k_4.insert(18, "ズレ時間", zure_time)
    overtime = df_jyusetu_k_4["残業時間    "] + df_jyusetu_k_4["深夜残業    "]
    df_jyusetu_k_4.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_jyusetu_k_4.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_jyusetu_k_4["法外休出    "]
    df_jyusetu_k_4.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_jyusetu_k_4.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_jyusetu_k_4["法定休出    "]
    df_jyusetu_k_4.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_jyusetu_k_4.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_jyusetu_k_4["６０Ｈ超    "]
    df_jyusetu_k_4.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_jyusetu_k_4["代休時間    "] + df_jyusetu_k_4["深夜代休    "]
    df_jyusetu_k_4.insert(31, "代休時間", holiday_time)
    df_jyusetu_k_4.insert(32, "応援時間", 0)
    total_work_time = (
        df_jyusetu_k_4["勤務時間    "]
        + df_jyusetu_k_4["残業時間    "]
        + df_jyusetu_k_4["法外休出    "]
        + df_jyusetu_k_4["法定休出    "]
    )
    df_jyusetu_k_4.insert(33, "総労働時間", total_work_time)
    basic_salary = df_jyusetu_k_4["基 本 給    "] + df_jyusetu_k_4["支給額"]
    df_jyusetu_k_4.insert(35, "基本給", basic_salary)
    post_allowance = df_jyusetu_k_4["役職手当    "]
    df_jyusetu_k_4.insert(36, "役職手当", post_allowance)
    sales_allowance = df_jyusetu_k_4["営業手当    "]
    df_jyusetu_k_4.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_jyusetu_k_4["地域手当    "]
    df_jyusetu_k_4.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_jyusetu_k_4["特殊手当    "]
    df_jyusetu_k_4.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_jyusetu_k_4["特別技技手当"]
    df_jyusetu_k_4.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_jyusetu_k_4["調整手当    "]
    df_jyusetu_k_4.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_jyusetu_k_4["別居手当    "]
    df_jyusetu_k_4.insert(42, "別居手当", sep_allowance)
    com_allowance = df_jyusetu_k_4["通勤手当    "]
    df_jyusetu_k_4.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_jyusetu_k_4.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_jyusetu_k_4["残業手当    "]
    df_jyusetu_k_4.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_jyusetu_k_4["休出手当    "]
    df_jyusetu_k_4.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_jyusetu_k_4["深夜勤務手当"]
    df_jyusetu_k_4.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_jyusetu_k_4["交替時差手当"]
    df_jyusetu_k_4.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_jyusetu_k_4["休業手当    "]
    df_jyusetu_k_4.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_jyusetu_k_4["休業控除    "]
    df_jyusetu_k_4.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_jyusetu_k_4["代 休 他    "]
    df_jyusetu_k_4.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = (
        df_jyusetu_k_4["欠勤控除    "] + df_jyusetu_k_4["遅早控除    "]
    )
    df_jyusetu_k_4.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_jyusetu_k_4["精 算 分    "]
    df_jyusetu_k_4.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_jyusetu_k_4.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_jyusetu_k_4.insert(55, "総支給額", total)
    df_jyusetu_k_4.insert(56, "応援時間額", 0)
    df_jyusetu_k_4.insert(57, "役員振替", 0)
    df_jyusetu_k_4.insert(58, "部門振替", 0)
    df_jyusetu_k_4.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("所属2", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("所属3", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("社員ｺｰﾄﾞ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("区分", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("出勤日数    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("有休日数    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("欠勤日数    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("残業時間    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("深夜残業    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("法外休出    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("法定休出    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("代休時間    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("深夜代休    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("６０Ｈ超    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("別居手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("通勤手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("特別技技手当", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("特殊手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("地域手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("営業手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("役職手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("調整手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("基 本 給    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("残業手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("休出手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("深夜勤務手当", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("交替時差手当", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("休業手当    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("代 休 他    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("欠勤控除    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("遅早控除    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("精 算 分    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("支給合計額  ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("休業控除    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("支給額", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("ズレ時間    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("雑費・食事代", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("雑費・衣靴代", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("雑費        ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("受診料・他  ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("雑費・会費等", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("勤務時間    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("遅早時間    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("特休日数    ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.drop("集計区分－２        ", axis=1)
    df_jyusetu_k_4 = df_jyusetu_k_4.sum()
    df_jyusetu_k_4.to_csv(
        "c:/test/住設/B.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )
    # 住設_間接6
    df_jyusetu_k_6 = df_jyusetu_m.groupby("区分").get_group("間接6")
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("所属1", axis=1)
    member = df_jyusetu_k_6["所属2"] > 0
    df_jyusetu_k_6.insert(0, "在籍者", member)
    member = len(member)
    under_mastar = df_jyusetu_k_6["所属3"] > 0
    df_jyusetu_k_6.insert(1, "在籍者主幹以下人数", under_mastar)
    under_mastar = (under_mastar == True).sum()
    real_member = df_jyusetu_k_6["出勤日数    "]
    df_jyusetu_k_6.insert(2, "実在籍者", real_member)
    time_yukyu = df_jyusetu_k_6["有休日数    "] * 8
    df_jyusetu_k_6.insert(3, "有休時間", time_yukyu)
    ave_yukyu = time_yukyu / member
    ave_yukyu = ave_yukyu.round(2)
    df_jyusetu_k_6.insert(4, "有休時間在籍者平均", ave_yukyu)
    abs_time = df_jyusetu_k_6["欠勤日数    "] * 8
    df_jyusetu_k_6.insert(5, "欠勤時間", abs_time)
    work_time = df_jyusetu_k_6["勤務時間    "]
    df_jyusetu_k_6.insert(6, "勤務時間", work_time)
    late_early_time = df_jyusetu_k_6["遅早時間    "]
    df_jyusetu_k_6.insert(7, "遅早時間", late_early_time)
    work_rate1 = (real_member * 8) - df_jyusetu_k_6["遅早時間    "]
    work_rate1 = work_rate1.sum()
    work_rate2 = (real_member * 8) * 100
    work_rate2 = work_rate2.sum()
    work_rate = (work_rate1 / work_rate2) * 100
    work_rate = work_rate / member
    df_jyusetu_k_6.insert(16, "出勤率", work_rate)
    real_work_time = real_member * 8
    df_jyusetu_k_6.insert(17, "実労働時間", real_work_time)
    zure_time = df_jyusetu_k_6["ズレ時間    "]
    df_jyusetu_k_6.insert(18, "ズレ時間", zure_time)
    overtime = df_jyusetu_k_6["残業時間    "] + df_jyusetu_k_6["深夜残業    "]
    df_jyusetu_k_6.insert(24, "残業時間", overtime.round(2))
    under_mater_overtime = overtime / under_mastar
    df_jyusetu_k_6.insert(25, "残業時間主幹以下平均", under_mater_overtime.round(2))
    non_leave_time = df_jyusetu_k_6["法外休出    "]
    df_jyusetu_k_6.insert(26, "法定外休出時間", non_leave_time)
    ave_non_leave_time = non_leave_time / under_mastar
    df_jyusetu_k_6.insert(27, "法定外主幹以下平均", ave_non_leave_time.round(2))
    legal_leave_time = df_jyusetu_k_6["法定休出    "]
    df_jyusetu_k_6.insert(28, "法定休出時間", legal_leave_time)
    ave_legal_leave_time = legal_leave_time / under_mastar
    df_jyusetu_k_6.insert(29, "法定主幹以下平均", ave_legal_leave_time.round(2))
    overtime_60 = df_jyusetu_k_6["６０Ｈ超    "]
    df_jyusetu_k_6.insert(30, "時間外60時間超え", overtime_60)
    holiday_time = df_jyusetu_k_6["代休時間    "] + df_jyusetu_k_6["深夜代休    "]
    df_jyusetu_k_6.insert(31, "代休時間", holiday_time)
    df_jyusetu_k_6.insert(32, "応援時間", 0)
    total_work_time = (
        df_jyusetu_k_6["勤務時間    "]
        + df_jyusetu_k_6["残業時間    "]
        + df_jyusetu_k_6["法外休出    "]
        + df_jyusetu_k_6["法定休出    "]
    )
    df_jyusetu_k_6.insert(33, "総労働時間", total_work_time)
    basic_salary = df_jyusetu_k_6["基 本 給    "] + df_jyusetu_k_6["支給額"]
    df_jyusetu_k_6.insert(35, "基本給", basic_salary)
    post_allowance = df_jyusetu_k_6["役職手当    "]
    df_jyusetu_k_6.insert(36, "役職手当", post_allowance)
    sales_allowance = df_jyusetu_k_6["営業手当    "]
    df_jyusetu_k_6.insert(37, "営業手当", sales_allowance)
    aria_allowance = df_jyusetu_k_6["地域手当    "]
    df_jyusetu_k_6.insert(38, "地域手当", aria_allowance)
    spe_allowance = df_jyusetu_k_6["特殊手当    "]
    df_jyusetu_k_6.insert(39, "特別手当", spe_allowance)
    spe_tec_allowance = df_jyusetu_k_6["特別技技手当"]
    df_jyusetu_k_6.insert(40, "特別技技手当 ", spe_tec_allowance)
    adjust_allowance = df_jyusetu_k_6["調整手当    "]
    df_jyusetu_k_6.insert(41, "調整手当", adjust_allowance)
    sep_allowance = df_jyusetu_k_6["別居手当    "]
    df_jyusetu_k_6.insert(42, "別居手当", sep_allowance)
    com_allowance = df_jyusetu_k_6["通勤手当    "]
    df_jyusetu_k_6.insert(43, "通勤手当", com_allowance)
    sub_total_1 = (
        basic_salary
        + post_allowance
        + sales_allowance
        + aria_allowance
        + spe_allowance
        + adjust_allowance
        + sep_allowance
        + com_allowance
    )
    df_jyusetu_k_6.insert(44, "小計 1", sub_total_1)
    overtime_allowance = df_jyusetu_k_6["残業手当    "]
    df_jyusetu_k_6.insert(45, "残業手当", overtime_allowance)
    vacation_allowance = df_jyusetu_k_6["休出手当    "]
    df_jyusetu_k_6.insert(46, "休出手当", vacation_allowance)
    night_work_allowance = df_jyusetu_k_6["深夜勤務手当"]
    df_jyusetu_k_6.insert(47, "深夜勤務手当　", night_work_allowance)
    time_difference_allowance = df_jyusetu_k_6["交替時差手当"]
    df_jyusetu_k_6.insert(48, "交替時差手当　", time_difference_allowance)
    closed_allowance = df_jyusetu_k_6["休業手当    "]
    df_jyusetu_k_6.insert(49, "休業手当", closed_allowance)
    closed_deduction = df_jyusetu_k_6["休業控除    "]
    df_jyusetu_k_6.insert(50, "休業控除", closed_deduction)
    compny_leave_etc = df_jyusetu_k_6["代 休 他    "]
    df_jyusetu_k_6.insert(51, "代休他", compny_leave_etc)
    abs_early_deduction = (
        df_jyusetu_k_6["欠勤控除    "] + df_jyusetu_k_6["遅早控除    "]
    )
    df_jyusetu_k_6.insert(52, "欠勤・遅早控除", abs_early_deduction)
    settlement = df_jyusetu_k_6["精 算 分    "]
    df_jyusetu_k_6.insert(53, "精算分", settlement)
    sub_total_2 = abs_early_deduction + settlement
    df_jyusetu_k_6.insert(54, "小計 2", sub_total_2)
    total = sub_total_1 - sub_total_2
    df_jyusetu_k_6.insert(55, "総支給額", total)
    df_jyusetu_k_6.insert(56, "応援時間額", 0)
    df_jyusetu_k_6.insert(57, "役員振替", 0)
    df_jyusetu_k_6.insert(58, "部門振替", 0)
    df_jyusetu_k_6.insert(59, "合計", 0)
    # 不要フィールドの削除
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("所属2", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("所属3", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("社員ｺｰﾄﾞ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("区分", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("出勤日数    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("有休日数    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("欠勤日数    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("残業時間    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("深夜残業    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("法外休出    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("法定休出    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("代休時間    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("深夜代休    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("６０Ｈ超    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("別居手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("通勤手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("特別技技手当", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("特殊手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("地域手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("営業手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("役職手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("調整手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("基 本 給    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("残業手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("休出手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("深夜勤務手当", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("交替時差手当", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("休業手当    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("代 休 他    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("欠勤控除    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("遅早控除    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("精 算 分    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("支給合計額  ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("休業控除    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("支給額", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("ズレ時間    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("雑費・食事代", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("雑費・衣靴代", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("雑費        ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("受診料・他  ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("雑費・会費等", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("勤務時間    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("遅早時間    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("特休日数    ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.drop("集計区分－２        ", axis=1)
    df_jyusetu_k_6 = df_jyusetu_k_6.sum()
    df_jyusetu_k_6.to_csv(
        "c:/test/住設/C.csv",
        header=True,
        index=False,
        encoding="shift-jis",
    )

    # csvファイルの結合
    # 一般管理
    # パスで指定したファイルの一覧をリスト形式で取得
    csv_files_ipan = glob.glob("c:/test/一般管理/*.csv")

    # CSVファイルの中身を追加していくリストを準備
    data_list_ipan = []

    # 読み込むファイルのリストをスキャンして取得
    for file in csv_files_ipan:
        data_list_ipan.append(pd.read_csv(file))

    # リストを全て列方向に結合
    df_ipan_t = pd.concat(data_list_ipan, axis=1, sort=True)

    # columnsパラメータで列名を設定
    feature_ipan = [
        "役員1",
        "一般間接1",
        "一般間接2",
        "一般間接3",
        "一般間接6",
        "一般販売1",
        "一般販売2",
    ]
    df_ipan_t.columns = feature_ipan

    # 行名の設定
    df_ipan_t = df_ipan_t.rename(
        {
            0: "在籍者",
            1: "在籍者主幹以下人数",
            2: "実在籍者",
            3: "有休時間",
            4: "有休時間在籍者平均",
            5: "欠勤時間",
            6: "勤務時間",
            7: "遅早時間",
            8: "出勤率",
            9: "実労働時間",
            10: "ズレ時間",
            11: "残業時間",
            12: "残業時間主幹以下平均",
            13: "法定外休出時間",
            14: "法定外休出主幹以下平均",
            15: "法定休出時間",
            16: "法定休出主幹以下平均",
            17: "時間外60時間超",
            18: "代休時間",
            19: "応援時間",
            20: "総労働時間",
            21: "基本給",
            22: "役職手当",
            23: "営業手当",
            24: "地域手当",
            25: "特別手当",
            26: "特別技技手当",
            27: "調整手当",
            28: "別居手当",
            29: "通勤手当",
            30: "小計1",
            31: "残業手当",
            32: "休出手当",
            33: "深夜勤務手当",
            34: "交替時差手当",
            35: "休業手当",
            36: "休業控除",
            37: "代休他",
            38: "欠勤・遅早控除",
            39: "精算分",
            40: "小計2",
            41: "総支給額",
            42: "応援時間額",
            43: "役員振替",
            44: "部門振替",
            45: "合計",
        }
    )

    # 整形---------------------
    # ワークブックの生成
    wb = Workbook()
    # ワークシートの生成
    ws = wb.active
    ws.title = "一般管理"  # type:ignore

    # DataFrameを行単位のデータにする
    rows = dataframe_to_rows(df_ipan_t, index=True, header=True)

    # 1セルずつ処理を実行する
    for row_no, row in enumerate(rows, 1):
        for col_no, value in enumerate(row, 1):
            # データを書き込む
            ws.cell(row=row_no, column=col_no, value=value)  # type:ignore

    # 不要な行の削除
    ws.delete_rows(2)  # type:ignore
    # 表示倍率の設定
    ws.sheet_view.zoomScale = 100  # type:ignore
    # 列幅の設定
    ws.column_dimensions["A"].width = 25  # type:ignore
    ws.column_dimensions["B"].width = 12  # type:ignore
    ws.column_dimensions["C"].width = 12  # type:ignore
    ws.column_dimensions["D"].width = 12  # type:ignore
    ws.column_dimensions["E"].width = 12  # type:ignore
    ws.column_dimensions["F"].width = 12  # type:ignore
    ws.column_dimensions["G"].width = 12  # type:ignore
    ws.column_dimensions["H"].width = 12  # type:ignore
    ws.column_dimensions["I"].width = 12  # type:ignore
    ws.column_dimensions["J"].width = 12  # type:ignore
    ws.column_dimensions["K"].width = 12  # type:ignore
    ws.column_dimensions["L"].width = 12  # type:ignore

    # 表示形式
    format = "#,##0"
    for row in ws["B2:L2"]:  # type:ignore
        for cell in row:
            cell.number_format = format

    format = "0.00"
    for row in ws["B4:L9"]:  # type:ignore
        for cell in row:
            cell.number_format = format

    format = "0.00%"
    for row in ws["B10:L10"]:  # type:ignore
        for cell in row:
            cell.number_format = format

    format = "0.00"
    for row in ws["B11:L12"]:  # type:ignore
        for cell in row:
            cell.number_format = format

    format = "0.00"
    for row in ws["B13:L22"]:  # type:ignore
        for cell in row:
            cell.number_format = format

    format = "#,##0"
    for row in ws["B23:L23"]:  # type:ignore
        for cell in row:
            cell.number_format = format

    format = "0.00"
    for row in ws["B24:L26"]:  # type:ignore
        for cell in row:
            cell.number_format = format

    format = "#,##0"
    for row in ws["B23:L48"]:  # type:ignore
        for cell in row:
            cell.number_format = format

    # ヘッダー行のスタイル設定
    header = ws[1]  # type:ignore
    for header_cell in header:
        # フォントを設定する
        header_cell.fill = PatternFill(
            patternType="solid", fgColor="008000"  # type:ignore
        )  # type:ignore
        # 罫線を設定する
        header_cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
    # 中央揃え
    for row in ws["B1:H1"]:  # type:ignore
        for cell in row:
            cell.alignment = Alignment(horizontal="centerContinuous")
    ws["L1"].alignment = Alignment(horizontal="centerContinuous")  # type:ignore
    # 指定した行の背景色を黄色にする
    mylist = [21, 44, 45, 46]

    for list in mylist:
        for row in ws.iter_rows():  # type:ignore
            for cell in row:
                if cell.row == list:
                    cell.fill = PatternFill(
                        fgColor="FFFF00",  # type:ignore
                        bgColor="FFFF00",  # type:ignore
                        fill_type="solid",  # type:ignore
                    )

    # 小計1
    ws["B32"] = "=SUM(B23:B31)"  # type:ignore
    ws["C32"] = "=SUM(C23:C31)"  # type:ignore
    ws["D32"] = "=SUM(D23:D31)"  # type:ignore
    ws["E32"] = "=SUM(E23:E31)"  # type:ignore
    ws["F32"] = "=SUM(F23:F31)"  # type:ignore
    ws["G32"] = "=SUM(G23:G31)"  # type:ignore
    ws["H32"] = "=SUM(H23:H31)"  # type:ignore

    # 小計2
    ws["B42"] = "=SUM(B33:B41)"  # type:ignore
    ws["C42"] = "=SUM(C33:C41)"  # type:ignore
    ws["D42"] = "=SUM(D33:D41)"  # type:ignore
    ws["E42"] = "=SUM(E33:E41)"  # type:ignore
    ws["F42"] = "=SUM(F33:F41)"  # type:ignore
    ws["G42"] = "=SUM(G33:G41)"  # type:ignore
    ws["H42"] = "=SUM(H33:H41)"  # type:ignore

    # 総支給額
    ws["B43"] = "=B32 + B42"  # type:ignore
    ws["C43"] = "=C32 + C42"  # type:ignore
    ws["D43"] = "=D32 + D42"  # type:ignore
    ws["E43"] = "=E32 + E42"  # type:ignore
    ws["F43"] = "=F32 + F42"  # type:ignore
    ws["G43"] = "=G32 + G42"  # type:ignore
    ws["H43"] = "=H32 + H42"  # type:ignore

    # 列合計の埋め込み
    ws["B47"] = "=B43+B44+B45+B46"  # type:ignore
    ws["C47"] = "=C43+C44+C45+C46"  # type:ignore
    ws["D47"] = "=D43+D44+D45+D46"  # type:ignore
    ws["E47"] = "=E43+E44+E45+E46"  # type:ignore
    ws["F47"] = "=F43+F44+F45+F46"  # type:ignore
    ws["G47"] = "=G43+G44+G45+G46"  # type:ignore
    ws["H47"] = "=H43+H44+H45+H46"  # type:ignore

    # 合計列の追加
    ws["L1"] = "【合計】"  # type:ignore

    side1 = Side(border_style="thin", color="000000")
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)

    for row in ws["A2:L47"]:  # type:ignore
        for cell in row:
            cell.border = border_aro

    # 行合計の埋め込み
    ws["L2"] = "=SUM(B2:H2)"  # type:ignore
    ws["L3"] = "=SUM(B3:H3)"  # type:ignore
    ws["L4"] = "=SUM(B4:H4)"  # type:ignore
    ws["L5"] = "=SUM(B5:H5)"  # type:ignore
    ws["L6"] = "=SUM(B6:H6) / 7"  # type:ignore
    ws["L7"] = "=SUM(B7:H7)"  # type:ignore
    ws["L8"] = "=SUM(B8:H8)"  # type:ignore
    ws["L9"] = "=SUM(B9:H9)"  # type:ignore
    ws["L10"] = "=SUM(B10:H10) / 7"  # type:ignore
    ws["L11"] = "=SUM(B11:H11)"  # type:ignore
    ws["L12"] = "=SUM(B12:H12)"  # type:ignore
    ws["L13"] = "=SUM(B13:H13)"  # type:ignore
    ws["L14"] = "=SUM(B14:H14) / 7"  # type:ignore
    ws["L15"] = "=SUM(B15:H15)"  # type:ignore
    ws["L16"] = "=SUM(B16:H16) / 7"  # type:ignore
    ws["L17"] = "=SUM(B17:H17)"  # type:ignore
    ws["L18"] = "=SUM(B18:H18) / 7"  # type:ignore
    ws["L19"] = "=SUM(B19:H19)"  # type:ignore
    ws["L20"] = "=SUM(B20:H20)"  # type:ignore
    ws["L21"] = "=SUM(B21:H21)"  # type:ignore
    ws["L22"] = "=SUM(B22:H22)"  # type:ignore
    ws["L23"] = "=SUM(B23:H23)"  # type:ignore
    ws["L24"] = "=SUM(B24:H24)"  # type:ignore
    ws["L25"] = "=SUM(B25:H25)"  # type:ignore
    ws["L26"] = "=SUM(B26:H26)"  # type:ignore
    ws["L27"] = "=SUM(B27:H27)"  # type:ignore
    ws["L28"] = "=SUM(B28:H28)"  # type:ignore
    ws["L29"] = "=SUM(B29:H29)"  # type:ignore
    ws["L30"] = "=SUM(B30:H30)"  # type:ignore
    ws["L31"] = "=SUM(B31:H31)"  # type:ignore
    ws["L32"] = "=SUM(B32:H32)"  # type:ignore
    ws["L33"] = "=SUM(B33:H33)"  # type:ignore
    ws["L34"] = "=SUM(B34:H34)"  # type:ignore
    ws["L35"] = "=SUM(B35:H35)"  # type:ignore
    ws["L36"] = "=SUM(B36:H36)"  # type:ignore
    ws["L37"] = "=SUM(B37:H37)"  # type:ignore
    ws["L38"] = "=SUM(B38:H38)"  # type:ignore
    ws["L39"] = "=SUM(B39:H39)"  # type:ignore
    ws["L40"] = "=SUM(B40:H40)"  # type:ignore
    ws["L41"] = "=SUM(B41:H41)"  # type:ignore
    ws["L42"] = "=SUM(B42:H42)"  # type:ignore
    ws["L43"] = "=SUM(B43:H43)"  # type:ignore
    ws["L44"] = "=SUM(B44:H44)"  # type:ignore
    ws["L45"] = "=SUM(B45:H45)"  # type:ignore
    ws["L46"] = "=SUM(B46:H46)"  # type:ignore
    ws["L47"] = "=L43+L44+L45+L46"  # type:ignore

    ws.delete_rows(48, 49)  # type:ignore

    # Excelファイルを出力
    wb.save("c:/test/一般管理/一般管理.xlsx")

    # 鍛造

    # パスで指定したファイルの一覧をリスト形式で取得
    csv_files_tanzo = glob.glob("c:/test/鍛造/*.csv")

    # CSVファイルの中身を追加していくリストを表示
    data_list_tanzo = []

    # 読み込むファイルのリストをスキャン
    for file in csv_files_tanzo:
        data_list_tanzo.append(pd.read_csv(file))

    # リストを全て列方向に結合
    df_tanzo_t = pd.concat(data_list_tanzo, axis=1, sort=True)

    # columnsパラメータで列名を設定
    feature_tanzo = [
        "間接1",
        "間接2",
        "間接3",
        "間接4",
        "間接5",
        "間接6",
        "直接1",
        "直接2",
    ]
    df_tanzo_t.columns = feature_tanzo

    # 行名の設定
    df_tanzo_t = df_tanzo_t.rename(
        {
            0: "在籍者",
            1: "在籍者主幹以下人数",
            2: "実在籍者",
            3: "有休時間",
            4: "有休時間在籍者平均",
            5: "欠勤時間",
            6: "勤務時間",
            7: "遅早時間",
            8: "出勤率",
            9: "実労働時間",
            10: "ズレ時間",
            11: "残業時間",
            12: "残業時間主幹以下平均",
            13: "法定外休出時間",
            14: "法定外休出主幹以下平均",
            15: "法定休出時間",
            16: "法定休出主幹以下平均",
            17: "時間外60時間超",
            18: "代休時間",
            19: "応援時間",
            20: "総労働時間",
            21: "基本給",
            22: "役職手当",
            23: "営業手当",
            24: "地域手当",
            25: "特別手当",
            26: "特別技技手当",
            27: "調整手当",
            28: "別居手当",
            29: "通勤手当",
            30: "小計1",
            31: "残業手当",
            32: "休出手当",
            33: "深夜勤務手当",
            34: "交替時差手当",
            35: "休業手当",
            36: "休業控除",
            37: "代休他",
            38: "欠勤・遅早控除",
            39: "精算分",
            40: "小計2",
            41: "総支給額",
            42: "応援時間額",
            43: "役員振替",
            44: "部門振替",
            45: "合計",
        }
    )

    # 整形---------------------
    # ワークブックの生成
    wb = Workbook()  # type:ignore
    # ワークシートの生成
    ws = wb.active
    ws.title = "鍛造工場"  # type:ignore

    # DataFrameを行単位のデータにする
    rows = dataframe_to_rows(df_tanzo_t, index=True, header=True)

    # 1セルずつ処理を実行する
    for row_no, row in enumerate(rows, 1):
        for col_no, value in enumerate(row, 1):
            # データを書き込む
            ws.cell(row=row_no, column=col_no, value=value)  # type:ignore

    # 不要な行の削除
    ws.delete_rows(2)  # type:ignore
    # 列の追加
    ws.insert_cols(8)  # type:ignore
    # 表示倍率の設定
    ws.sheet_view.zoomScale = 100  # type:ignore
    # 列幅の設定
    ws.column_dimensions["A"].width = 25  # type:ignore
    ws.column_dimensions["B"].width = 12  # type:ignore
    ws.column_dimensions["C"].width = 12  # type:ignore
    ws.column_dimensions["D"].width = 12  # type:ignore
    ws.column_dimensions["E"].width = 12  # type:ignore
    ws.column_dimensions["F"].width = 12  # type:ignore
    ws.column_dimensions["G"].width = 12  # type:ignore
    ws.column_dimensions["H"].width = 12  # type:ignore
    ws.column_dimensions["I"].width = 12  # type:ignore
    ws.column_dimensions["J"].width = 12  # type:ignore
    ws.column_dimensions["K"].width = 12  # type:ignore
    ws.column_dimensions["L"].width = 12  # type:ignore
    # 表示形式
    format = "#,##0"
    for row in ws["B2:L2"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B4:L9"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00%"
    for row in ws["B10:L10"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B11:L12"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B13:L22"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B23:L23"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B24:L26"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B27:L51"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    # ヘッダー行のスタイル設定
    header = ws[1]  # type:ignore
    for header_cell in header:
        # フォントを設定する
        header_cell.fill = PatternFill(
            patternType="solid", fgColor="008000"  # type:ignore
        )  # type:ignore
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
    # 中央揃え
    for row in ws["B1:L1"]:  # type:ignore
        for cell in row:
            cell.alignment = Alignment(horizontal="centerContinuous")
        # 指定した行の背景色を黄色にする
        mylist = [21, 44, 45, 46]
        for list in mylist:
            for row in ws.iter_rows():  # type:ignore
                for cell in row:
                    if cell.row == list:
                        cell.fill = PatternFill(
                            fgColor="FFFF00",  # type:ignore
                            bgColor="FFFF00",  # type:ignore
                            fill_type="solid",  # type:ignore
                        )

    # 小計1
    ws["B32"] = "=SUM(B23:B31)"  # type:ignore
    ws["C32"] = "=SUM(C23:C31)"  # type:ignore
    ws["D32"] = "=SUM(D23:D31)"  # type:ignore
    ws["E32"] = "=SUM(E23:E31)"  # type:ignore
    ws["F32"] = "=SUM(F23:F31)"  # type:ignore
    ws["G32"] = "=SUM(G23:G31)"  # type:ignore
    ws["H32"] = "=SUM(H23:H31)"  # type:ignore
    ws["I32"] = "=SUM(I23:I31)"  # type:ignore
    ws["J32"] = "=SUM(J23:J31)"  # type:ignore
    ws["K32"] = "=SUM(K23:K31)"  # type:ignore
    # 小計2
    ws["B42"] = "=SUM(B33:B41)"  # type:ignore
    ws["C42"] = "=SUM(C33:C41)"  # type:ignore
    ws["D42"] = "=SUM(D33:D41)"  # type:ignore
    ws["E42"] = "=SUM(E33:E41)"  # type:ignore
    ws["F42"] = "=SUM(F33:F41)"  # type:ignore
    ws["G42"] = "=SUM(G33:G41)"  # type:ignore
    ws["H42"] = "=SUM(H33:H41)"  # type:ignore
    ws["I42"] = "=SUM(I33:I41)"  # type:ignore
    ws["J42"] = "=SUM(J33:J41)"  # type:ignore
    ws["K42"] = "=SUM(K33:K41)"  # type:ignore
    # 総支給額
    ws["B43"] = "=B32 + B42"  # type:ignore
    ws["C43"] = "=C32 + C42"  # type:ignore
    ws["D43"] = "=D32 + D42"  # type:ignore
    ws["E43"] = "=E32 + E42"  # type:ignore
    ws["F43"] = "=F32 + F42"  # type:ignore
    ws["G43"] = "=G32 + G42"  # type:ignore
    ws["H43"] = "=H32 + H42"  # type:ignore
    ws["I43"] = "=I32 + I42"  # type:ignore
    ws["J43"] = "=J32 + J42"  # type:ignore
    ws["K43"] = "=K32 + K42"  # type:ignore
    # 列合計の埋め込み
    ws["B47"] = "=B43+B44+B45+B46"  # type:ignore
    ws["C47"] = "=C43+C44+C45+C46"  # type:ignore
    ws["D47"] = "=D43+D44+D45+D46"  # type:ignore
    ws["E47"] = "=E43+E44+E45+E46"  # type:ignore
    ws["F47"] = "=F43+F44+F45+F46"  # type:ignore
    ws["G47"] = "=G43+G44+G45+G46"  # type:ignore
    ws["H47"] = "=H43+H44+H45+H46"  # type:ignore
    ws["I47"] = "=I43+I44+I45+I46"  # type:ignore
    ws["J47"] = "=J43+J44+J45+J46"  # type:ignore
    ws["K47"] = "=K43+K44+K45+L46"  # type:ignore

    # 間接計列の追加
    ws["H1"] = "【間接計】"  # type:ignore
    # 間接計列への計算式の埋め込み
    ws["H2"] = "=SUM(B2:G2)"  # type:ignore
    ws["H3"] = "=SUM(B3:G3)"  # type:ignore
    ws["H4"] = "=SUM(B4:G4)"  # type:ignore
    ws["H5"] = "=SUM(B5:G5)"  # type:ignore
    ws["H6"] = "=SUM(B6:G6) / 6"  # type:ignore
    ws["H7"] = "=SUM(B7:G7)"  # type:ignore
    ws["H8"] = "=SUM(B8:G8)"  # type:ignore
    ws["H9"] = "=SUM(B9:G9)"  # type:ignore
    ws["H10"] = "=SUM(B10:G10) / 6"  # type:ignore
    ws["H11"] = "=SUM(B11:G11)"  # type:ignore
    ws["H12"] = "=SUM(B12:G12)"  # type:ignore
    ws["H13"] = "=SUM(B13:G13)"  # type:ignore
    ws["H14"] = "=SUM(B14:G14) / 6"  # type:ignore
    ws["H15"] = "=SUM(B15:G15)"  # type:ignore
    ws["H16"] = "=SUM(B16:G16) / 6"  # type:ignore
    ws["H17"] = "=SUM(B17:G17)"  # type:ignore
    ws["H18"] = "=SUM(B18:G18) / 6"  # type:ignore
    ws["H19"] = "=SUM(B19:G19)"  # type:ignore
    ws["H20"] = "=SUM(B20:G20)"  # type:ignore
    ws["H21"] = "=SUM(B21:G21)"  # type:ignore
    ws["H22"] = "=SUM(B22:G22)"  # type:ignore
    ws["H23"] = "=SUM(B23:G23)"  # type:ignore
    ws["H24"] = "=SUM(B24:G24)"  # type:ignore
    ws["H25"] = "=SUM(B25:G25)"  # type:ignore
    ws["H26"] = "=SUM(B26:G26)"  # type:ignore
    ws["H27"] = "=SUM(B27:G27)"  # type:ignore
    ws["H28"] = "=SUM(B28:G28)"  # type:ignore
    ws["H29"] = "=SUM(B29:G29)"  # type:ignore
    ws["H30"] = "=SUM(B30:G30)"  # type:ignore
    ws["H31"] = "=SUM(B31:G31)"  # type:ignore
    ws["H32"] = "=SUM(B32:G32)"  # type:ignore
    ws["H33"] = "=SUM(B33:G33)"  # type:ignore
    ws["H34"] = "=SUM(B34:G34)"  # type:ignore
    ws["H35"] = "=SUM(B35:G35)"  # type:ignore
    ws["H36"] = "=SUM(B36:G36)"  # type:ignore
    ws["H37"] = "=SUM(B37:G37)"  # type:ignore
    ws["H38"] = "=SUM(B38:G38)"  # type:ignore
    ws["H39"] = "=SUM(B39:G39)"  # type:ignore
    ws["H40"] = "=SUM(B40:G40)"  # type:ignore
    ws["H41"] = "=SUM(B41:G41)"  # type:ignore
    ws["H42"] = "=SUM(B42:G42)"  # type:ignore
    ws["H43"] = "=SUM(B43:G43)"  # type:ignore
    ws["H44"] = "=SUM(B44:G44)"  # type:ignore
    ws["H45"] = "=SUM(B45:G45)"  # type:ignore
    ws["H46"] = "=SUM(B46:G46)"  # type:ignore
    ws["H47"] = "=H43+H44+H45+H46"  # type:ignore
    # 直接計列の追加
    ws["K1"] = "【直接計】"  # type:ignore
    # 直接計列への計算式の埋め込み
    ws["K2"] = "=SUM(I2:J2)"  # type:ignore
    ws["K3"] = "=SUM(I3:J3)"  # type:ignore
    ws["K4"] = "=SUM(I4:J4)"  # type:ignore
    ws["K5"] = "=SUM(I5:J5)"  # type:ignore
    ws["K6"] = "=SUM(I6:J6) / 2"  # type:ignore
    ws["K7"] = "=SUM(I7:J7)"  # type:ignore
    ws["K8"] = "=SUM(I8:J8)"  # type:ignore
    ws["K9"] = "=SUM(I9:J9)"  # type:ignore
    ws["K10"] = "=SUM(I10:J10) / 2"  # type:ignore
    ws["K11"] = "=SUM(I11:J11)"  # type:ignore
    ws["K12"] = "=SUM(I12:J12)"  # type:ignore
    ws["K13"] = "=SUM(I13:J13)"  # type:ignore
    ws["K14"] = "=SUM(I14:J14) / 2"  # type:ignore
    ws["K15"] = "=SUM(I15:J15)"  # type:ignore
    ws["K16"] = "=SUM(I16:J16) / 2"  # type:ignore
    ws["K17"] = "=SUM(I17:J17)"  # type:ignore
    ws["K18"] = "=SUM(I18:J18) / 2"  # type:ignore
    ws["K19"] = "=SUM(I19:J19)"  # type:ignore
    ws["K20"] = "=SUM(I20:J20)"  # type:ignore
    ws["K21"] = "=SUM(I21:J21)"  # type:ignore
    ws["K22"] = "=SUM(I22:J22)"  # type:ignore
    ws["K23"] = "=SUM(I23:J23)"  # type:ignore
    ws["K24"] = "=SUM(I24:J24)"  # type:ignore
    ws["K25"] = "=SUM(I25:J25)"  # type:ignore
    ws["K26"] = "=SUM(I26:J26)"  # type:ignore
    ws["K27"] = "=SUM(I27:J27)"  # type:ignore
    ws["K28"] = "=SUM(I28:J28)"  # type:ignore
    ws["K29"] = "=SUM(I29:J29)"  # type:ignore
    ws["K30"] = "=SUM(I30:J30)"  # type:ignore
    ws["K31"] = "=SUM(I31:J31)"  # type:ignore
    ws["K32"] = "=SUM(I32:J32)"  # type:ignore
    ws["K33"] = "=SUM(I33:J33)"  # type:ignore
    ws["K34"] = "=SUM(I34:J34)"  # type:ignore
    ws["K35"] = "=SUM(I35:J35)"  # type:ignore
    ws["K36"] = "=SUM(I36:J36)"  # type:ignore
    ws["K37"] = "=SUM(I37:J37)"  # type:ignore
    ws["K38"] = "=SUM(I38:J38)"  # type:ignore
    ws["K39"] = "=SUM(I39:J39)"  # type:ignore
    ws["K40"] = "=SUM(I40:J40)"  # type:ignore
    ws["K41"] = "=SUM(I41:J41)"  # type:ignore
    ws["K42"] = "=SUM(I42:J42)"  # type:ignore
    ws["K43"] = "=SUM(I43:J43)"  # type:ignore
    ws["K44"] = "=SUM(I44:J44)"  # type:ignore
    ws["K45"] = "=SUM(I45:J45)"  # type:ignore
    ws["K46"] = "=SUM(I46:J46)"  # type:ignore
    ws["K47"] = "=K43+K44+K45+K46"  # type:ignore
    # 直接列ヘッダーの書式
    fill = PatternFill(patternType="solid", fgColor="008000")  # type:ignore
    ws["K1"].fill = fill  # type:ignore
    ws["K1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 合計列の追加とヘッダーの書式
    ws["L1"] = "【合計】"  # type:ignore
    ws["L1"].fill = fill  # type:ignore
    ws["L1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 罫線
    side1 = Side(border_style="thin", color="000000")
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
    for row in ws["A2:L49"]:  # type:ignore
        for cell in row:
            cell.border = border_aro
    # 行合計の埋め込み
    ws["L2"] = "=SUM(H2,K2)"  # type:ignore
    ws["L3"] = "=SUM(H3,K3)"  # type:ignore
    ws["L4"] = "=SUM(H4,K4)"  # type:ignore
    ws["L5"] = "=SUM(H5,K5)"  # type:ignore
    ws["L6"] = "=SUM(H6,K6) / 2"  # type:ignore
    ws["L7"] = "=SUM(H7,K7)"  # type:ignore
    ws["L8"] = "=SUM(H8,K8)"  # type:ignore
    ws["L9"] = "=SUM(H9,K9)"  # type:ignore
    ws["L10"] = "=(H10+K10) / 2"  # type:ignore
    ws["L11"] = "=SUM(H11,K11)"  # type:ignore
    ws["L12"] = "=SUM(K12,H12)"  # type:ignore
    ws["L13"] = "=SUM(H13,K13)"  # type:ignore
    ws["L14"] = "=SUM(H14,K14) / 2"  # type:ignore
    ws["L15"] = "=SUM(H15,K15)"  # type:ignore
    ws["L16"] = "=SUM(H16,K16) / 2"  # type:ignore
    ws["L17"] = "=SUM(H17,K17)"  # type:ignore
    ws["L18"] = "=SUM(H18,K18) / 2"  # type:ignore
    ws["L19"] = "=SUM(H19,K19)"  # type:ignore
    ws["L20"] = "=SUM(H20,K20)"  # type:ignore
    ws["L21"] = "=SUM(H21,K21)"  # type:ignore
    ws["L22"] = "=SUM(H22,K22)"  # type:ignore
    ws["L23"] = "=SUM(H23,K23)"  # type:ignore
    ws["L24"] = "=SUM(H24,K24)"  # type:ignore
    ws["L25"] = "=SUM(H25,K25)"  # type:ignore
    ws["L26"] = "=SUM(H26,K26)"  # type:ignore
    ws["L27"] = "=SUM(H27,K27)"  # type:ignore
    ws["L28"] = "=SUM(H28,K28)"  # type:ignore
    ws["L29"] = "=SUM(H29,K29)"  # type:ignore
    ws["L30"] = "=SUM(H30,K30)"  # type:ignore
    ws["L31"] = "=SUM(H31,K31)"  # type:ignore
    ws["L32"] = "=SUM(H32,K32)"  # type:ignore
    ws["L33"] = "=SUM(H33,K33)"  # type:ignore
    ws["L34"] = "=SUM(H34,K34)"  # type:ignore
    ws["L35"] = "=SUM(H35,K35)"  # type:ignore
    ws["L36"] = "=SUM(H36,K36)"  # type:ignore
    ws["L37"] = "=SUM(H37,K37)"  # type:ignore
    ws["L38"] = "=SUM(H38,K38)"  # type:ignore
    ws["L39"] = "=SUM(H39,K39)"  # type:ignore
    ws["L40"] = "=SUM(H40,K40)"  # type:ignore
    ws["L41"] = "=SUM(H41,K41)"  # type:ignore
    ws["L42"] = "=SUM(H42,K42)"  # type:ignore
    ws["L43"] = "=SUM(H43,K43)"  # type:ignore
    ws["L44"] = "=SUM(H44,K44)"  # type:ignore
    ws["L45"] = "=SUM(H45,K45)"  # type:ignore
    ws["L46"] = "=SUM(H46,K46)"  # type:ignore
    ws["L47"] = "=L43+L44+L45+L46"  # type:ignore
    ws.delete_rows(48, 49)  # type:ignore
    # Excelファイルを出力
    wb.save("c:/test/鍛造/鍛造.xlsx")

    # 切削
    # パスで指定したファイルの一覧をリスト形式で取得
    csv_files_sesaku = glob.glob("c:/test/切削/*.csv")
    # CSVファイルの中身を追加していくリストを表示
    data_list_sesaku = []
    # 読み込むファイルのリストをスキャン
    for file in csv_files_sesaku:
        data_list_sesaku.append(pd.read_csv(file))
    # リストを全て列方向に結合
    df_sesaku_t = pd.concat(data_list_sesaku, axis=1, sort=True)
    # columnsパラメータで列名を設定
    feature_sesaku = [
        "間接1",
        "間接2",
        "間接4",
        "間接5",
        "間接6",
        "直接1",
        "直接2",
        "直接4",
    ]
    df_sesaku_t.columns = feature_sesaku
    # 行名の設定
    df_sesaku_t = df_sesaku_t.rename(
        {
            0: "在籍者",
            1: "在籍者主幹以下人数",
            2: "実在籍者",
            3: "有休時間",
            4: "有休時間在籍者平均",
            5: "欠勤時間",
            6: "勤務時間",
            7: "遅早時間",
            8: "出勤率",
            9: "実労働時間",
            10: "ズレ時間",
            11: "残業時間",
            12: "残業時間主幹以下平均",
            13: "法定外休出時間",
            14: "法定外休出主幹以下平均",
            15: "法定休出時間",
            16: "法定休出主幹以下平均",
            17: "時間外60時間超",
            18: "代休時間",
            19: "応援時間",
            20: "総労働時間",
            21: "基本給",
            22: "役職手当",
            23: "営業手当",
            24: "地域手当",
            25: "特別手当",
            26: "特別技技手当",
            27: "調整手当",
            28: "別居手当",
            29: "通勤手当",
            30: "小計1",
            31: "残業手当",
            32: "休出手当",
            33: "深夜勤務手当",
            34: "交替時差手当",
            35: "休業手当",
            36: "休業控除",
            37: "代休他",
            38: "欠勤・遅早控除",
            39: "精算分",
            40: "小計2",
            41: "総支給額",
            42: "応援時間額",
            43: "役員振替",
            44: "部門振替",
            45: "合計",
        }
    )

    # 整形---------------------
    # ワークブックの生成
    wb = Workbook()
    # ワークシートの生成
    ws = wb.active
    ws.title = "切削工場"  # type:ignore
    # DataFrameを行単位のデータにする
    rows = dataframe_to_rows(df_sesaku_t, index=True, header=True)
    # 1セルずつ処理を実行する
    for row_no, row in enumerate(rows, 1):
        for col_no, value in enumerate(row, 1):
            # データを書き込む
            ws.cell(row=row_no, column=col_no, value=value)  # type:ignore
    # 不要な行の削除
    ws.delete_rows(2)  # type:ignore
    # 列の追加
    ws.insert_cols(7)  # type:ignore
    # 表示倍率の設定
    ws.sheet_view.zoomScale = 100  # type:ignore
    # 列幅の設定
    ws.column_dimensions["A"].width = 25  # type:ignore
    ws.column_dimensions["B"].width = 12  # type:ignore
    ws.column_dimensions["C"].width = 12  # type:ignore
    ws.column_dimensions["D"].width = 12  # type:ignore
    ws.column_dimensions["E"].width = 12  # type:ignore
    ws.column_dimensions["F"].width = 12  # type:ignore
    ws.column_dimensions["G"].width = 12  # type:ignore
    ws.column_dimensions["H"].width = 12  # type:ignore
    ws.column_dimensions["I"].width = 12  # type:ignore
    ws.column_dimensions["J"].width = 12  # type:ignore
    ws.column_dimensions["K"].width = 12  # type:ignore
    ws.column_dimensions["L"].width = 12  # type:ignore
    # 表示形式
    format = "#,##0"
    for row in ws["B2:L2"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B4:L9"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00%"
    for row in ws["B10:L10"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B11:L12"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B13:L22"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B23:L23"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B24:L26"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B25:L51"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    # ヘッダー行のスタイル設定
    header = ws[1]  # type:ignore
    for header_cell in header:
        # フォントを設定する
        header_cell.fill = PatternFill(
            patternType="solid", fgColor="008000"  # type:ignore
        )  # type:ignore
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
    # 中央揃え
    for row in ws["B1:L1"]:  # type:ignore
        for cell in row:
            cell.alignment = Alignment(horizontal="centerContinuous")
    # 指定した行の背景色を黄色にする
    mylist = [21, 44, 45, 46]
    for list in mylist:
        for row in ws.iter_rows():  # type:ignore
            for cell in row:
                if cell.row == list:
                    cell.fill = PatternFill(
                        fgColor="FFFF00",  # type:ignore
                        bgColor="FFFF00",  # type:ignore
                        fill_type="solid",  # type:ignore
                    )

    # 小計1
    ws["B32"] = "=SUM(B23:B31)"  # type:ignore
    ws["C32"] = "=SUM(C23:C31)"  # type:ignore
    ws["D32"] = "=SUM(D23:D31)"  # type:ignore
    ws["E32"] = "=SUM(E23:E31)"  # type:ignore
    ws["F32"] = "=SUM(F23:F31)"  # type:ignore
    ws["G32"] = "=SUM(G23:G31)"  # type:ignore
    ws["H32"] = "=SUM(H23:H31)"  # type:ignore
    ws["I32"] = "=SUM(I23:I31)"  # type:ignore
    ws["J32"] = "=SUM(J23:J31)"  # type:ignore
    ws["K32"] = "=SUM(K23:K31)"  # type:ignore
    # 小計2
    ws["B42"] = "=SUM(B33:B41)"  # type:ignore
    ws["C42"] = "=SUM(C33:C41)"  # type:ignore
    ws["D42"] = "=SUM(D33:D41)"  # type:ignore
    ws["E42"] = "=SUM(E33:E41)"  # type:ignore
    ws["F42"] = "=SUM(F33:F41)"  # type:ignore
    ws["G42"] = "=SUM(G33:G41)"  # type:ignore
    ws["H42"] = "=SUM(H33:H41)"  # type:ignore
    ws["I42"] = "=SUM(I33:I41)"  # type:ignore
    ws["J42"] = "=SUM(J33:J41)"  # type:ignore
    ws["K42"] = "=SUM(K33:K41)"  # type:ignore
    # 総支給額
    ws["B43"] = "=B32 + B42"  # type:ignore
    ws["C43"] = "=C32 + C42"  # type:ignore
    ws["D43"] = "=D32 + D42"  # type:ignore
    ws["E43"] = "=E32 + E42"  # type:ignore
    ws["F43"] = "=F32 + F42"  # type:ignore
    ws["G43"] = "=G32 + G42"  # type:ignore
    ws["H43"] = "=H32 + H42"  # type:ignore
    ws["I43"] = "=I32 + I42"  # type:ignore
    ws["J43"] = "=J32 + J42"  # type:ignore
    ws["K43"] = "=K32 + K42"  # type:ignore
    # 列合計の埋め込み
    ws["B47"] = "=B43+B44+B45+B46"  # type:ignore
    ws["C47"] = "=C43+C44+C45+C46"  # type:ignore
    ws["D47"] = "=D43+D44+D45+D46"  # type:ignore
    ws["E47"] = "=E43+E44+E45+E46"  # type:ignore
    ws["F47"] = "=F43+F44+F45+F46"  # type:ignore
    ws["G47"] = "=G43+G44+G45+G46"  # type:ignore
    ws["H47"] = "=H43+H44+H45+H46"  # type:ignore
    ws["I47"] = "=I43+I44+I45+I46"  # type:ignore
    ws["J47"] = "=J43+J44+J45+J46"  # type:ignore
    ws["K47"] = "=K43+K44+K45+K46"  # type:ignore

    # 間接計列の追加
    ws["G1"] = "【間接計】"  # type:ignore
    # 間接計列への計算式の埋め込み
    ws["G2"] = "=SUM(B2:F2)"  # type:ignore
    ws["G3"] = "=SUM(B3:F3)"  # type:ignore
    ws["G4"] = "=SUM(B4:F4)"  # type:ignore
    ws["G5"] = "=SUM(B5:F5)"  # type:ignore
    ws["G6"] = "=SUM(B6:F6) / 5"  # type:ignore
    ws["G7"] = "=SUM(B7:F7)"  # type:ignore
    ws["G8"] = "=SUM(B8:F8)"  # type:ignore
    ws["G9"] = "=SUM(B9:F9)"  # type:ignore
    ws["G10"] = "=SUM(B10:F10) / 5"  # type:ignore
    ws["G11"] = "=SUM(B11:F11)"  # type:ignore
    ws["G12"] = "=SUM(B12:F12)"  # type:ignore
    ws["G13"] = "=SUM(B13:F13)"  # type:ignore
    ws["G14"] = "=SUM(B14:F14) / 5"  # type:ignore
    ws["G15"] = "=SUM(B15:F15)"  # type:ignore
    ws["G16"] = "=SUM(B16:F16) / 5"  # type:ignore
    ws["G17"] = "=SUM(B17:F17)"  # type:ignore
    ws["G18"] = "=SUM(B18:F18) / 5"  # type:ignore
    ws["G19"] = "=SUM(B19:F19)"  # type:ignore
    ws["G20"] = "=SUM(B20:F20)"  # type:ignore
    ws["G21"] = "=SUM(B21:F21)"  # type:ignore
    ws["G22"] = "=SUM(B22:F22)"  # type:ignore
    ws["G23"] = "=SUM(B23:F23)"  # type:ignore
    ws["G24"] = "=SUM(B24:F24)"  # type:ignore
    ws["G25"] = "=SUM(B25:F25)"  # type:ignore
    ws["G26"] = "=SUM(B26:F26)"  # type:ignore
    ws["G27"] = "=SUM(B27:F27)"  # type:ignore
    ws["G28"] = "=SUM(B28:F28)"  # type:ignore
    ws["G29"] = "=SUM(B29:F29)"  # type:ignore
    ws["G30"] = "=SUM(B30:F30)"  # type:ignore
    ws["G31"] = "=SUM(B31:F31)"  # type:ignore
    ws["G32"] = "=SUM(B32:F32)"  # type:ignore
    ws["G33"] = "=SUM(B33:F33)"  # type:ignore
    ws["G34"] = "=SUM(B34:F34)"  # type:ignore
    ws["G35"] = "=SUM(B35:F35)"  # type:ignore
    ws["G36"] = "=SUM(B36:F36)"  # type:ignore
    ws["G37"] = "=SUM(B37:F37)"  # type:ignore
    ws["G38"] = "=SUM(B38:F38)"  # type:ignore
    ws["G39"] = "=SUM(B39:F39)"  # type:ignore
    ws["G40"] = "=SUM(B40:F40)"  # type:ignore
    ws["G41"] = "=SUM(B41:F41)"  # type:ignore
    ws["G42"] = "=SUM(B42:F42)"  # type:ignore
    ws["G43"] = "=SUM(B43:F43)"  # type:ignore
    ws["G44"] = "=SUM(B44:F44)"  # type:ignore
    ws["G45"] = "=SUM(B45:F45)"  # type:ignore
    ws["G46"] = "=SUM(B46:F46)"  # type:ignore
    ws["G47"] = "=G43+G44+G45+G46"  # type:ignore
    # ヘッダーの書式
    fill = PatternFill(patternType="solid", fgColor="008000")  # type:ignore
    ws["K1"].fill = fill  # type:ignore
    ws["K1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 合計列の追加とヘッダーの書式
    ws["K1"] = "【合計】"  # type:ignore
    ws["K1"].fill = fill  # type:ignore
    ws["K1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 直接計列の追加
    ws["K1"] = "【直接計】"  # type:ignore
    # 直接計列への計算式の埋め込み
    ws["K2"] = "=SUM(H2:J2)"  # type:ignore
    ws["K3"] = "=SUM(H3:J3)"  # type:ignore
    ws["K4"] = "=SUM(H4:J4)"  # type:ignore
    ws["K5"] = "=SUM(H5:J5)"  # type:ignore
    ws["K6"] = "=SUM(H6:J6) / 3"  # type:ignore
    ws["K7"] = "=SUM(H7:J7)"  # type:ignore
    ws["K8"] = "=SUM(H8:J8)"  # type:ignore
    ws["K9"] = "=SUM(H9:J9)"  # type:ignore
    ws["K10"] = "=SUM(H10:J10) / 3"  # type:ignore
    ws["K11"] = "=SUM(H11:J11)"  # type:ignore
    ws["K12"] = "=SUM(H12:J12)"  # type:ignore
    ws["K13"] = "=SUM(H13:J13)"  # type:ignore
    ws["K14"] = "=SUM(H14:J14) / 3"  # type:ignore
    ws["K15"] = "=SUM(H15:J15)"  # type:ignore
    ws["K16"] = "=SUM(H16:J16) / 3"  # type:ignore
    ws["K17"] = "=SUM(H17:J17)"  # type:ignore
    ws["K18"] = "=SUM(H18:J18) / 3"  # type:ignore
    ws["K19"] = "=SUM(H19:J19)"  # type:ignore
    ws["K20"] = "=SUM(H20:J20)"  # type:ignore
    ws["K21"] = "=SUM(H21:J21)"  # type:ignore
    ws["K22"] = "=SUM(H22:J22)"  # type:ignore
    ws["K23"] = "=SUM(H23:J23)"  # type:ignore
    ws["K24"] = "=SUM(H24:J24)"  # type:ignore
    ws["K25"] = "=SUM(H25:J25)"  # type:ignore
    ws["K26"] = "=SUM(H26:J26)"  # type:ignore
    ws["K27"] = "=SUM(H27:J27)"  # type:ignore
    ws["K28"] = "=SUM(H28:J28)"  # type:ignore
    ws["K29"] = "=SUM(H29:J29)"  # type:ignore
    ws["K30"] = "=SUM(H30:J30)"  # type:ignore
    ws["K31"] = "=SUM(H31:J31)"  # type:ignore
    ws["K32"] = "=SUM(H32:J32)"  # type:ignore
    ws["K33"] = "=SUM(H33:J33)"  # type:ignore
    ws["K34"] = "=SUM(H34:J34)"  # type:ignore
    ws["K35"] = "=SUM(H35:J35)"  # type:ignore
    ws["K36"] = "=SUM(H36:J36)"  # type:ignore
    ws["K37"] = "=SUM(H37:J37)"  # type:ignore
    ws["K38"] = "=SUM(H38:J38)"  # type:ignore
    ws["K39"] = "=SUM(H39:J39)"  # type:ignore
    ws["K40"] = "=SUM(H40:J40)"  # type:ignore
    ws["K41"] = "=SUM(H41:J41)"  # type:ignore
    ws["K42"] = "=SUM(H42:J42)"  # type:ignore
    ws["K43"] = "=SUM(H43:J43)"  # type:ignore
    ws["K44"] = "=SUM(H44:J44)"  # type:ignore
    ws["K45"] = "=SUM(H45:J45)"  # type:ignore
    ws["K46"] = "=SUM(H46:J46)"  # type:ignore
    ws["K47"] = "=K43+K44+K45+K46"  # type:ignore
    # ヘッダーの書式
    fill = PatternFill(patternType="solid", fgColor="008000")  # type:ignore
    ws["L1"].fill = fill  # type:ignore
    ws["L1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 合計列の追加とヘッダーの書式
    ws["L1"] = "【合計】"  # type:ignore
    ws["L1"].fill = fill  # type:ignore
    ws["L1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 罫線
    side1 = Side(border_style="thin", color="000000")
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
    for row in ws["A2:L49"]:  # type:ignore
        for cell in row:
            cell.border = border_aro
    # 行合計の計算式埋め込み
    ws["L2"] = "=G2+K2"  # type:ignore
    ws["L3"] = "=G3+K3"  # type:ignore
    ws["L4"] = "=G4+K4"  # type:ignore
    ws["L5"] = "=G5+K5"  # type:ignore
    ws["L6"] = "=(G6+K6) / 2"  # type:ignore
    ws["L7"] = "=G7+K7"  # type:ignore
    ws["L8"] = "=G8+K8"  # type:ignore
    ws["L9"] = "=G9+K9"  # type:ignore
    ws["L10"] = "=(G10+K10) / 2"  # type:ignore
    ws["L11"] = "=G11+K11"  # type:ignore
    ws["L12"] = "=G12+K12"  # type:ignore
    ws["L13"] = "=G13+K13"  # type:ignore
    ws["L14"] = "=(G14+K14) / 2"  # type:ignore
    ws["L15"] = "=G15+K15"  # type:ignore
    ws["L16"] = "=(G16+K16) / 2"  # type:ignore
    ws["L17"] = "=G17+K17"  # type:ignore
    ws["L18"] = "=(G18+K18) / 2"  # type:ignore
    ws["L19"] = "=G19+K19"  # type:ignore
    ws["L20"] = "=G20+K20"  # type:ignore
    ws["L21"] = "=G21+K21"  # type:ignore
    ws["L22"] = "=G22+K22"  # type:ignore
    ws["L23"] = "=G23+K23"  # type:ignore
    ws["L24"] = "=G24+K24"  # type:ignore
    ws["L25"] = "=G25+K25"  # type:ignore
    ws["L26"] = "=G26+K26"  # type:ignore
    ws["L27"] = "=G27+K27"  # type:ignore
    ws["L28"] = "=G28+K28"  # type:ignore
    ws["L29"] = "=G29+K29"  # type:ignore
    ws["L30"] = "=G30+K30"  # type:ignore
    ws["L31"] = "=G31+K31"  # type:ignore
    ws["L32"] = "=G32+K32"  # type:ignore
    ws["L33"] = "=G33+K33"  # type:ignore
    ws["L34"] = "=G34+K34"  # type:ignore
    ws["L35"] = "=G35+K35"  # type:ignore
    ws["L36"] = "=G36+K36"  # type:ignore
    ws["L37"] = "=G37+K37"  # type:ignore
    ws["L38"] = "=G38+K38"  # type:ignore
    ws["L39"] = "=G39+K39"  # type:ignore
    ws["L40"] = "=G40+K40"  # type:ignore
    ws["L41"] = "=G41+K41"  # type:ignore
    ws["L42"] = "=G42+K42"  # type:ignore
    ws["L43"] = "=G43+K43"  # type:ignore
    ws["L44"] = "=G44+K44"  # type:ignore
    ws["L45"] = "=G45+K45"  # type:ignore
    ws["L46"] = "=G46+K46"  # type:ignore
    ws["L47"] = "=L43+L44+L45+L46"  # type:ignore
    ws.delete_rows(48, 49)  # type:ignore
    # Excelファイルを出力
    wb.save("c:/test/切削/切削.xlsx")

    # AC
    # パスで指定したファイルの一覧をリスト形式で取得
    csv_files_ac = glob.glob("c:/test/AC/*.csv")

    # CSVファイルの中身を追加していくリストを表示
    data_list_ac = []
    # 読み込むファイルのリストをスキャン
    for file in csv_files_ac:
        data_list_ac.append(pd.read_csv(file))
    # リストを全て列方向に結合
    df_ac_t = pd.concat(data_list_ac, axis=1, sort=True)
    # columnsパラメータで列名を設定
    feature_ac = ["間接1", "間接2", "間接4", "間接5", "直接1", "直接4"]
    df_ac_t.columns = feature_ac
    # 行名の設定
    df_ac_t = df_ac_t.rename(
        {
            0: "在籍者",
            1: "在籍者主幹以下人数",
            2: "実在籍者",
            3: "有休時間",
            4: "有休時間在籍者平均",
            5: "欠勤時間",
            6: "勤務時間",
            7: "遅早時間",
            8: "出勤率",
            9: "実労働時間",
            10: "ズレ時間",
            11: "残業時間",
            12: "残業時間主幹以下平均",
            13: "法定外休出時間",
            14: "法定外休出主幹以下平均",
            15: "法定休出時間",
            16: "法定休出主幹以下平均",
            17: "時間外60時間超",
            18: "代休時間",
            19: "応援時間",
            20: "総労働時間",
            21: "基本給",
            22: "役職手当",
            23: "営業手当",
            24: "地域手当",
            25: "特別手当",
            26: "特別技技手当",
            27: "調整手当",
            28: "別居手当",
            29: "通勤手当",
            30: "小計1",
            31: "残業手当",
            32: "休出手当",
            33: "深夜勤務手当",
            34: "交替時差手当",
            35: "休業手当",
            36: "休業控除",
            37: "代休他",
            38: "欠勤・遅早控除",
            39: "精算分",
            40: "小計2",
            41: "総支給額",
            42: "応援時間額",
            43: "役員振替",
            44: "部門振替",
            45: "合計",
        }
    )

    # 整形---------------------
    # ワークブックの生成
    wb = Workbook()
    # ワークシートの生成
    ws = wb.active
    ws.title = "AC工場"  # type:ignore
    # DataFrameを行単位のデータにする
    rows = dataframe_to_rows(df_ac_t, index=True, header=True)
    # 1セルずつ処理を実行する
    for row_no, row in enumerate(rows, 1):
        for col_no, value in enumerate(row, 1):
            # データを書き込む
            ws.cell(row=row_no, column=col_no, value=value)  # type:ignore

    # 不要な行の削除
    ws.delete_rows(2)  # type:ignore
    # 列の追加
    ws.insert_cols(6)  # type:ignore
    # 表示倍率の設定
    ws.sheet_view.zoomScale = 100  # type:ignore
    # 列幅の設定
    ws.column_dimensions["A"].width = 25  # type:ignore
    ws.column_dimensions["B"].width = 12  # type:ignore
    ws.column_dimensions["C"].width = 12  # type:ignore
    ws.column_dimensions["D"].width = 12  # type:ignore
    ws.column_dimensions["E"].width = 12  # type:ignore
    ws.column_dimensions["F"].width = 12  # type:ignore
    ws.column_dimensions["G"].width = 12  # type:ignore
    ws.column_dimensions["H"].width = 12  # type:ignore
    ws.column_dimensions["I"].width = 12  # type:ignore
    ws.column_dimensions["J"].width = 12  # type:ignore
    ws.column_dimensions["K"].width = 12  # type:ignore
    ws.column_dimensions["L"].width = 12  # type:ignore
    # 表示形式
    format = "#,##0"
    for row in ws["B2:L2"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B4:L9"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00%"
    for row in ws["B10:L10"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B11:L12"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B13:L22"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B23:L23"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B24:L26"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B23:L48"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    # ヘッダー行のスタイル設定
    header = ws[1]  # type:ignore
    for header_cell in header:
        # フォントを設定する
        header_cell.fill = PatternFill(
            patternType="solid", fgColor="008000"  # type:ignore
        )  # type:ignore
        # 罫線を設定する
        header_cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
    # 中央揃え
    for row in ws["B1:F1"]:  # type:ignore
        for cell in row:
            cell.alignment = Alignment(horizontal="centerContinuous")
        ws["H1"].alignment = Alignment(horizontal="centerContinuous")  # type:ignore
        ws["G1"].alignment = Alignment(horizontal="centerContinuous")  # type:ignore
        ws["L1"].alignment = Alignment(horizontal="centerContinuous")  # type:ignore
    # 指定した行の背景色を黄色にする
    mylist = [21, 44, 45, 46]
    for list in mylist:
        for row in ws.iter_rows():  # type:ignore
            for cell in row:
                if cell.row == list:
                    cell.fill = PatternFill(
                        fgColor="FFFF00",  # type:ignore
                        bgColor="FFFF00",  # type:ignore
                        fill_type="solid",  # type:ignore
                    )
    # 小計1
    ws["B32"] = "=SUM(B23:B31)"  # type:ignore
    ws["C32"] = "=SUM(C23:C31)"  # type:ignore
    ws["D32"] = "=SUM(D23:D31)"  # type:ignore
    ws["E32"] = "=SUM(E23:E31)"  # type:ignore
    ws["F32"] = "=SUM(F23:F31)"  # type:ignore
    ws["G32"] = "=SUM(G23:G31)"  # type:ignore
    ws["H32"] = "=SUM(H23:H31)"  # type:ignore
    # 小計2
    ws["B42"] = "=SUM(B33:B41)"  # type:ignore
    ws["C42"] = "=SUM(C33:C41)"  # type:ignore
    ws["D42"] = "=SUM(D33:D41)"  # type:ignore
    ws["E42"] = "=SUM(E33:E41)"  # type:ignore
    ws["F42"] = "=SUM(F33:F41)"  # type:ignore
    ws["G42"] = "=SUM(G33:G41)"  # type:ignore
    ws["H42"] = "=SUM(H33:H41)"  # type:ignore
    # 総支給額
    ws["B43"] = "=B32 + B42"  # type:ignore
    ws["C43"] = "=C32 + C42"  # type:ignore
    ws["D43"] = "=D32 + D42"  # type:ignore
    ws["E43"] = "=E32 + E42"  # type:ignore
    ws["F43"] = "=F32 + F42"  # type:ignore
    ws["G43"] = "=G32 + G42"  # type:ignore
    ws["H43"] = "=H32 + H42"  # type:ignore
    # 列合計の埋め込み
    ws["B47"] = "=B43+B44+B45+B46"  # type:ignore
    ws["C47"] = "=C43+C44+C45+C46"  # type:ignore
    ws["D47"] = "=D43+D44+D45+D46"  # type:ignore
    ws["E47"] = "=E43+E44+E45+E46"  # type:ignore
    ws["F47"] = "=F43+F44+F45+F46"  # type:ignore
    ws["G47"] = "=G43+G44+G45+G46"  # type:ignore
    ws["H47"] = "=H43+H44+H45+H46"  # type:ignore
    # 間接計列の追加
    ws["F1"] = "【間接計】"  # type:ignore
    # 間接計列への計算式の埋め込み
    ws["F2"] = "=SUM(B2:E2)"  # type:ignore
    ws["F3"] = "=SUM(B3:E3)"  # type:ignore
    ws["F4"] = "=SUM(B4:E4)"  # type:ignore
    ws["F5"] = "=SUM(B5:E5)"  # type:ignore
    ws["F6"] = "=SUM(B6:E6) / 4"  # type:ignore
    ws["F7"] = "=SUM(B7:E7)"  # type:ignore
    ws["F8"] = "=SUM(B8:E8)"  # type:ignore
    ws["F9"] = "=SUM(B9:E9)"  # type:ignore
    ws["F10"] = "=SUM(B10:E10) / 4"  # type:ignore
    ws["F11"] = "=SUM(B11:E11)"  # type:ignore
    ws["F12"] = "=SUM(B12:E12)"  # type:ignore
    ws["F13"] = "=SUM(B13:E13)"  # type:ignore
    ws["F14"] = "=SUM(B14:E14) / 4"  # type:ignore
    ws["F15"] = "=SUM(B15:E15)"  # type:ignore
    ws["F16"] = "=SUM(B16:E16) / 4"  # type:ignore
    ws["F17"] = "=SUM(B17:E17)"  # type:ignore
    ws["F18"] = "=SUM(B18:E18) / 4"  # type:ignore
    ws["F19"] = "=SUM(B19:E19)"  # type:ignore
    ws["F20"] = "=SUM(B20:E20)"  # type:ignore
    ws["F21"] = "=SUM(B21:E21)"  # type:ignore
    ws["F22"] = "=SUM(B22:E22)"  # type:ignore
    ws["F23"] = "=SUM(B23:E23)"  # type:ignore
    ws["F24"] = "=SUM(B24:E24)"  # type:ignore
    ws["F25"] = "=SUM(B25:E25)"  # type:ignore
    ws["F26"] = "=SUM(B26:E26)"  # type:ignore
    ws["F27"] = "=SUM(B27:E27)"  # type:ignore
    ws["F28"] = "=SUM(B28:E28)"  # type:ignore
    ws["F29"] = "=SUM(B29:E29)"  # type:ignore
    ws["F30"] = "=SUM(B30:E30)"  # type:ignore
    ws["F31"] = "=SUM(B31:E31)"  # type:ignore
    ws["F32"] = "=SUM(B32:E32)"  # type:ignore
    ws["F33"] = "=SUM(B33:E33)"  # type:ignore
    ws["F34"] = "=SUM(B34:E34)"  # type:ignore
    ws["F35"] = "=SUM(B35:E35)"  # type:ignore
    ws["F36"] = "=SUM(B36:E36)"  # type:ignore
    ws["F37"] = "=SUM(B37:E37)"  # type:ignore
    ws["F38"] = "=SUM(B38:E38)"  # type:ignore
    ws["F39"] = "=SUM(B39:E39)"  # type:ignore
    ws["F40"] = "=SUM(B40:E40)"  # type:ignore
    ws["F41"] = "=SUM(B41:E41)"  # type:ignore
    ws["F42"] = "=SUM(B42:E42)"  # type:ignore
    ws["F43"] = "=SUM(B43:E43)"  # type:ignore
    ws["F44"] = "=SUM(B44:E44)"  # type:ignore
    ws["F45"] = "=SUM(B45:E45)"  # type:ignore
    ws["F46"] = "=SUM(B46:E46)"  # type:ignore
    ws["F47"] = "=E43+E44+E45+E46"  # type:ignore
    # ヘッダーの書式
    fill = PatternFill(patternType="solid", fgColor="008000")  # type:ignore

    # 直接計列の追加
    ws["I1"] = "【直接計】"  # type:ignore
    # 直接計列への計算式の埋め込み
    ws["I2"] = "=SUM(G2:H2)"  # type:ignore
    ws["I3"] = "=SUM(G3:H3)"  # type:ignore
    ws["I4"] = "=SUM(G4:H4)"  # type:ignore
    ws["I5"] = "=SUM(G5:H5)"  # type:ignore
    ws["I6"] = "=SUM(G6:H6) / 2"  # type:ignore
    ws["I7"] = "=SUM(G7:H7)"  # type:ignore
    ws["I8"] = "=SUM(G8:H8)"  # type:ignore
    ws["I9"] = "=SUM(G9:H9)"  # type:ignore
    ws["I10"] = "=SUM(G10:H10) / 2"  # type:ignore
    ws["I11"] = "=SUM(G11:H11)"  # type:ignore
    ws["I12"] = "=SUM(G12:H12)"  # type:ignore
    ws["I13"] = "=SUM(G13:H13)"  # type:ignore
    ws["I14"] = "=SUM(G14:H14) / 2"  # type:ignore
    ws["I15"] = "=SUM(G15:H15)"  # type:ignore
    ws["I16"] = "=SUM(G16:H16) / 2"  # type:ignore
    ws["I17"] = "=SUM(G17:H17)"  # type:ignore
    ws["I18"] = "=SUM(G18:H18) / 2"  # type:ignore
    ws["I19"] = "=SUM(G19:H19)"  # type:ignore
    ws["I20"] = "=SUM(G20:H20)"  # type:ignore
    ws["I21"] = "=SUM(G21:H21)"  # type:ignore
    ws["I22"] = "=SUM(G22:H22)"  # type:ignore
    ws["I23"] = "=SUM(G23:H23)"  # type:ignore
    ws["I24"] = "=SUM(G24:H24)"  # type:ignore
    ws["I25"] = "=SUM(G25:H25)"  # type:ignore
    ws["I26"] = "=SUM(G26:H26)"  # type:ignore
    ws["I27"] = "=SUM(G27:H27)"  # type:ignore
    ws["I28"] = "=SUM(G28:H28)"  # type:ignore
    ws["I29"] = "=SUM(G29:H29)"  # type:ignore
    ws["I30"] = "=SUM(G30:H30)"  # type:ignore
    ws["I31"] = "=SUM(G31:H31)"  # type:ignore
    ws["I32"] = "=SUM(G32:H32)"  # type:ignore
    ws["I33"] = "=SUM(G33:H33)"  # type:ignore
    ws["I34"] = "=SUM(G34:H34)"  # type:ignore
    ws["I35"] = "=SUM(G35:H35)"  # type:ignore
    ws["I36"] = "=SUM(G36:H36)"  # type:ignore
    ws["I37"] = "=SUM(G37:H37)"  # type:ignore
    ws["I38"] = "=SUM(G38:H38)"  # type:ignore
    ws["I39"] = "=SUM(G39:H39)"  # type:ignore
    ws["I40"] = "=SUM(G40:H40)"  # type:ignore
    ws["I41"] = "=SUM(G41:H41)"  # type:ignore
    ws["I42"] = "=SUM(G42:H42)"  # type:ignore
    ws["I43"] = "=SUM(G43:H43)"  # type:ignore
    ws["I44"] = "=SUM(G44:H44)"  # type:ignore
    ws["I45"] = "=SUM(G45:H45)"  # type:ignore
    ws["I46"] = "=SUM(G46:H46)"  # type:ignore
    ws["I47"] = "=H43+H44+H45+H46"  # type:ignore
    # ヘッダーの書式
    fill = PatternFill(patternType="solid", fgColor="008000")  # type:ignore
    ws["L1"].fill = fill  # type:ignore
    ws["L1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 合計列の追加とヘッダーの書式
    ws["L1"] = "【合計】"  # type:ignore
    ws["L1"].fill = fill  # type:ignore
    ws["L1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 罫線
    side1 = Side(border_style="thin", color="000000")
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
    for row in ws["A2:L49"]:  # type:ignore
        for cell in row:
            cell.border = border_aro
    # 行合計の計算式埋め込み
    ws["L2"] = "=F2+I2"  # type:ignore
    ws["L3"] = "=F3+I3"  # type:ignore
    ws["L4"] = "=F4+I4"  # type:ignore
    ws["L5"] = "=F5+I5"  # type:ignore
    ws["L6"] = "=(F6+I6) / 2"  # type:ignore
    ws["L7"] = "=F7+I7"  # type:ignore
    ws["L8"] = "=F8+I8"  # type:ignore
    ws["L9"] = "=F9+I9"  # type:ignore
    ws["L10"] = "=(F10+I10) / 2"  # type:ignore
    ws["L11"] = "=F11+I11"  # type:ignore
    ws["L12"] = "=F12+I12"  # type:ignore
    ws["L13"] = "=F13+I13"  # type:ignore
    ws["L14"] = "=(F14+I14) / 2"  # type:ignore
    ws["L15"] = "=F15+I15"  # type:ignore
    ws["L16"] = "=(F16+I16) / 2 "  # type:ignore
    ws["L17"] = "=F17+I17"  # type:ignore
    ws["L18"] = "=(F18+I18) / 2"  # type:ignore
    ws["L19"] = "=F19+I19"  # type:ignore
    ws["L20"] = "=F20+I20"  # type:ignore
    ws["L21"] = "=F21+I21"  # type:ignore
    ws["L22"] = "=F22+I22"  # type:ignore
    ws["L23"] = "=F23+I23"  # type:ignore
    ws["L24"] = "=F24+I24"  # type:ignore
    ws["L25"] = "=F25+I25"  # type:ignore
    ws["L26"] = "=F26+I26"  # type:ignore
    ws["L27"] = "=F27+I27"  # type:ignore
    ws["L28"] = "=F28+I28"  # type:ignore
    ws["L29"] = "=F29+I29"  # type:ignore
    ws["L30"] = "=F30+I30"  # type:ignore
    ws["L31"] = "=F31+I31"  # type:ignore
    ws["L32"] = "=F32+I32"  # type:ignore
    ws["L33"] = "=F33+I33"  # type:ignore
    ws["L34"] = "=F34+I34"  # type:ignore
    ws["L35"] = "=F35+I35"  # type:ignore
    ws["L36"] = "=F36+I36"  # type:ignore
    ws["L37"] = "=F37+I37"  # type:ignore
    ws["L38"] = "=F38+I38"  # type:ignore
    ws["L39"] = "=F39+I39"  # type:ignore
    ws["L40"] = "=F40+I40"  # type:ignore
    ws["L41"] = "=F41+I41"  # type:ignore
    ws["L42"] = "=F42+I42"  # type:ignore
    ws["L43"] = "=F43+I43"  # type:ignore
    ws["L44"] = "=F44+I44"  # type:ignore
    ws["L45"] = "=F45+I45"  # type:ignore
    ws["L46"] = "=F46+I46"  # type:ignore
    ws["L47"] = "=L43+L44+L45+L46"  # type:ignore
    ws.delete_rows(48, 49)  # type:ignore
    # Excelファイルを出力
    wb.save("c:/test/AC/AC.xlsx")

    # PC
    # パスで指定したファイルの一覧をリスト形式で取得
    csv_files_pc = glob.glob("c:/test/PC/*.csv")
    # CSVファイルの中身を追加していくリストを表示
    data_list_pc = []
    # 読み込むファイルのリストをスキャン
    for file in csv_files_pc:
        data_list_pc.append(pd.read_csv(file))
        # リストを全て列方向に結合
        df_pc_t = pd.concat(data_list_pc, axis=1, sort=True)
    # columnsパラメータで列名を設定
    feature_pc = ["間接1", "間接2", "間接4", "間接5", "間接6", "直接1", "直接4"]
    df_pc_t.columns = feature_pc
    # 行名の設定
    df_pc_t = df_pc_t.rename(
        {
            0: "在籍者",
            1: "在籍者主幹以下人数",
            2: "実在籍者",
            3: "有休時間",
            4: "有休時間在籍者平均",
            5: "欠勤時間",
            6: "勤務時間",
            7: "遅早時間",
            8: "出勤率",
            9: "実労働時間",
            10: "ズレ時間",
            11: "残業時間",
            12: "残業時間主幹以下平均",
            13: "法定外休出時間",
            14: "法定外休出主幹以下平均",
            15: "法定休出時間",
            16: "法定休出主幹以下平均",
            17: "時間外60時間超",
            18: "代休時間",
            19: "応援時間",
            20: "総労働時間",
            21: "基本給",
            22: "役職手当",
            23: "営業手当",
            24: "地域手当",
            25: "特別手当",
            26: "特別技技手当",
            27: "調整手当",
            28: "別居手当",
            29: "通勤手当",
            30: "小計1",
            31: "残業手当",
            32: "休出手当",
            33: "深夜勤務手当",
            34: "交替時差手当",
            35: "休業手当",
            36: "休業控除",
            37: "代休他",
            38: "欠勤・遅早控除",
            39: "精算分",
            40: "小計2",
            41: "総支給額",
            42: "応援時間額",
            43: "役員振替",
            44: "部門振替",
            45: "合計",
        }
    )

    # 整形---------------------
    # ワークブックの生成
    wb = Workbook()
    # ワークシートの生成
    ws = wb.active
    ws.title = "PC工場"  # type:ignore
    # DataFrameを行単位のデータにする
    rows = dataframe_to_rows(df_pc_t, index=True, header=True)
    # 1セルずつ処理を実行する
    for row_no, row in enumerate(rows, 1):
        for col_no, value in enumerate(row, 1):
            # データを書き込む
            ws.cell(row=row_no, column=col_no, value=value)  # type:ignore
    # 不要な行の削除
    ws.delete_rows(2)  # type:ignore
    # 列の追加
    ws.insert_cols(7)  # type:ignore
    # 表示倍率の設定
    ws.sheet_view.zoomScale = 100  # type:ignore
    # 列幅の設定
    ws.column_dimensions["A"].width = 25  # type:ignore
    ws.column_dimensions["B"].width = 12  # type:ignore
    ws.column_dimensions["C"].width = 12  # type:ignore
    ws.column_dimensions["D"].width = 12  # type:ignore
    ws.column_dimensions["E"].width = 12  # type:ignore
    ws.column_dimensions["F"].width = 12  # type:ignore
    ws.column_dimensions["G"].width = 12  # type:ignore
    ws.column_dimensions["H"].width = 12  # type:ignore
    ws.column_dimensions["I"].width = 12  # type:ignore
    ws.column_dimensions["J"].width = 12  # type:ignore
    ws.column_dimensions["K"].width = 12  # type:ignore
    ws.column_dimensions["L"].width = 12  # type:ignore
    # 表示形式
    format = "#,##0"
    for row in ws["B2:L2"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B4:L9"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00%"
    for row in ws["B10:L10"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B11:L12"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B13:L22"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B23:L23"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B24:L26"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B25:L49"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    # ヘッダー行のスタイル設定
    header = ws[1]  # type:ignore
    for header_cell in header:
        # フォントを設定する
        header_cell.fill = PatternFill(
            patternType="solid", fgColor="008000"  # type:ignore
        )  # type:ignore
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
    # 中央揃え
    for row in ws["B1:J1"]:  # type:ignore
        for cell in row:
            cell.alignment = Alignment(horizontal="centerContinuous")
        ws["L1"].alignment = Alignment(horizontal="centerContinuous")  # type:ignore
    # 指定した行の背景色を黄色にする
    mylist = [21, 44, 45, 46]
    for list in mylist:
        for row in ws.iter_rows():  # type:ignore
            for cell in row:
                if cell.row == list:
                    cell.fill = PatternFill(
                        fgColor="FFFF00",  # type:ignore
                        bgColor="FFFF00",  # type:ignore
                        fill_type="solid",  # type:ignore
                    )
    # 小計1
    ws["B32"] = "=SUM(B23:B31)"  # type:ignore
    ws["C32"] = "=SUM(C23:C31)"  # type:ignore
    ws["D32"] = "=SUM(D23:D31)"  # type:ignore
    ws["E32"] = "=SUM(E23:E31)"  # type:ignore
    ws["F32"] = "=SUM(F23:F31)"  # type:ignore
    ws["G32"] = "=SUM(G23:G31)"  # type:ignore
    ws["H32"] = "=SUM(H23:H31)"  # type:ignore
    ws["I32"] = "=SUM(I23:I31)"  # type:ignore
    ws["J32"] = "=SUM(J23:J31)"  # type:ignore
    # 小計2
    ws["B42"] = "=SUM(B33:B41)"  # type:ignore
    ws["C42"] = "=SUM(C33:C41)"  # type:ignore
    ws["D42"] = "=SUM(D33:D41)"  # type:ignore
    ws["E42"] = "=SUM(E33:E41)"  # type:ignore
    ws["F42"] = "=SUM(F33:F41)"  # type:ignore
    ws["G42"] = "=SUM(G33:G41)"  # type:ignore
    ws["H42"] = "=SUM(H33:H41)"  # type:ignore
    ws["I42"] = "=SUM(I33:I41)"  # type:ignore
    ws["J42"] = "=SUM(J33:J41)"  # type:ignore
    # 総支給額
    ws["B43"] = "=B32 + B42"  # type:ignore
    ws["C43"] = "=C32 + C42"  # type:ignore
    ws["D43"] = "=D32 + D42"  # type:ignore
    ws["E43"] = "=E32 + E42"  # type:ignore
    ws["F43"] = "=F32 + F42"  # type:ignore
    ws["G43"] = "=G32 + G42"  # type:ignore
    ws["H43"] = "=H32 + H42"  # type:ignore
    ws["I43"] = "=I32 + I42"  # type:ignore
    ws["J43"] = "=J32 + J42"  # type:ignore
    # 列合計の埋め込み
    ws["B47"] = "=B43+B44+B45+B46"  # type:ignore
    ws["C47"] = "=C43+C44+C45+C46"  # type:ignore
    ws["D47"] = "=D43+D44+D45+D46"  # type:ignore
    ws["E47"] = "=E43+E44+E45+E46"  # type:ignore
    ws["F47"] = "=F43+F44+F45+F46"  # type:ignore
    ws["G47"] = "=G43+G44+G45+G46"  # type:ignore
    ws["H47"] = "=H43+H44+H45+H46"  # type:ignore
    ws["I47"] = "=I43+I44+I45+I46"  # type:ignore
    ws["J47"] = "=J43+J44+J45+J46"  # type:ignore
    # 間接計列の追加
    ws["G1"] = "【間接計】"  # type:ignore
    # 間接計列への計算式の埋め込み
    ws["G2"] = "=SUM(B2:F2)"  # type:ignore
    ws["G3"] = "=SUM(B3:F3)"  # type:ignore
    ws["G4"] = "=SUM(B4:F4)"  # type:ignore
    ws["G5"] = "=SUM(B5:F5)"  # type:ignore
    ws["G6"] = "=SUM(B6:F6) / 5"  # type:ignore
    ws["G7"] = "=SUM(B7:F7)"  # type:ignore
    ws["G8"] = "=SUM(B8:F8)"  # type:ignore
    ws["G9"] = "=SUM(B9:F9)"  # type:ignore
    ws["G10"] = "=SUM(B10:F10) / 5"  # type:ignore
    ws["G11"] = "=SUM(B11:F11)"  # type:ignore
    ws["G12"] = "=SUM(B12:F12)"  # type:ignore
    ws["G13"] = "=SUM(B13:F13)"  # type:ignore
    ws["G14"] = "=SUM(B14:F14) / 5"  # type:ignore
    ws["G15"] = "=SUM(B15:F15)"  # type:ignore
    ws["G16"] = "=SUM(B16:F16) / 5"  # type:ignore
    ws["G17"] = "=SUM(B17:F17)"  # type:ignore
    ws["G18"] = "=SUM(B18:F18) / 5"  # type:ignore
    ws["G19"] = "=SUM(B19:F19)"  # type:ignore
    ws["G20"] = "=SUM(B20:F20)"  # type:ignore
    ws["G21"] = "=SUM(B21:F21)"  # type:ignore
    ws["G22"] = "=SUM(B22:F22)"  # type:ignore
    ws["G23"] = "=SUM(B23:F23)"  # type:ignore
    ws["G24"] = "=SUM(B24:F24)"  # type:ignore
    ws["G25"] = "=SUM(B25:F25)"  # type:ignore
    ws["G26"] = "=SUM(B26:F26)"  # type:ignore
    ws["G27"] = "=SUM(B27:F27)"  # type:ignore
    ws["G28"] = "=SUM(B28:F28)"  # type:ignore
    ws["G29"] = "=SUM(B29:F29)"  # type:ignore
    ws["G30"] = "=SUM(B30:F30)"  # type:ignore
    ws["G31"] = "=SUM(B31:F31)"  # type:ignore
    ws["G32"] = "=SUM(B32:F32)"  # type:ignore
    ws["G33"] = "=SUM(B33:F33)"  # type:ignore
    ws["G34"] = "=SUM(B34:F34)"  # type:ignore
    ws["G35"] = "=SUM(B35:F35)"  # type:ignore
    ws["G36"] = "=SUM(B36:F36)"  # type:ignore
    ws["G37"] = "=SUM(B37:F37)"  # type:ignore
    ws["G38"] = "=SUM(B38:F38)"  # type:ignore
    ws["G39"] = "=SUM(B39:F39)"  # type:ignore
    ws["G40"] = "=SUM(B40:F40)"  # type:ignore
    ws["G41"] = "=SUM(B41:F41)"  # type:ignore
    ws["G42"] = "=SUM(B42:F42)"  # type:ignore
    ws["G43"] = "=SUM(B43:F43)"  # type:ignore
    ws["G44"] = "=SUM(B44:F44)"  # type:ignore
    ws["G45"] = "=SUM(B45:F45)"  # type:ignore
    ws["G46"] = "=SUM(B46:F46)"  # type:ignore
    ws["G47"] = "=G43+G44+G45+G46"  # type:ignore
    # 直接計列の追加
    fill = PatternFill(patternType="solid", fgColor="008000")  # type:ignore
    ws["J1"].fill = fill  # type:ignore
    ws["J1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    ws["J1"] = "【直接計】"  # type:ignore
    # 直接計列への計算式の埋め込み
    ws["J2"] = "=SUM(H2:I2)"  # type:ignore
    ws["J3"] = "=SUM(H3:I3)"  # type:ignore
    ws["J4"] = "=SUM(H4:I4)"  # type:ignore
    ws["J5"] = "=SUM(H5:I5)"  # type:ignore
    ws["J6"] = "=SUM(H6:I6) / 2"  # type:ignore
    ws["J7"] = "=SUM(H7:I7)"  # type:ignore
    ws["J8"] = "=SUM(H8:I8)"  # type:ignore
    ws["J9"] = "=SUM(H9:I9)"  # type:ignore
    ws["J10"] = "=SUM(H10:I10) / 2"  # type:ignore
    ws["J11"] = "=SUM(H11:I11)"  # type:ignore
    ws["J12"] = "=SUM(H12:I12)"  # type:ignore
    ws["J13"] = "=SUM(H13:I13)"  # type:ignore
    ws["J14"] = "=SUM(H14:I14) / 2"  # type:ignore
    ws["J15"] = "=SUM(H15:I15)"  # type:ignore
    ws["J16"] = "=SUM(H16:I16) / 2"  # type:ignore
    ws["J17"] = "=SUM(H17:I17)"  # type:ignore
    ws["J18"] = "=SUM(H18:I18) / 2"  # type:ignore
    ws["J19"] = "=SUM(H19:I19)"  # type:ignore
    ws["J20"] = "=SUM(H20:I20)"  # type:ignore
    ws["J21"] = "=SUM(H21:I21)"  # type:ignore
    ws["J22"] = "=SUM(H22:I22)"  # type:ignore
    ws["J23"] = "=SUM(H23:I23)"  # type:ignore
    ws["J24"] = "=SUM(H24:I24)"  # type:ignore
    ws["J25"] = "=SUM(H25:I25)"  # type:ignore
    ws["J26"] = "=SUM(H26:I26)"  # type:ignore
    ws["J27"] = "=SUM(H27:I27)"  # type:ignore
    ws["J28"] = "=SUM(H28:I28)"  # type:ignore
    ws["J29"] = "=SUM(H29:I29)"  # type:ignore
    ws["J30"] = "=SUM(H30:I30)"  # type:ignore
    ws["J31"] = "=SUM(H31:I31)"  # type:ignore
    ws["J32"] = "=SUM(H32:I32)"  # type:ignore
    ws["J33"] = "=SUM(H33:I33)"  # type:ignore
    ws["J34"] = "=SUM(H34:I34)"  # type:ignore
    ws["J35"] = "=SUM(H35:I35)"  # type:ignore
    ws["J36"] = "=SUM(H36:I36)"  # type:ignore
    ws["J37"] = "=SUM(H37:I37)"  # type:ignore
    ws["J38"] = "=SUM(H38:I38)"  # type:ignore
    ws["J39"] = "=SUM(H39:I39)"  # type:ignore
    ws["J40"] = "=SUM(H40:I40)"  # type:ignore
    ws["J41"] = "=SUM(H41:I41)"  # type:ignore
    ws["J42"] = "=SUM(H42:I42)"  # type:ignore
    ws["J43"] = "=SUM(H43:I43)"  # type:ignore
    ws["J44"] = "=SUM(H44:I44)"  # type:ignore
    ws["J45"] = "=SUM(H45:I45)"  # type:ignore
    ws["J46"] = "=SUM(H46:I46)"  # type:ignore
    ws["J47"] = "=J43+J44+J45+J46"  # type:ignore
    # ヘッダーの書式
    fill = PatternFill(patternType="solid", fgColor="008000")  # type:ignore
    ws["L1"].fill = fill  # type:ignore
    ws["L1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 合計列の追加とヘッダーの書式
    ws["L1"] = "【合計】"  # type:ignore
    ws["L1"].fill = fill  # type:ignore
    ws["L1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 罫線
    side1 = Side(border_style="thin", color="000000")  # type:ignore
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
    for row in ws["A2:L49"]:  # type:ignore
        for cell in row:
            cell.border = border_aro
    # 行合計の計算式埋め込み
    ws["L2"] = "=G2+J2"  # type:ignore
    ws["L3"] = "=G3+J3"  # type:ignore
    ws["L4"] = "=G4+J4"  # type:ignore
    ws["L5"] = "=G5+J5"  # type:ignore
    ws["L6"] = "=(G6+J6) / 2"  # type:ignore
    ws["L7"] = "=G7+J7"  # type:ignore
    ws["L8"] = "=G8+J8"  # type:ignore
    ws["L9"] = "=G9+J9"  # type:ignore
    ws["L10"] = "=(G10+J10) / 2"  # type:ignore
    ws["L11"] = "=G11+J11"  # type:ignore
    ws["L12"] = "=G12+J12"  # type:ignore
    ws["L13"] = "=G13+J13"  # type:ignore
    ws["L14"] = "=(G14+J14) / 2"  # type:ignore
    ws["L15"] = "=G15+J15"  # type:ignore
    ws["L16"] = "=(G16+J16) / 2"  # type:ignore
    ws["L17"] = "=G17+J17"  # type:ignore
    ws["L18"] = "=(G18+J18) / 2"  # type:ignore
    ws["L19"] = "=G19+J19"  # type:ignore
    ws["L20"] = "=G20+J20"  # type:ignore
    ws["L21"] = "=G21+J21"  # type:ignore
    ws["L22"] = "=G22+J22"  # type:ignore
    ws["L23"] = "=G23+J23"  # type:ignore
    ws["L24"] = "=G24+J24"  # type:ignore
    ws["L25"] = "=G25+J25"  # type:ignore
    ws["L26"] = "=G26+J26"  # type:ignore
    ws["L27"] = "=G27+J27"  # type:ignore
    ws["L28"] = "=G28+J28"  # type:ignore
    ws["L29"] = "=G29+J29"  # type:ignore
    ws["L30"] = "=G30+J30"  # type:ignore
    ws["L31"] = "=G31+J31"  # type:ignore
    ws["L32"] = "=G32+J32"  # type:ignore
    ws["L33"] = "=G33+J33"  # type:ignore
    ws["L34"] = "=G34+J34"  # type:ignore
    ws["L35"] = "=G35+J35"  # type:ignore
    ws["L36"] = "=G36+J36"  # type:ignore
    ws["L37"] = "=G37+J37"  # type:ignore
    ws["L38"] = "=G38+J38"  # type:ignore
    ws["L39"] = "=G39+J39"  # type:ignore
    ws["L40"] = "=G40+J40"  # type:ignore
    ws["L41"] = "=G41+J41"  # type:ignore
    ws["L42"] = "=G42+J42"  # type:ignore
    ws["L43"] = "=G43+J43"  # type:ignore
    ws["L44"] = "=G44+J44"  # type:ignore
    ws["L45"] = "=G45+J45"  # type:ignore
    ws["L46"] = "=G46+J46"  # type:ignore
    ws["L47"] = "=L43+L44+L45+L46"  # type:ignore
    ws.delete_rows(48, 49)  # type:ignore
    # Excelファイルを出力
    wb.save("c:/test/PC/PC.xlsx")

    # 宮城
    # パスで指定したファイルの一覧をリスト形式で取得
    csv_files_miyagi = glob.glob("c:/test/宮城/*.csv")
    # CSVファイルの中身を追加していくリストを表示
    data_list_miyagi = []
    # 読み込むファイルのリストをスキャン
    for file in csv_files_miyagi:
        data_list_miyagi.append(pd.read_csv(file))
    # リストを全て列方向に結合
    df_miyagi_t = pd.concat(data_list_miyagi, axis=1, sort=True)
    # columnsパラメータで列名を設定
    feature_miyagi = ["間接1", "間接2", "間接4", "間接6", "直接1"]
    df_miyagi_t.columns = feature_miyagi
    # 行名の設定
    df_miyagi_t = df_miyagi_t.rename(
        {
            0: "在籍者",
            1: "在籍者主幹以下人数",
            2: "実在籍者",
            3: "有休時間",
            4: "有休時間在籍者平均",
            5: "欠勤時間",
            6: "勤務時間",
            7: "遅早時間",
            8: "出勤率",
            9: "実労働時間",
            10: "ズレ時間",
            11: "残業時間",
            12: "残業時間主幹以下平均",
            13: "法定外休出時間",
            14: "法定外休出主幹以下平均",
            15: "法定休出時間",
            16: "法定休出主幹以下平均",
            17: "時間外60時間超",
            18: "代休時間",
            19: "応援時間",
            20: "総労働時間",
            21: "基本給",
            22: "役職手当",
            23: "営業手当",
            24: "地域手当",
            25: "特別手当",
            26: "特別技技手当",
            27: "調整手当",
            28: "別居手当",
            29: "通勤手当",
            30: "小計1",
            31: "残業手当",
            32: "休出手当",
            33: "深夜勤務手当",
            34: "交替時差手当",
            35: "休業手当",
            36: "休業控除",
            37: "代休他",
            38: "欠勤・遅早控除",
            39: "精算分",
            40: "小計2",
            41: "総支給額",
            42: "応援時間額",
            43: "役員振替",
            44: "部門振替",
            45: "合計",
        }
    )
    # 整形---------------------
    # ワークブックの生成
    wb = Workbook()
    # ワークシートの生成
    ws = wb.active
    ws.title = "宮城工場"  # type:ignore
    # DataFrameを行単位のデータにする
    rows = dataframe_to_rows(df_miyagi_t, index=True, header=True)
    # 1セルずつ処理を実行する
    for row_no, row in enumerate(rows, 1):
        for col_no, value in enumerate(row, 1):
            # データを書き込む
            ws.cell(row=row_no, column=col_no, value=value)  # type:ignore
    # 不要な行の削除
    ws.delete_rows(2)  # type:ignore
    # 列の追加
    ws.insert_cols(6)  # type:ignore
    # 表示倍率の設定
    ws.sheet_view.zoomScale = 100  # type:ignore
    # 列幅の設定
    ws.column_dimensions["A"].width = 25  # type:ignore
    ws.column_dimensions["B"].width = 12  # type:ignore
    ws.column_dimensions["C"].width = 12  # type:ignore
    ws.column_dimensions["D"].width = 12  # type:ignore
    ws.column_dimensions["E"].width = 12  # type:ignore
    ws.column_dimensions["F"].width = 12  # type:ignore
    ws.column_dimensions["G"].width = 12  # type:ignore
    ws.column_dimensions["H"].width = 12  # type:ignore
    ws.column_dimensions["I"].width = 12  # type:ignore
    ws.column_dimensions["J"].width = 12  # type:ignore
    ws.column_dimensions["K"].width = 12  # type:ignore
    ws.column_dimensions["L"].width = 12  # type:ignore
    # 表示形式
    format = "#,##0"
    for row in ws["B2:L2"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B4:L9"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00%"
    for row in ws["B10:L10"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B11:L12"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B13:L22"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B23:L23"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B24:L26"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B25:L49"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    # ヘッダー行のスタイル設定
    header = ws[1]  # type:ignore
    for header_cell in header:
        # フォントを設定する
        header_cell.fill = PatternFill(
            patternType="solid", fgColor="008000"  # type:ignore
        )  # type:ignore
        # 罫線を設定する
        header_cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
    # 中央揃え
    for row in ws["B1:H1"]:  # type:ignore
        for cell in row:
            cell.alignment = Alignment(horizontal="centerContinuous")
    ws["L1"].alignment = Alignment(horizontal="centerContinuous")  # type:ignore
    # 指定した行の背景色を黄色にする
    mylist = [21, 44, 45, 46]
    for list in mylist:
        for row in ws.iter_rows():  # type:ignore
            for cell in row:
                if cell.row == list:
                    cell.fill = PatternFill(
                        fgColor="FFFF00",  # type:ignore
                        bgColor="FFFF00",  # type:ignore
                        fill_type="solid",  # type:ignore
                    )
    # 小計1
    ws["B32"] = "=SUM(B23:B31)"  # type:ignore
    ws["C32"] = "=SUM(C23:C31)"  # type:ignore
    ws["D32"] = "=SUM(D23:D31)"  # type:ignore
    ws["E32"] = "=SUM(E23:E31)"  # type:ignore
    ws["F32"] = "=SUM(F23:F31)"  # type:ignore
    ws["G32"] = "=SUM(G23:G31)"  # type:ignore
    ws["H32"] = "=SUM(H23:H31)"  # type:ignore
    # 小計2
    ws["B42"] = "=SUM(B33:B41)"  # type:ignore
    ws["C42"] = "=SUM(C33:C41)"  # type:ignore
    ws["D42"] = "=SUM(D33:D41)"  # type:ignore
    ws["E42"] = "=SUM(E33:E41)"  # type:ignore
    ws["F42"] = "=SUM(F33:F41)"  # type:ignore
    ws["G42"] = "=SUM(G33:G41)"  # type:ignore
    ws["H42"] = "=SUM(H33:h41)"  # type:ignore
    # 総支給額
    ws["B43"] = "=B32 + B42"  # type:ignore
    ws["C43"] = "=C32 + C42"  # type:ignore
    ws["D43"] = "=D32 + D42"  # type:ignore
    ws["E43"] = "=E32 + E42"  # type:ignore
    ws["F43"] = "=F32 + F42"  # type:ignore
    ws["G43"] = "=G32 + G42"  # type:ignore
    ws["H43"] = "=H32 + H42"  # type:ignore
    # 列合計の埋め込み
    ws["B47"] = "=B43+B44+B45+B46"  # type:ignore
    ws["C47"] = "=C43+C44+C45+C46"  # type:ignore
    ws["D47"] = "=D43+D44+D45+D46"  # type:ignore
    ws["E47"] = "=E43+E44+E45+E46"  # type:ignore
    ws["F47"] = "=F43+F44+F45+F46"  # type:ignore
    ws["G47"] = "=G43+G44+G45+G46"  # type:ignore
    ws["H47"] = "=H43+H44+H45+H46"  # type:ignore
    # 間接計列の追加
    ws["F1"] = "【間接計】"  # type:ignore
    # 間接計列への計算式の埋め込み
    ws["F2"] = "=SUM(B2:E2)"  # type:ignore
    ws["F3"] = "=SUM(B3:E3)"  # type:ignore
    ws["F4"] = "=SUM(B4:E4)"  # type:ignore
    ws["F5"] = "=SUM(B5:E5)"  # type:ignore
    ws["F6"] = "=SUM(B6:E6) / 4"  # type:ignore
    ws["F7"] = "=SUM(B7:E7)"  # type:ignore
    ws["F8"] = "=SUM(B8:E8)"  # type:ignore
    ws["F9"] = "=SUM(B9:E9)"  # type:ignore
    ws["F10"] = "=SUM(B10:E10) / 4"  # type:ignore
    ws["F11"] = "=SUM(B11:E11)"  # type:ignore
    ws["F12"] = "=SUM(B12:E12)"  # type:ignore
    ws["F13"] = "=SUM(B13:E13)"  # type:ignore
    ws["F14"] = "=SUM(B14:E14) / 4"  # type:ignore
    ws["F15"] = "=SUM(B15:E15)"  # type:ignore
    ws["F16"] = "=SUM(B16:E16) / 4"  # type:ignore
    ws["F17"] = "=SUM(B17:E17)"  # type:ignore
    ws["F18"] = "=SUM(B18:E18) / 4"  # type:ignore
    ws["F19"] = "=SUM(B19:E19)"  # type:ignore
    ws["F20"] = "=SUM(B20:E20)"  # type:ignore
    ws["F21"] = "=SUM(B21:E21)"  # type:ignore
    ws["F22"] = "=SUM(B22:E22)"  # type:ignore
    ws["F23"] = "=SUM(B23:E23)"  # type:ignore
    ws["F24"] = "=SUM(B24:E24)"  # type:ignore
    ws["F25"] = "=SUM(B25:E25)"  # type:ignore
    ws["F26"] = "=SUM(B26:E26)"  # type:ignore
    ws["F27"] = "=SUM(B27:E27)"  # type:ignore
    ws["F28"] = "=SUM(B28:E28)"  # type:ignore
    ws["F29"] = "=SUM(B29:E29)"  # type:ignore
    ws["F30"] = "=SUM(B30:E30)"  # type:ignore
    ws["F31"] = "=SUM(B31:E31)"  # type:ignore
    ws["F32"] = "=SUM(B32:E32)"  # type:ignore
    ws["F33"] = "=SUM(B33:E33)"  # type:ignore
    ws["F34"] = "=SUM(B34:E34)"  # type:ignore
    ws["F35"] = "=SUM(B35:E35)"  # type:ignore
    ws["F36"] = "=SUM(B36:E36)"  # type:ignore
    ws["F37"] = "=SUM(B37:E37)"  # type:ignore
    ws["F38"] = "=SUM(B38:E38)"  # type:ignore
    ws["F39"] = "=SUM(B39:E39)"  # type:ignore
    ws["F40"] = "=SUM(B40:E40)"  # type:ignore
    ws["F41"] = "=SUM(B41:E41)"  # type:ignore
    ws["F42"] = "=SUM(B42:E42)"  # type:ignore
    ws["F43"] = "=SUM(B43:E43)"  # type:ignore
    ws["F44"] = "=SUM(B44:E44)"  # type:ignore
    ws["F45"] = "=SUM(B45:E45)"  # type:ignore
    ws["F46"] = "=SUM(B46:E46)"  # type:ignore
    ws["F47"] = "=F43+F44+F45+F46"  # type:ignore
    # 直接計列の追加
    ws["H1"] = "【直接計】"  # type:ignore
    # 直接計列への計算式の埋め込み
    ws["H2"] = "=G2"  # type:ignore
    ws["H3"] = "=G3"  # type:ignore
    ws["H4"] = "=G4"  # type:ignore
    ws["H5"] = "=G5"  # type:ignore
    ws["H6"] = "=G6"  # type:ignore
    ws["H7"] = "=G7"  # type:ignore
    ws["H8"] = "=G8"  # type:ignore
    ws["H9"] = "=G9"  # type:ignore
    ws["H10"] = "=G10"  # type:ignore
    ws["H11"] = "=G11"  # type:ignore
    ws["H12"] = "=G12"  # type:ignore
    ws["H13"] = "=G13"  # type:ignore
    ws["H14"] = "=G14"  # type:ignore
    ws["H15"] = "=G15"  # type:ignore
    ws["H16"] = "=G16"  # type:ignore
    ws["H17"] = "=G17"  # type:ignore
    ws["H18"] = "=G18"  # type:ignore
    ws["H19"] = "=G19"  # type:ignore
    ws["H20"] = "=G20"  # type:ignore
    ws["H21"] = "=G21"  # type:ignore
    ws["H22"] = "=G22"  # type:ignore
    ws["H23"] = "=G23"  # type:ignore
    ws["H24"] = "=G24"  # type:ignore
    ws["H25"] = "=G25"  # type:ignore
    ws["H26"] = "=G26"  # type:ignore
    ws["H27"] = "=G27"  # type:ignore
    ws["H28"] = "=G28"  # type:ignore
    ws["H29"] = "=G29"  # type:ignore
    ws["H30"] = "=G30"  # type:ignore
    ws["H31"] = "=G31"  # type:ignore
    ws["H32"] = "=G32"  # type:ignore
    ws["H33"] = "=G33"  # type:ignore
    ws["H34"] = "=G34"  # type:ignore
    ws["H35"] = "=G35"  # type:ignore
    ws["H36"] = "=G36"  # type:ignore
    ws["H37"] = "=G37"  # type:ignore
    ws["H38"] = "=G38"  # type:ignore
    ws["H39"] = "=G39"  # type:ignore
    ws["H40"] = "=G40"  # type:ignore
    ws["H41"] = "=G41"  # type:ignore
    ws["H42"] = "=G42"  # type:ignore
    ws["H43"] = "=G43"  # type:ignore
    ws["H44"] = "=G44"  # type:ignore
    ws["H45"] = "=G45"  # type:ignore
    ws["H46"] = "=G46"  # type:ignore
    ws["H47"] = "=H43+H44+H45+H46"  # type:ignore
    # 直接列ヘッダーの書式
    fill = PatternFill(patternType="solid", fgColor="008000")  # type:ignore
    # 合計列の追加とヘッダーの書式
    ws["L1"] = "【合計】"  # type:ignore
    ws["L1"].fill = fill  # type:ignore
    ws["L1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 罫線
    side1 = Side(border_style="thin", color="000000")
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
    for row in ws["A2:L49"]:  # type:ignore
        for cell in row:
            cell.border = border_aro
    # 行合計の埋め込み
    ws["L2"] = "=SUM(F2,H2)"  # type:ignore
    ws["L3"] = "=SUM(F3,H3)"  # type:ignore
    ws["L4"] = "=SUM(F4,H4)"  # type:ignore
    ws["L5"] = "=SUM(F5,H5)"  # type:ignore
    ws["L6"] = "=SUM(F6,H6) / 2"  # type:ignore
    ws["L7"] = "=SUM(F7,H7)"  # type:ignore
    ws["L8"] = "=SUM(F8,H8)"  # type:ignore
    ws["L9"] = "=SUM(F9,H9)"  # type:ignore
    ws["L10"] = "=SUM(F10,H10) / 2"  # type:ignore
    ws["L11"] = "=SUM(F11,H11)"  # type:ignore
    ws["L12"] = "=SUM(F12,H12)"  # type:ignore
    ws["L13"] = "=SUM(F13,H13)"  # type:ignore
    ws["L14"] = "=SUM(F14,H14) / 2"  # type:ignore
    ws["L15"] = "=SUM(F15,H15)"  # type:ignore
    ws["L16"] = "=SUM(F16,H16) / 2"  # type:ignore
    ws["L17"] = "=SUM(F17,H17)"  # type:ignore
    ws["L18"] = "=SUM(F18,H18) / 2"  # type:ignore
    ws["L19"] = "=SUM(F19,H19)"  # type:ignore
    ws["L20"] = "=SUM(F20,H20)"  # type:ignore
    ws["L21"] = "=SUM(F21,H21)"  # type:ignore
    ws["L22"] = "=SUM(F22,H22)"  # type:ignore
    ws["L23"] = "=SUM(F23,H23)"  # type:ignore
    ws["L24"] = "=SUM(F24,H24)"  # type:ignore
    ws["L25"] = "=SUM(F25,H25)"  # type:ignore
    ws["L26"] = "=SUM(F26,H26)"  # type:ignore
    ws["L27"] = "=SUM(F27,H27)"  # type:ignore
    ws["L28"] = "=SUM(F28,H28)"  # type:ignore
    ws["L29"] = "=SUM(F29,H29)"  # type:ignore
    ws["L30"] = "=SUM(F30,H30)"  # type:ignore
    ws["L31"] = "=SUM(F31,H31)"  # type:ignore
    ws["L32"] = "=SUM(F32,H32)"  # type:ignore
    ws["L33"] = "=SUM(F33,H33)"  # type:ignore
    ws["L34"] = "=SUM(F34,H34)"  # type:ignore
    ws["L35"] = "=SUM(F35,H35)"  # type:ignore
    ws["L36"] = "=SUM(F36,H36)"  # type:ignore
    ws["L37"] = "=SUM(F37,H37)"  # type:ignore
    ws["L38"] = "=SUM(F38,H38)"  # type:ignore
    ws["L39"] = "=SUM(F39,H39)"  # type:ignore
    ws["L40"] = "=SUM(F40,H40)"  # type:ignore
    ws["L41"] = "=SUM(F41,H41)"  # type:ignore
    ws["L42"] = "=SUM(F42,H42)"  # type:ignore
    ws["L43"] = "=SUM(F43,H43)"  # type:ignore
    ws["L44"] = "=SUM(F44,H44)"  # type:ignore
    ws["L45"] = "=SUM(F45,H45)"  # type:ignore
    ws["L46"] = "=SUM(F46,H46)"  # type:ignore
    ws["L47"] = "=L43+L44+L45+L46"  # type:ignore
    ws.delete_rows(48, 49)  # type:ignore
    # Excelファイルを出力
    wb.save("c:/test/宮城/宮城.xlsx")

    # 住設
    # パスで指定したファイルの一覧をリスト形式で取得
    csv_files_jyusetu = glob.glob("c:/test/住設/*.csv")
    # CSVファイルの中身を追加していくリストを表示
    data_list_jyusetu = []
    # 読み込むファイルのリストをスキャン
    for file in csv_files_jyusetu:
        data_list_jyusetu.append(pd.read_csv(file))
    # リストを全て列方向に結合
    df_jyusetu_t = pd.concat(data_list_jyusetu, axis=1, sort=True)
    # columnsパラメータで列名を設定
    feature_jyusetu = ["間接2", "間接4", "間接6"]
    df_jyusetu_t.columns = feature_jyusetu
    # 行名の設定
    df_jyusetu_t = df_jyusetu_t.rename(
        {
            0: "在籍者",
            1: "在籍者主幹以下人数",
            2: "実在籍者",
            3: "有休時間",
            4: "有休時間在籍者平均",
            5: "欠勤時間",
            6: "勤務時間",
            7: "遅早時間",
            8: "出勤率",
            9: "実労働時間",
            10: "ズレ時間",
            11: "残業時間",
            12: "残業時間主幹以下平均",
            13: "法定外休出時間",
            14: "法定外休出主幹以下平均",
            15: "法定休出時間",
            16: "法定休出主幹以下平均",
            17: "時間外60時間超",
            18: "代休時間",
            19: "応援時間",
            20: "総労働時間",
            21: "基本給",
            22: "役職手当",
            23: "営業手当",
            24: "地域手当",
            25: "特別手当",
            26: "特別技技手当",
            27: "調整手当",
            28: "別居手当",
            29: "通勤手当",
            30: "小計1",
            31: "残業手当",
            32: "休出手当",
            33: "深夜勤務手当",
            34: "交替時差手当",
            35: "休業手当",
            36: "休業控除",
            37: "代休他",
            38: "欠勤・遅早控除",
            39: "精算分",
            40: "小計2",
            41: "総支給額",
            42: "応援時間額",
            43: "役員振替",
            44: "部門振替",
            45: "合計",
        }
    )

    # 整形---------------------
    # ワークブックの生成
    wb = Workbook()
    # ワークシートの生成
    ws = wb.active
    ws.title = "住設"  # type:ignore
    # DataFrameを行単位のデータにする
    rows = dataframe_to_rows(df_jyusetu_t, index=True, header=True)
    # 1セルずつ処理を実行する
    for row_no, row in enumerate(rows, 1):
        for col_no, value in enumerate(row, 1):
            # データを書き込む
            ws.cell(row=row_no, column=col_no, value=value)  # type:ignore

    # 不要な行の削除
    ws.delete_rows(2)  # type:ignore
    # 表示倍率の設定
    ws.sheet_view.zoomScale = 100  # type:ignore
    # 列幅の設定
    ws.column_dimensions["A"].width = 25  # type:ignore
    ws.column_dimensions["B"].width = 12  # type:ignore
    ws.column_dimensions["C"].width = 12  # type:ignore
    ws.column_dimensions["D"].width = 12  # type:ignore
    ws.column_dimensions["E"].width = 12  # type:ignore
    ws.column_dimensions["F"].width = 12  # type:ignore
    ws.column_dimensions["G"].width = 12  # type:ignore
    ws.column_dimensions["H"].width = 12  # type:ignore
    ws.column_dimensions["I"].width = 12  # type:ignore
    ws.column_dimensions["J"].width = 12  # type:ignore
    ws.column_dimensions["K"].width = 12  # type:ignore
    ws.column_dimensions["L"].width = 12  # type:ignore
    # 表示形式
    format = "#,##0"
    for row in ws["B2:L2"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B4:L9"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00%"
    for row in ws["B10:L10"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B11:L12"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "0.00"
    for row in ws["B13:L22"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B23:L23"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B24:L26"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    format = "#,##0"
    for row in ws["B25:L49"]:  # type:ignore
        for cell in row:
            cell.number_format = format
    # ヘッダー行のスタイル設定
    header = ws[1]  # type:ignore
    for header_cell in header:
        # フォントを設定する
        header_cell.fill = PatternFill(
            patternType="solid", fgColor="008000"  # type:ignore
        )  # type:ignore
    # 罫線を設定する
    header_cell.border = Border(
        top=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    # 背景色を設定する
    header_cell.font = Font(bold=True, color="FFFFFF")
    # 中央揃え
    for row in ws["B1:E1"]:  # type:ignore
        for cell in row:
            cell.alignment = Alignment(horizontal="centerContinuous")
    ws["L1"].alignment = Alignment(horizontal="centerContinuous")  # type:ignore
    # 指定した行の背景色を黄色にする
    mylist = [21, 44, 45, 46]
    for list in mylist:
        for row in ws.iter_rows():  # type:ignore
            for cell in row:
                if cell.row == list:
                    cell.fill = PatternFill(
                        fgColor="FFFF00",  # type:ignore
                        bgColor="FFFF00",  # type:ignore
                        fill_type="solid",  # type:ignore
                    )
    # 小計1
    ws["B32"] = "=SUM(B23:B31)"  # type:ignore
    ws["C32"] = "=SUM(C23:C31)"  # type:ignore
    ws["D32"] = "=SUM(D23:D31)"  # type:ignore
    ws["E32"] = "=SUM(E23:E31)"  # type:ignore
    # 小計2
    ws["B42"] = "=SUM(B33:B41)"  # type:ignore
    ws["C42"] = "=SUM(C33:C41)"  # type:ignore
    ws["D42"] = "=SUM(D33:D41)"  # type:ignore
    ws["E42"] = "=SUM(E33:E41)"  # type:ignore
    # 総支給額
    ws["B43"] = "=B32 + B42"  # type:ignore
    ws["C43"] = "=C32 + C42"  # type:ignore
    ws["D43"] = "=D32 + D42"  # type:ignore
    ws["E43"] = "=E32 + E42"  # type:ignore
    # 列合計の埋め込み
    ws["B47"] = "=B43+B44+B45+B46"  # type:ignore
    ws["C47"] = "=C43+C44+C45+C46"  # type:ignore
    ws["D47"] = "=D43+D44+D45+D46"  # type:ignore
    ws["E47"] = "=E43+E44+E45+E46"  # type:ignore
    # 間接計列の追加
    ws["E1"] = "【間接計】"  # type:ignore
    # 間接計列への計算式の埋め込み
    ws["E2"] = "=SUM(B2:D2)"  # type:ignore
    ws["E3"] = "=SUM(B3:D3)"  # type:ignore
    ws["E4"] = "=SUM(B4:D4)"  # type:ignore
    ws["E5"] = "=SUM(B5:D5)"  # type:ignore
    ws["E6"] = "=SUM(B6:D6) / 3"  # type:ignore
    ws["E7"] = "=SUM(B7:D7)"  # type:ignore
    ws["E8"] = "=SUM(B8:D8)"  # type:ignore
    ws["E9"] = "=SUM(B9:D9)"  # type:ignore
    ws["E10"] = "=SUM(B10:D10) / 3"  # type:ignore
    ws["E11"] = "=SUM(B11:D11)"  # type:ignore
    ws["E12"] = "=SUM(B12:D12)"  # type:ignore
    ws["E13"] = "=SUM(B13:D13)"  # type:ignore
    ws["E14"] = "=SUM(B14:D14) / 3"  # type:ignore
    ws["E15"] = "=SUM(B15:D15)"  # type:ignore
    ws["E16"] = "=SUM(B16:D16) / 3"  # type:ignore
    ws["E17"] = "=SUM(B17:D17)"  # type:ignore
    ws["E18"] = "=SUM(B18:D18) / 3"  # type:ignore
    ws["E19"] = "=SUM(B19:D19)"  # type:ignore
    ws["E20"] = "=SUM(B20:D20)"  # type:ignore
    ws["E21"] = "=SUM(B21:D21)"  # type:ignore
    ws["E22"] = "=SUM(B22:D22)"  # type:ignore
    ws["E23"] = "=SUM(B23:D23)"  # type:ignore
    ws["E24"] = "=SUM(B24:D24)"  # type:ignore
    ws["E25"] = "=SUM(B25:D25)"  # type:ignore
    ws["E26"] = "=SUM(B26:D26)"  # type:ignore
    ws["E27"] = "=SUM(B27:D27)"  # type:ignore
    ws["E28"] = "=SUM(B28:D28)"  # type:ignore
    ws["E29"] = "=SUM(B29:D29)"  # type:ignore
    ws["E30"] = "=SUM(B30:D30)"  # type:ignore
    ws["E31"] = "=SUM(B31:D31)"  # type:ignore
    ws["E32"] = "=SUM(B32:D32)"  # type:ignore
    ws["E33"] = "=SUM(B33:D33)"  # type:ignore
    ws["E34"] = "=SUM(B34:D34)"  # type:ignore
    ws["E35"] = "=SUM(B35:D35)"  # type:ignore
    ws["E36"] = "=SUM(B36:D36)"  # type:ignore
    ws["E37"] = "=SUM(B37:D37)"  # type:ignore
    ws["E38"] = "=SUM(B38:D38)"  # type:ignore
    ws["E39"] = "=SUM(B39:D39)"  # type:ignore
    ws["E40"] = "=SUM(B40:D40)"  # type:ignore
    ws["E41"] = "=SUM(B41:D41)"  # type:ignore
    ws["E42"] = "=SUM(B42:D42)"  # type:ignore
    ws["E43"] = "=SUM(B43:D43)"  # type:ignore
    ws["E44"] = "=SUM(B44:D44)"  # type:ignore
    ws["E45"] = "=SUM(B45:D45)"  # type:ignore
    ws["E46"] = "=SUM(B46:D46)"  # type:ignore
    ws["E47"] = "=E43+E44+E45+E46"  # type:ignore
    # ヘッダーの書式
    fill = PatternFill(patternType="solid", fgColor="008000")  # type:ignore
    # 合計列の追加とヘッダーの書式
    ws["L1"] = "【合計】"  # type:ignore
    ws["L1"].fill = fill  # type:ignore
    ws["L1"].font = Font(bold=True, color="FFFFFF")  # type:ignore
    # 罫線
    side1 = Side(border_style="thin", color="000000")
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
    for row in ws["A2:L49"]:  # type:ignore
        for cell in row:
            cell.border = border_aro
    # 行合計の埋め込み
    ws["L2"] = "=E2"  # type:ignore
    ws["L3"] = "=E3"  # type:ignore
    ws["L4"] = "=E4"  # type:ignore
    ws["L5"] = "=E5"  # type:ignore
    ws["L6"] = "=E6"  # type:ignore
    ws["L7"] = "=E7"  # type:ignore
    ws["L8"] = "=E8"  # type:ignore
    ws["L9"] = "=E9"  # type:ignore
    ws["L10"] = "=E10"  # type:ignore
    ws["L11"] = "=E11"  # type:ignore
    ws["L12"] = "=E12"  # type:ignore
    ws["L13"] = "=E13"  # type:ignore
    ws["L14"] = "=E14"  # type:ignore
    ws["L15"] = "=E15"  # type:ignore
    ws["L16"] = "=E16"  # type:ignore
    ws["L17"] = "=E17"  # type:ignore
    ws["L18"] = "=E18"  # type:ignore
    ws["L19"] = "=E19"  # type:ignore
    ws["L20"] = "=E20"  # type:ignore
    ws["L21"] = "=E21"  # type:ignore
    ws["L22"] = "=E22"  # type:ignore
    ws["L23"] = "=E23"  # type:ignore
    ws["L24"] = "=E24"  # type:ignore
    ws["L25"] = "=E25"  # type:ignore
    ws["L26"] = "=E26"  # type:ignore
    ws["L27"] = "=E27"  # type:ignore
    ws["L28"] = "=E28"  # type:ignore
    ws["L29"] = "=E29"  # type:ignore
    ws["L30"] = "=E30"  # type:ignore
    ws["L31"] = "=E31"  # type:ignore
    ws["L32"] = "=E32"  # type:ignore
    ws["L33"] = "=E33"  # type:ignore
    ws["L34"] = "=E34"  # type:ignore
    ws["L35"] = "=E35"  # type:ignore
    ws["L36"] = "=E36"  # type:ignore
    ws["L37"] = "=E37"  # type:ignore
    ws["L38"] = "=E38"  # type:ignore
    ws["L39"] = "=E39"  # type:ignore
    ws["L40"] = "=E40"  # type:ignore
    ws["L41"] = "=E41"  # type:ignore
    ws["L42"] = "=E42"  # type:ignore
    ws["L43"] = "=E43"  # type:ignore
    ws["L44"] = "=E44"  # type:ignore
    ws["L45"] = "=E45"  # type:ignore
    ws["L46"] = "=E46"  # type:ignore
    ws["L47"] = "=L43+L44+L45+L46"  # type:ignore

    ws.delete_rows(48, 49)  # type:ignore
    # Excelファイルを出力
    wb.save("c:/test/住設/住設.xlsx")
